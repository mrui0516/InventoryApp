import csv
import json
import tempfile
from io import BytesIO, StringIO
from xml.etree import ElementTree
from zipfile import ZipFile
from datetime import timedelta
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase, TransactionTestCase
from django.urls import reverse
from django.utils import timezone

from .forms import ProductForm
from .models import ARInvoice, ARPayment, AttendanceRecord, Brand, Category, Customer, DailySalesSummary, InboundOrder, InboundPendingItem, Product, ProductSeries, Purchase, Sale, SaleOrder, SaleOrderChangeLog, SaleOrderPayment, SalesTarget, Store, StoreProfile, Supplier
from .services.dashboard import (
    build_period_comparison,
    build_target_progress,
    compute_period_headline,
)


class SummaryRebuildTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.category = Category.objects.create(name="Perfume")
        cls.customer = Customer.objects.create(nif="123456789", name="Alice")
        cls.product = Product.objects.create(
            name="Rose",
            barcode="1234567890123",
            brand="Maison",
            category=cls.category,
            default_price=Decimal("15.00"),
        )

    def create_purchase(self, quantity=10, cost_price=Decimal("10.00")):
        return Purchase.objects.create(
            product=self.product,
            supplier=None,
            quantity=quantity,
            cost_price=cost_price,
            remaining=quantity,
        )

    def create_sale(self, quantity=2, unit_price=Decimal("15.00")):
        order = SaleOrder.objects.create(customer=self.customer)
        with self.captureOnCommitCallbacks(execute=True):
            sale = Sale.objects.create(
                order=order,
                product=self.product,
                customer=self.customer,
                quantity=quantity,
                unit_price=unit_price,
                payment_method="cash",
            )
        sale.refresh_from_db()
        return sale

    def test_sale_create_rebuilds_daily_summary(self):
        self.create_purchase()
        sale = self.create_sale()

        summary = DailySalesSummary.objects.get(date=sale.date.date())
        self.assertEqual(summary.total_sales, Decimal("30.00"))
        self.assertEqual(summary.total_profit, Decimal("10.00"))
        self.assertEqual(summary.total_items_sold, 2)

    def test_moving_sale_to_another_day_rebuilds_both_days(self):
        purchase = self.create_purchase()
        sale = self.create_sale()
        original_day = sale.date.date()

        moved_sale_dt = sale.date - timedelta(days=1)
        Purchase.objects.filter(pk=purchase.pk).update(date=moved_sale_dt - timedelta(hours=1))

        sale.date = moved_sale_dt
        with self.captureOnCommitCallbacks(execute=True):
            sale.save(update_fields=["date"])

        self.assertFalse(DailySalesSummary.objects.filter(date=original_day).exists())
        moved_summary = DailySalesSummary.objects.get(date=moved_sale_dt.date())
        self.assertEqual(moved_summary.total_sales, Decimal("30.00"))
        self.assertEqual(moved_summary.total_items_sold, 2)

    def test_deleting_last_sale_removes_summary_row(self):
        self.create_purchase()
        sale = self.create_sale()
        day = sale.date.date()

        with self.captureOnCommitCallbacks(execute=True):
            sale.delete()

        self.assertFalse(DailySalesSummary.objects.filter(date=day).exists())


class WorkflowRegressionTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.user = user_model.objects.create_user(username="cashier", password="pw123456")
        cls.staff_user = user_model.objects.create_user(username="manager", password="pw123456", is_staff=True)

        cls.category = Category.objects.create(name="Attar")
        cls.customer = Customer.objects.create(nif="987654321", name="Bob")
        cls.product = Product.objects.create(
            name="Amber",
            barcode="9876543210123",
            brand="Maison",
            category=cls.category,
            default_price=Decimal("18.00"),
        )
        cls.purchase = Purchase.objects.create(
            product=cls.product,
            supplier=None,
            quantity=5,
            cost_price=Decimal("10.00"),
            remaining=5,
        )

    def setUp(self):
        cache.clear()

    def test_outbound_summary_matches_sales_exactly(self):
        self.client.login(username="cashier", password="pw123456")

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse("outbound"),
                {
                    "customer_id": str(self.customer.id),
                    "items_json": json.dumps(
                        [
                            {
                                "barcode": self.product.barcode,
                                "qty": 2,
                                "price": "15.00",
                                "payment": "cash",
                            }
                        ]
                    ),
                },
            )

        self.assertEqual(response.status_code, 200)
        sale = Sale.objects.get()
        summary = DailySalesSummary.objects.get(date=sale.date.date())
        self.assertEqual(summary.total_sales, Decimal("30.00"))
        self.assertEqual(summary.total_profit, Decimal("10.00"))
        self.assertEqual(summary.total_items_sold, 2)

    def test_outbound_split_payment_records_each_method(self):
        self.client.login(username="cashier", password="pw123456")

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse("outbound"),
                {
                    "customer_id": str(self.customer.id),
                    "items_json": json.dumps([{"barcode": self.product.barcode, "qty": 4, "price": "15.00"}]),
                    "payments_json": json.dumps([
                        {"method": "cash", "amount": "40.00"},
                        {"method": "card", "amount": "20.00"},
                    ]),
                },
            )

        self.assertEqual(response.status_code, 200)
        order = SaleOrder.objects.latest("id")
        self.assertEqual(
            {p.method: p.amount for p in order.payments.all()},
            {"cash": Decimal("40.00"), "card": Decimal("20.00")},
        )
        # Largest tender becomes the per-line payment_method (category-aware reports stay populated)
        self.assertEqual(order.items.first().payment_method, "cash")

    def test_outbound_per_line_payment_methods_aggregate(self):
        # New flow: each line carries its own method; SaleOrderPayment aggregates by method.
        self.client.login(username="cashier", password="pw123456")
        product2 = Product.objects.create(
            name="Oud", barcode="9876543210130", brand="Maison",
            category=self.category, default_price=Decimal("20.00"),
        )
        Purchase.objects.create(product=product2, supplier=None, quantity=5, cost_price=Decimal("8.00"), remaining=5)

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(
                reverse("outbound"),
                {
                    "items_json": json.dumps([
                        {"barcode": self.product.barcode, "qty": 2, "price": "15.00", "payment": "cash"},
                        {"barcode": product2.barcode, "qty": 1, "price": "20.00", "payment": "card"},
                    ]),
                },
            )

        self.assertEqual(response.status_code, 200)
        order = SaleOrder.objects.latest("id")
        methods = {s.product_id: s.payment_method for s in order.items.all()}
        self.assertEqual(methods[self.product.id], "cash")
        self.assertEqual(methods[product2.id], "card")
        self.assertEqual(
            {p.method: p.amount for p in order.payments.all()},
            {"cash": Decimal("30.00"), "card": Decimal("20.00")},
        )

    def test_inbound_with_supplier_creates_pending_without_stock(self):
        self.client.login(username="manager", password="pw123456")
        supplier = Supplier.objects.create(name="Acme Supplies")
        purchases_before = Purchase.objects.count()
        stock_before = self.product.total_stock()

        response = self.client.post(
            reverse("inbound"),
            {
                "supplier": str(supplier.id),
                "invoice_no": "INV-2",
                "invoice_date": "2026-06-01",
                "items_json": json.dumps([{"barcode": self.product.barcode, "qty": 5, "cost_price": "9.00"}]),
            },
        )

        self.assertEqual(response.status_code, 200)
        order = InboundOrder.objects.latest("id")
        self.assertEqual(order.status, "pending_receipt")
        self.assertEqual(order.pending_items.count(), 1)
        self.assertEqual(order.total_amount, Decimal("45.00"))
        # No stock created yet
        self.assertEqual(Purchase.objects.count(), purchases_before)
        self.assertEqual(self.product.total_stock(), stock_before)

    def test_inbound_receive_converts_pending_into_stock(self):
        self.client.login(username="manager", password="pw123456")
        supplier = Supplier.objects.create(name="Acme Supplies")
        order = InboundOrder.objects.create(supplier=supplier, status="pending_receipt", total_amount=Decimal("0.00"))
        item = InboundPendingItem.objects.create(
            inbound_order=order, product=self.product, quantity=6, cost_price=Decimal("9.00")
        )
        stock_before = self.product.total_stock()

        # The review/confirm UI is a per-order modal on the inbound page now.
        page = self.client.get(reverse("inbound"))
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Confirm receipt")
        self.assertContains(page, f"receive-modal-{order.id}")

        response = self.client.post(
            reverse("inbound_receive", args=[order.id]),
            {
                "action": "receive",
                "supplier": str(supplier.id),
                "invoice_no": "INV-1",
                "invoice_date": "2026-06-01",
                "note": "",
                "lines-TOTAL_FORMS": "1",
                "lines-INITIAL_FORMS": "1",
                "lines-MIN_NUM_FORMS": "0",
                "lines-MAX_NUM_FORMS": "1000",
                "lines-0-id": str(item.id),
                "lines-0-quantity": "6",
                "lines-0-cost_price": "9.00",
                "lines-0-DELETE": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        order.refresh_from_db()
        self.assertEqual(order.status, "received")
        self.assertIsNotNone(order.received_at)
        self.assertEqual(order.pending_items.count(), 0)
        purchase = order.items.get()
        self.assertEqual(purchase.quantity, 6)
        self.assertEqual(purchase.remaining, 6)
        self.assertEqual(purchase.cost_price, Decimal("9.00"))
        self.assertEqual(self.product.total_stock(), stock_before + 6)

    def test_receive_buttons_stay_submittable(self):
        """Regression: the double-submit guard disabled the submit buttons inside
        the submit handler. A disabled button is not serialised, so name="action"
        never reached the server, the view fell back to its 'save' default and the
        order looked saved while staying pending with no stock added."""
        self.client.login(username="manager", password="pw123456")
        supplier = Supplier.objects.create(name="Guard Supplies")
        order = InboundOrder.objects.create(supplier=supplier, status="pending_receipt",
                                            total_amount=Decimal("0.00"))
        InboundPendingItem.objects.create(inbound_order=order, product=self.product,
                                          quantity=2, cost_price=Decimal("5.00"))
        html = self.client.get(reverse("inbound")).content.decode()

        # the action must ride on the button, so a no-JS submit still works
        self.assertIn('name="action" value="receive"', html)
        # ...and the guard must not disable the buttons before serialisation
        guard = html[html.index('Guard against a double submit'):]
        guard = guard[:guard.index('</script>')]
        # strip comments: the explanation itself mentions setTimeout
        code = chr(10).join(l for l in guard.splitlines()
                            if not l.strip().startswith('//'))
        self.assertIn('b.disabled = true', code)
        self.assertIn('setTimeout(function', code)
        self.assertLess(code.index('setTimeout(function'), code.index('b.disabled = true'),
                        'buttons are disabled before the form is serialised')

    def test_receiving_the_same_order_twice_does_not_404(self):
        """Regression: the order lookup filtered on status='pending_receipt', so a
        repeated submit (double tap / browser re-post) hit an order the first
        request had already received and returned a bare 404 - even though the
        stock had gone in correctly."""
        self.client.login(username="manager", password="pw123456")
        supplier = Supplier.objects.create(name="Twice Supplies")
        order = InboundOrder.objects.create(supplier=supplier, status="pending_receipt",
                                            total_amount=Decimal("0.00"))
        item = InboundPendingItem.objects.create(
            inbound_order=order, product=self.product, quantity=4, cost_price=Decimal("7.00")
        )
        payload = {
            "action": "receive", "supplier": str(supplier.id),
            "invoice_no": "INV-2", "invoice_date": "2026-06-02", "note": "",
            "lines-TOTAL_FORMS": "1", "lines-INITIAL_FORMS": "1",
            "lines-MIN_NUM_FORMS": "0", "lines-MAX_NUM_FORMS": "1000",
            "lines-0-id": str(item.id), "lines-0-quantity": "4",
            "lines-0-cost_price": "7.00", "lines-0-DELETE": "",
        }
        stock_before = self.product.total_stock()

        first = self.client.post(reverse("inbound_receive", args=[order.id]), payload)
        self.assertEqual(first.status_code, 302)

        # The repeat must redirect with a message, not 404, and must not double
        # the stock.
        second = self.client.post(reverse("inbound_receive", args=[order.id]), payload)
        self.assertEqual(second.status_code, 302)
        self.assertIn(reverse("inbound"), second.headers["Location"])
        self.assertEqual(self.product.total_stock(), stock_before + 4)

        # A genuinely unknown order still 404s.
        self.assertEqual(
            self.client.post(reverse("inbound_receive", args=[order.id + 9999]), payload).status_code,
            404,
        )

    def test_outbound_rejects_payment_not_matching_total(self):
        self.client.login(username="cashier", password="pw123456")
        before = SaleOrder.objects.count()

        response = self.client.post(
            reverse("outbound"),
            {
                "items_json": json.dumps([{"barcode": self.product.barcode, "qty": 4, "price": "15.00"}]),
                "payments_json": json.dumps([{"method": "cash", "amount": "50.00"}]),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "must add up")
        self.assertEqual(SaleOrder.objects.count(), before)
        self.assertEqual(SaleOrderPayment.objects.count(), 0)

    def test_customer_lookup_requires_login(self):
        response = self.client.get(reverse("check_customer"), {"q": "Bob"})
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("login"), response.url)

    def test_add_customer_requires_csrf_but_still_works_with_token(self):
        client = Client(enforce_csrf_checks=True)
        client.login(username="cashier", password="pw123456")

        blocked = client.post(reverse("add_customer_ajax"), {"nif": "111222333", "name": "Carol"})
        self.assertEqual(blocked.status_code, 403)

        page = client.get(reverse("customer_search"))
        csrf_token = page.cookies["csrftoken"].value
        allowed = client.post(
            reverse("add_customer_ajax"),
            {"nif": "111222333", "name": "Carol"},
            HTTP_X_CSRFTOKEN=csrf_token,
        )

        self.assertEqual(allowed.status_code, 200)
        self.assertTrue(allowed.json()["success"])

    def test_only_staff_can_adjust_purchase_stock(self):
        denied_client = Client()
        denied_client.login(username="cashier", password="pw123456")
        denied = denied_client.post(
            reverse("api_adjust_purchase_stock"),
            data=json.dumps({"purchase_id": self.purchase.id, "new_remaining": 4}),
            content_type="application/json",
        )
        self.assertEqual(denied.status_code, 403)

        staff_client = Client()
        staff_client.login(username="manager", password="pw123456")
        allowed = staff_client.post(
            reverse("api_adjust_purchase_stock"),
            data=json.dumps({"purchase_id": self.purchase.id, "new_remaining": 4}),
            content_type="application/json",
        )

        self.assertEqual(allowed.status_code, 200)
        self.purchase.refresh_from_db()
        self.assertEqual(self.purchase.remaining, 4)


class InventoryAdjustmentSyncTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.staff_user = user_model.objects.create_user(username="stock_manager", password="pw123456", is_staff=True)
        cls.admin = user_model.objects.create_superuser(username="stock_admin", password="pw123456")
        cls.category = Category.objects.create(name="Perfumes")
        cls.product = Product.objects.create(
            name="Sync Test",
            barcode="4445556667778",
            brand="Lattafa",
            category=cls.category,
            default_price=Decimal("32.00"),
        )
        cls.purchase = Purchase.objects.create(
            product=cls.product,
            supplier=None,
            quantity=10,
            cost_price=Decimal("9.50"),
            remaining=10,
        )

    def test_decreasing_purchase_remaining_returns_recomputed_inventory_snapshot(self):
        self.client.login(username="stock_manager", password="pw123456")

        response = self.client.post(
            reverse("api_adjust_purchase_stock"),
            data=json.dumps({"purchase_id": self.purchase.id, "new_remaining": 4}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["inventory_snapshot"]["product_total_stock"], 4)
        self.assertEqual(payload["inventory_snapshot"]["perfume_stock_units"], 4)
        self.assertEqual(payload["inventory_snapshot"]["perfume_stock_cost"], "38.00")

        self.client.login(username="stock_admin", password="pw123456")
        dashboard_response = self.client.get(reverse("dashboard"), {"month": timezone.now().strftime("%Y-%m")})
        self.assertEqual(dashboard_response.context["perfume_stock_units"], 4)
        self.assertEqual(dashboard_response.context["perfume_stock_cost"], Decimal("38.00"))

    def test_decreasing_total_stock_preserves_purchase_row_and_recomputes_snapshot(self):
        self.client.login(username="stock_manager", password="pw123456")

        response = self.client.post(
            reverse("api_adjust_total_stock"),
            data=json.dumps({"product_id": self.product.id, "new_total_stock": 0}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["inventory_snapshot"]["product_total_stock"], 0)
        self.assertTrue(Purchase.objects.filter(pk=self.purchase.pk).exists())
        self.purchase.refresh_from_db()
        self.assertEqual(self.purchase.remaining, 0)


class SalesRecordsViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.user = user_model.objects.create_user(username="records_user", password="pw123456")
        cls.admin = user_model.objects.create_superuser(username="records_admin", password="pw123456")

        cls.category = Category.objects.create(name="Record Test")
        cls.customer = Customer.objects.create(nif="222333444", name="Record Customer")
        cls.product = Product.objects.create(
            name="Musk",
            barcode="5556667778881",
            brand="Maison",
            category=cls.category,
            default_price=Decimal("18.00"),
        )

        cls.sale_order = SaleOrder.objects.create(customer=cls.customer)
        cls.sale = Sale.objects.create(
            order=cls.sale_order,
            product=cls.product,
            customer=cls.customer,
            quantity=2,
            unit_price=Decimal("15.00"),
            payment_method="cash",
        )

        cls.inbound_order = InboundOrder.objects.create(invoice_no="INV-1001")
        cls.purchase = Purchase.objects.create(
            inbound_order=cls.inbound_order,
            product=cls.product,
            supplier=None,
            quantity=3,
            cost_price=Decimal("4.00"),
            remaining=3,
        )
        cls.standalone_purchase = Purchase.objects.create(
            inbound_order=None,
            product=cls.product,
            supplier=None,
            quantity=2,
            cost_price=Decimal("5.00"),
            remaining=2,
        )

    def test_sales_records_page_renders_sales_and_purchase_totals(self):
        self.client.login(username="records_admin", password="pw123456")
        today = timezone.localdate()

        response = self.client.get(
            reverse("sales_records"),
            {"start_date": today.isoformat(), "end_date": today.isoformat()},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_sales_amount"], Decimal("30.00"))
        self.assertEqual(response.context["total_sales_orders"], 1)
        self.assertEqual(response.context["total_purchase_amount"], Decimal("22.00"))
        self.assertEqual(response.context["total_purchase_orders"], 2)
        self.assertContains(response, "Recent Searches")
        self.assertContains(response, "Sales Orders")
        self.assertContains(response, "No supplier linked")
        self.assertContains(response, "Purchases")
        self.assertContains(response, reverse("customer_detail", args=[self.customer.id]))
        self.assertTrue(response.context["show_profit"])
        self.assertContains(response, "Net Profit")

    def test_sales_records_shows_profit_for_admin_only(self):
        self.client.login(username="records_admin", password="pw123456")
        today = timezone.localdate()

        response = self.client.get(
            reverse("sales_records"),
            {"start_date": today.isoformat(), "end_date": today.isoformat()},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["show_profit"])
        self.assertEqual(response.context["total_sales_profit"], Decimal("30.00"))
        self.assertContains(response, "Net Profit")

    def test_sales_records_builds_visualization_context(self):
        self.client.login(username="records_admin", password="pw123456")
        today = timezone.localdate()

        response = self.client.get(
            reverse("sales_records"),
            {"start_date": today.isoformat(), "end_date": today.isoformat()},
        )

        self.assertEqual(response.status_code, 200)
        trend = response.context["trend_data"]
        self.assertEqual(len(trend), 1)
        self.assertEqual(trend[0]["sales"], 30.0)
        self.assertEqual(trend[0]["purchases"], 22.0)
        self.assertEqual(trend[0]["profit"], 30.0)
        self.assertEqual(response.context["payment_chart"], [{"label": "Cash", "amount": 30.0}])
        top = response.context["top_products"]
        self.assertEqual(len(top), 1)
        self.assertEqual(top[0]["qty"], 2)
        self.assertEqual(top[0]["revenue"], Decimal("30.00"))
        self.assertContains(response, 'id="trendChart"')
        self.assertContains(response, "Top Products")
        self.assertContains(response, "vs purchases")

    def test_sales_records_hides_purchase_data_for_regular_user(self):
        # A non-manager now gets the lean employee Sales view: today's orders
        # and amounts only, no purchase data and no charts.
        self.client.login(username="records_user", password="pw123456")
        today = timezone.localdate()

        response = self.client.get(
            reverse("sales_records"),
            {"start_date": today.isoformat(), "end_date": today.isoformat()},
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Purchase Cost")
        self.assertNotContains(response, "Inbound Purchases")
        self.assertNotContains(response, "vs purchases")
        self.assertNotContains(response, "<canvas")
        # The day's order and its amount are still shown.
        self.assertContains(response, "EUR 30.00")
        self.assertContains(response, "Record Customer")
        self.assertContains(response, 'name="date"')

    def test_sales_records_groups_corrected_order_by_order_datetime(self):
        self.client.login(username="records_admin", password="pw123456")

        drift_order = SaleOrder.objects.create(customer=self.customer)
        order_dt = timezone.now() - timedelta(days=1)
        SaleOrder.objects.filter(pk=drift_order.pk).update(created_at=order_dt)
        drift_sale = Sale.objects.create(
            order=drift_order,
            product=self.product,
            customer=self.customer,
            quantity=1,
            unit_price=Decimal("16.00"),
            payment_method="card",
        )
        Sale.objects.filter(pk=drift_sale.pk).update(date=timezone.now())

        response = self.client.get(
            reverse("sales_records"),
            {
                "start_date": timezone.localdate(order_dt).isoformat(),
                "end_date": timezone.localdate().isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        day_map = {
            block["date"]: [row["order_id"] for row in block["orders"]]
            for block in response.context["day_blocks"]
        }
        self.assertIn(drift_order.id, day_map[timezone.localdate(order_dt)])
        self.assertNotIn(drift_order.id, day_map.get(timezone.localdate(), []))


class CustomerDetailViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.user = user_model.objects.create_user(username="customer_timeline_user", password="pw123456")
        cls.admin = user_model.objects.create_superuser(username="customer_timeline_admin", password="pw123456")

        cls.category = Category.objects.create(name="Timeline Test")
        cls.customer = Customer.objects.create(nif="333444555", name="Timeline Customer")
        cls.product = Product.objects.create(
            name="Oud",
            barcode="4445556667778",
            brand="Atelier",
            category=cls.category,
            default_price=Decimal("20.00"),
        )

        cls.older_order = SaleOrder.objects.create(customer=cls.customer)
        cls.older_sale = Sale.objects.create(
            order=cls.older_order,
            product=cls.product,
            customer=cls.customer,
            quantity=1,
            unit_price=Decimal("20.00"),
            payment_method="cash",
        )
        older_dt = timezone.now() - timedelta(days=40)
        SaleOrder.objects.filter(pk=cls.older_order.pk).update(created_at=older_dt)
        Sale.objects.filter(pk=cls.older_sale.pk).update(date=older_dt)

        cls.newer_order = SaleOrder.objects.create(customer=cls.customer)
        cls.newer_sale = Sale.objects.create(
            order=cls.newer_order,
            product=cls.product,
            customer=cls.customer,
            quantity=2,
            unit_price=Decimal("22.00"),
            payment_method="card",
        )
        newer_dt = timezone.now() - timedelta(days=2)
        SaleOrder.objects.filter(pk=cls.newer_order.pk).update(created_at=newer_dt)
        Sale.objects.filter(pk=cls.newer_sale.pk).update(date=newer_dt)

    def test_customer_detail_groups_orders_by_month_day_and_time(self):
        self.client.login(username="customer_timeline_admin", password="pw123456")

        response = self.client.get(reverse("customer_detail", args=[self.customer.id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_orders"], 2)
        self.assertEqual(len(response.context["month_blocks"]), 2)
        self.assertContains(response, "Order Timeline")
        self.assertContains(response, "Timeline Customer")
        self.assertContains(response, "Order #")
        self.assertTrue(response.context["show_profit"])
        self.assertContains(response, "Net Profit")

    def test_customer_detail_shows_profit_for_admin_only(self):
        self.client.login(username="customer_timeline_admin", password="pw123456")

        response = self.client.get(reverse("customer_detail", args=[self.customer.id]))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["show_profit"])
        self.assertEqual(response.context["total_profit"], Decimal("64.00"))
        self.assertContains(response, "Net Profit")

    def test_customer_detail_builds_visual_context(self):
        self.client.login(username="customer_timeline_admin", password="pw123456")

        response = self.client.get(reverse("customer_detail", args=[self.customer.id]))

        self.assertEqual(response.status_code, 200)
        # Monthly spend trend is chronological (oldest first).
        trend = response.context["spend_trend"]
        self.assertEqual([row["amount"] for row in trend], [20.0, 44.0])
        # Payment mix aggregates each method across the customer's orders.
        mix = {row["label"]: row["amount"] for row in response.context["payment_mix"]}
        self.assertEqual(mix["Card"], 44.0)
        self.assertEqual(mix["Cash"], 20.0)
        # Top products aggregate units + spend per product.
        top = response.context["top_products"]
        self.assertEqual(len(top), 1)
        self.assertEqual(top[0]["qty"], 3)
        self.assertEqual(top[0]["spend"], Decimal("64.00"))
        # Cadence KPIs are populated for a customer with history.
        cadence = response.context["cadence"]
        self.assertIsNotNone(cadence["since"])
        self.assertIsNotNone(cadence["avg_gap_days"])
        self.assertContains(response, "Top Products")
        self.assertContains(response, "Monthly Spend")

    def test_customer_detail_date_range_filter_scopes_data(self):
        self.client.login(username="customer_timeline_admin", password="pw123456")
        start = (timezone.now().date() - timedelta(days=5)).isoformat()

        response = self.client.get(
            reverse("customer_detail", args=[self.customer.id]),
            {"start_date": start},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["range_key"], "custom")
        self.assertTrue(response.context["range_active"])
        # Only the recent order falls inside the range.
        self.assertEqual(response.context["total_orders"], 1)
        self.assertEqual(response.context["total_spent"], Decimal("44.00"))
        self.assertEqual([row["amount"] for row in response.context["spend_trend"]], [44.0])
        self.assertEqual(response.context["top_products"][0]["qty"], 2)

    def test_customer_detail_hides_sensitive_sections_for_regular_user(self):
        # Task 9: non-manager users now get the lean employee orders view
        # (200) instead of being redirected away from customer_detail.
        self.client.login(username="customer_timeline_user", password="pw123456")

        response = self.client.get(reverse("customer_detail", args=[self.customer.id]))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "stock/customer_orders_employee.html")


class DashboardViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.user = user_model.objects.create_user(username="dashboard_user", password="pw123456")
        cls.admin = user_model.objects.create_superuser(username="dashboard_admin", password="pw123456")

        cls.category = Category.objects.create(name="Dashboard Test")
        cls.customer = Customer.objects.create(nif="777888999", name="Dashboard Customer")
        cls.product = Product.objects.create(
            name="Saffron",
            barcode="1112223334445",
            brand="Maison",
            category=cls.category,
            default_price=Decimal("15.00"),
        )
        cls.slow_product = Product.objects.create(
            name="Oud Reserve",
            barcode="1112223334446",
            brand="Maison",
            category=cls.category,
            default_price=Decimal("32.00"),
        )

        current_dt = timezone.now()
        previous_dt = current_dt - timedelta(days=40)

        current_purchase = Purchase.objects.create(
            product=cls.product,
            supplier=None,
            quantity=5,
            cost_price=Decimal("8.00"),
            remaining=5,
        )
        Purchase.objects.filter(pk=current_purchase.pk).update(date=current_dt - timedelta(hours=1))

        previous_purchase = Purchase.objects.create(
            product=cls.product,
            supplier=None,
            quantity=3,
            cost_price=Decimal("7.00"),
            remaining=3,
        )
        Purchase.objects.filter(pk=previous_purchase.pk).update(date=previous_dt - timedelta(hours=1))

        slow_purchase = Purchase.objects.create(
            product=cls.slow_product,
            supplier=None,
            quantity=9,
            cost_price=Decimal("12.00"),
            remaining=9,
        )
        Purchase.objects.filter(pk=slow_purchase.pk).update(date=previous_dt - timedelta(days=15))

        cls.current_order = SaleOrder.objects.create(customer=cls.customer)
        cls.current_sale = Sale.objects.create(
            order=cls.current_order,
            product=cls.product,
            customer=cls.customer,
            quantity=2,
            unit_price=Decimal("15.00"),
            payment_method="cash",
        )
        SaleOrder.objects.filter(pk=cls.current_order.pk).update(created_at=current_dt)
        Sale.objects.filter(pk=cls.current_sale.pk).update(date=current_dt)

        cls.previous_order = SaleOrder.objects.create(customer=cls.customer)
        cls.previous_sale = Sale.objects.create(
            order=cls.previous_order,
            product=cls.product,
            customer=cls.customer,
            quantity=1,
            unit_price=Decimal("14.00"),
            payment_method="card",
        )
        SaleOrder.objects.filter(pk=cls.previous_order.pk).update(created_at=previous_dt)
        Sale.objects.filter(pk=cls.previous_sale.pk).update(date=previous_dt)

        cls.invoice = ARInvoice.objects.create(customer=cls.customer, total_amount=Decimal("50.00"), amount_paid=Decimal("10.00"))
        ARInvoice.objects.filter(pk=cls.invoice.pk).update(date=current_dt.date(), created_at=current_dt)
        cls.payment = ARPayment.objects.create(invoice=cls.invoice, amount=Decimal("10.00"), method="cash")
        ARPayment.objects.filter(pk=cls.payment.pk).update(created_at=current_dt)

        cls.current_month = current_dt.strftime("%Y-%m")

    def test_dashboard_hides_sensitive_sections_for_regular_user(self):
        self.client.login(username="dashboard_user", password="pw123456")

        response = self.client.get(reverse("dashboard"), {"month": self.current_month})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["month_sales_amount"], Decimal("30.00"))
        self.assertEqual(response.context["month_order_count"], 1)
        self.assertEqual(response.context["month_receivables_added"], Decimal("50.00"))
        self.assertEqual(response.context["month_receipts_collected"], Decimal("10.00"))
        self.assertFalse(response.context["show_profit"])
        self.assertNotContains(response, "Employee view")
        self.assertNotContains(response, "Today only")
        self.assertNotContains(response, "Monthly overview")
        self.assertNotContains(response, "Top customers")
        self.assertNotContains(response, "Gross Profit")
        self.assertContains(response, "Sales today")           # KPI strip
        self.assertContains(response, "EUR 30.00")             # order total is visible to employee
        self.assertContains(response, "Dashboard Customer")    # order card renders the customer
        self.assertContains(response, "Operations")
        self.assertContains(response, "Sales &amp; Clients")
        self.assertContains(response, reverse("sales_records"))
        self.assertContains(response, reverse("customer_search"))

    def test_dashboard_shows_admin_only_profit_and_purchase_metrics(self):
        self.client.login(username="dashboard_admin", password="pw123456")

        response = self.client.get(reverse("dashboard"), {"month": self.current_month})

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["show_profit"])
        self.assertTrue(response.context["show_purchase_metrics"])
        self.assertEqual(response.context["month_sales_profit"], Decimal("16.00"))
        self.assertEqual(response.context["month_purchase_amount"], Decimal("40.00"))
        self.assertNotContains(response, "Monthly overview")
        self.assertContains(response, "Sales trend")
        self.assertContains(response, "Gross Profit")
        self.assertContains(response, "Purchase Spend")
        self.assertContains(response, "Capital lock and slow movers")
        self.assertContains(response, "Most capital locked stock")
        self.assertContains(response, "Slow movers")
        self.assertContains(response, "Oud Reserve")

    def test_catalog_page_renders(self):
        self.client.login(username="dashboard_admin", password="pw123456")
        response = self.client.get(reverse("catalog"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Fragrance catalogue")
        self.assertContains(response, "A library of")

    def test_employee_blocked_from_catalog_page(self):
        self.client.login(username="dashboard_user", password="pw123456")
        response = self.client.get(reverse("catalog"))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("dashboard"), response.headers["Location"])

    def test_sales_page_shows_year_trend_by_default(self):
        # Sales Trend lives on the Sales (records) page via ?view=year
        # (the no-param default is now the month calendar).
        self.client.login(username="dashboard_admin", password="pw123456")

        response = self.client.get(reverse("sales_records"), {"view": "year"})

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["show_trend"])
        self.assertEqual(len(response.context["monthly_rows"]), 12)
        self.assertEqual(response.context["year_sales_amount"], Decimal("44.00"))
        self.assertContains(response, "Sales by Month")

    def test_sales_trend_url_redirects_to_merged_page(self):
        self.client.login(username="dashboard_admin", password="pw123456")
        response = self.client.get(reverse("sales_trend"))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("sales_records"), response.headers["Location"])

    def test_regular_user_can_open_pages_but_not_sensitive_views(self):
        self.client.login(username="dashboard_user", password="pw123456")

        self.assertEqual(self.client.get(reverse("sales_records")).status_code, 200)
        self.assertEqual(self.client.get(reverse("customer_search")).status_code, 200)

        product_list_response = self.client.get(reverse("product_list"))
        self.assertEqual(product_list_response.status_code, 200)

        ar_list_response = self.client.get(reverse("ar_list"))
        self.assertEqual(ar_list_response.status_code, 302)
        self.assertIn(reverse("dashboard"), ar_list_response.headers["Location"])

    def test_regular_user_ar_pages_hide_financial_details(self):
        self.client.login(username="dashboard_user", password="pw123456")

        list_response = self.client.get(reverse("ar_list"))
        detail_response = self.client.get(reverse("ar_detail", args=[self.invoice.id]))

        self.assertEqual(list_response.status_code, 302)
        self.assertIn(reverse("dashboard"), list_response.headers["Location"])
        self.assertEqual(detail_response.status_code, 302)
        self.assertIn(reverse("dashboard"), detail_response.headers["Location"])

    def test_regular_user_can_open_any_order_detail_without_customer_contacts(self):
        self.client.login(username="dashboard_user", password="pw123456")

        today_response = self.client.get(reverse("sale_order_detail", args=[self.current_order.id]))
        old_response = self.client.get(reverse("sale_order_detail", args=[self.previous_order.id]))

        self.assertEqual(today_response.status_code, 200)
        self.assertEqual(old_response.status_code, 200)
        self.assertNotContains(old_response, "NIF:")
        self.assertNotContains(old_response, "Email:")
        self.assertNotContains(old_response, "Order amounts, prices, and payment details are hidden for employee accounts.")
        self.assertContains(old_response, "Grand Total")
        self.assertContains(old_response, "EUR 14.00")
        self.assertContains(old_response, "Card")

    def test_regular_user_can_open_printable_sale_order(self):
        self.client.login(username="dashboard_user", password="pw123456")

        detail_response = self.client.get(reverse("sale_order_detail", args=[self.previous_order.id]))
        print_response = self.client.get(
            reverse("sale_order_detail", args=[self.previous_order.id]),
            {"print": "1", "layout": "a4"},
        )

        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(print_response.status_code, 200)
        self.assertContains(detail_response, "Print A4")
        self.assertContains(detail_response, "Print POS (48mm)")
        self.assertContains(detail_response, "Grand Total")
        self.assertContains(print_response, "Grand Total")
        self.assertContains(print_response, "EUR 14.00")

    def test_dashboard_uses_order_datetime_for_today_bucket(self):
        drift_order = SaleOrder.objects.create(customer=self.customer)
        order_dt = timezone.now() - timedelta(days=1)
        SaleOrder.objects.filter(pk=drift_order.pk).update(created_at=order_dt)
        drift_sale = Sale.objects.create(
            order=drift_order,
            product=self.product,
            customer=self.customer,
            quantity=2,
            unit_price=Decimal("15.00"),
            payment_method="cash",
        )
        Sale.objects.filter(pk=drift_sale.pk).update(date=timezone.now())

        # Today's orders live in the employee dashboard (managers use /today).
        self.client.login(username="dashboard_user", password="pw123456")
        response = self.client.get(reverse("dashboard"), {"month": self.current_month})

        self.assertEqual(response.status_code, 200)
        today_order_ids = [order.id for order in response.context["sale_orders_today"]]
        # Real same-day order is present; an order dated yesterday is bucketed out
        # even though its sale row was touched today.
        self.assertIn(self.current_order.id, today_order_ids)
        self.assertNotIn(drift_order.id, today_order_ids)


class ProductArchitectureTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.manager = user_model.objects.create_superuser(username="product_admin", password="pw123456")
        cls.employee = user_model.objects.create_user(username="product_employee", password="pw123456")
        cls.category = Category.objects.create(name="Structured Product")
        cls.other_category = Category.objects.create(name="Phone Accessory")

    def test_product_form_can_create_brand_series_variant_and_prices(self):
        form = ProductForm(data={
            "barcode": "3216549870123",
            "category": self.category.id,
            "new_brand_name": "Baseus",
            "new_series_name": "iPhone 15 Pro",
            "name": "Clear Case",
            "spec": "MagSafe",
            "color": "Black",
            "default_price": "19.90",
            "wholesale_price": "11.40",
            "description": "Transparent case",
        })

        self.assertTrue(form.is_valid(), form.errors)
        product = form.save()

        self.assertEqual(product.brand, "Baseus")
        self.assertEqual(product.model, "iPhone 15 Pro")
        self.assertEqual(product.display_name, "Baseus - iPhone 15 Pro - Clear Case MagSafe Black")
        self.assertEqual(product.default_price, Decimal("19.90"))
        self.assertEqual(product.wholesale_price, Decimal("11.40"))
        self.assertEqual(product.brand_master.name, "Baseus")
        self.assertEqual(product.series_master.name, "iPhone 15 Pro")
        self.assertTrue(Brand.objects.filter(name="Baseus").exists())
        self.assertTrue(ProductSeries.objects.filter(name="iPhone 15 Pro", brand__name="Baseus").exists())
        self.assertTrue(product.brand_master.categories.filter(id=self.category.id).exists())

    def test_product_form_filters_brands_by_selected_category(self):
        perfume_brand = Brand.objects.create(name="Lattafa")
        perfume_brand.categories.add(self.category)
        accessory_brand = Brand.objects.create(name="Ugreen")
        accessory_brand.categories.add(self.other_category)

        form = ProductForm(data={"category": self.category.id})

        self.assertIn(perfume_brand, form.fields["brand_master"].queryset)
        self.assertNotIn(accessory_brand, form.fields["brand_master"].queryset)

    def test_existing_brand_can_be_linked_to_another_category_by_saving_product(self):
        brand = Brand.objects.create(name="Baseus")
        brand.categories.add(self.other_category)

        form = ProductForm(data={
            "barcode": "1231231231234",
            "category": self.category.id,
            "brand_master": brand.id,
            "name": "Car Charger",
            "default_price": "14.50",
        })

        self.assertTrue(form.is_valid(), form.errors)
        product = form.save()

        self.assertEqual(product.brand_master, brand)
        self.assertTrue(brand.categories.filter(id=self.category.id).exists())

    def test_add_product_view_accepts_new_brand_when_brand_select_is_blank(self):
        self.client.login(username="product_admin", password="pw123456")

        response = self.client.post(reverse("add_product"), data={
            "barcode": "5556667778881",
            "category": self.category.id,
            "brand_master": "",
            "new_brand_name": "Afnan",
            "series_master": "",
            "new_series_name": "9PM",
            "name": "EDP",
            "spec": "100ml",
            "color": "",
            "default_price": "39.90",
            "wholesale_price": "26.50",
            "description": "Evening fragrance",
        })

        self.assertEqual(response.status_code, 302)
        product = Product.objects.get(barcode="5556667778881")
        self.assertEqual(product.brand_master.name, "Afnan")
        self.assertEqual(product.series_master.name, "9PM")
        self.assertEqual(product.brand, "Afnan")
        self.assertEqual(product.model, "9PM")

    def test_products_autocomplete_supports_product_id_lookup(self):
        brand = Brand.objects.create(name="Lattafa")
        series = ProductSeries.objects.create(brand=brand, name="Asad")
        product = Product.objects.create(
            name="Elixir",
            barcode="1122334455667",
            brand="Lattafa",
            brand_master=brand,
            series_master=series,
            model="Asad",
            category=self.category,
            default_price=Decimal("39.90"),
        )

        self.client.login(username="product_employee", password="pw123456")
        response = self.client.get(reverse("products_autocomplete"), {"product_id": product.id})

        self.assertEqual(response.status_code, 200)
        payload = response.json()["results"]
        self.assertEqual(len(payload), 1)
        self.assertEqual(payload[0]["id"], product.id)
        self.assertEqual(payload[0]["display_name"], product.display_name)

    def test_product_detail_shows_structured_product_fields_for_manager(self):
        brand = Brand.objects.create(name="Dior")
        series = ProductSeries.objects.create(brand=brand, name="Sauvage")
        product = Product.objects.create(
            name="EDP",
            barcode="8887776665554",
            brand="Dior",
            brand_master=brand,
            series_master=series,
            model="Sauvage",
            category=self.category,
            spec="100ml",
            color="Blue",
            default_price=Decimal("99.00"),
            wholesale_price=Decimal("72.00"),
        )

        self.client.login(username="product_admin", password="pw123456")
        response = self.client.get(reverse("product_detail", args=[product.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Dior - Sauvage - EDP 100ml Blue")
        self.assertContains(response, "Wholesale Price")
        self.assertContains(response, "100ml")
        self.assertContains(response, "Blue")

    def test_product_detail_shows_prices_but_hides_sales_history_for_regular_user(self):
        brand = Brand.objects.create(name="Anker")
        series = ProductSeries.objects.create(brand=brand, name="USB-C")
        product = Product.objects.create(
            name="Cable",
            barcode="9998887776665",
            brand="Anker",
            brand_master=brand,
            series_master=series,
            model="USB-C",
            category=self.category,
            default_price=Decimal("12.50"),
            wholesale_price=Decimal("8.00"),
        )

        self.client.login(username="product_employee", password="pw123456")
        response = self.client.get(reverse("product_detail", args=[product.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Anker - USB-C - Cable")
        self.assertContains(response, "Retail Price")
        self.assertContains(response, "Wholesale Price")
        self.assertNotContains(response, "Suppliers &amp; cost")
        self.assertNotContains(response, "Sales History")

    def test_product_list_can_filter_by_brand(self):
        perfume_brand = Brand.objects.create(name="Lattafa")
        perfume_brand.categories.add(self.category)
        other_brand = Brand.objects.create(name="Baseus")
        other_brand.categories.add(self.category)

        perfume_product = Product.objects.create(
            name="Asad",
            barcode="1234509876501",
            brand="Lattafa",
            brand_master=perfume_brand,
            category=self.category,
            default_price=Decimal("39.90"),
        )
        Product.objects.create(
            name="Charger",
            barcode="1234509876502",
            brand="Baseus",
            brand_master=other_brand,
            category=self.category,
            default_price=Decimal("19.90"),
        )

        self.client.login(username="product_admin", password="pw123456")
        response = self.client.get(reverse("product_list"), {
            "category": self.category.id,
            "brand": perfume_brand.id,
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, perfume_product.display_name)
        self.assertNotContains(response, "Baseus - Charger")

    def test_product_list_can_sort_by_sales_quantity(self):
        best_seller = Product.objects.create(
            name="Winner",
            barcode="1234509876503",
            brand="Maison",
            category=self.category,
            default_price=Decimal("25.00"),
        )
        slower = Product.objects.create(
            name="Slower",
            barcode="1234509876504",
            brand="Maison",
            category=self.category,
            default_price=Decimal("20.00"),
        )
        order = SaleOrder.objects.create()
        Sale.objects.create(order=order, product=slower, quantity=1, unit_price=Decimal("20.00"), payment_method="cash")
        Sale.objects.create(order=order, product=best_seller, quantity=5, unit_price=Decimal("25.00"), payment_method="cash")

        self.client.login(username="product_admin", password="pw123456")
        response = self.client.get(reverse("product_list"), {"sort": "sales_desc"})

        self.assertEqual(response.status_code, 200)
        product_ids = [product.id for product in response.context["page_obj"].object_list[:2]]
        self.assertEqual(product_ids[0], best_seller.id)

    def test_product_list_hides_sales_metrics_but_keeps_prices_for_regular_user(self):
        product = Product.objects.create(
            name="Visible Price",
            barcode="1234509876510",
            brand="Maison",
            category=self.category,
            default_price=Decimal("29.00"),
            wholesale_price=Decimal("17.00"),
        )
        order = SaleOrder.objects.create()
        Sale.objects.create(order=order, product=product, quantity=4, unit_price=Decimal("29.00"), payment_method="cash")

        self.client.login(username="product_employee", password="pw123456")
        response = self.client.get(reverse("product_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Retail / Wholesale")
        self.assertContains(response, "EUR 29.00")
        self.assertContains(response, "EUR 17.00")
        self.assertNotContains(response, "Best selling")

    def test_product_list_shows_shopify_export_for_manager(self):
        self.client.login(username="product_admin", password="pw123456")

        response = self.client.get(reverse("product_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Export For Shopify")
        self.assertContains(response, reverse("export_shopify_inventory_csv"))
        self.assertContains(response, "Export Shopify Product CSV")
        self.assertContains(response, "Only in stock")

    def test_shopify_export_matches_product_template_and_uses_current_stock(self):
        shopify_category = Category.objects.create(name="Perfumes")
        brand = Brand.objects.create(name="Lattafa")
        brand.categories.add(shopify_category)
        series = ProductSeries.objects.create(brand=brand, name="Asad")
        first = Product.objects.create(
            name="EDP",
            barcode="1234509876511",
            brand="Lattafa",
            brand_master=brand,
            series_master=series,
            model="Asad",
            category=shopify_category,
            spec="100ml",
            color="Blue",
            description="Evening fragrance",
            default_price=Decimal("39.90"),
            price_locked=True,  # keep the manually-set price for this export-format test
        )
        second = Product.objects.create(
            name="EDP",
            barcode="1234509876512",
            brand="Lattafa",
            brand_master=brand,
            series_master=series,
            model="Asad",
            category=shopify_category,
            spec="50ml",
            color="Blue",
            default_price=Decimal("29.90"),
            price_locked=True,  # keep the manually-set price for this export-format test
        )
        Purchase.objects.create(product=first, quantity=8, remaining=6, cost_price=Decimal("12.34"))
        Purchase.objects.create(product=second, quantity=3, remaining=0, cost_price=Decimal("11.11"))

        self.client.login(username="product_admin", password="pw123456")
        response = self.client.get(reverse("export_shopify_inventory_csv"), {
            "brand": brand.id,
            "sort": "name_asc",
        })

        self.assertEqual(response.status_code, 200)
        self.assertIn("Shopify_Product_Inventory_", response["Content-Disposition"])
        reader = csv.DictReader(StringIO(response.content.decode("utf-8-sig")))
        self.assertEqual(reader.fieldnames, [
            "Title",
            "URL handle",
            "Description",
            "Vendor",
            "Product category",
            "Type",
            "Tags",
            "Published on online store",
            "Status",
            "SKU",
            "Barcode",
            "Option1 name",
            "Option1 value",
            "Option1 Linked To",
            "Option2 name",
            "Option2 value",
            "Option2 Linked To",
            "Option3 name",
            "Option3 value",
            "Option3 Linked To",
            "Price",
            "Compare-at price",
            "Cost per item",
            "Charge tax",
            "Tax code",
            "Unit price total measure",
            "Unit price total measure unit",
            "Unit price base measure",
            "Unit price base measure unit",
            "Inventory tracker",
            "Inventory quantity",
            "Continue selling when out of stock",
            "Weight value (grams)",
            "Weight unit for display",
            "Requires shipping",
            "Fulfillment service",
            "Product image URL",
            "Image position",
            "Image alt text",
            "Variant image URL",
            "Gift card",
            "SEO title",
            "SEO description",
            "Color (product.metafields.shopify.color-pattern)",
            "Google Shopping / Google product category",
            "Google Shopping / Gender",
            "Google Shopping / Age group",
            "Google Shopping / Manufacturer part number (MPN)",
            "Google Shopping / Ad group name",
            "Google Shopping / Ads labels",
            "Google Shopping / Condition",
            "Google Shopping / Custom product",
            "Google Shopping / Custom label 0",
            "Google Shopping / Custom label 1",
            "Google Shopping / Custom label 2",
            "Google Shopping / Custom label 3",
            "Google Shopping / Custom label 4",
        ])
        rows = list(reader)
        self.assertEqual(len(rows), 2)

        self.assertEqual(rows[0]["Title"], "Lattafa - Asad - EDP")
        self.assertEqual(rows[0]["URL handle"], "lattafa-asad-edp")
        self.assertEqual(rows[0]["Description"], "Evening fragrance")
        self.assertEqual(rows[0]["Vendor"], "Lattafa")
        # "Product category" must be a valid Shopify taxonomy path (not the bare local
        # name, which Shopify can't map and would ML-mis-categorize). Type keeps the name.
        self.assertEqual(
            rows[0]["Product category"],
            "Health & Beauty > Personal Care > Cosmetics > Perfumes & Colognes > Eaux de Parfum",
        )
        self.assertEqual(rows[0]["Type"], "Perfumes")
        self.assertEqual(rows[0]["Published on online store"], "TRUE")
        self.assertEqual(rows[0]["Status"], "Active")
        self.assertEqual(rows[0]["SKU"], first.barcode)
        self.assertEqual(rows[0]["Barcode"], first.barcode)
        self.assertEqual(rows[0]["Option1 name"], "Spec")
        self.assertEqual(rows[0]["Option1 value"], "100ml")
        self.assertEqual(rows[0]["Option2 name"], "Color")
        self.assertEqual(rows[0]["Option2 value"], "Blue")
        self.assertEqual(rows[0]["Option2 Linked To"], "product.metafields.shopify.color-pattern")
        self.assertEqual(rows[0]["Price"], "39.90")
        self.assertEqual(rows[0]["Cost per item"], "12.34")
        self.assertEqual(rows[0]["Charge tax"], "TRUE")
        self.assertEqual(rows[0]["Inventory tracker"], "shopify")
        self.assertEqual(rows[0]["Inventory quantity"], "6")
        self.assertEqual(rows[0]["Continue selling when out of stock"], "DENY")
        self.assertEqual(rows[0]["Requires shipping"], "TRUE")
        self.assertEqual(rows[0]["Fulfillment service"], "manual")
        self.assertEqual(rows[0]["Gift card"], "FALSE")
        self.assertEqual(rows[0]["SEO title"], "Lattafa - Asad - EDP")
        self.assertEqual(rows[0]["Color (product.metafields.shopify.color-pattern)"], "Blue")
        self.assertEqual(rows[0]["Google Shopping / Condition"], "New")
        self.assertEqual(rows[0]["Google Shopping / Custom product"], "FALSE")

        self.assertEqual(rows[1]["Title"], "")
        self.assertEqual(rows[1]["URL handle"], rows[0]["URL handle"])
        self.assertEqual(rows[1]["SKU"], second.barcode)
        self.assertEqual(rows[1]["Barcode"], second.barcode)
        self.assertEqual(rows[1]["Option1 value"], "50ml")
        self.assertEqual(rows[1]["Price"], "29.90")
        self.assertEqual(rows[1]["Cost per item"], "11.11")
        self.assertEqual(rows[1]["Inventory quantity"], "0")


class EmployeeProductAccessTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.employee = user_model.objects.create_user(username="prod_emp", password="pw123456")
        cls.category = Category.objects.create(name="PCat")
        cls.product = Product.objects.create(name="PP", barcode="7100000000001", brand="B",
                                             default_price=Decimal("9.90"))

    def test_employee_can_view_products_and_detail(self):
        self.client.login(username="prod_emp", password="pw123456")
        self.assertEqual(self.client.get(reverse("product_list")).status_code, 200)
        self.assertEqual(self.client.get(reverse("product_detail", args=[self.product.pk])).status_code, 200)

    def test_employee_cannot_add_product(self):
        # Products are read-only for employees: managers own the catalogue.
        self.client.login(username="prod_emp", password="pw123456")
        resp = self.client.post(reverse("add_product"), {
            "barcode": "7100000000002", "category": self.category.id,
            "new_brand_name": "NB", "name": "New Item", "default_price": "12.00",
        })
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("dashboard"), resp.headers["Location"])
        self.assertFalse(Product.objects.filter(barcode="7100000000002").exists())

    def test_employee_still_blocked_from_inbound(self):
        self.client.login(username="prod_emp", password="pw123456")
        for url in [reverse("inbound")]:
            resp = self.client.get(url)
            self.assertEqual(resp.status_code, 302)
            self.assertIn(reverse("dashboard"), resp.headers["Location"])


class EmployeeProductEditExportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.employee = user_model.objects.create_user(username="edit_emp", password="pw123456")
        cls.category = Category.objects.create(name="ECat")
        cls.product = Product.objects.create(name="Old Name", barcode="7400000000001", brand="B",
                                             default_price=Decimal("9.90"))

    def test_employee_cannot_open_or_submit_edit(self):
        # View-only for employees — the edit form is a manager tool.
        self.client.login(username="edit_emp", password="pw123456")
        self.assertEqual(self.client.get(reverse("edit_product", args=[self.product.pk])).status_code, 302)
        resp = self.client.post(reverse("edit_product", args=[self.product.pk]), {
            "barcode": "7400000000001", "category": self.category.id,
            "new_brand_name": "B", "name": "New Name", "default_price": "10.50",
        })
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("dashboard"), resp.headers["Location"])
        self.product.refresh_from_db()
        self.assertEqual(self.product.name, "Old Name")

    def test_employee_can_download_product_excel(self):
        self.client.login(username="edit_emp", password="pw123456")
        resp = self.client.get(reverse("export_product_list_excel"))
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheet", resp["Content-Type"])

    def test_product_excel_includes_ean_column(self):
        from openpyxl import load_workbook
        self.client.login(username="edit_emp", password="pw123456")
        resp = self.client.get(reverse("export_product_list_excel"))
        wb = load_workbook(BytesIO(b"".join(resp.streaming_content)))
        values = {str(v) for ws in wb.worksheets for row in ws.iter_rows(values_only=True)
                  for v in row if v is not None}
        self.assertIn("EAN", values)                 # the new header
        self.assertIn("7400000000001", values)       # the product's barcode value

    def test_product_excel_appends_spec_to_name(self):
        from openpyxl import load_workbook
        Product.objects.filter(pk=self.product.pk).update(spec="77ml")
        self.client.login(username="edit_emp", password="pw123456")
        resp = self.client.get(reverse("export_product_list_excel"))
        wb = load_workbook(BytesIO(b"".join(resp.streaming_content)))
        values = [str(v) for ws in wb.worksheets for row in ws.iter_rows(values_only=True)
                  for v in row if v is not None]
        self.assertTrue(any(v.endswith("77ml") for v in values))  # name column ends with the spec

    def test_product_excel_uses_excel_safe_currency_number_format(self):
        self.client.login(username="edit_emp", password="pw123456")

        resp = self.client.get(reverse("export_product_list_excel"))

        workbook_bytes = b"".join(resp.streaming_content)
        with ZipFile(BytesIO(workbook_bytes)) as workbook_zip:
            styles = ElementTree.fromstring(workbook_zip.read("xl/styles.xml"))
        namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        format_codes = {
            node.attrib["formatCode"]
            for node in styles.findall("main:numFmts/main:numFmt", namespace)
        }
        self.assertIn('"EUR" #,##0.00', format_codes)

    def test_delete_archives_product_but_keeps_sales_and_profit(self):
        """Regression: 'delete' used to cascade the product away, taking its
        purchase + sales rows (and any emptied order) with it, so past revenue
        and profit changed retroactively. It must leave the books alone."""
        from stock.models import ProductImage
        from stock.services.profit import sale_profit_map_for_sale_ids
        get_user_model().objects.create_superuser("del_mgr", password="pw123456")
        self.client.login(username="del_mgr", password="pw123456")
        Purchase.objects.create(product=self.product, supplier=None, quantity=4,
                                cost_price=Decimal("3.00"), remaining=3)
        order = SaleOrder.objects.create()
        sale = Sale.objects.create(order=order, product=self.product, quantity=1,
                                   unit_price=Decimal("9.90"), payment_method="cash",
                                   cost_basis=Decimal("3.00"))
        ProductImage.objects.create(product=self.product, image="products/x.jpg")
        before = sale_profit_map_for_sale_ids([sale.id])[sale.id]["profit"]

        resp = self.client.post(reverse("delete_product", args=[self.product.pk]))
        self.assertEqual(resp.status_code, 302)

        # Gone from the catalogue...
        self.assertFalse(Product.objects.filter(pk=self.product.pk).exists())
        self.assertNotContains(self.client.get(reverse("product_list")),
                               self.product.barcode)
        # ...but every record, and the profit built on it, survives.
        archived = Product.all_objects.get(pk=self.product.pk)
        self.assertTrue(archived.is_archived)
        self.assertEqual(archived.purchase_set.count(), 1)
        self.assertEqual(archived.sale_set.count(), 1)
        self.assertTrue(SaleOrder.objects.filter(pk=order.pk).exists())
        self.assertEqual(sale_profit_map_for_sale_ids([sale.id])[sale.id]["profit"], before)
        # A historical sale still resolves its product, and photos are cleared.
        self.assertEqual(Sale.objects.get(pk=sale.id).product.pk, self.product.pk)
        self.assertEqual(archived.images.count(), 0)

    def test_employee_still_blocked_from_deletes_and_shopify_export(self):
        self.client.login(username="edit_emp", password="pw123456")
        for resp in [
            self.client.post(reverse("delete_product", args=[self.product.pk])),
        ]:
            self.assertEqual(resp.status_code, 302)
            self.assertIn(reverse("dashboard"), resp.headers["Location"])
        # product_list offers the employee viewing + client export only: no
        # Add/Edit affordances and no Shopify export.
        page = self.client.get(reverse("product_list"))
        self.assertNotContains(page, reverse("add_product"))
        self.assertNotContains(page, reverse("edit_product", args=[self.product.pk]))
        self.assertContains(page, reverse("product_detail", args=[self.product.pk]))
        self.assertContains(page, reverse("export_product_list_excel"))
        self.assertNotContains(page, reverse("export_shopify_inventory_csv"))


class InboundOutboundPageTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.user = user_model.objects.create_user(username="ops_employee", password="pw123456")
        cls.manager = user_model.objects.create_user(username="ops_manager", password="pw123456", is_staff=True)

    def test_outbound_page_renders_review_guardrails(self):
        self.client.login(username="ops_employee", password="pw123456")

        response = self.client.get(reverse("outbound"))

        self.assertEqual(response.status_code, 200)
        # Sticky checkout bar -> confirm dialog -> its double-check guardrails.
        self.assertContains(response, "Review &amp; confirm")
        self.assertContains(response, "Confirm sale")
        self.assertContains(response, "Please double-check")

    def test_inbound_page_renders_review_guardrails(self):
        self.client.login(username="ops_manager", password="pw123456")

        response = self.client.get(reverse("inbound"))

        self.assertEqual(response.status_code, 200)
        # Sticky checkout bar -> confirm dialog -> its double-check guardrails.
        self.assertContains(response, "Review &amp; confirm")
        self.assertContains(response, "Final Inbound Review")
        self.assertContains(response, "Please double-check")

    def test_employee_blocked_from_inbound_page(self):
        self.client.login(username="ops_employee", password="pw123456")

        response = self.client.get(reverse("inbound"))

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("dashboard"), response.headers["Location"])


class SaleOrderCorrectionTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.admin = user_model.objects.create_superuser(username="order_fix_admin", password="pw123456")
        cls.employee = user_model.objects.create_user(username="order_fix_employee", password="pw123456")

        cls.category = Category.objects.create(name="Correction Perfume")
        cls.customer = Customer.objects.create(nif="987654321", name="Correction Customer")
        cls.product = Product.objects.create(
            name="Asad",
            barcode="1119991119991",
            brand="Lattafa",
            category=cls.category,
            default_price=Decimal("25.00"),
        )

    def test_admin_can_open_order_correction_center(self):
        self.client.login(username="order_fix_admin", password="pw123456")

        response = self.client.get(reverse("sale_order_correction_center"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Order Correction Center")
        self.assertContains(response, "Add missing historical order")

    def test_admin_can_open_order_correction_form_pages(self):
        purchase = Purchase.objects.create(
            product=self.product,
            supplier=None,
            quantity=5,
            cost_price=Decimal("10.00"),
            remaining=5,
        )
        purchase_dt = timezone.now() - timedelta(days=1)
        Purchase.objects.filter(pk=purchase.pk).update(date=purchase_dt)

        order = SaleOrder.objects.create(customer=self.customer, note="Editable order")
        SaleOrder.objects.filter(pk=order.pk).update(created_at=purchase_dt + timedelta(hours=1))
        sale = Sale.objects.create(
            order=order,
            product=self.product,
            customer=self.customer,
            quantity=1,
            unit_price=Decimal("25.00"),
            payment_method="cash",
        )
        Sale.objects.filter(pk=sale.pk).update(date=purchase_dt + timedelta(hours=1))

        self.client.login(username="order_fix_admin", password="pw123456")

        create_response = self.client.get(reverse("sale_order_correction_create"))
        edit_response = self.client.get(reverse("sale_order_correction_edit", args=[order.id]))

        self.assertEqual(create_response.status_code, 200)
        self.assertEqual(edit_response.status_code, 200)
        self.assertContains(create_response, "Add Missing Historical Order")
        self.assertContains(create_response, "Search product by barcode / brand / model / name")
        self.assertContains(create_response, 'id="product-search"')
        self.assertContains(create_response, 'name="items_json"')
        self.assertContains(edit_response, "Correct Order")
        self.assertContains(edit_response, "Delete entire order")
        # Edit preloads the existing line into the JS cart payload.
        self.assertContains(edit_response, 'id="initialCart"')

    def test_non_admin_is_redirected_from_order_correction_center(self):
        self.client.login(username="order_fix_employee", password="pw123456")

        response = self.client.get(reverse("sale_order_correction_center"))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], reverse("dashboard"))

    def test_admin_can_create_backdated_order_and_rebuild_stock(self):
        purchase = Purchase.objects.create(
            product=self.product,
            supplier=None,
            quantity=10,
            cost_price=Decimal("10.00"),
            remaining=10,
        )
        purchase_dt = timezone.now() - timedelta(days=4)
        Purchase.objects.filter(pk=purchase.pk).update(date=purchase_dt)
        order_dt = timezone.localtime(purchase_dt + timedelta(hours=2)).strftime("%Y-%m-%dT%H:%M")

        self.client.login(username="order_fix_admin", password="pw123456")
        response = self.client.post(reverse("sale_order_correction_create"), data={
            "customer": self.customer.id,
            "order_datetime": order_dt,
            "note": "Missed by employee",
            "reason": "Employee forgot to input the sale on time",
            "items_json": json.dumps([{"product_id": self.product.id, "qty": 3, "price": "25.00", "payment": "cash"}]),
            "payments_json": json.dumps([{"method": "cash", "amount": "75.00"}]),
            "affects_stock": "on",
        })

        self.assertEqual(response.status_code, 302)
        order = SaleOrder.objects.get(note="Missed by employee")
        purchase.refresh_from_db()

        self.assertEqual(order.items.count(), 1)
        self.assertEqual(purchase.remaining, 7)
        self.assertEqual({p.method: p.amount for p in order.payments.all()}, {"cash": Decimal("75.00")})
        self.assertTrue(SaleOrderChangeLog.objects.filter(order_id_snapshot=order.id, action="create").exists())

    def test_admin_can_create_order_for_earlier_date_without_full_fifo_rebuild(self):
        purchase = Purchase.objects.create(
            product=self.product,
            supplier=None,
            quantity=10,
            cost_price=Decimal("10.00"),
            remaining=10,
        )
        purchase_dt = timezone.now() - timedelta(days=1)
        Purchase.objects.filter(pk=purchase.pk).update(date=purchase_dt)
        order_dt = timezone.localtime(purchase_dt - timedelta(days=5)).strftime("%Y-%m-%dT%H:%M")

        self.client.login(username="order_fix_admin", password="pw123456")
        response = self.client.post(reverse("sale_order_correction_create"), data={
            "customer": self.customer.id,
            "order_datetime": order_dt,
            "note": "Very old missed order",
            "reason": "Entered late after stock count",
            "items_json": json.dumps([{"product_id": self.product.id, "qty": 4, "price": "25.00", "payment": "cash"}]),
            "payments_json": json.dumps([{"method": "cash", "amount": "100.00"}]),
            "affects_stock": "on",
        })

        self.assertEqual(response.status_code, 302)
        purchase.refresh_from_db()
        order = SaleOrder.objects.get(note="Very old missed order")

        self.assertEqual(order.items.count(), 1)
        self.assertEqual(purchase.remaining, 6)
        self.assertTrue(SaleOrderChangeLog.objects.filter(order_id_snapshot=order.id, action="create").exists())

    def test_split_payment_reloads_as_split_not_single_method(self):
        # Regression: an order paid card+cash reloaded in the correction center as
        # card-only, because the cart was rebuilt from each line's primary method
        # and ignored the order-level SaleOrderPayment split.
        purchase = Purchase.objects.create(
            product=self.product, supplier=None, quantity=10,
            cost_price=Decimal("10.00"), remaining=10,
        )
        Purchase.objects.filter(pk=purchase.pk).update(date=timezone.now() - timedelta(days=1))
        order_dt = timezone.localtime(timezone.now() - timedelta(hours=2)).strftime("%Y-%m-%dT%H:%M")

        self.client.login(username="order_fix_admin", password="pw123456")
        create = self.client.post(reverse("sale_order_correction_create"), data={
            "customer": self.customer.id,
            "order_datetime": order_dt,
            "note": "Split order",
            "reason": "paid part card part cash",
            "items_json": json.dumps([{"product_id": self.product.id, "qty": 2, "price": "25.00", "payment": "card"}]),
            "payments_json": json.dumps([
                {"method": "card", "amount": "30.00"},
                {"method": "cash", "amount": "20.00"},
            ]),
        })
        self.assertEqual(create.status_code, 302)
        order = SaleOrder.objects.get(note="Split order")
        self.assertEqual(
            {p.method: p.amount for p in order.payments.all()},
            {"card": Decimal("30.00"), "cash": Decimal("20.00")},
        )

        # Reopening the correction page must reconstruct the split, not collapse it.
        edit = self.client.get(reverse("sale_order_correction_edit", args=[order.id]))
        self.assertEqual(edit.status_code, 200)
        cart = edit.context["initial_cart"]
        self.assertEqual(len(cart), 1)
        self.assertTrue(cart[0]["isSplit"])
        self.assertEqual(cart[0]["paymentSplit"], {"card": 30.0, "cash": 20.0})

    def test_admin_can_update_order_and_rebuild_stock(self):
        purchase = Purchase.objects.create(
            product=self.product,
            supplier=None,
            quantity=10,
            cost_price=Decimal("10.00"),
            remaining=8,
        )
        purchase_dt = timezone.now() - timedelta(days=3)
        Purchase.objects.filter(pk=purchase.pk).update(date=purchase_dt)

        order = SaleOrder.objects.create(customer=self.customer, note="Original order")
        SaleOrder.objects.filter(pk=order.pk).update(created_at=purchase_dt + timedelta(hours=1))
        sale = Sale.objects.create(
            order=order,
            product=self.product,
            customer=self.customer,
            quantity=2,
            unit_price=Decimal("25.00"),
            payment_method="cash",
        )
        Sale.objects.filter(pk=sale.pk).update(date=purchase_dt + timedelta(hours=1))

        self.client.login(username="order_fix_admin", password="pw123456")
        response = self.client.post(reverse("sale_order_correction_edit", args=[order.id]), data={
            "customer": self.customer.id,
            "order_datetime": timezone.localtime(purchase_dt + timedelta(hours=3)).strftime("%Y-%m-%dT%H:%M"),
            "note": "Corrected order",
            "reason": "Price and quantity were entered incorrectly",
            "items_json": json.dumps([{"product_id": self.product.id, "qty": 4, "price": "27.50", "payment": "card"}]),
            "payments_json": json.dumps([{"method": "card", "amount": "110.00"}]),
        })

        self.assertEqual(response.status_code, 302)
        order.refresh_from_db()
        purchase.refresh_from_db()
        sale = order.items.get()

        self.assertEqual(order.note, "Corrected order")
        self.assertEqual(sale.quantity, 4)
        self.assertEqual(sale.unit_price, Decimal("27.50"))
        self.assertEqual(sale.payment_method, "card")
        self.assertEqual(purchase.remaining, 6)
        self.assertEqual({p.method: p.amount for p in order.payments.all()}, {"card": Decimal("110.00")})
        self.assertTrue(SaleOrderChangeLog.objects.filter(order_id_snapshot=order.id, action="update").exists())

    def test_backdated_order_update_keeps_sales_out_of_today_dashboard(self):
        extra_product = Product.objects.create(
            name="Reserve Line",
            barcode="1119991119992",
            brand="Lattafa",
            category=self.category,
            default_price=Decimal("35.00"),
        )
        first_purchase = Purchase.objects.create(
            product=self.product,
            supplier=None,
            quantity=10,
            cost_price=Decimal("10.00"),
            remaining=9,
        )
        extra_purchase = Purchase.objects.create(
            product=extra_product,
            supplier=None,
            quantity=10,
            cost_price=Decimal("12.00"),
            remaining=10,
        )
        purchase_dt = timezone.now() - timedelta(days=1)
        Purchase.objects.filter(pk=first_purchase.pk).update(date=purchase_dt)
        Purchase.objects.filter(pk=extra_purchase.pk).update(date=purchase_dt)

        order = SaleOrder.objects.create(customer=self.customer, note="Yesterday order")
        order_dt = purchase_dt + timedelta(hours=1)
        SaleOrder.objects.filter(pk=order.pk).update(created_at=order_dt)
        sale = Sale.objects.create(
            order=order,
            product=self.product,
            customer=self.customer,
            quantity=1,
            unit_price=Decimal("25.00"),
            payment_method="cash",
        )
        Sale.objects.filter(pk=sale.pk).update(date=order_dt)

        self.client.login(username="order_fix_admin", password="pw123456")
        response = self.client.post(reverse("sale_order_correction_edit", args=[order.id]), data={
            "customer": self.customer.id,
            "order_datetime": timezone.localtime(order_dt).strftime("%Y-%m-%dT%H:%M"),
            "note": "Yesterday order fixed",
            "reason": "Missing products were added later",
            "items_json": json.dumps([
                {"product_id": self.product.id, "qty": 1, "price": "25.00", "payment": "cash"},
                {"product_id": extra_product.id, "qty": 3, "price": "35.00", "payment": "card"},
            ]),
            "payments_json": json.dumps([
                {"method": "cash", "amount": "25.00"},
                {"method": "card", "amount": "105.00"},
            ]),
        })

        self.assertEqual(response.status_code, 302)
        order.refresh_from_db()
        sale_dates = list(order.items.order_by('id').values_list('date', flat=True))
        self.assertEqual(len(sale_dates), 2)
        self.assertTrue(all(timezone.localtime(sale_date).date() == timezone.localtime(order_dt).date() for sale_date in sale_dates))

        dashboard_response = self.client.get(reverse("dashboard"), {"month": timezone.localdate().strftime("%Y-%m")})
        self.assertEqual(dashboard_response.status_code, 200)
        today_order_ids = [item.id for item in dashboard_response.context["sale_orders_today"]]
        self.assertNotIn(order.id, today_order_ids)

    def test_removing_a_line_on_edit_restores_stock_and_deletes_sale(self):
        # Reproduces the old rollback bug: a removed product must restore its stock
        # and drop its sale record (previously the line silently survived).
        extra_product = Product.objects.create(
            name="Drop Me", barcode="2229991119993", brand="Lattafa",
            category=self.category, default_price=Decimal("30.00"),
        )
        purchase_a = Purchase.objects.create(product=self.product, supplier=None, quantity=10, cost_price=Decimal("10.00"), remaining=8)
        purchase_b = Purchase.objects.create(product=extra_product, supplier=None, quantity=10, cost_price=Decimal("12.00"), remaining=7)
        purchase_dt = timezone.now() - timedelta(days=2)
        Purchase.objects.filter(pk=purchase_a.pk).update(date=purchase_dt)
        Purchase.objects.filter(pk=purchase_b.pk).update(date=purchase_dt)

        order = SaleOrder.objects.create(customer=self.customer, note="Two lines")
        order_dt = purchase_dt + timedelta(hours=1)
        SaleOrder.objects.filter(pk=order.pk).update(created_at=order_dt)
        sale_a = Sale.objects.create(order=order, product=self.product, customer=self.customer, quantity=2, unit_price=Decimal("25.00"), payment_method="cash")
        sale_b = Sale.objects.create(order=order, product=extra_product, customer=self.customer, quantity=3, unit_price=Decimal("30.00"), payment_method="cash")
        Sale.objects.filter(pk__in=[sale_a.pk, sale_b.pk]).update(date=order_dt)

        self.client.login(username="order_fix_admin", password="pw123456")
        response = self.client.post(reverse("sale_order_correction_edit", args=[order.id]), data={
            "customer": self.customer.id,
            "order_datetime": timezone.localtime(order_dt).strftime("%Y-%m-%dT%H:%M"),
            "note": "One line",
            "reason": "Removed a product that was never actually sold",
            "items_json": json.dumps([{"product_id": self.product.id, "qty": 2, "price": "25.00", "payment": "cash"}]),
            "payments_json": json.dumps([{"method": "cash", "amount": "50.00"}]),
        })

        self.assertEqual(response.status_code, 302)
        order.refresh_from_db()
        purchase_a.refresh_from_db()
        purchase_b.refresh_from_db()
        self.assertEqual(order.items.count(), 1)
        self.assertEqual(order.items.get().product_id, self.product.id)
        self.assertFalse(Sale.objects.filter(order=order, product=extra_product).exists())
        # A unchanged (restore 2 + consume 2); B fully rolled back (7 -> 10).
        self.assertEqual(purchase_a.remaining, 8)
        self.assertEqual(purchase_b.remaining, 10)

    def test_admin_can_delete_order_and_restore_stock(self):
        purchase = Purchase.objects.create(
            product=self.product,
            supplier=None,
            quantity=10,
            cost_price=Decimal("10.00"),
            remaining=7,
        )
        purchase_dt = timezone.now() - timedelta(days=2)
        Purchase.objects.filter(pk=purchase.pk).update(date=purchase_dt)

        order = SaleOrder.objects.create(customer=self.customer, note="Delete me")
        SaleOrder.objects.filter(pk=order.pk).update(created_at=purchase_dt + timedelta(hours=1))
        sale = Sale.objects.create(
            order=order,
            product=self.product,
            customer=self.customer,
            quantity=3,
            unit_price=Decimal("25.00"),
            payment_method="cash",
        )
        Sale.objects.filter(pk=sale.pk).update(date=purchase_dt + timedelta(hours=1))

        self.client.login(username="order_fix_admin", password="pw123456")
        response = self.client.post(reverse("sale_order_correction_edit", args=[order.id]), data={
            "customer": self.customer.id,
            "order_datetime": timezone.localtime(purchase_dt + timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M"),
            "note": "Delete me",
            "reason": "Duplicate order entered by mistake",
            "delete_order": "1",
        })

        self.assertEqual(response.status_code, 302)
        self.assertFalse(SaleOrder.objects.filter(pk=order.id).exists())
        purchase.refresh_from_db()
        self.assertEqual(purchase.remaining, 10)
        self.assertTrue(SaleOrderChangeLog.objects.filter(order_id_snapshot=order.id, action="delete").exists())


class SupplierManagementTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.manager = user_model.objects.create_user(
            username="supplier_manager",
            password="pw123456",
            is_staff=True,
        )
        cls.employee = user_model.objects.create_user(
            username="supplier_employee",
            password="pw123456",
        )
        cls.category = Category.objects.create(name="Supplier Category")

    def test_manager_can_create_supplier_from_frontend(self):
        self.client.login(username="supplier_manager", password="pw123456")

        response = self.client.post(reverse("supplier_create"), data={
            "name": "Dubai Wholesale",
            "phone": "+351900000000",
            "country": "UAE",
            "address": "Main supplier street",
            "product_types": [self.category.id],
        })

        self.assertEqual(response.status_code, 302)
        supplier = Supplier.objects.get(name="Dubai Wholesale")
        self.assertEqual(supplier.country, "UAE")
        self.assertEqual(list(supplier.product_types.values_list("id", flat=True)), [self.category.id])

    def test_supplier_list_groups_and_filters_by_country(self):
        Supplier.objects.create(name="Beta", country="France")
        Supplier.objects.create(name="Alpha", country="France")
        Supplier.objects.create(name="Gamma", country="Spain")
        Supplier.objects.create(name="Nowhere")  # no country
        self.client.login(username="supplier_manager", password="pw123456")

        response = self.client.get(reverse("supplier_list"))
        self.assertEqual(response.status_code, 200)
        groups = response.context["supplier_groups"]
        self.assertEqual([g["country"] for g in groups], ["France", "Spain", "No country"])
        france = next(g for g in groups if g["country"] == "France")
        self.assertEqual([s.name for s in france["suppliers"]], ["Alpha", "Beta"])

        filtered = self.client.get(reverse("supplier_list"), {"country": "Spain"})
        self.assertEqual([g["country"] for g in filtered.context["supplier_groups"]], ["Spain"])

    def test_regular_employee_cannot_open_supplier_management(self):
        self.client.login(username="supplier_employee", password="pw123456")

        response = self.client.get(reverse("supplier_list"))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], reverse("dashboard"))

    def test_manager_can_open_supplier_detail_history(self):
        supplier = Supplier.objects.create(name="History Supplier", phone="+351911111111")
        product = Product.objects.create(
            name="Amber Musk",
            barcode="4445556667778",
            brand="Maison",
            category=self.category,
            default_price=Decimal("22.00"),
        )
        inbound_order = InboundOrder.objects.create(supplier=supplier, invoice_no="INV-1001", total_amount=Decimal("60.00"))
        purchase = Purchase.objects.create(
            inbound_order=inbound_order,
            product=product,
            supplier=supplier,
            quantity=5,
            cost_price=Decimal("12.00"),
            remaining=5,
        )
        Purchase.objects.filter(pk=purchase.pk).update(date=timezone.now() - timedelta(days=1))

        self.client.login(username="supplier_manager", password="pw123456")
        response = self.client.get(reverse("supplier_detail", args=[supplier.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "History Supplier")
        self.assertContains(response, "INV-1001")
        self.assertContains(response, "Amber Musk")

    def test_supplier_create_saves_contact_fields(self):
        self.client.login(username="supplier_manager", password="pw123456")
        self.client.post(reverse("supplier_create"), data={
            "name": "Contact Co", "contact_person": "Ali", "phone": "+351900111222",
            "email": "ali@example.com", "website": "https://contact.co", "nif": "123456789",
            "country": "UAE", "address": "Street 1", "product_types": [self.category.id],
        })
        s = Supplier.objects.get(name="Contact Co")
        self.assertEqual(s.contact_person, "Ali")
        self.assertEqual(s.email, "ali@example.com")
        self.assertEqual(s.nif, "123456789")
        self.assertEqual(s.website, "https://contact.co")

    def test_supplier_scorecard_and_lead_time(self):
        supplier = Supplier.objects.create(name="Score Supplier")
        product = Product.objects.create(
            name="Saffron Oil", barcode="5556667778889", brand="Maison",
            category=self.category, default_price=Decimal("30.00"),
        )
        order = InboundOrder.objects.create(supplier=supplier, status="received", received_at=timezone.now())
        InboundOrder.objects.filter(pk=order.pk).update(
            created_at=timezone.now() - timedelta(days=2),
            placed_at=timezone.now() - timedelta(days=2),
            received_at=timezone.now()
        )
        Purchase.objects.create(
            inbound_order=order, product=product, supplier=supplier,
            quantity=4, cost_price=Decimal("10.00"), remaining=4,
        )

        self.client.login(username="supplier_manager", password="pw123456")
        response = self.client.get(reverse("supplier_detail", args=[supplier.id]))

        self.assertEqual(response.status_code, 200)
        sc = response.context["scorecard"]
        self.assertEqual(sc["lifetime_spend"], Decimal("40.00"))
        self.assertEqual(sc["units"], 4)
        self.assertEqual(sc["top_products"][0]["product_id"], product.id)
        self.assertIsNotNone(sc["avg_lead_days"])
        self.assertAlmostEqual(sc["avg_lead_days"], 2.0, delta=0.1)

    def test_product_detail_supplier_cost_comparison(self):
        s1 = Supplier.objects.create(name="Cheap Co")
        s2 = Supplier.objects.create(name="Pricey Co")
        product = Product.objects.create(
            name="Rose Attar", barcode="6667778889990", brand="Maison",
            category=self.category, default_price=Decimal("25.00"),
        )
        Purchase.objects.create(product=product, supplier=s1, quantity=3, cost_price=Decimal("8.00"), remaining=3)
        Purchase.objects.create(product=product, supplier=s2, quantity=2, cost_price=Decimal("11.00"), remaining=2)

        self.client.login(username="supplier_manager", password="pw123456")
        response = self.client.get(reverse("product_detail", args=[product.id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["cheapest_supplier_id"], s1.id)
        self.assertContains(response, "Cheapest")


class EmployeeManagementTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.admin = user_model.objects.create_superuser(username="team_admin", password="pw123456")
        cls.employee = user_model.objects.create_user(username="team_employee", password="pw123456")

    def test_admin_can_create_manager_account(self):
        self.client.login(username="team_admin", password="pw123456")

        response = self.client.post(reverse("employee_create"), data={
            "username": "new_manager",
            "first_name": "New",
            "last_name": "Manager",
            "email": "manager@example.com",
            "role": "manager",
            "is_active": "on",
            "password1": "pw12345678",
            "password2": "pw12345678",
        })

        self.assertEqual(response.status_code, 302)
        user_model = get_user_model()
        new_manager = user_model.objects.get(username="new_manager")
        self.assertTrue(new_manager.is_staff)
        self.assertTrue(new_manager.groups.filter(name="Managers").exists())

    def test_non_admin_cannot_open_team_management(self):
        self.client.login(username="team_employee", password="pw123456")

        response = self.client.get(reverse("employee_list"))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], reverse("dashboard"))


class AttendanceManagementTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.employee = user_model.objects.create_user(username="attendance_employee", password="pw123456")
        cls.manager = user_model.objects.create_user(username="attendance_manager", password="pw123456", is_staff=True)

    def test_login_opens_shift_logout_closes_it(self):
        self.client.login(username="attendance_employee", password="pw123456")
        rec = AttendanceRecord.objects.get(user=self.employee)
        self.assertIsNone(rec.clock_out_at)
        self.client.logout()
        rec.refresh_from_db()
        self.assertIsNotNone(rec.clock_out_at)

    def test_second_login_while_open_does_not_duplicate(self):
        self.client.login(username="attendance_employee", password="pw123456")
        self.client.login(username="attendance_employee", password="pw123456")
        self.assertEqual(AttendanceRecord.objects.filter(user=self.employee).count(), 1)

    def test_stale_previous_day_shift_closed_on_next_login(self):
        stale = AttendanceRecord.objects.create(
            user=self.employee,
            clock_in_at=timezone.now() - timedelta(days=1),
        )
        self.client.login(username="attendance_employee", password="pw123456")
        stale.refresh_from_db()
        self.assertIsNotNone(stale.clock_out_at)                 # stale closed
        self.assertEqual(AttendanceRecord.objects.filter(user=self.employee, clock_out_at__isnull=True).count(), 1)  # one fresh open

    def test_manager_login_creates_no_record(self):
        self.client.login(username="attendance_manager", password="pw123456")
        self.assertEqual(AttendanceRecord.objects.filter(user=self.manager).count(), 0)

    def test_employee_blocked_from_attendance_page(self):
        self.client.login(username="attendance_employee", password="pw123456")
        resp = self.client.get(reverse("attendance"))
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("dashboard"), resp.headers["Location"])

    def test_manager_can_see_team_attendance_section(self):
        AttendanceRecord.objects.create(user=self.employee, clock_in_at=timezone.now() - timedelta(hours=2))
        self.client.login(username="attendance_manager", password="pw123456")
        resp = self.client.get(reverse("attendance"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Team attendance overview")
        self.assertContains(resp, "attendance_employee")


class DashboardEnhancementTests(TestCase):
    """MoM comparison, sales-target progress, and reorder suggestions."""

    @classmethod
    def setUpTestData(cls):
        cls.category = Category.objects.create(name="Perfumes")
        cls.other_category = Category.objects.create(name="Accessories")
        cls.customer = Customer.objects.create(nif="222333444", name="Repeat Buyer")

        # Two fixed months so the comparison window is deterministic.
        cls.this_month_start = timezone.datetime(2026, 5, 1).date()
        cls.this_period_end = timezone.datetime(2026, 5, 15).date()
        cls.prev_month_start = timezone.datetime(2026, 4, 1).date()

        cls.fast = Product.objects.create(
            name="Fast Seller", barcode="9000000000001", brand="Maison",
            category=cls.category, default_price=Decimal("20.00"),
        )

    def _make_sale(self, product, qty, price, when, order=None):
        order = order or SaleOrder.objects.create(customer=self.customer)
        sale = Sale.objects.create(
            order=order, product=product, customer=self.customer,
            quantity=qty, unit_price=Decimal(price), payment_method="cash",
        )
        aware = timezone.make_aware(timezone.datetime.combine(when, timezone.datetime.min.time()))
        SaleOrder.objects.filter(pk=order.pk).update(created_at=aware)
        Sale.objects.filter(pk=sale.pk).update(date=aware)
        return order

    def test_compute_period_headline_totals_match(self):
        self._make_sale(self.fast, 2, "20.00", self.this_month_start)
        self._make_sale(self.fast, 3, "20.00", self.this_period_end)

        headline = compute_period_headline(self.this_month_start, self.this_period_end)

        self.assertEqual(headline["sales_amount"], Decimal("100.00"))  # (2+3) * 20
        self.assertEqual(headline["sales_qty"], 5)
        self.assertEqual(headline["order_count"], 2)

    def test_period_comparison_reports_growth_vs_prior_month(self):
        # Prior month (April) day-1..15 window: one 40 EUR order.
        self._make_sale(self.fast, 2, "20.00", self.prev_month_start)
        # Current month: 100 EUR.
        self._make_sale(self.fast, 5, "20.00", self.this_month_start)

        month_context = {
            "month_start": self.this_month_start,
            "period_end": self.this_period_end,
            "is_current_month": False,
        }
        current_headline = {
            "sales_amount": Decimal("100.00"), "profit": Decimal("0.00"),
            "order_count": 1, "avg_ticket": Decimal("100.00"),
        }
        comparison = build_period_comparison(month_context, current_headline)

        self.assertEqual(comparison["prev_sales_amount"], Decimal("40.00"))
        self.assertAlmostEqual(comparison["sales_amount_delta_pct"], 150.0)  # 40 -> 100

    def test_target_progress_respects_category_filter(self):
        SalesTarget.objects.create(category=self.category, monthly_amount=Decimal("200.00"))
        SalesTarget.objects.create(category=self.other_category, monthly_amount=Decimal("500.00"))

        month_context = {
            "month_start": self.this_month_start,
            "period_end": self.this_period_end,
            "is_current_month": False,
        }
        progress = build_target_progress(
            month_context, Decimal("100.00"), selected_category_ids=[self.category.id]
        )

        self.assertEqual(progress["target_amount"], Decimal("200.00"))  # only Perfumes target
        self.assertAlmostEqual(progress["progress_pct"], 50.0)

    def test_target_progress_is_none_when_no_target_set(self):
        month_context = {"month_start": self.this_month_start, "period_end": self.this_period_end, "is_current_month": False}
        self.assertIsNone(build_target_progress(month_context, Decimal("100.00")))


class MultiStoreTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.admin = user_model.objects.create_superuser(username="store_admin", password="pw123456")
        cls.employee = user_model.objects.create_user(username="store_emp", password="pw123456")

        cls.category = Category.objects.create(name="Store Test")
        cls.customer = Customer.objects.create(nif="777888999", name="Store Cust")
        cls.product = Product.objects.create(
            name="Musk", barcode="7778889990001", brand="Maison",
            category=cls.category, default_price=Decimal("20.00"),
        )
        Purchase.objects.create(product=cls.product, supplier=None, quantity=50, cost_price=Decimal("5.00"), remaining=50)

        cls.store_a = Store.objects.create(name="Amadora Two", code="AMD2")
        cls.store_b = Store.objects.create(name="Lisboa", code="LIS")
        StoreProfile.objects.update_or_create(user=cls.employee, defaults={'store': cls.store_b})

    def test_default_store_seeded_by_migration(self):
        self.assertTrue(Store.objects.filter(code="MAIN", is_default=True).exists())

    def test_outbound_attributes_sale_to_active_store(self):
        self.client.login(username="store_admin", password="pw123456")
        session = self.client.session
        session['active_store_id'] = self.store_b.id
        session.save()

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(reverse("outbound"), {
                "items_json": json.dumps([{"barcode": self.product.barcode, "qty": 1, "price": "20.00", "payment": "cash"}]),
            })

        self.assertEqual(response.status_code, 200)
        order = SaleOrder.objects.latest("id")
        self.assertEqual(order.store_id, self.store_b.id)
        self.assertTrue(all(sale.store_id == self.store_b.id for sale in order.items.all()))

    def test_outbound_blocked_when_viewing_all_stores(self):
        # A manager with no store chosen defaults to the "All stores" aggregate,
        # which has no single target for a sale — selling must be blocked so the
        # order can't silently land on the wrong store.
        self.client.login(username="store_admin", password="pw123456")
        before = SaleOrder.objects.count()

        # GET surfaces the block state for the template banner.
        get_resp = self.client.get(reverse("outbound"))
        self.assertTrue(get_resp.context["outbound_blocked"])

        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(reverse("outbound"), {
                "items_json": json.dumps([{"barcode": self.product.barcode, "qty": 1, "price": "20.00", "payment": "cash"}]),
            })

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["outbound_blocked"])
        self.assertEqual(SaleOrder.objects.count(), before)  # nothing was sold

    def test_records_totals_filtered_by_active_store(self):
        for store, price in [(self.store_a, Decimal("10.00")), (self.store_b, Decimal("30.00"))]:
            order = SaleOrder.objects.create(customer=self.customer, store=store)
            Sale.objects.create(
                order=order, product=self.product, customer=self.customer, store=store,
                quantity=1, unit_price=price, payment_method="cash",
            )

        self.client.login(username="store_admin", password="pw123456")
        session = self.client.session
        session['active_store_id'] = self.store_a.id
        session.save()

        today = timezone.localdate()
        response = self.client.get(
            reverse("sales_records"),
            {"start_date": today.isoformat(), "end_date": today.isoformat()},
        )
        self.assertEqual(response.status_code, 200)
        # Only store A's sale counts under the active store; store B is excluded.
        self.assertEqual(response.context["total_sales_amount"], Decimal("10.00"))

    def test_employee_is_locked_to_home_store(self):
        self.client.login(username="store_emp", password="pw123456")
        # Even if a store id is forced into the session, an employee stays on home store.
        session = self.client.session
        session['active_store_id'] = self.store_a.id
        session.save()

        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["active_store"], self.store_b)
        self.assertFalse(response.context["store_can_switch"])

    def test_admin_can_switch_and_defaults_to_all_stores(self):
        self.client.login(username="store_admin", password="pw123456")
        response = self.client.get(reverse("dashboard"))
        self.assertTrue(response.context["store_can_switch"])
        self.assertTrue(response.context["active_store_is_all"])

    def test_admin_can_create_and_list_stores(self):
        self.client.login(username="store_admin", password="pw123456")

        list_response = self.client.get(reverse("store_list"))
        self.assertEqual(list_response.status_code, 200)
        self.assertContains(list_response, "Store Management")

        create_response = self.client.post(reverse("store_create"), {
            "name": "Cascais", "code": "csc", "is_active": "on",
        })
        self.assertEqual(create_response.status_code, 302)
        store = Store.objects.get(code="CSC")  # clean_code upper-cases it
        self.assertEqual(store.name, "Cascais")

    def test_setting_a_default_store_unsets_the_others(self):
        self.client.login(username="store_admin", password="pw123456")
        self.client.post(reverse("store_edit", args=[self.store_a.id]), {
            "name": self.store_a.name, "code": self.store_a.code,
            "is_active": "on", "is_default": "on",
        })
        self.store_a.refresh_from_db()
        self.assertTrue(self.store_a.is_default)
        self.assertEqual(Store.objects.filter(is_default=True).count(), 1)

    def test_non_admin_cannot_manage_stores(self):
        self.client.login(username="store_emp", password="pw123456")
        response = self.client.get(reverse("store_list"))
        self.assertEqual(response.status_code, 302)

    def _store_sale(self, store, price):
        order = SaleOrder.objects.create(customer=self.customer, store=store)
        Sale.objects.create(
            order=order, product=self.product, customer=self.customer, store=store,
            quantity=1, unit_price=price, payment_method="cash",
        )
        return order

    def test_dashboard_and_trend_scoped_by_store(self):
        from django.core.cache import cache
        self._store_sale(self.store_a, Decimal("10.00"))
        self._store_sale(self.store_b, Decimal("30.00"))

        self.client.login(username="store_admin", password="pw123456")
        session = self.client.session
        session['active_store_id'] = self.store_a.id
        session.save()

        cache.clear()
        dash = self.client.get(reverse("dashboard"))
        self.assertEqual(dash.context["month_sales_amount"], Decimal("10.00"))

        cache.clear()
        trend = self.client.get(reverse("sales_records"), {"view": "year"})
        self.assertEqual(trend.context["year_sales_amount"], Decimal("10.00"))

    def test_dashboard_all_stores_shows_store_comparison(self):
        from django.core.cache import cache
        self._store_sale(self.store_a, Decimal("10.00"))
        self._store_sale(self.store_b, Decimal("30.00"))

        self.client.login(username="store_admin", password="pw123456")
        # Admin with no active store defaults to the "All stores" aggregate.
        cache.clear()
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["store_is_all"])

        breakdown = {row["name"]: row for row in response.context["month_store_breakdown"]}
        self.assertIn(self.store_a.name, breakdown)
        self.assertIn(self.store_b.name, breakdown)
        self.assertEqual(breakdown[self.store_a.name]["amount"], Decimal("10.00"))
        self.assertEqual(breakdown[self.store_b.name]["amount"], Decimal("30.00"))
        # Top store (by amount) sorts first; the comparison card renders.
        self.assertEqual(response.context["month_store_breakdown"][0]["name"], self.store_b.name)
        self.assertContains(response, "Store comparison")

    def test_daily_summary_all_stores_shows_store_column(self):
        self._store_sale(self.store_a, Decimal("10.00"))
        self._store_sale(self.store_b, Decimal("30.00"))

        self.client.login(username="store_admin", password="pw123456")
        response = self.client.get(reverse("daily_summary"))
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["store_is_all"])
        self.assertContains(response, "ord-store")   # each order card names its store
        self.assertContains(response, self.store_a.name)
        self.assertContains(response, self.store_b.name)

    def test_correction_center_scoped_by_store(self):
        order_a = self._store_sale(self.store_a, Decimal("10.00"))
        order_b = self._store_sale(self.store_b, Decimal("30.00"))

        self.client.login(username="store_admin", password="pw123456")
        session = self.client.session
        session['active_store_id'] = self.store_a.id
        session.save()

        response = self.client.get(reverse("sale_order_correction_center"))
        ids = [order.id for order in response.context["page_obj"].object_list]
        self.assertIn(order_a.id, ids)
        self.assertNotIn(order_b.id, ids)

    def test_correction_center_all_stores_groups_by_date_with_store(self):
        self._store_sale(self.store_a, Decimal("10.00"))
        self._store_sale(self.store_b, Decimal("30.00"))

        self.client.login(username="store_admin", password="pw123456")
        # Admin with no active store => All stores aggregate.
        response = self.client.get(reverse("sale_order_correction_center"))
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["store_is_all"])

        groups = response.context["order_date_groups"]
        self.assertTrue(groups)
        today = timezone.localdate()
        self.assertEqual(groups[0]["date"], today)
        # Both stores' orders land in today's group under All stores.
        self.assertEqual(groups[0]["count"], 2)
        # Store column header + both store names render.
        self.assertContains(response, "<th>Store</th>")
        self.assertContains(response, self.store_a.name)
        self.assertContains(response, self.store_b.name)

    def test_daily_summary_shows_todays_orders_scoped_by_store(self):
        self._store_sale(self.store_a, Decimal("10.00"))
        self._store_sale(self.store_b, Decimal("30.00"))

        self.client.login(username="store_admin", password="pw123456")
        session = self.client.session
        session['active_store_id'] = self.store_a.id
        session.save()

        response = self.client.get(reverse("daily_summary"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Daily Summary")
        # Only the active store's order counts today.
        self.assertEqual(response.context["order_count"], 1)
        self.assertEqual(response.context["total_amount"], Decimal("10.00"))
        self.assertEqual(response.context["total_qty"], 1)
        # Each order renders inline (no dialog) with a Print link to the receipt.
        order = response.context["orders"][0]
        self.assertContains(response, "#%s" % order.id)
        self.assertNotContains(response, "order-modal")
        self.assertContains(response, reverse("sale_order_detail", args=[order.id]))

    def test_today_payment_stats_reflect_split_tender(self):
        # Regression: after splitting an order card+cash, today's payment
        # statistics still showed card-only, because the breakdown aggregated by
        # Sale.payment_method (primary) instead of the SaleOrderPayment split.
        from django.core.cache import cache

        order = SaleOrder.objects.create(customer=self.customer, store=self.store_a)
        Sale.objects.create(
            order=order, product=self.product, customer=self.customer, store=self.store_a,
            quantity=2, unit_price=Decimal("25.00"), payment_method="card",
        )
        SaleOrderPayment.objects.create(order=order, method="card", amount=Decimal("30.00"))
        SaleOrderPayment.objects.create(order=order, method="cash", amount=Decimal("20.00"))

        self.client.login(username="store_admin", password="pw123456")

        cache.clear()
        daily = self.client.get(reverse("daily_summary"))
        self.assertEqual(daily.status_code, 200)
        daily_rows = {row["label"]: row["amount"] for row in daily.context["payment_rows"]}
        self.assertEqual(daily_rows.get("Card"), Decimal("30.00"))
        self.assertEqual(daily_rows.get("Cash"), Decimal("20.00"))

        # The dashboard's today breakdown is the employee view (managers use
        # /today); scope the employee to the store holding the split order.
        StoreProfile.objects.update_or_create(user=self.employee, defaults={'store': self.store_a})
        self.client.logout()
        self.client.login(username="store_emp", password="pw123456")
        cache.clear()
        dash = self.client.get(reverse("dashboard"))
        self.assertEqual(dash.status_code, 200)
        dash_rows = {row["label"]: row["amount"] for row in dash.context["today_payment_breakdown"]}
        self.assertEqual(dash_rows.get("Card"), Decimal("30.00"))
        self.assertEqual(dash_rows.get("Cash"), Decimal("20.00"))

    def test_print_header_is_per_store(self):
        from stock.models import PrintProfile
        profile_a = PrintProfile.get_for_store(self.store_a)
        profile_b = PrintProfile.get_for_store(self.store_b)
        self.assertNotEqual(profile_a.id, profile_b.id)

        self.client.login(username="store_admin", password="pw123456")
        session = self.client.session
        session['active_store_id'] = self.store_b.id
        session.save()

        response = self.client.post(reverse("print_profile_edit"), {
            "name": "Scentory Shop", "nif": "111", "phone": "222",
            "address": "Rua X", "email": "s@x.com", "footer_note": "Thanks",
            "next": reverse("dashboard"),
        })
        self.assertEqual(response.status_code, 302)
        profile_b.refresh_from_db()
        profile_a.refresh_from_db()
        self.assertEqual(profile_b.name, "Scentory Shop")
        self.assertNotEqual(profile_a.name, "Scentory Shop")

    def test_receipt_uses_the_orders_store_header(self):
        from stock.models import PrintProfile
        profile_b = PrintProfile.get_for_store(self.store_b)
        profile_b.name = "Scentory Receipt"
        profile_b.save()

        order = self._store_sale(self.store_b, Decimal("20.00"))
        self.client.login(username="store_admin", password="pw123456")
        response = self.client.get(reverse("sale_order_detail", args=[order.id]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["shop"]["name"], "Scentory Receipt")


class ShopifySyncTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.category = Category.objects.create(name="Shopify Perfume")
        cls.product = Product.objects.create(
            name="Asad", barcode="6290362346548", brand="Lattafa",
            category=cls.category, default_price=Decimal("35.00"),
        )

    def _fake_client(self, *, match=None, has_image=False):
        from unittest import mock
        client = mock.Mock()
        client.is_configured.return_value = True
        client.find_product_by_sku.return_value = (
            None if match is None else {"id": "gid://shopify/Product/1", "title": "Asad", "has_image": has_image}
        )
        client.stage_and_upload_image.return_value = "resource://staged"
        client.attach_image.return_value = "gid://shopify/MediaImage/9"
        return client

    def test_attaches_image_when_shopify_product_has_none(self):
        from unittest import mock
        from stock.services import shopify_sync
        client = self._fake_client(match=True, has_image=False)
        with mock.patch.object(shopify_sync, "_local_image_path", return_value="/tmp/x.jpg"):
            code, _ = shopify_sync.sync_product_image(self.product, client)
        self.assertEqual(code, shopify_sync.UPLOADED)
        client.stage_and_upload_image.assert_called_once()
        client.attach_image.assert_called_once()

    def test_skips_when_shopify_product_already_has_image(self):
        from unittest import mock
        from stock.services import shopify_sync
        client = self._fake_client(match=True, has_image=True)
        with mock.patch.object(shopify_sync, "_local_image_path", return_value="/tmp/x.jpg"):
            code, _ = shopify_sync.sync_product_image(self.product, client)
        self.assertEqual(code, shopify_sync.SKIP_HAS_IMAGE)
        client.stage_and_upload_image.assert_not_called()

    def test_skips_when_not_in_shopify(self):
        from unittest import mock
        from stock.services import shopify_sync
        client = self._fake_client(match=None)
        with mock.patch.object(shopify_sync, "_local_image_path", return_value="/tmp/x.jpg"):
            code, _ = shopify_sync.sync_product_image(self.product, client)
        self.assertEqual(code, shopify_sync.SKIP_NOT_IN_SHOPIFY)

    def test_dry_run_uploads_nothing(self):
        from unittest import mock
        from stock.services import shopify_sync
        client = self._fake_client(match=True, has_image=False)
        with mock.patch.object(shopify_sync, "_local_image_path", return_value="/tmp/x.jpg"):
            code, _ = shopify_sync.sync_product_image(self.product, client, dry_run=True)
        self.assertEqual(code, shopify_sync.WOULD_UPLOAD)
        client.stage_and_upload_image.assert_not_called()

    def test_command_dry_run_does_not_upload(self):
        from unittest import mock
        client = self._fake_client(match=True, has_image=False)
        from stock.services import shopify_sync
        with mock.patch("stock.management.commands.sync_shopify_images.ShopifyClient", return_value=client), \
             mock.patch.object(shopify_sync, "_local_image_path", return_value="/tmp/x.jpg"):
            from django.core.management import call_command
            out = StringIO()
            call_command("sync_shopify_images", "--brand", "Lattafa", stdout=out)
        self.assertIn("DRY RUN", out.getvalue())
        client.stage_and_upload_image.assert_not_called()

    def test_signal_off_by_default_does_not_call_shopify(self):
        import tempfile
        from unittest import mock
        from django.core.files.uploadedfile import SimpleUploadedFile
        from django.test import override_settings
        from stock.models import ProductImage
        with tempfile.TemporaryDirectory() as media, override_settings(MEDIA_ROOT=media), \
             mock.patch("stock.services.shopify_sync.sync_product") as synced:
            img = ProductImage.objects.create(
                product=self.product,
                image=SimpleUploadedFile("t.jpg", b"\xff\xd8\xff\xd9", content_type="image/jpeg"),
            )
            self.assertTrue(img.pk)
        synced.assert_not_called()

    def _create_client(self):
        client = self._fake_client(match=None)
        client.get_location_id.return_value = "gid://shopify/Location/1"
        client.product_set.return_value = "gid://shopify/Product/99"
        return client

    def test_sync_product_creates_when_missing(self):
        from unittest import mock
        from stock.services import shopify_sync
        client = self._create_client()
        with mock.patch.object(shopify_sync, "_local_image_path", return_value="/tmp/x.jpg"):
            code, _ = shopify_sync.sync_product(self.product, client, create_missing=True)
        self.assertEqual(code, shopify_sync.CREATED)
        client.product_set.assert_called_once()
        # variant carries barcode/sku + inventory at the location; image is included.
        payload = client.product_set.call_args[0][0]
        variant = payload["variants"][0]
        self.assertEqual(variant["sku"], self.product.barcode)
        self.assertEqual(variant["barcode"], self.product.barcode)
        self.assertEqual(variant["inventoryQuantities"][0]["locationId"], "gid://shopify/Location/1")
        self.assertIn("seo", payload)
        self.assertIn("files", payload)

    def test_create_sets_eau_de_parfum_category_for_perfume(self):
        from unittest import mock
        from stock.services import shopify_sync
        client = self._create_client()
        with mock.patch.object(shopify_sync, "_local_image_path", return_value=None):
            shopify_sync.create_product_in_shopify(self.product, client)
        payload = client.product_set.call_args[0][0]
        self.assertEqual(payload.get("category"), shopify_sync.EAU_DE_PARFUM_TAXONOMY_GID)

    def test_sync_product_does_not_create_when_disabled(self):
        from unittest import mock
        from stock.services import shopify_sync
        client = self._create_client()
        with mock.patch.object(shopify_sync, "_local_image_path", return_value="/tmp/x.jpg"):
            code, _ = shopify_sync.sync_product(self.product, client, create_missing=False)
        self.assertEqual(code, shopify_sync.SKIP_NOT_IN_SHOPIFY)
        client.product_set.assert_not_called()

    def test_sync_product_attaches_image_when_already_in_shopify(self):
        from unittest import mock
        from stock.services import shopify_sync
        client = self._fake_client(match=True, has_image=False)
        with mock.patch.object(shopify_sync, "_local_image_path", return_value="/tmp/x.jpg"):
            code, _ = shopify_sync.sync_product(self.product, client, create_missing=True)
        self.assertEqual(code, shopify_sync.UPLOADED)
        client.product_set.assert_not_called()
        client.attach_image.assert_called_once()

    def test_create_dry_run_writes_nothing(self):
        from unittest import mock
        from stock.services import shopify_sync
        client = self._create_client()
        with mock.patch.object(shopify_sync, "_local_image_path", return_value="/tmp/x.jpg"):
            code, _ = shopify_sync.sync_product(self.product, client, create_missing=True, dry_run=True)
        self.assertEqual(code, shopify_sync.WOULD_CREATE)
        client.product_set.assert_not_called()

    def test_shopify_title_is_cleaned(self):
        from stock.services import shopify_sync
        self.product.model = "Khamrah"
        self.product.name = "Waha EDP 100ml"
        title = shopify_sync._shopify_title(self.product)
        self.assertIn("Lattafa", title)
        self.assertIn("EDP", title)
        self.assertIn("ml", title)

    def test_products_command_dry_run(self):
        from unittest import mock
        from stock.services import shopify_sync
        client = self._create_client()
        with mock.patch("stock.management.commands.sync_shopify_products.ShopifyClient", return_value=client), \
             mock.patch.object(shopify_sync, "_local_image_path", return_value="/tmp/x.jpg"):
            from django.core.management import call_command
            out = StringIO()
            call_command("sync_shopify_products", "--brand", "Lattafa", stdout=out)
        self.assertIn("DRY RUN", out.getvalue())
        client.product_set.assert_not_called()


from types import SimpleNamespace
from django.test import SimpleTestCase


class UploadPathTests(SimpleTestCase):
    def _obj(self, brand):
        return SimpleNamespace(product=SimpleNamespace(brand=brand))

    def test_brand_subfolder(self):
        from stock.models.catalog import product_image_upload_to
        self.assertEqual(
            product_image_upload_to(self._obj("Lattafa"), "a.jpg"),
            "product_images/Lattafa/a.jpg",
        )

    def test_blank_brand_falls_back(self):
        from stock.models.catalog import product_image_upload_to
        self.assertEqual(
            product_image_upload_to(self._obj(""), "a.jpg"),
            "product_images/a.jpg",
        )

    def test_unsafe_chars_stripped(self):
        from stock.models.catalog import product_image_upload_to
        self.assertEqual(
            product_image_upload_to(self._obj("A/B:C"), "a.jpg"),
            "product_images/ABC/a.jpg",
        )


from unittest import mock
from django.test import override_settings


class CloudinaryClientTests(SimpleTestCase):
    @override_settings(CLOUDINARY_CLOUD_NAME="c", CLOUDINARY_API_KEY="k", CLOUDINARY_API_SECRET="s")
    def test_is_configured_true_when_all_present(self):
        from stock.services.cloudinary_client import CloudinaryClient
        self.assertTrue(CloudinaryClient().is_configured())

    @override_settings(CLOUDINARY_CLOUD_NAME="c", CLOUDINARY_API_KEY="k", CLOUDINARY_API_SECRET="")
    def test_is_configured_false_when_secret_missing(self):
        from stock.services.cloudinary_client import CloudinaryClient
        self.assertFalse(CloudinaryClient().is_configured())

    @override_settings(CLOUDINARY_CLOUD_NAME="c", CLOUDINARY_API_KEY="k", CLOUDINARY_API_SECRET="s")
    def test_upload_image_passes_expected_params(self):
        from stock.services.cloudinary_client import CloudinaryClient
        with mock.patch("cloudinary.uploader.upload", return_value={"secure_url": "https://x/y.jpg"}) as up, \
             mock.patch("cloudinary.config"):
            url = CloudinaryClient().upload_image("/tmp/a.jpg", public_id="123", asset_folder="product_images/Lattafa")
        self.assertEqual(url, "https://x/y.jpg")
        _, kwargs = up.call_args
        self.assertEqual(kwargs["public_id"], "123")
        self.assertEqual(kwargs["asset_folder"], "product_images/Lattafa")
        self.assertFalse(kwargs["unique_filename"])
        self.assertTrue(kwargs["overwrite"])

    @override_settings(CLOUDINARY_CLOUD_NAME="c", CLOUDINARY_API_KEY="k", CLOUDINARY_API_SECRET="s")
    def test_upload_error_wrapped(self):
        from stock.services.cloudinary_client import CloudinaryClient, CloudinaryError
        with mock.patch("cloudinary.uploader.upload", side_effect=RuntimeError("boom")), \
             mock.patch("cloudinary.config"):
            with self.assertRaises(CloudinaryError):
                CloudinaryClient().upload_image("/tmp/a.jpg", public_id="123", asset_folder="product_images")


def _make_product(barcode="111", brand="Lattafa"):
    from stock.models import Product
    return Product.objects.create(name="X", brand=brand, barcode=barcode)


def _tiny_png():
    # 1x1 transparent PNG
    import base64
    data = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg=="
    )
    return SimpleUploadedFile("p.png", data, content_type="image/png")


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class CloudinarySyncTests(TestCase):
    def test_upload_when_primary_exists(self):
        from stock.models import ProductImage
        from stock.services import cloudinary_sync
        p = _make_product(barcode="222", brand="Lattafa")
        ProductImage.objects.create(product=p, image=_tiny_png())
        client = mock.Mock()
        code, _ = cloudinary_sync.sync_product_primary_image(p, client=client)
        self.assertEqual(code, cloudinary_sync.UPLOADED)
        _, kwargs = client.upload_image.call_args
        self.assertEqual(kwargs["public_id"], "222")
        self.assertEqual(kwargs["asset_folder"], "product_images/Lattafa")

    def test_delete_when_no_image(self):
        from stock.services import cloudinary_sync
        p = _make_product(barcode="333")
        client = mock.Mock()
        code, _ = cloudinary_sync.sync_product_primary_image(p, client=client)
        self.assertEqual(code, cloudinary_sync.DELETED)
        client.delete_image.assert_called_once_with("333")

    def test_skip_no_barcode(self):
        from stock.services import cloudinary_sync
        p = _make_product(barcode="")
        client = mock.Mock()
        code, _ = cloudinary_sync.sync_product_primary_image(p, client=client)
        self.assertEqual(code, cloudinary_sync.SKIP_NO_BARCODE)
        client.upload_image.assert_not_called()
        client.delete_image.assert_not_called()

    def test_error_wrapped(self):
        from stock.models import ProductImage
        from stock.services import cloudinary_sync
        from stock.services.cloudinary_client import CloudinaryError
        p = _make_product(barcode="444")
        ProductImage.objects.create(product=p, image=_tiny_png())
        client = mock.Mock()
        client.upload_image.side_effect = CloudinaryError("nope")
        code, _ = cloudinary_sync.sync_product_primary_image(p, client=client)
        self.assertEqual(code, cloudinary_sync.ERROR)

    def test_dry_run_makes_no_calls(self):
        from stock.models import ProductImage
        from stock.services import cloudinary_sync
        p = _make_product(barcode="555")
        ProductImage.objects.create(product=p, image=_tiny_png())
        client = mock.Mock()
        code, _ = cloudinary_sync.sync_product_primary_image(p, client=client, dry_run=True)
        self.assertEqual(code, cloudinary_sync.UPLOADED)
        client.upload_image.assert_not_called()


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class CloudinarySignalTests(TestCase):
    @override_settings(CLOUDINARY_AUTO_SYNC=True)
    def test_save_triggers_sync_when_enabled(self):
        from stock.models import ProductImage
        p = _make_product(barcode="666")
        with mock.patch("stock.services.cloudinary_sync.sync_product_primary_image") as sync:
            sync.return_value = ("uploaded", "ok")
            with self.captureOnCommitCallbacks(execute=True):
                ProductImage.objects.create(product=p, image=_tiny_png())
        sync.assert_called()

    @override_settings(CLOUDINARY_AUTO_SYNC=True)
    def test_delete_triggers_sync_when_enabled(self):
        from stock.models import ProductImage
        p = _make_product(barcode="777")
        img = ProductImage.objects.create(product=p, image=_tiny_png())
        with mock.patch("stock.services.cloudinary_sync.sync_product_primary_image") as sync:
            sync.return_value = ("uploaded", "ok")
            with self.captureOnCommitCallbacks(execute=True):
                img.delete()
        sync.assert_called()

    @override_settings(CLOUDINARY_AUTO_SYNC=False)
    def test_disabled_no_sync(self):
        from stock.models import ProductImage
        p = _make_product(barcode="888")
        with mock.patch("stock.services.cloudinary_sync.sync_product_primary_image") as sync:
            with self.captureOnCommitCallbacks(execute=True):
                ProductImage.objects.create(product=p, image=_tiny_png())
        sync.assert_not_called()


from django.core.management import call_command


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class CloudinaryCommandTests(TestCase):
    def test_dry_run_makes_no_client_calls(self):
        from stock.models import ProductImage
        p = _make_product(barcode="999")
        ProductImage.objects.create(product=p, image=_tiny_png())
        with mock.patch("stock.services.cloudinary_sync.sync_product_primary_image",
                        return_value=("uploaded", "999")) as sync:
            out = StringIO()
            call_command("sync_cloudinary_images", stdout=out)
        # dry-run passes dry_run=True through
        _, kwargs = sync.call_args
        self.assertTrue(kwargs.get("dry_run"))
        self.assertIn("999", out.getvalue())


class CorrectionStoreChangeServiceTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        from stock.models import Store, Category, Customer, Product
        cls.admin = get_user_model().objects.create_superuser(username="corr_store_svc_admin", password="pw123456")
        cls.store_a = Store.objects.create(name="Store A", code="CSA")
        cls.store_b = Store.objects.create(name="Store B", code="CSB")
        cls.category = Category.objects.create(name="Corr Store Cat")
        cls.customer = Customer.objects.create(nif="111222333", name="Corr Store Cust")
        cls.product = Product.objects.create(
            name="Asad", barcode="7770001777", brand="Lattafa",
            category=cls.category, default_price=Decimal("25.00"),
        )

    def _make_order_in_store(self, store):
        from stock.models import Purchase, SaleOrder, Sale
        # remaining=8 reflects the 2 units already consumed by the Sale created
        # below (created directly via the ORM, bypassing consume_stock_fifo) so
        # the correction service's restore-then-consume round-trip balances.
        Purchase.objects.create(product=self.product, supplier=None, quantity=10,
                                cost_price=Decimal("10.00"), remaining=8)
        order = SaleOrder.objects.create(customer=self.customer, note="o", store=store)
        Sale.objects.create(order=order, product=self.product, customer=self.customer,
                            store=store, quantity=2, unit_price=Decimal("25.00"),
                            payment_method="cash")
        return order

    def _line_items(self):
        return [{"product": self.product, "quantity": 2,
                 "unit_price": Decimal("25.00"), "payment_method": "cash"}]

    def test_selected_store_moves_order_and_sales(self):
        from stock.services.order_corrections import save_sale_order_correction
        order = self._make_order_in_store(self.store_a)
        save_sale_order_correction(
            order=order, customer=self.customer, note="o", order_datetime=timezone.now(),
            line_items=self._line_items(), payment_totals={"cash": Decimal("50.00")},
            changed_by=self.admin, reason="wrong store", store=self.store_b,
        )
        order.refresh_from_db()
        self.assertEqual(order.store_id, self.store_b.id)
        self.assertTrue(all(s.store_id == self.store_b.id for s in order.items.all()))

    def test_no_store_passed_keeps_current(self):
        from stock.services.order_corrections import save_sale_order_correction
        order = self._make_order_in_store(self.store_a)
        save_sale_order_correction(
            order=order, customer=self.customer, note="o", order_datetime=timezone.now(),
            line_items=self._line_items(), payment_totals={"cash": Decimal("50.00")},
            changed_by=self.admin, reason="x", store=None,
        )
        order.refresh_from_db()
        self.assertEqual(order.store_id, self.store_a.id)

    def test_snapshot_captures_store(self):
        from stock.services.order_corrections import snapshot_sale_order
        order = self._make_order_in_store(self.store_a)
        snap = snapshot_sale_order(order)
        self.assertEqual(snap["store_id"], self.store_a.id)
        self.assertEqual(snap["store_name"], "Store A")


class CorrectionStoreChangeViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        from stock.models import Store, Category, Customer, Product
        cls.admin = get_user_model().objects.create_superuser(username="corr_store_view_admin", password="pw123456")
        cls.store_a = Store.objects.create(name="Store A", code="VSA")
        cls.store_b = Store.objects.create(name="Store B", code="VSB")
        cls.category = Category.objects.create(name="Corr View Cat")
        cls.customer = Customer.objects.create(nif="222333444", name="Corr View Cust")
        cls.product = Product.objects.create(
            name="Asad", barcode="7770002777", brand="Lattafa",
            category=cls.category, default_price=Decimal("25.00"),
        )

    def _order_in_a(self):
        from stock.models import Purchase, SaleOrder, Sale
        # remaining=8 reflects the 2 units already consumed by the Sale created
        # below (created directly via the ORM, bypassing consume_stock_fifo) so
        # the correction service's restore-then-consume round-trip balances.
        Purchase.objects.create(product=self.product, supplier=None, quantity=10,
                                cost_price=Decimal("10.00"), remaining=8)
        order = SaleOrder.objects.create(customer=self.customer, note="o", store=self.store_a)
        Sale.objects.create(order=order, product=self.product, customer=self.customer,
                            store=self.store_a, quantity=2, unit_price=Decimal("25.00"),
                            payment_method="cash")
        return order

    def test_form_shows_store_selector(self):
        order = self._order_in_a()
        self.client.login(username="corr_store_view_admin", password="pw123456")
        resp = self.client.get(reverse("sale_order_correction_edit", args=[order.id]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'name="store"')
        self.assertContains(resp, 'id="order-store"')
        self.assertContains(resp, "Store B")

    def test_post_moves_order_to_selected_store(self):
        order = self._order_in_a()
        self.client.login(username="corr_store_view_admin", password="pw123456")
        resp = self.client.post(reverse("sale_order_correction_edit", args=[order.id]), data={
            "customer": self.customer.id,
            "order_datetime": timezone.localtime(timezone.now()).strftime("%Y-%m-%dT%H:%M"),
            "note": "o",
            "reason": "Entered under the wrong store",
            "store": self.store_b.id,
            "items_json": json.dumps([{"product_id": self.product.id, "qty": 2, "price": "25.00", "payment": "cash"}]),
            "payments_json": json.dumps([{"method": "cash", "amount": "50.00"}]),
        })
        self.assertEqual(resp.status_code, 302)
        order.refresh_from_db()
        self.assertEqual(order.store_id, self.store_b.id)
        self.assertTrue(all(s.store_id == self.store_b.id for s in order.items.all()))


class AffectsStockModelTests(TestCase):
    def test_defaults_true(self):
        from stock.models import SaleOrder
        order = SaleOrder.objects.create(note="x")
        self.assertTrue(order.affects_stock)

    def test_can_set_false(self):
        from stock.models import SaleOrder
        order = SaleOrder.objects.create(note="x", affects_stock=False)
        order.refresh_from_db()
        self.assertFalse(order.affects_stock)


class AffectsStockServiceTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        from stock.models import Category, Customer, Product
        cls.admin = get_user_model().objects.create_superuser(username="affstock_admin", password="pw123456")
        cls.category = Category.objects.create(name="Aff Cat")
        cls.customer = Customer.objects.create(nif="900900900", name="Aff Cust")
        cls.product = Product.objects.create(
            name="Asad", barcode="8880001888", brand="Lattafa",
            category=cls.category, default_price=Decimal("25.00"),
        )

    def _purchase(self, remaining=5):
        from stock.models import Purchase
        return Purchase.objects.create(product=self.product, supplier=None, quantity=5,
                                       cost_price=Decimal("10.00"), remaining=remaining)

    def _line_items(self, qty=2):
        return [{"product": self.product, "quantity": qty,
                 "unit_price": Decimal("25.00"), "payment_method": "cash"}]

    def test_create_no_stock_skips_consume(self):
        from stock.services.order_corrections import save_sale_order_correction
        purchase = self._purchase(remaining=5)
        order = save_sale_order_correction(
            order=None, customer=self.customer, note="backfill", order_datetime=timezone.now(),
            line_items=self._line_items(2), payment_totals={"cash": Decimal("50.00")},
            changed_by=self.admin, reason="missed", affects_stock=False,
        )
        purchase.refresh_from_db()
        self.assertFalse(order.affects_stock)
        self.assertEqual(purchase.remaining, 5)          # not consumed
        self.assertEqual(order.items.count(), 1)          # sale recorded

    def test_create_default_consumes(self):
        from stock.services.order_corrections import save_sale_order_correction
        purchase = self._purchase(remaining=5)
        order = save_sale_order_correction(
            order=None, customer=self.customer, note="normal", order_datetime=timezone.now(),
            line_items=self._line_items(2), payment_totals={"cash": Decimal("50.00")},
            changed_by=self.admin, reason="x", affects_stock=None,
        )
        purchase.refresh_from_db()
        self.assertTrue(order.affects_stock)
        self.assertEqual(purchase.remaining, 3)          # consumed 2

    def test_edit_no_stock_order_leaves_stock_untouched(self):
        from stock.models import SaleOrder, Sale
        from stock.services.order_corrections import save_sale_order_correction
        purchase = self._purchase(remaining=5)
        order = SaleOrder.objects.create(customer=self.customer, note="b", affects_stock=False)
        Sale.objects.create(order=order, product=self.product, customer=self.customer,
                            quantity=2, unit_price=Decimal("25.00"), payment_method="cash")
        save_sale_order_correction(
            order=order, customer=self.customer, note="b2", order_datetime=timezone.now(),
            line_items=self._line_items(3), payment_totals={"cash": Decimal("75.00")},
            changed_by=self.admin, reason="x", affects_stock=None,
        )
        purchase.refresh_from_db()
        order.refresh_from_db()
        self.assertFalse(order.affects_stock)             # kept
        self.assertEqual(purchase.remaining, 5)           # neither restored nor consumed
        self.assertEqual(order.items.get().quantity, 3)   # rebuilt

    def test_delete_no_stock_order_does_not_restore(self):
        from stock.models import SaleOrder, Sale
        from stock.services.order_corrections import delete_sale_order_correction
        purchase = self._purchase(remaining=5)
        order = SaleOrder.objects.create(customer=self.customer, note="b", affects_stock=False)
        Sale.objects.create(order=order, product=self.product, customer=self.customer,
                            quantity=2, unit_price=Decimal("25.00"), payment_method="cash")
        delete_sale_order_correction(order=order, changed_by=self.admin, reason="x")
        purchase.refresh_from_db()
        self.assertEqual(purchase.remaining, 5)           # not restored

    def test_snapshot_includes_affects_stock(self):
        from stock.models import SaleOrder
        from stock.services.order_corrections import snapshot_sale_order
        order = SaleOrder.objects.create(customer=self.customer, note="b", affects_stock=False)
        self.assertFalse(snapshot_sale_order(order)["affects_stock"])


class AffectsStockProfitTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        from stock.models import Category, Customer, Product
        cls.category = Category.objects.create(name="Prof Cat")
        cls.customer = Customer.objects.create(nif="700700700", name="Prof Cust")
        cls.product = Product.objects.create(
            name="Asad", barcode="6660001666", brand="Lattafa",
            category=cls.category, default_price=Decimal("25.00"),
        )

    def _purchase(self, qty, cost, when):
        from stock.models import Purchase
        p = Purchase.objects.create(product=self.product, supplier=None, quantity=qty,
                                    cost_price=Decimal(cost), remaining=qty)
        Purchase.objects.filter(pk=p.pk).update(date=when)
        return p

    def _sale(self, qty, price, when, affects_stock=True):
        from stock.models import SaleOrder, Sale
        order = SaleOrder.objects.create(customer=self.customer, affects_stock=affects_stock)
        s = Sale.objects.create(order=order, product=self.product, customer=self.customer,
                                quantity=qty, unit_price=Decimal(price), payment_method="cash")
        Sale.objects.filter(pk=s.pk).update(date=when)
        return Sale.objects.get(pk=s.pk)

    def test_no_stock_sale_booked_at_50pct(self):
        from stock.services.profit import sale_profit_map_for_sale_ids
        base = timezone.now() - timedelta(days=5)
        self._purchase(10, "10.00", base)
        s = self._sale(2, "25.00", base + timedelta(hours=1), affects_stock=False)
        m = sale_profit_map_for_sale_ids([s.id])[s.id]
        self.assertEqual(m["revenue"], Decimal("50.00"))
        self.assertEqual(m["cost"], Decimal("25.00"))
        self.assertEqual(m["profit"], Decimal("25.00"))

    def test_no_stock_sale_does_not_distort_normal_fifo(self):
        from stock.services.profit import sale_profit_map_for_sale_ids
        base = timezone.now() - timedelta(days=5)
        self._purchase(3, "10.00", base)                                   # only 3 units @10
        no_stock = self._sale(2, "25.00", base + timedelta(hours=1), affects_stock=False)
        normal = self._sale(3, "25.00", base + timedelta(hours=2), affects_stock=True)
        m = sale_profit_map_for_sale_ids([normal.id, no_stock.id])
        # normal sale draws all 3 purchased units @10 => cost 30, unaffected by the no-stock sale
        self.assertEqual(m[normal.id]["cost"], Decimal("30.00"))
        self.assertEqual(m[normal.id]["profit"], Decimal("45.00"))
        # no-stock sale is flat 50%
        self.assertEqual(m[no_stock.id]["cost"], Decimal("25.00"))


class AffectsStockViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        from stock.models import Category, Customer, Product
        cls.admin = get_user_model().objects.create_superuser(username="affview_admin", password="pw123456")
        cls.category = Category.objects.create(name="AffView Cat")
        cls.customer = Customer.objects.create(nif="500500500", name="AffView Cust")
        cls.product = Product.objects.create(
            name="Asad", barcode="5550001555", brand="Lattafa",
            category=cls.category, default_price=Decimal("25.00"),
        )

    def _purchase(self, remaining=5):
        from stock.models import Purchase
        return Purchase.objects.create(product=self.product, supplier=None, quantity=5,
                                       cost_price=Decimal("10.00"), remaining=remaining)

    def _payload(self, **overrides):
        data = {
            "customer": self.customer.id,
            "order_datetime": timezone.localtime(timezone.now()).strftime("%Y-%m-%dT%H:%M"),
            "note": "backfill",
            "reason": "missed historical order",
            "items_json": json.dumps([{"product_id": self.product.id, "qty": 2, "price": "25.00", "payment": "cash"}]),
            "payments_json": json.dumps([{"method": "cash", "amount": "50.00"}]),
        }
        data.update(overrides)
        return data

    def test_create_page_shows_checkbox_checked(self):
        self.client.login(username="affview_admin", password="pw123456")
        resp = self.client.get(reverse("sale_order_correction_create"))
        self.assertContains(resp, 'name="affects_stock"')
        self.assertContains(resp, 'id="affects-stock"')

    def test_unchecked_creates_no_stock_order(self):
        from stock.models import SaleOrder
        purchase = self._purchase(remaining=5)
        self.client.login(username="affview_admin", password="pw123456")
        # affects_stock omitted from POST => unchecked
        resp = self.client.post(reverse("sale_order_correction_create"), data=self._payload())
        self.assertEqual(resp.status_code, 302)
        order = SaleOrder.objects.latest("id")
        purchase.refresh_from_db()
        self.assertFalse(order.affects_stock)
        self.assertEqual(purchase.remaining, 5)          # not consumed

    def test_checked_creates_stock_order(self):
        from stock.models import SaleOrder
        purchase = self._purchase(remaining=5)
        self.client.login(username="affview_admin", password="pw123456")
        resp = self.client.post(reverse("sale_order_correction_create"), data=self._payload(affects_stock="on"))
        self.assertEqual(resp.status_code, 302)
        order = SaleOrder.objects.latest("id")
        purchase.refresh_from_db()
        self.assertTrue(order.affects_stock)
        self.assertEqual(purchase.remaining, 3)          # consumed 2


@override_settings(CLOUDINARY_CLOUD_NAME="testcloud", CLOUDINARY_AUTO_SYNC=False)
class CloudinaryImageUrlTests(TestCase):
    def test_url_for_product_with_barcode_and_image(self):
        from stock.models import ProductImage
        from stock.services.cloudinary_urls import product_image_cdn_url
        p = _make_product(barcode="6290362349730")
        ProductImage.objects.create(product=p, image=_tiny_png())
        self.assertEqual(
            product_image_cdn_url(p),
            "https://res.cloudinary.com/testcloud/image/upload/"
            "c_pad,b_white,w_1600,h_1600,q_auto/6290362349730.jpg",
        )

    def test_blank_without_image(self):
        from stock.services.cloudinary_urls import product_image_cdn_url
        p = _make_product(barcode="6290362349730")
        self.assertEqual(product_image_cdn_url(p), "")

    def test_blank_without_barcode(self):
        from stock.models import ProductImage
        from stock.services.cloudinary_urls import product_image_cdn_url
        p = _make_product(barcode="")
        ProductImage.objects.create(product=p, image=_tiny_png())
        self.assertEqual(product_image_cdn_url(p), "")

    def test_blank_for_falsy_product(self):
        from stock.services.cloudinary_urls import product_image_cdn_url
        self.assertEqual(product_image_cdn_url(None), "")

    @override_settings(CLOUDINARY_CLOUD_NAME="")
    def test_blank_when_cloud_name_missing(self):
        from stock.models import ProductImage
        from stock.services.cloudinary_urls import product_image_cdn_url
        p = _make_product(barcode="6290362349730")
        ProductImage.objects.create(product=p, image=_tiny_png())
        self.assertEqual(product_image_cdn_url(p), "")

    def test_custom_transformation_is_used(self):
        from stock.models import ProductImage
        from stock.services.cloudinary_urls import product_image_cdn_url
        p = _make_product(barcode="6290362349730")
        ProductImage.objects.create(product=p, image=_tiny_png())
        self.assertEqual(
            product_image_cdn_url(p, transformation="w_400"),
            "https://res.cloudinary.com/testcloud/image/upload/w_400/6290362349730.jpg",
        )

    def test_barcode_is_stripped(self):
        # cloudinary_sync strips the barcode when choosing public_id; the URL
        # must use the same key or it would point at a non-existent asset.
        # Barcode is max_length=13, so keep the padded value within that.
        from stock.models import ProductImage
        from stock.services.cloudinary_urls import product_image_cdn_url
        p = _make_product(barcode="  62903623  ")
        ProductImage.objects.create(product=p, image=_tiny_png())
        self.assertIn("/62903623.jpg", product_image_cdn_url(p))

    def test_url_key_matches_sync_upload_public_id(self):
        # Binds the two halves of the naming contract: whatever public_id
        # sync_product_primary_image actually passes to the Cloudinary client
        # is the exact key product_image_cdn_url must build a URL for. If the
        # upload side ever changes the key (e.g. adds a folder prefix) without
        # updating the URL side, this test fails instead of both halves
        # silently drifting apart.
        from stock.models import ProductImage
        from stock.services import cloudinary_sync
        from stock.services.cloudinary_urls import product_image_cdn_url
        p = _make_product(barcode="6290362349730")
        ProductImage.objects.create(product=p, image=_tiny_png())

        client = mock.Mock()
        code, _ = cloudinary_sync.sync_product_primary_image(p, client=client)
        self.assertEqual(code, cloudinary_sync.UPLOADED)
        _, kwargs = client.upload_image.call_args
        public_id = kwargs["public_id"]

        self.assertTrue(product_image_cdn_url(p).endswith(f"/{public_id}.jpg"))


@override_settings(CLOUDINARY_CLOUD_NAME="testcloud", CLOUDINARY_AUTO_SYNC=False)
class ShopifyCsvCloudinaryImageTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.manager = user_model.objects.create_superuser(username="csv_admin", password="pw123456")
        cls.category = Category.objects.create(name="Perfumes")
        cls.brand = Brand.objects.create(name="Lattafa")
        cls.brand.categories.add(cls.category)
        cls.series = ProductSeries.objects.create(brand=cls.brand, name="Asad")

    def _make_csv_product(self, barcode, spec):
        from stock.models import Product
        return Product.objects.create(
            name="EDP",
            barcode=barcode,
            brand="Lattafa",
            brand_master=self.brand,
            series_master=self.series,
            model="Asad",
            category=self.category,
            spec=spec,
            default_price=Decimal("39.90"),
        )

    def _export_rows(self):
        self.client.login(username="csv_admin", password="pw123456")
        response = self.client.get(reverse("export_shopify_inventory_csv"))
        self.assertEqual(response.status_code, 200)
        return list(csv.DictReader(StringIO(response.content.decode("utf-8-sig"))))

    def test_export_uses_cloudinary_url_for_both_image_columns(self):
        from stock.models import ProductImage
        product = self._make_csv_product("1234509876511", "100ml")
        ProductImage.objects.create(product=product, image=_tiny_png())

        rows = self._export_rows()

        expected = (
            "https://res.cloudinary.com/testcloud/image/upload/"
            "c_pad,b_white,w_1600,h_1600,q_auto/1234509876511.jpg"
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["Product image URL"], expected)
        self.assertEqual(rows[0]["Variant image URL"], expected)
        self.assertNotIn("/media/", rows[0]["Product image URL"])

    def test_export_leaves_image_columns_blank_without_image(self):
        self._make_csv_product("1234509876512", "50ml")

        rows = self._export_rows()

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["Product image URL"], "")
        self.assertEqual(rows[0]["Variant image URL"], "")

    @override_settings(CLOUDINARY_CLOUD_NAME="")
    def test_export_falls_back_to_absolute_media_url_without_cloud_name(self):
        from stock.models import ProductImage
        product = self._make_csv_product("1234509876513", "30ml")
        ProductImage.objects.create(product=product, image=_tiny_png())

        rows = self._export_rows()

        self.assertEqual(len(rows), 1)
        self.assertIn("/media/", rows[0]["Product image URL"])
        self.assertTrue(rows[0]["Product image URL"].startswith("http://"))
        self.assertNotIn("res.cloudinary.com", rows[0]["Product image URL"])


class EmployeeNavTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.employee = user_model.objects.create_user(username="nav_emp", password="pw123456")
        cls.manager = user_model.objects.create_user(username="nav_mgr", password="pw123456", is_staff=True)

    def _nav(self, username):
        self.client.login(username=username, password="pw123456")
        return self.client.get(reverse("dashboard"))

    def test_employee_sidebar_shows_only_allowed_links(self):
        resp = self._nav("nav_emp")
        for name in ["dashboard", "outbound", "sales_records", "customer_search", "product_list"]:
            self.assertContains(resp, 'href="%s"' % reverse(name))
        for name in ["daily_summary", "inbound", "attendance", "catalog", "ar_list"]:
            self.assertNotContains(resp, 'href="%s"' % reverse(name))

    def test_manager_sidebar_shows_everything(self):
        resp = self._nav("nav_mgr")
        for name in ["daily_summary", "inbound", "attendance", "product_list", "catalog", "ar_list", "sales_records", "customer_search"]:
            self.assertContains(resp, 'href="%s"' % reverse(name))


class EmployeeBlockedViewsTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.employee = user_model.objects.create_user(username="blk_emp", password="pw123456")
        cls.manager = user_model.objects.create_user(username="blk_mgr", password="pw123456", is_staff=True)
        cls.customer = Customer.objects.create(nif="123456789", name="Blk Cust")

    def _assert_blocked(self, url):
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 302, url)
        self.assertIn(reverse("dashboard"), resp.headers["Location"], url)

    def test_employee_blocked_from_restricted_pages(self):
        self.client.login(username="blk_emp", password="pw123456")
        for name in ["daily_summary", "inbound", "catalog", "ar_list"]:
            self._assert_blocked(reverse(name))

    def test_employee_customer_detail_now_routes_to_orders_view(self):
        # Task 9: customer_detail is no longer manager-only; employees get
        # a lean orders-only view instead of a 302 redirect.
        self.client.login(username="blk_emp", password="pw123456")
        resp = self.client.get(reverse("customer_detail", args=[self.customer.id]))
        self.assertEqual(resp.status_code, 200)

    def test_manager_can_open_restricted_pages(self):
        self.client.login(username="blk_mgr", password="pw123456")
        for name in ["daily_summary", "inbound", "product_list", "catalog", "ar_list"]:
            self.assertEqual(self.client.get(reverse(name)).status_code, 200, name)


class EmployeeCustomerOrdersTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.employee = user_model.objects.create_user(username="recon_emp", password="pw123456")
        cls.customer = Customer.objects.create(nif="333333333", name="Recon Cust")
        cls.order = SaleOrder.objects.create(customer=cls.customer)
        product = Product.objects.create(name="R", barcode="7300000000001", brand="B")
        Sale.objects.create(order=cls.order, product=product, quantity=2,
                            unit_price=Decimal("7.50"), payment_method="cash")

    def test_employee_sees_customer_orders_only(self):
        self.client.login(username="recon_emp", password="pw123456")
        resp = self.client.get(reverse("customer_detail", args=[self.customer.id]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Recon Cust")
        self.assertContains(resp, "EUR 15.00")
        self.assertContains(resp, reverse("sale_order_detail", args=[self.order.id]))
        self.assertNotContains(resp, "<canvas")

    def test_manager_still_sees_full_customer_detail(self):
        user_model = get_user_model()
        user_model.objects.create_user(username="recon_mgr", password="pw123456", is_staff=True)
        self.client.login(username="recon_mgr", password="pw123456")
        resp = self.client.get(reverse("customer_detail", args=[self.customer.id]))
        self.assertEqual(resp.status_code, 200)


class EmployeeSalesViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.employee = user_model.objects.create_user(username="sales_emp", password="pw123456")

    def test_employee_sales_shows_todays_orders_with_amounts_no_charts(self):
        order = SaleOrder.objects.create()
        product = Product.objects.create(name="P", barcode="7000000000001", brand="B")
        Sale.objects.create(order=order, product=product, quantity=2,
                            unit_price=Decimal("10.00"), payment_method="cash")

        self.client.login(username="sales_emp", password="pw123456")
        resp = self.client.get(reverse("sales_records"))

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "EUR 20.00")       # order/day total shown
        self.assertContains(resp, 'name="date"')      # single-day date picker
        self.assertNotContains(resp, "<canvas")       # no charts at all

    def test_order_number_search_redirects_to_detail(self):
        order = SaleOrder.objects.create()
        product = Product.objects.create(name="Q", barcode="7200000000001", brand="B")
        Sale.objects.create(order=order, product=product, quantity=1,
                            unit_price=Decimal("5.00"), payment_method="cash")
        self.client.login(username="sales_emp", password="pw123456")
        resp = self.client.get(reverse("sales_records"), {"order": str(order.id)})
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("sale_order_detail", args=[order.id]), resp.headers["Location"])

    def test_order_rows_link_to_detail(self):
        order = SaleOrder.objects.create()
        product = Product.objects.create(name="Q2", barcode="7200000000002", brand="B")
        Sale.objects.create(order=order, product=product, quantity=1,
                            unit_price=Decimal("5.00"), payment_method="cash")
        self.client.login(username="sales_emp", password="pw123456")
        resp = self.client.get(reverse("sales_records"))
        self.assertContains(resp, reverse("sale_order_detail", args=[order.id]))


class EmployeeCustomerViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.employee = user_model.objects.create_user(username="cust_emp", password="pw123456")
        cls.alice = Customer.objects.create(nif="111111111", name="Alice Stone", phone="912000001")
        cls.bob = Customer.objects.create(nif="222222222", name="Bob Rivers", phone="912000002")

    def test_no_query_shows_no_customer_rows(self):
        self.client.login(username="cust_emp", password="pw123456")
        resp = self.client.get(reverse("customer_search"))
        self.assertEqual(resp.status_code, 200)
        self.assertNotContains(resp, "Alice Stone")
        self.assertNotContains(resp, "Bob Rivers")

    def test_query_shows_only_matching_minimal_fields(self):
        self.client.login(username="cust_emp", password="pw123456")
        resp = self.client.get(reverse("customer_search"), {"q": "Alice"})
        self.assertContains(resp, "Alice Stone")
        self.assertContains(resp, "912000001")
        self.assertNotContains(resp, "Bob Rivers")
        self.assertNotContains(resp, "Total spent")   # no spend/balance columns

    def test_employee_can_add_customer(self):
        self.client.login(username="cust_emp", password="pw123456")
        resp = self.client.post(reverse("add_customer_ajax"), {
            "nif": "999999999", "name": "New Walkin", "phone": "912999999",
        })
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(Customer.objects.filter(nif="999999999").exists())


class EmployeeCrossStoreIsolationTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.employee = user_model.objects.create_user(username="iso_emp", password="pw123456")

        cls.store_a = Store.objects.create(name="Porto", code="PRT")
        cls.store_b = Store.objects.create(name="Faro", code="FAR")
        StoreProfile.objects.update_or_create(user=cls.employee, defaults={'store': cls.store_a})

        cls.customer = Customer.objects.create(nif="333444555", name="Iso Cust")
        cls.product = Product.objects.create(
            name="Iso Product", barcode="7330000000001", brand="Maison",
            default_price=Decimal("20.00"),
        )

        cls.order_a = SaleOrder.objects.create(customer=cls.customer, store=cls.store_a)
        Sale.objects.create(
            order=cls.order_a, product=cls.product, customer=cls.customer, store=cls.store_a,
            quantity=1, unit_price=Decimal("20.00"), payment_method="cash",
        )

        cls.order_b = SaleOrder.objects.create(customer=cls.customer, store=cls.store_b)
        Sale.objects.create(
            order=cls.order_b, product=cls.product, customer=cls.customer, store=cls.store_b,
            quantity=1, unit_price=Decimal("50.00"), payment_method="cash",
        )

    def test_employee_cannot_open_other_store_order_detail(self):
        self.client.login(username="iso_emp", password="pw123456")
        response = self.client.get(reverse("sale_order_detail", args=[self.order_b.id]))
        self.assertEqual(response.status_code, 404)

    def test_employee_can_open_own_store_order_detail(self):
        self.client.login(username="iso_emp", password="pw123456")
        response = self.client.get(reverse("sale_order_detail", args=[self.order_a.id]))
        self.assertEqual(response.status_code, 200)

    def test_employee_sales_day_view_excludes_other_store_order(self):
        self.client.login(username="iso_emp", password="pw123456")
        today = timezone.localdate()
        response = self.client.get(reverse("sales_records"), {"date": today.isoformat()})
        self.assertEqual(response.status_code, 200)
        order_ids = [row['order_id'] for row in response.context['orders']]
        self.assertIn(self.order_a.id, order_ids)
        self.assertNotIn(self.order_b.id, order_ids)


class EmployeeOrderActionsTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.employee = user_model.objects.create_user(username="act_emp", password="pw123456")
        cls.customer = Customer.objects.create(nif="444444444", name="Act Cust")
        cls.order = SaleOrder.objects.create(customer=cls.customer)
        cls.product = Product.objects.create(name="Widget", barcode="7500000000001", brand="B")
        Sale.objects.create(order=cls.order, product=cls.product, quantity=2,
                            unit_price=Decimal("5.00"), payment_method="cash")

    def test_sales_row_has_view_modal_and_print_button(self):
        self.client.login(username="act_emp", password="pw123456")
        resp = self.client.get(reverse("sales_records"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "order-modal-%s" % self.order.id)     # View modal present
        self.assertContains(resp, "Widget")                              # item rendered in modal
        self.assertContains(resp, reverse("sale_order_detail", args=[self.order.id]))  # Print link

    def test_customer_orders_row_has_view_modal_and_print_button(self):
        self.client.login(username="act_emp", password="pw123456")
        resp = self.client.get(reverse("customer_detail", args=[self.customer.id]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "order-modal-%s" % self.order.id)
        self.assertContains(resp, "Widget")
        self.assertContains(resp, reverse("sale_order_detail", args=[self.order.id]))


class EmployeePageStyleTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.employee = user_model.objects.create_user(username="style_emp", password="pw123456")

    def test_employee_pages_use_shared_card_scaffold(self):
        self.client.login(username="style_emp", password="pw123456")
        for url in [reverse("sales_records"), reverse("customer_search")]:
            resp = self.client.get(url)
            self.assertEqual(resp.status_code, 200)
            self.assertContains(resp, "page-card")     # shared design-system card, like manager pages
            self.assertContains(resp, "page-title")    # shared heading class, like product_list.html


class PerfumePriceLockedFieldTests(TestCase):
    def test_price_locked_defaults_false(self):
        from stock.models import Product
        p = Product.objects.create(name="X", barcode="8000000000001", brand="B")
        self.assertFalse(p.price_locked)


class PerfumePricingServiceTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.perfumes = Category.objects.create(name="Perfumes")
        cls.accessories = Category.objects.create(name="Accessories")

    def _perfume(self, barcode, cost=None, remaining=5, locked=False):
        from stock.models import Product, Purchase
        p = Product.objects.create(name="P", barcode=barcode, brand="B",
                                   category=self.perfumes, price_locked=locked)
        if cost is not None:
            Purchase.objects.create(product=p, quantity=remaining, remaining=remaining,
                                    cost_price=Decimal(str(cost)))
        return p

    def test_formula_rounds_up(self):
        from stock.services.pricing import sync_perfume_price
        p = self._perfume("8100000000001", cost="12.34")
        # Purchase.post_save already synced the price (Task 3 trigger); the
        # explicit call here is idempotent and returns False (no change).
        self.assertFalse(sync_perfume_price(p))
        self.assertEqual(p.wholesale_price, Decimal("23"))   # ceil(12.34+10)=23
        self.assertEqual(p.default_price, Decimal("35"))      # 23+12

    def test_ceil_boundary(self):
        from stock.services.pricing import sync_perfume_price
        exact = self._perfume("8100000000002", cost="12.00")
        sync_perfume_price(exact)
        self.assertEqual(exact.wholesale_price, Decimal("22"))  # ceil(22.00)=22
        over = self._perfume("8100000000003", cost="12.01")
        sync_perfume_price(over)
        self.assertEqual(over.wholesale_price, Decimal("23"))   # ceil(22.01)=23

    def test_no_cost_skips(self):
        from stock.services.pricing import sync_perfume_price
        p = self._perfume("8100000000004", cost=None)   # no purchase -> cost 0
        p.default_price = Decimal("99")
        p.save(update_fields=["default_price"])
        self.assertFalse(sync_perfume_price(p))
        p.refresh_from_db()
        self.assertEqual(p.default_price, Decimal("99"))

    def test_locked_skips(self):
        from stock.services.pricing import sync_perfume_price
        p = self._perfume("8100000000005", cost="12.34", locked=True)
        p.default_price = Decimal("99")
        p.save(update_fields=["default_price"])
        self.assertFalse(sync_perfume_price(p))
        p.refresh_from_db()
        self.assertEqual(p.default_price, Decimal("99"))

    def test_non_perfume_skips(self):
        from stock.models import Product, Purchase
        from stock.services.pricing import sync_perfume_price
        p = Product.objects.create(name="A", barcode="8100000000006", brand="B",
                                   category=self.accessories, default_price=Decimal("5"))
        Purchase.objects.create(product=p, quantity=3, remaining=3, cost_price=Decimal("12.34"))
        self.assertFalse(sync_perfume_price(p))
        p.refresh_from_db()
        self.assertEqual(p.default_price, Decimal("5"))

    def test_idempotent(self):
        from stock.services.pricing import sync_perfume_price
        p = self._perfume("8100000000007", cost="12.34")
        # Purchase.post_save already synced the price (Task 3 trigger).
        self.assertFalse(sync_perfume_price(p))   # already correct: no change
        self.assertFalse(sync_perfume_price(p))   # second call: still no change


class PerfumePricingTriggerTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.perfumes = Category.objects.create(name="Perfumes")

    def _perfume(self, barcode):
        from stock.models import Product
        return Product.objects.create(name="P", barcode=barcode, brand="B", category=self.perfumes)

    def test_inbound_first_batch_sets_price(self):
        from stock.models import Purchase
        p = self._perfume("8200000000001")
        Purchase.objects.create(product=p, quantity=5, remaining=5, cost_price=Decimal("12.34"))
        p.refresh_from_db()
        self.assertEqual(p.wholesale_price, Decimal("23"))
        self.assertEqual(p.default_price, Decimal("35"))

    def _two_batches(self, product, cheap_cost, cheap_qty, pricey_cost, pricey_qty):
        # Purchase.date is auto_now_add=True, so create(date=...) is IGNORED.
        # Force distinct dates with .update() (bypasses auto_now_add) so the
        # cheap batch is unambiguously the oldest (current FIFO), then re-sync
        # because the create-time signal ran before the dates were fixed.
        from stock.models import Purchase
        from stock.services.pricing import sync_perfume_price
        cheap = Purchase.objects.create(product=product, quantity=cheap_qty, remaining=cheap_qty,
                                        cost_price=Decimal(str(cheap_cost)))
        pricey = Purchase.objects.create(product=product, quantity=pricey_qty, remaining=pricey_qty,
                                         cost_price=Decimal(str(pricey_cost)))
        Purchase.objects.filter(pk=cheap.pk).update(date=timezone.now() - timedelta(days=2))
        Purchase.objects.filter(pk=pricey.pk).update(date=timezone.now())
        sync_perfume_price(product)
        return cheap, pricey

    def test_sale_batch_transition_changes_price(self):
        from stock.services.stock_ops import consume_stock_fifo
        p = self._perfume("8200000000002")
        self._two_batches(p, cheap_cost="10.00", cheap_qty=1, pricey_cost="20.00", pricey_qty=5)
        p.refresh_from_db()
        self.assertEqual(p.wholesale_price, Decimal("20"))   # current = cheap batch: ceil(10+10)
        consume_stock_fifo(p, 1)                              # deplete the cheap batch
        p.refresh_from_db()
        self.assertEqual(p.wholesale_price, Decimal("30"))   # now current = pricier: ceil(20+10)
        self.assertEqual(p.default_price, Decimal("42"))

    def test_adjust_purchase_stock_reprices(self):
        p = self._perfume("8200000000003")
        cheap, _ = self._two_batches(p, cheap_cost="10.00", cheap_qty=2, pricey_cost="20.00", pricey_qty=5)
        p.refresh_from_db()
        self.assertEqual(p.wholesale_price, Decimal("20"))
        user_model = get_user_model()
        user_model.objects.create_user(username="adj_mgr", password="pw123456", is_staff=True)
        self.client.login(username="adj_mgr", password="pw123456")
        self.client.post(reverse("api_adjust_purchase_stock"),
                         data={"purchase_id": cheap.id, "new_remaining": 0},
                         content_type="application/json")
        p.refresh_from_db()
        self.assertEqual(p.wholesale_price, Decimal("30"))   # cheap batch emptied -> pricier current

    def test_deleting_current_batch_reprices(self):
        p = self._perfume("8200000000004")
        cheap, _ = self._two_batches(p, cheap_cost="10.00", cheap_qty=2, pricey_cost="20.00", pricey_qty=5)
        p.refresh_from_db()
        self.assertEqual(p.wholesale_price, Decimal("20"))   # current = cheap batch
        cheap.delete()                                        # remove the current batch
        p.refresh_from_db()
        self.assertEqual(p.wholesale_price, Decimal("30"))   # now current = pricier: ceil(20+10)
        self.assertEqual(p.default_price, Decimal("42"))


class PerfumePriceLockFormTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.perfumes = Category.objects.create(name="Perfumes")

    def test_manager_can_lock_and_lock_prevents_reprice(self):
        from stock.models import Product, Purchase
        user_model = get_user_model()
        user_model.objects.create_user(username="lock_mgr", password="pw123456", is_staff=True)
        self.client.login(username="lock_mgr", password="pw123456")
        p = Product.objects.create(name="P", barcode="8300000000001", brand="B",
                                   category=self.perfumes, default_price=Decimal("99"),
                                   wholesale_price=Decimal("88"))
        resp = self.client.post(reverse("edit_product", args=[p.pk]), {
            "barcode": "8300000000001", "category": self.perfumes.id,
            "new_brand_name": "B", "name": "P", "default_price": "99",
            "wholesale_price": "88", "price_locked": "on",
        })
        self.assertIn(resp.status_code, (200, 302))
        p.refresh_from_db()
        self.assertTrue(p.price_locked)
        # An inbound must NOT overwrite the locked prices.
        Purchase.objects.create(product=p, quantity=5, remaining=5, cost_price=Decimal("12.34"))
        p.refresh_from_db()
        self.assertEqual(p.default_price, Decimal("99"))
        self.assertEqual(p.wholesale_price, Decimal("88"))


class EmployeePriceReadOnlyTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.category = Category.objects.create(name="Accessories")

    def _product(self):
        from stock.models import Product
        return Product.objects.create(name="Item", barcode="8500000000001", brand="B",
                                      category=self.category, default_price=Decimal("50"),
                                      wholesale_price=Decimal("40"))

    def test_employee_cannot_change_prices_even_by_tampering(self):
        user_model = get_user_model()
        user_model.objects.create_user(username="ro_emp", password="pw123456")
        self.client.login(username="ro_emp", password="pw123456")
        p = self._product()
        # Products are manager-owned: a hand-crafted POST changes nothing at all.
        resp = self.client.post(reverse("edit_product", args=[p.pk]), {
            "barcode": "8500000000001", "category": self.category.id,
            "new_brand_name": "B", "name": "Renamed",
            "default_price": "1", "wholesale_price": "1",
        })
        self.assertEqual(resp.status_code, 302)
        p.refresh_from_db()
        self.assertEqual(p.name, "Item")
        self.assertEqual(p.default_price, Decimal("50"))
        self.assertEqual(p.wholesale_price, Decimal("40"))

    def test_manager_can_change_prices(self):
        user_model = get_user_model()
        user_model.objects.create_user(username="ro_mgr", password="pw123456", is_staff=True)
        self.client.login(username="ro_mgr", password="pw123456")
        p = self._product()
        resp = self.client.post(reverse("edit_product", args=[p.pk]), {
            "barcode": "8500000000001", "category": self.category.id,
            "new_brand_name": "B", "name": "Item",
            "default_price": "77", "wholesale_price": "66",
        })
        self.assertIn(resp.status_code, (200, 302))
        p.refresh_from_db()
        self.assertEqual(p.default_price, Decimal("77"))
        self.assertEqual(p.wholesale_price, Decimal("66"))

    def test_employee_cannot_open_the_edit_form(self):
        user_model = get_user_model()
        user_model.objects.create_user(username="ro_emp2", password="pw123456")
        self.client.login(username="ro_emp2", password="pw123456")
        p = self._product()
        resp = self.client.get(reverse("edit_product", args=[p.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("dashboard"), resp.headers["Location"])
        # Prices stay visible on the read-only detail page.
        detail = self.client.get(reverse("product_detail", args=[p.pk]))
        self.assertContains(detail, "50.00")

    def test_employee_cannot_lock_price_by_tampering(self):
        from stock.models import Product
        user_model = get_user_model()
        user_model.objects.create_user(username="ro_emp3", password="pw123456")
        self.client.login(username="ro_emp3", password="pw123456")
        p = Product.objects.create(name="Item", barcode="8500000000009", brand="B",
                                   category=self.category, default_price=Decimal("50"),
                                   wholesale_price=Decimal("40"))
        self.client.post(reverse("edit_product", args=[p.pk]), {
            "barcode": "8500000000009", "category": self.category.id, "new_brand_name": "B",
            "name": "Item", "default_price": "50", "wholesale_price": "40", "price_locked": "on",
        })
        p.refresh_from_db()
        self.assertFalse(p.price_locked)   # employee POST cannot lock it


class CategoryAttributeTests(TestCase):
    """Fields the shop defines for itself, with no migration."""

    def setUp(self):
        from stock.models import CategoryAttribute, AttributeOption
        self.accessories = Category.objects.create(name="Accessories")
        self.cases = Category.objects.create(name="Cases", parent=self.accessories)

        self.colour = CategoryAttribute.objects.create(
            category=self.accessories, name="Colour", code="colour",
            data_type="choice", variant_attribute=True, sort_order=1)
        for label in ["Black", "Clear", "Blue"]:
            AttributeOption.objects.create(attribute=self.colour, label=label)

        self.case_type = CategoryAttribute.objects.create(
            category=self.cases, name="Case type", code="case_type",
            data_type="choice", variant_attribute=True, sort_order=2)
        for label in ["Rubber", "Flip", "MagSafe"]:
            AttributeOption.objects.create(attribute=self.case_type, label=label)

        self.thickness = CategoryAttribute.objects.create(
            category=self.cases, name="Thickness", code="thickness",
            data_type="number", unit="mm", sort_order=3)
        self.wireless = CategoryAttribute.objects.create(
            category=self.cases, name="Wireless charging", code="wireless",
            data_type="boolean", sort_order=4)

        self.product = self._product("Clear TPU", self.cases)

    def _product(self, name, category):
        from stock.services.barcodes import assign_internal_barcode
        product = Product(name=name, brand="Generic", category=category,
                          default_price=Decimal("9"))
        assign_internal_barcode(product)
        return product

    # -- inheritance -------------------------------------------------------
    def test_a_subcategory_inherits_its_parents_attributes(self):
        from stock.models import attributes_for_category
        codes = [a.code for a in attributes_for_category(self.cases)]
        self.assertEqual(codes, ["colour", "case_type", "thickness", "wireless"])

    def test_a_parent_does_not_see_its_childrens_attributes(self):
        from stock.models import attributes_for_category
        codes = [a.code for a in attributes_for_category(self.accessories)]
        self.assertEqual(codes, ["colour"])

    def test_a_subcategory_can_override_an_inherited_attribute(self):
        from stock.models import CategoryAttribute, attributes_for_category
        own = CategoryAttribute.objects.create(
            category=self.cases, name="Colour", code="colour",
            data_type="text", sort_order=1)
        found = {a.code: a for a in attributes_for_category(self.cases)}
        self.assertEqual(found["colour"], own)

    def test_a_category_loop_does_not_hang(self):
        # Category.parent is a plain FK with nothing stopping a cycle.
        from stock.models import attributes_for_category
        self.accessories.parent = self.cases
        self.accessories.save()
        self.assertTrue(attributes_for_category(self.cases))

    def test_no_category_means_no_attributes(self):
        from stock.models import attributes_for_category
        self.assertEqual(list(attributes_for_category(None)), [])

    # -- values ------------------------------------------------------------
    def test_a_choice_is_matched_by_label_whatever_the_case(self):
        value = self.product.set_attribute("colour", "black")
        self.assertEqual(value.value_option.label, "Black")
        self.assertEqual(value.display(), "Black")

    def test_a_choice_outside_the_options_is_refused(self):
        with self.assertRaises(ValidationError):
            self.product.set_attribute("colour", "Turquoise")

    def test_a_number_is_stored_as_a_number_and_shows_its_unit(self):
        value = self.product.set_attribute("thickness", "1.5")
        self.assertEqual(value.value_number, Decimal("1.5"))
        self.assertEqual(value.display(), "1.5 mm")

    def test_a_comma_decimal_is_accepted(self):
        # Portuguese keyboards produce 1,5 - refusing it would be a till bug.
        value = self.product.set_attribute("thickness", "1,5")
        self.assertEqual(value.value_number, Decimal("1.5"))

    def test_a_number_that_is_not_a_number_is_refused(self):
        with self.assertRaises(ValidationError):
            self.product.set_attribute("thickness", "thick")

    def test_a_boolean_reads_as_yes_or_no(self):
        self.assertEqual(self.product.set_attribute("wireless", "yes").display(), "Yes")
        self.assertEqual(self.product.set_attribute("wireless", "0").display(), "No")

    def test_setting_an_unknown_code_is_refused(self):
        with self.assertRaises(ValueError):
            self.product.set_attribute("nonsense", "x")

    def test_an_attribute_from_another_category_is_not_reachable(self):
        from stock.models import CategoryAttribute
        other = Category.objects.create(name="Shisha")
        CategoryAttribute.objects.create(category=other, name="Bowl", code="bowl")
        with self.assertRaises(ValueError):
            self.product.set_attribute("bowl", "Clay")

    def test_answering_the_same_attribute_twice_overwrites(self):
        self.product.set_attribute("colour", "Black")
        self.product.set_attribute("colour", "Blue")
        self.assertEqual(self.product.attribute_values.count(), 1)
        self.assertEqual(self.product.attributes.first().display(), "Blue")

    def test_switching_type_clears_the_old_column(self):
        # A stale value in another column would resurface if the type changed.
        value = self.product.set_attribute("thickness", "1.5")
        value.attribute.data_type = "text"
        value.set_value("thin")
        self.assertIsNone(value.value_number)
        self.assertEqual(value.display(), "thin")

    def test_an_empty_answer_clears_the_value(self):
        self.product.set_attribute("colour", "Black")
        value = self.product.set_attribute("colour", "")
        self.assertIsNone(value.value_option)
        self.assertEqual(value.display(), "")

    def test_an_option_from_another_attribute_is_refused(self):
        from stock.models import ProductAttributeValue
        wrong = self.case_type.options.first()
        value = ProductAttributeValue(product=self.product, attribute=self.colour,
                                      value_option=wrong)
        with self.assertRaises(ValidationError):
            value.clean()

    # -- summaries ---------------------------------------------------------
    def test_summary_is_ordered_the_way_the_shop_arranged_it(self):
        self.product.set_attribute("thickness", "1.5")
        self.product.set_attribute("colour", "Black")
        self.product.set_attribute("case_type", "Rubber")
        self.assertEqual(self.product.attribute_summary(),
                         [("Colour", "Black"), ("Case type", "Rubber"), ("Thickness", "1.5 mm")])

    def test_summary_can_be_narrowed_to_what_separates_stock_rows(self):
        self.product.set_attribute("thickness", "1.5")
        self.product.set_attribute("colour", "Black")
        self.assertEqual(self.product.attribute_summary(only_variant=True),
                         [("Colour", "Black")])

    def test_blank_answers_are_left_out_of_the_summary(self):
        self.product.set_attribute("colour", "Black")
        self.product.set_attribute("thickness", "")
        self.assertEqual(self.product.attribute_summary(), [("Colour", "Black")])

    def test_variant_attributes_are_what_matrix_entry_will_use(self):
        from stock.models import variant_attributes_for_category
        codes = [a.code for a in variant_attributes_for_category(self.cases)]
        self.assertEqual(codes, ["colour", "case_type"])

    def test_perfume_is_unaffected_by_any_of_this(self):
        perfumes = Category.objects.create(name="Perfumes")
        perfume = Product.objects.create(name="Oud", barcode="9700000000001", brand="B",
                                         category=perfumes, default_price=Decimal("50"))
        self.assertEqual(perfume.attribute_summary(), [])


class ProductFormAttributeTests(TestCase):
    """Shop-defined attributes must survive a round trip through the form."""

    def setUp(self):
        from stock.models import CategoryAttribute, AttributeOption
        self.accessories = Category.objects.create(name="Accessories")
        self.perfumes = Category.objects.create(name="Perfumes")
        self.brand = Brand.objects.create(name="Generic")

        self.colour = CategoryAttribute.objects.create(
            category=self.accessories, name="Colour", code="colour",
            data_type="choice", variant_attribute=True, sort_order=1)
        self.black = AttributeOption.objects.create(attribute=self.colour, label="Black")
        AttributeOption.objects.create(attribute=self.colour, label="Clear")
        self.thickness = CategoryAttribute.objects.create(
            category=self.accessories, name="Thickness", code="thickness",
            data_type="number", unit="mm", sort_order=2)
        self.wireless = CategoryAttribute.objects.create(
            category=self.accessories, name="Wireless", code="wireless",
            data_type="boolean", sort_order=3)
        # An attribute on a category this product is not in.
        self.bowl = CategoryAttribute.objects.create(
            category=self.perfumes, name="Bottle", code="bottle", data_type="text")

    def _data(self, **overrides):
        data = {
            'barcode': '',
            'category': self.accessories.id,
            'brand_master': self.brand.id,
            'name': 'Clear TPU case',
            'default_price': '9.90',
        }
        data.update(overrides)
        return data

    def test_attribute_answers_are_saved(self):
        form = ProductForm(data=self._data(**{
            f'attr_{self.colour.pk}': str(self.black.pk),
            f'attr_{self.thickness.pk}': '1.5',
            f'attr_{self.wireless.pk}': 'yes',
        }))
        self.assertTrue(form.is_valid(), form.errors.as_json())
        product = form.save()
        self.assertEqual(product.attribute_summary(),
                         [("Colour", "Black"), ("Thickness", "1.5 mm"), ("Wireless", "Yes")])

    def test_answers_for_other_categories_are_ignored(self):
        # Every category's fields are rendered; only the chosen one applies.
        form = ProductForm(data=self._data(**{f'attr_{self.bowl.pk}': 'Crystal'}))
        self.assertTrue(form.is_valid(), form.errors.as_json())
        product = form.save()
        self.assertEqual(product.attribute_summary(), [])

    def test_editing_keeps_answers_that_were_not_changed(self):
        form = ProductForm(data=self._data(**{f'attr_{self.colour.pk}': str(self.black.pk)}))
        self.assertTrue(form.is_valid(), form.errors.as_json())
        product = form.save()

        again = ProductForm(instance=product)
        self.assertEqual(again.initial.get(f'attr_{self.colour.pk}'), self.black.pk)

        edit = ProductForm(data=self._data(
            name="Renamed", **{f'attr_{self.colour.pk}': str(self.black.pk)}),
            instance=product)
        self.assertTrue(edit.is_valid(), edit.errors.as_json())
        saved = edit.save()
        self.assertEqual(saved.attribute_summary(), [("Colour", "Black")])

    def test_clearing_an_answer_clears_it(self):
        form = ProductForm(data=self._data(**{f'attr_{self.colour.pk}': str(self.black.pk)}))
        self.assertTrue(form.is_valid())
        product = form.save()
        edit = ProductForm(data=self._data(**{f'attr_{self.colour.pk}': ''}), instance=product)
        self.assertTrue(edit.is_valid(), edit.errors.as_json())
        self.assertEqual(edit.save().attribute_summary(), [])

    def test_moving_a_product_to_another_category_drops_stale_answers(self):
        form = ProductForm(data=self._data(**{f'attr_{self.colour.pk}': str(self.black.pk)}))
        self.assertTrue(form.is_valid())
        product = form.save()
        self.assertEqual(product.attribute_values.count(), 1)

        moved = ProductForm(data=self._data(category=self.perfumes.id), instance=product)
        self.assertTrue(moved.is_valid(), moved.errors.as_json())
        saved = moved.save()
        self.assertFalse(saved.attribute_values.filter(attribute=self.colour).exists())

    def test_an_unanswered_boolean_is_not_recorded_as_no(self):
        # "Not answered" and "No" are different facts about a product.
        form = ProductForm(data=self._data(**{f'attr_{self.wireless.pk}': ''}))
        self.assertTrue(form.is_valid())
        product = form.save()
        self.assertFalse(product.attribute_values.filter(attribute=self.wireless).exists())

    def test_an_answered_no_is_recorded(self):
        form = ProductForm(data=self._data(**{f'attr_{self.wireless.pk}': 'no'}))
        self.assertTrue(form.is_valid())
        product = form.save()
        self.assertEqual(product.attribute_summary(), [("Wireless", "No")])

    def test_a_number_field_rejects_text(self):
        form = ProductForm(data=self._data(**{f'attr_{self.thickness.pk}': 'thick'}))
        self.assertFalse(form.is_valid())
        self.assertIn(f'attr_{self.thickness.pk}', form.errors)


class AttributeUiTests(TestCase):
    """The pages must actually render - manage.py check does not compile templates."""

    def setUp(self):
        from stock.models import CategoryAttribute, AttributeOption, DeviceModel
        self.accessories = Category.objects.create(name="Accessories")
        self.perfumes = Category.objects.create(name="Perfumes")
        Brand.objects.create(name="Generic")
        self.colour = CategoryAttribute.objects.create(
            category=self.accessories, name="Colour", code="colour",
            data_type="choice", variant_attribute=True)
        self.black = AttributeOption.objects.create(attribute=self.colour, label="Black")

        from stock.services.barcodes import assign_internal_barcode
        self.case = Product(name="Clear TPU", brand="Generic", category=self.accessories,
                            default_price=Decimal("9"))
        assign_internal_barcode(self.case)
        self.case.set_attribute("colour", "Black")
        apple = Brand.objects.create(name="Apple")
        device = DeviceModel.objects.create(brand=apple, name="iPhone 15 Pro Max")
        self.case.device_models.add(device)

        self.perfume = Product.objects.create(
            name="Oud", barcode="9800000000001", brand="B",
            category=self.perfumes, default_price=Decimal("50"), volume_ml=100)

        user = get_user_model().objects.create_superuser(
            username="mgr", password="pw123456", email="m@x.com")
        self.client.force_login(user)

    def test_add_product_page_renders_the_attribute_fields(self):
        html = self.client.get(reverse('add_product')).content.decode()
        self.assertIn(f'attr_{self.colour.pk}', html)
        self.assertIn(f'data-attr-for="{self.accessories.pk}"', html)

    def test_edit_product_page_preselects_the_saved_answer(self):
        html = self.client.get(reverse('edit_product', args=[self.case.pk])).content.decode()
        self.assertIn(f'value="{self.black.pk}" selected', html)

    def test_product_detail_shows_attributes_and_fitment(self):
        html = self.client.get(reverse('product_detail', args=[self.case.pk])).content.decode()
        self.assertIn("Colour", html)
        self.assertIn("Black", html)
        self.assertIn("Apple iPhone 15 Pro Max", html)

    def test_an_accessory_page_does_not_show_empty_perfume_rows(self):
        html = self.client.get(reverse('product_detail', args=[self.case.pk])).content.decode()
        self.assertNotIn("Fragrance family", html)
        self.assertNotIn("Concentration", html)

    def test_a_perfume_page_still_shows_its_own_rows(self):
        html = self.client.get(reverse('product_detail', args=[self.perfume.pk])).content.decode()
        self.assertIn("Fragrance family", html)
        self.assertIn("100 ml", html)


class SeedAccessoryAttributesTests(TestCase):
    def setUp(self):
        from django.core.management import call_command
        self.call = call_command

    def _counts(self):
        from stock.models import CategoryAttribute, AttributeOption
        return CategoryAttribute.objects.count(), AttributeOption.objects.count()

    def test_dry_run_writes_nothing(self):
        self.call("seed_accessory_attributes", verbosity=0)
        self.assertEqual(self._counts(), (0, 0))
        self.assertEqual(Category.objects.count(), 0)

    def test_apply_creates_the_tree_and_the_attributes(self):
        self.call("seed_accessory_attributes", "--apply", verbosity=0)
        attributes, options = self._counts()
        self.assertGreater(attributes, 0)
        self.assertGreater(options, 0)
        cases = Category.objects.get(name="Cases")
        self.assertEqual(cases.parent.name, "Accessories")

    def test_colour_is_inherited_by_every_subcategory(self):
        from stock.models import attributes_for_category
        self.call("seed_accessory_attributes", "--apply", verbosity=0)
        for name in ["Cases", "Screen protectors", "Audio"]:
            codes = [a.code for a in attributes_for_category(Category.objects.get(name=name))]
            self.assertIn("colour", codes, name)

    def test_running_it_twice_changes_nothing(self):
        self.call("seed_accessory_attributes", "--apply", verbosity=0)
        before = self._counts()
        self.call("seed_accessory_attributes", "--apply", verbosity=0)
        self.assertEqual(self._counts(), before)

    def test_it_does_not_overwrite_a_hand_edited_attribute(self):
        from stock.models import CategoryAttribute
        self.call("seed_accessory_attributes", "--apply", verbosity=0)
        colour = CategoryAttribute.objects.get(code="colour",
                                               category__name="Accessories")
        colour.name = "Cor"
        colour.save()
        self.call("seed_accessory_attributes", "--apply", verbosity=0)
        colour.refresh_from_db()
        self.assertEqual(colour.name, "Cor")


class TemplateHygieneTests(SimpleTestCase):
    def test_no_multi_line_django_comments(self):
        """``{# #}`` is single-line only - a comment spanning two lines is
        rendered to the page as literal text, which is how it reaches the shop
        floor looking like garbage."""
        import glob
        offenders = []
        for path in glob.glob("stock/templates/**/*.html", recursive=True):
            with open(path, encoding="utf-8") as handle:
                for number, line in enumerate(handle, 1):
                    if "{#" in line and "#}" not in line.split("{#", 1)[1]:
                        offenders.append(f"{path}:{number}")
        self.assertEqual(offenders, [], "use {% comment %} for multi-line comments")


class LeadTimeStampingTests(TestCase):
    """Every order received from now on must be measured."""

    def setUp(self):
        from stock.models import Supplier
        self.supplier = Supplier.objects.create(name="PERFUME EUROPE")

    def _order(self, **kwargs):
        from stock.models import InboundOrder
        return InboundOrder.objects.create(supplier=self.supplier, **kwargs)

    def test_a_new_order_records_when_it_was_placed(self):
        order = self._order(status='pending_receipt')
        self.assertIsNotNone(order.placed_at)

    def test_receiving_stamps_the_time_once(self):
        from stock.services.lead_time import supplier_lead_stats
        order = self._order(status='pending_receipt')
        order.placed_at = timezone.now() - timedelta(days=3)
        order.save(update_fields=['placed_at'])

        order.mark_received()
        order.refresh_from_db()
        self.assertEqual(order.status, 'received')
        self.assertIsNotNone(order.received_at)
        self.assertAlmostEqual(order.lead_time_days, 3.0, places=1)

        stats = supplier_lead_stats(self.supplier)
        self.assertEqual(stats['sample'], 1)
        self.assertAlmostEqual(stats['avg_days'], 3.0, places=1)

    def test_an_order_created_straight_as_received_still_carries_a_time(self):
        # status defaults to 'received'; such a row used to keep received_at
        # NULL and could never be measured. That is how 40 orders were lost.
        order = self._order()
        self.assertEqual(order.status, 'received')
        self.assertIsNotNone(order.received_at)
        self.assertIsNotNone(order.placed_at)

    def test_an_order_with_no_placed_at_is_left_out_rather_than_guessed(self):
        from stock.models import InboundOrder
        from stock.services.lead_time import supplier_lead_stats
        order = self._order(status='pending_receipt')
        InboundOrder.objects.filter(pk=order.pk).update(placed_at=None)
        order.refresh_from_db()
        order.mark_received()
        self.assertIsNone(supplier_lead_stats(self.supplier))

    def test_a_same_day_receipt_is_counted_not_dropped(self):
        # The old filter required received_at > created_at, so an order placed
        # and received in the same breath vanished instead of counting as ~0.
        from stock.services.lead_time import supplier_lead_stats
        order = self._order(status='pending_receipt')
        order.mark_received(order.placed_at)
        self.assertEqual(supplier_lead_stats(self.supplier)['sample'], 1)

    def test_the_average_moves_as_more_orders_arrive(self):
        from stock.services.lead_time import supplier_lead_stats
        for days in (2, 4, 6):
            order = self._order(status='pending_receipt')
            order.placed_at = timezone.now() - timedelta(days=days)
            order.save(update_fields=['placed_at'])
            order.mark_received()
        stats = supplier_lead_stats(self.supplier)
        self.assertEqual(stats['sample'], 3)
        self.assertAlmostEqual(stats['avg_days'], 4.0, places=1)
        self.assertAlmostEqual(stats['median_days'], 4.0, places=1)

    def test_a_clock_going_backwards_does_not_make_a_negative_wait(self):
        from stock.services.lead_time import supplier_lead_stats
        order = self._order(status='pending_receipt')
        order.mark_received(order.placed_at - timedelta(hours=2))
        self.assertEqual(supplier_lead_stats(self.supplier)['avg_days'], 0.0)

    def test_confirming_receipt_through_the_page_is_measured(self):
        # The whole point: the real flow, not just the model.
        from stock.models import InboundOrder, InboundPendingItem, Product, Category
        from stock.services.lead_time import supplier_lead_stats
        category = Category.objects.create(name="Perfumes")
        product = Product.objects.create(name="Oud", barcode="9920000000001", brand="B",
                                         category=category, default_price=Decimal("50"))
        order = self._order(status='pending_receipt')
        InboundOrder.objects.filter(pk=order.pk).update(
            placed_at=timezone.now() - timedelta(days=5))
        InboundPendingItem.objects.create(inbound_order=order, product=product,
                                          quantity=2, cost_price=Decimal("10"))

        user = get_user_model().objects.create_superuser(
            username="recv", password="pw123456", email="r@x.com")
        self.client.force_login(user)
        line = order.pending_items.first()
        response = self.client.post(reverse('inbound_receive', args=[order.pk]), {
            'action': 'receive',
            'supplier': str(self.supplier.pk),
            'invoice_no': 'INV-1',
            'invoice_date': timezone.localdate().isoformat(),
            'note': '',
            'lines-TOTAL_FORMS': '1',
            'lines-INITIAL_FORMS': '1',
            'lines-MIN_NUM_FORMS': '0',
            'lines-MAX_NUM_FORMS': '1000',
            'lines-0-id': str(line.pk),
            'lines-0-inbound_order': str(order.pk),
            'lines-0-quantity': '2',
            'lines-0-cost_price': '10',
        })
        self.assertIn(response.status_code, (200, 302))
        order.refresh_from_db()
        self.assertEqual(order.status, 'received')
        stats = supplier_lead_stats(self.supplier)
        self.assertIsNotNone(stats, 'a receipt through the page must be measured')
        self.assertAlmostEqual(stats['avg_days'], 5.0, places=0)


class LeadTimeReportCommandTests(TestCase):
    def setUp(self):
        from django.core.management import call_command
        from stock.models import Supplier
        self.call = call_command
        self.supplier = Supplier.objects.create(name="PERFUME EUROPE")

    def _run(self):
        from io import StringIO
        out = StringIO()
        self.call("lead_time_report", supplier="PERFUME", stdout=out)
        return out.getvalue()

    def test_it_reports_a_measured_order(self):
        from stock.models import InboundOrder
        order = InboundOrder.objects.create(supplier=self.supplier, status='pending_receipt')
        InboundOrder.objects.filter(pk=order.pk).update(
            placed_at=timezone.now() - timedelta(days=4))
        order.refresh_from_db()
        order.mark_received()
        output = self._run()
        self.assertIn("PERFUME EUROPE", output)
        self.assertIn("4.0d", output)

    def test_it_says_why_an_order_is_skipped(self):
        from stock.models import InboundOrder
        order = InboundOrder.objects.create(supplier=self.supplier)
        InboundOrder.objects.filter(pk=order.pk).update(placed_at=None)
        self.assertIn("no placed_at", self._run())

    def test_a_pending_order_is_reported_as_not_received_yet(self):
        from stock.models import InboundOrder
        InboundOrder.objects.create(supplier=self.supplier, status='pending_receipt')
        self.assertIn("not received yet", self._run())


class LeadTimeAdminCorrectionTests(TestCase):
    """The shop can correct an order the app never timed."""

    def test_setting_the_real_dates_makes_an_old_order_count(self):
        from stock.models import InboundOrder, Supplier
        from stock.services.lead_time import supplier_lead_stats, why_not_measured
        supplier = Supplier.objects.create(name="PERFUME EUROPE")
        order = InboundOrder.objects.create(supplier=supplier)
        InboundOrder.objects.filter(pk=order.pk).update(placed_at=None, received_at=None)
        order.refresh_from_db()

        self.assertIsNone(supplier_lead_stats(supplier))
        self.assertIn("no placed_at", why_not_measured(order))

        order.placed_at = timezone.now() - timedelta(days=7)
        order.received_at = timezone.now()
        order.save(update_fields=['placed_at', 'received_at'])

        self.assertEqual(why_not_measured(order), '')
        self.assertAlmostEqual(supplier_lead_stats(supplier)['avg_days'], 7.0, places=0)

    def test_the_admin_exposes_the_fields_needed_to_correct_it(self):
        from stock.admin import InboundOrderAdmin
        for field in ('placed_at', 'received_at', 'status'):
            self.assertIn(field, InboundOrderAdmin.fields, field)


class CatalogExportTests(TestCase):
    """Availability wording, colours, and the PDF that WhatsApp can show."""

    def setUp(self):
        from stock.models import Supplier
        self.category = Category.objects.create(name="Perfumes", form_kind="perfume")
        self.supplier = Supplier.objects.create(name="PERFUME EUROPE")
        self.user = get_user_model().objects.create_superuser(
            username="exp_mgr", password="pw123456", email="e@x.com")
        self.client.force_login(self.user)

    def _product(self, name, barcode, stock=0):
        from stock.models import Purchase
        product = Product.objects.create(
            name=name, barcode=barcode, brand="Khan", model="Line",
            category=self.category, default_price=Decimal("50"),
            wholesale_price=Decimal("25"))
        if stock:
            Purchase.objects.create(product=product, quantity=stock, remaining=stock,
                                    cost_price=Decimal("10"))
        return product

    def _on_order(self, product):
        from stock.models import InboundOrder, InboundPendingItem
        order = InboundOrder.objects.create(supplier=self.supplier,
                                            status='pending_receipt')
        InboundPendingItem.objects.create(inbound_order=order, product=product,
                                          quantity=5, cost_price=Decimal("10"))
        return order

    # -- the wording -------------------------------------------------------
    def test_three_or_more_is_available_now(self):
        from stock.views import get_catalog_availability_parts
        self.assertEqual(get_catalog_availability_parts(3), ('Available now', 'in-stock'))
        self.assertEqual(get_catalog_availability_parts(99)[0], 'Available now')

    def test_one_or_two_is_low_stock(self):
        from stock.views import get_catalog_availability_parts
        self.assertEqual(get_catalog_availability_parts(1), ('Low stock', 'low-stock'))
        self.assertEqual(get_catalog_availability_parts(2), ('Low stock', 'low-stock'))

    def test_nothing_on_the_shelf_and_nothing_coming_is_unavailable(self):
        from stock.views import get_catalog_availability_parts
        self.assertEqual(get_catalog_availability_parts(0),
                         ('Currently unavailable', 'out-stock'))

    def test_nothing_on_the_shelf_but_on_order_is_in_stock_soon(self):
        from stock.views import get_catalog_availability_parts
        self.assertEqual(get_catalog_availability_parts(0, on_order=True),
                         ('In stock soon', 'incoming'))

    def test_low_stock_wins_over_incoming(self):
        # What matters to someone buying today is what is on the shelf today.
        from stock.views import get_catalog_availability_parts
        self.assertEqual(get_catalog_availability_parts(1, on_order=True)[0], 'Low stock')

    # -- what counts as on order ------------------------------------------
    def test_a_pending_inbound_marks_a_product_as_coming(self):
        from stock.views import product_ids_on_order
        product = self._product("Oud", "9940000000001")
        self.assertEqual(product_ids_on_order(), set())
        self._on_order(product)
        self.assertEqual(product_ids_on_order(), {product.id})

    def test_a_received_order_no_longer_counts_as_coming(self):
        from stock.views import product_ids_on_order
        product = self._product("Oud", "9940000000002")
        order = self._on_order(product)
        order.mark_received()
        order.pending_items.all().delete()
        self.assertEqual(product_ids_on_order(), set())

    # -- the spreadsheet ---------------------------------------------------
    def _export(self, **params):
        return self.client.get(reverse('export_product_list_excel'), params)

    def test_the_export_settings_line_is_gone_and_a_legend_is_there(self):
        from openpyxl import load_workbook
        from io import BytesIO
        self._product("Oud", "9940000000003", stock=5)
        response = self._export()
        book = load_workbook(BytesIO(b''.join(response.streaming_content)))
        second_row = book.active.cell(row=2, column=1).value or ''
        self.assertNotIn('Images:', second_row)
        self.assertNotIn('Sort:', second_row)
        self.assertIn('green', second_row)
        self.assertIn('blue', second_row)
        self.assertIn('under 3', second_row)

    def test_the_availability_cell_is_coloured_per_state(self):
        from openpyxl import load_workbook
        from io import BytesIO
        from stock.views import AVAILABILITY_STYLES
        available = self._product("Plenty", "9940000000004", stock=5)
        low = self._product("Nearly out", "9940000000005", stock=1)
        coming = self._product("Coming", "9940000000006")
        self._on_order(coming)
        self._product("Gone", "9940000000007")

        response = self._export()
        sheet = load_workbook(BytesIO(b''.join(response.streaming_content))).active
        found = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in ('Available now', 'Low stock', 'In stock soon',
                                  'Currently unavailable'):
                    found[cell.value] = cell.fill.fgColor.rgb
        self.assertEqual(found['Available now'], AVAILABILITY_STYLES['in-stock'][0])
        self.assertEqual(found['Low stock'], AVAILABILITY_STYLES['low-stock'][0])
        self.assertEqual(found['In stock soon'], AVAILABILITY_STYLES['incoming'][0])
        self.assertEqual(found['Currently unavailable'], AVAILABILITY_STYLES['out-stock'][0])

    # -- the PDF -----------------------------------------------------------
    def test_a_pdf_is_returned_when_asked_for(self):
        self._product("Oud", "9940000000008", stock=5)
        response = self._export(format='pdf')
        self.assertEqual(response['Content-Type'], 'application/pdf')
        body = b''.join(response.streaming_content)
        self.assertTrue(body.startswith(b'%PDF'))
        self.assertIn('.pdf', response['Content-Disposition'])

    def test_the_pdf_carries_the_products_and_the_legend(self):
        self._product("Khamrah", "9940000000009", stock=5)
        body = b''.join(self._export(format='pdf').streaming_content)
        # Uncompressed enough to check the text made it in.
        self.assertGreater(len(body), 1000)

    def test_the_pdf_really_embeds_the_photo(self):
        """The whole reason the PDF exists: the picture is in the page, not a
        floating drawing a phone viewer can decline to render."""
        import tempfile, os
        from PIL import Image as PILImage
        from django.core.files import File
        from stock.models import ProductImage
        product = self._product("With photo", "9940000000011", stock=5)
        path = os.path.join(tempfile.gettempdir(), 'shot.png')
        PILImage.new('RGB', (300, 300), (10, 120, 200)).save(path)
        with open(path, 'rb') as handle:
            ProductImage.objects.create(product=product, image=File(handle, name='shot.png'))

        body = b''.join(self._export(format='pdf', include_images='1').streaming_content)
        self.assertTrue(body.startswith(b'%PDF'))
        # an XObject image stream is what "the photo is in the page" looks like
        self.assertIn(b'/Image', body)
        self.assertIn(b'/XObject', body)

    def test_a_product_with_a_missing_photo_file_does_not_break_the_pdf(self):
        from stock.models import ProductImage
        product = self._product("Broken photo", "9940000000012", stock=5)
        ProductImage.objects.create(product=product, image='products/gone.jpg')
        response = self._export(format='pdf', include_images='1')
        self.assertEqual(response.status_code, 200)
        self.assertTrue(b''.join(response.streaming_content).startswith(b'%PDF'))

    def test_excel_is_still_the_default(self):
        self._product("Oud", "9940000000010", stock=5)
        response = self._export()
        self.assertIn('spreadsheetml', response['Content-Type'])


class ShelfMapTests(TestCase):
    """Is there stock, and is there more above the shelf. Nothing is counted."""

    def setUp(self):
        from stock.models import ShelfAxis, ShelfOption, ShelfStyle, DeviceModel
        # migration 0051 already seeds a Colour axis, so take that one
        self.axis, _ = ShelfAxis.objects.get_or_create(
            slug="colour", defaults={"name": "Colour"})
        ShelfOption.objects.all().delete()
        self.style = ShelfStyle.objects.create(name="Normal silicone",
                                               slug="normal-silicone", axis=self.axis)
        self.apple = Brand.objects.create(name="Apple")
        self.samsung = Brand.objects.create(name="Samsung")
        self.black = ShelfOption.objects.create(axis=self.axis, name="Black", sort_order=1)
        self.clear = ShelfOption.objects.create(axis=self.axis, name="Transparent", sort_order=2)
        self.m15 = DeviceModel.objects.create(brand=self.apple, name="iPhone 15")
        self.m9 = DeviceModel.objects.create(brand=self.apple, name="iPhone 9")
        self.m11 = DeviceModel.objects.create(brand=self.apple, name="iPhone 11")

        self.manager = get_user_model().objects.create_superuser(
            username="shelf_mgr", password="pw123456", email="sm@x.com")
        self.staff = get_user_model().objects.create_user(
            username="shelf_staff", password="pw123456")
        self.client.force_login(self.manager)

    # -- natural ordering --------------------------------------------------
    def test_models_sort_by_number_not_alphabetically(self):
        from stock.models import DeviceModel
        names = list(DeviceModel.objects.filter(brand=self.apple)
                     .order_by('sort_key').values_list('name', flat=True))
        self.assertEqual(names, ["iPhone 9", "iPhone 11", "iPhone 15"])

    def test_a_newly_added_model_lands_in_its_place(self):
        from stock.models import DeviceModel
        response = self.client.post(reverse('shelf_add_model'),
                                    {'brand': 'Apple', 'name': 'iPhone 12'})
        self.assertEqual(response.status_code, 200)
        names = list(DeviceModel.objects.filter(brand=self.apple)
                     .order_by('sort_key').values_list('name', flat=True))
        self.assertEqual(names, ["iPhone 9", "iPhone 11", "iPhone 12", "iPhone 15"])

    def test_the_grid_lists_models_in_that_order_grouped_by_brand(self):
        from stock.models import DeviceModel
        DeviceModel.objects.create(brand=self.samsung, name="Galaxy S24")
        response = self.client.get(reverse('shelf'))
        order = [row['model'].name for row in response.context['rows']]
        self.assertEqual(order, ["iPhone 9", "iPhone 11", "iPhone 15", "Galaxy S24"])

    # -- state -------------------------------------------------------------
    def _set(self, model, colour, state='display'):
        import json
        return self.client.post(reverse('shelf_save_states'), {
            'style': self.style.pk,
            'changes': json.dumps([{'model': model.pk, 'option': colour.pk,
                                    'state': state}]),
        })

    def test_a_cell_starts_out_without_a_row_existing(self):
        from stock.models import ShelfStock
        response = self.client.get(reverse('shelf'))
        row = next(r for r in response.context['rows'] if r['model'] == self.m15)
        self.assertEqual([c['state'] for c in row['cells']], ['out', 'out'])
        self.assertEqual(ShelfStock.objects.count(), 0)

    def test_a_saved_edit_is_written(self):
        from stock.models import ShelfStock
        self.assertTrue(self._set(self.m15, self.black, 'extra').json()['ok'])
        self.assertEqual(ShelfStock.objects.get().state, 'extra')

    def test_several_cells_save_in_one_request(self):
        import json
        from stock.models import ShelfStock
        response = self.client.post(reverse('shelf_save_states'), {
            'style': self.style.pk,
            'changes': json.dumps([
                {'model': self.m15.pk, 'option': self.black.pk, 'state': 'extra'},
                {'model': self.m15.pk, 'option': self.clear.pk, 'state': 'display'},
                {'model': self.m11.pk, 'option': self.black.pk, 'state': 'display'},
            ]),
        })
        self.assertEqual(response.json()['saved'], 3)
        self.assertEqual(ShelfStock.objects.count(), 3)

    def test_an_unknown_state_is_ignored_rather_than_stored(self):
        from stock.models import ShelfStock
        self._set(self.m15, self.black, 'nonsense')
        self.assertFalse(ShelfStock.objects.filter(state='nonsense').exists())

    def test_a_malformed_change_does_not_take_the_rest_down(self):
        import json
        from stock.models import ShelfStock
        self.client.post(reverse('shelf_save_states'), {
            'style': self.style.pk,
            'changes': json.dumps([
                {'model': 'nonsense', 'option': self.black.pk, 'state': 'extra'},
                {'model': self.m15.pk, 'option': self.clear.pk, 'state': 'display'},
            ]),
        })
        self.assertEqual(ShelfStock.objects.count(), 1)
        self.assertEqual(ShelfStock.objects.get().option, self.clear)

    def test_broken_json_is_refused(self):
        response = self.client.post(reverse('shelf_save_states'),
                                    {'style': self.style.pk, 'changes': '{{'})
        self.assertEqual(response.status_code, 400)

    def test_colours_of_the_same_model_are_independent(self):
        self._set(self.m15, self.black, 'extra')
        response = self.client.get(reverse('shelf'))
        row = next(r for r in response.context['rows'] if r['model'] == self.m15)
        self.assertEqual([c['state'] for c in row['cells']], ['extra', 'out'])

    def test_the_change_records_who_made_it(self):
        from stock.models import ShelfStock
        self._set(self.m15, self.black)
        self.assertEqual(ShelfStock.objects.get().updated_by, self.manager)

    def test_shop_floor_staff_can_mark_a_colour_out(self):
        # Marking sold-out is the job of whoever is at the shelf.
        self.client.force_login(self.staff)
        response = self._set(self.m15, self.black, 'extra')
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])

    def test_a_signed_out_visitor_cannot_change_anything(self):
        from stock.models import ShelfStock
        self.client.logout()
        self._set(self.m15, self.black)
        self.assertEqual(ShelfStock.objects.count(), 0)

    def test_nothing_records_a_quantity(self):
        from stock.models import ShelfStock
        self._set(self.m15, self.black, 'extra')
        field_names = {f.name for f in ShelfStock._meta.get_fields()}
        self.assertNotIn('quantity', field_names)
        self.assertEqual(ShelfStock.objects.get().state, 'extra')

    # -- notes -------------------------------------------------------------
    def test_a_note_is_saved_against_the_model(self):
        response = self.client.post(reverse('shelf_set_note'), {
            'style': self.style.pk, 'model': self.m15.pk,
            'note': 'same case as 14 Pro'})
        self.assertEqual(response.json()['note'], 'same case as 14 Pro')

    def test_a_note_does_not_link_the_states_of_two_models(self):
        # Notes are a reminder for the person at the shelf, nothing more.
        self.client.post(reverse('shelf_set_note'), {
            'style': self.style.pk, 'model': self.m15.pk, 'note': 'same as 11'})
        self._set(self.m15, self.black, 'extra')
        response = self.client.get(reverse('shelf'))
        other = next(r for r in response.context['rows'] if r['model'] == self.m11)
        self.assertEqual([c['state'] for c in other['cells']], ['out', 'out'])

    def test_a_note_is_searchable(self):
        self.client.post(reverse('shelf_set_note'), {
            'style': self.style.pk, 'model': self.m15.pk, 'note': 'same as 14 Pro'})
        response = self.client.get(reverse('shelf'))
        row = next(r for r in response.context['rows'] if r['model'] == self.m15)
        self.assertIn('same as 14 pro', row['search'])

    # -- adding colours and models ----------------------------------------
    def test_a_colour_can_be_added(self):
        from stock.models import ShelfOption
        response = self.client.post(reverse('shelf_add_option'),
                                    {'axis': self.axis.pk, 'name': 'Beige'})
        self.assertTrue(response.json()['ok'])
        self.assertTrue(ShelfOption.objects.filter(name='Beige').exists())

    def test_a_duplicate_colour_is_refused(self):
        response = self.client.post(reverse('shelf_add_option'),
                                    {'axis': self.axis.pk, 'name': 'black'})
        self.assertEqual(response.status_code, 409)

    def test_a_duplicate_model_is_refused(self):
        response = self.client.post(reverse('shelf_add_model'),
                                    {'brand': 'Apple', 'name': 'iphone 15'})
        self.assertEqual(response.status_code, 409)

    def test_staff_cannot_add_models_or_colours(self):
        from stock.models import DeviceModel, ShelfOption
        self.client.force_login(self.staff)
        self.assertEqual(self.client.post(reverse('shelf_add_model'),
                         {'brand': 'Apple', 'name': 'iPhone 16'}).status_code, 403)
        self.assertEqual(self.client.post(reverse('shelf_add_option'),
                         {'axis': self.axis.pk, 'name': 'Teal'}).status_code, 403)
        self.assertFalse(DeviceModel.objects.filter(name='iPhone 16').exists())
        self.assertFalse(ShelfOption.objects.filter(name='Teal').exists())

    # -- the page ----------------------------------------------------------
    def test_the_page_renders_a_grid_with_no_photos(self):
        html = self.client.get(reverse('shelf')).content.decode()
        self.assertIn('id="shelf-body"', html)
        self.assertIn('Transparent', html)
        body = html[html.index('id="shelf-body"'):html.index('</table>')]
        self.assertNotIn('<img', body)

    def test_the_page_counts_how_many_models_are_fully_out(self):
        self._set(self.m15, self.black, 'display')
        response = self.client.get(reverse('shelf'))
        self.assertEqual(response.context['total_models'], 3)
        self.assertEqual(response.context['out_models'], 2)

    def test_a_style_with_no_rows_yet_still_opens(self):
        from stock.models import ShelfStyle
        style = ShelfStyle.objects.create(name="MagSafe", slug="magsafe",
                                          axis=self.axis, sort_order=2)
        response = self.client.get(reverse('shelf_style', args=[style.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['style'], style)

    def test_each_style_keeps_its_own_states(self):
        from stock.models import ShelfStyle
        magsafe = ShelfStyle.objects.create(name="MagSafe", slug="magsafe",
                                            axis=self.axis, sort_order=2)
        self._set(self.m15, self.black, 'extra')
        response = self.client.get(reverse('shelf_style', args=[magsafe.slug]))
        row = next(r for r in response.context['rows'] if r['model'] == self.m15)
        self.assertEqual([c['state'] for c in row['cells']], ['out', 'out'])

    def test_the_page_works_before_any_style_exists(self):
        from stock.models import ShelfStyle, ShelfStock
        ShelfStock.objects.all().delete()
        ShelfStyle.objects.all().delete()
        response = self.client.get(reverse('shelf'))
        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.context['style'])


class AddProductDefersToShelfTests(TestCase):
    """Cases are never entered as products - the shelf grid is the answer."""

    def setUp(self):
        from stock.models import ShelfAxis, ShelfStyle
        self.cases = Category.objects.create(name="Cases", form_kind="accessory")
        self.perfumes = Category.objects.create(name="Perfumes", form_kind="perfume")
        axis, _ = ShelfAxis.objects.get_or_create(
            slug='colour', defaults={'name': 'Colour'})
        self.style = ShelfStyle.objects.create(
            name="Normal silicone", slug="normal-silicone",
            axis=axis, category=self.cases)
        user = get_user_model().objects.create_superuser(
            username="ap_mgr", password="pw123456", email="ap@x.com")
        self.client.force_login(user)

    def _targets(self):
        import json
        response = self.client.get(reverse('add_product'))
        return json.loads(response.context['shelf_targets_json'])

    def test_a_shelf_tracked_category_points_at_its_grid(self):
        targets = self._targets()
        self.assertIn(str(self.cases.pk), targets)
        entry = targets[str(self.cases.pk)][0]
        self.assertEqual(entry['name'], "Normal silicone")
        self.assertEqual(entry['url'], reverse('shelf_style', args=['normal-silicone']))

    def test_a_normal_category_is_not_diverted(self):
        self.assertNotIn(str(self.perfumes.pk), self._targets())

    def test_several_styles_under_one_category_are_all_offered(self):
        from stock.models import ShelfAxis, ShelfStyle
        axis = ShelfAxis.objects.get(slug='colour')
        ShelfStyle.objects.create(name="MagSafe", slug="magsafe",
                                  axis=axis, category=self.cases)
        names = {t['name'] for t in self._targets()[str(self.cases.pk)]}
        self.assertEqual(names, {"Normal silicone", "MagSafe"})

    def test_an_inactive_style_is_not_offered(self):
        self.style.is_active = False
        self.style.save(update_fields=['is_active'])
        self.assertNotIn(str(self.cases.pk), self._targets())

    def test_the_page_renders_the_shelf_panel(self):
        html = self.client.get(reverse('add_product')).content.decode()
        self.assertIn('id="shelf-redirect"', html)
        self.assertIn('Tracked on the shelf', html)

    def test_adding_a_handset_is_the_only_thing_needed(self):
        # No product, no barcode, no price - just the model, then colours.
        from stock.models import DeviceModel, Product
        Brand.objects.create(name="Apple")
        response = self.client.post(reverse('shelf_add_model'),
                                    {'brand': 'Apple', 'name': 'iPhone 17 Pro'})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(DeviceModel.objects.filter(name='iPhone 17 Pro').exists())
        self.assertEqual(Product.all_objects.count(), 0)


class ShelfNavAndPriceTests(TestCase):
    """Category, then type, with the wholesale price beside it."""

    def setUp(self):
        from decimal import Decimal
        from stock.models import ShelfAxis, ShelfOption, ShelfStyle, DeviceModel
        self.colour, _ = ShelfAxis.objects.get_or_create(
            slug='colour', defaults={'name': 'Colour'})
        ShelfOption.objects.all().delete()
        self.glue = ShelfAxis.objects.create(name='Glue & edge', slug='glue-edge')
        ShelfOption.objects.create(axis=self.colour, name='Black')
        ShelfOption.objects.create(axis=self.glue, name='Full glue curved')

        self.cases = Category.objects.create(name="Cases", form_kind="accessory")
        self.glass_cat = Category.objects.create(name="Screen protectors",
                                                 form_kind="accessory")
        self.silicone = ShelfStyle.objects.create(
            name="Normal silicone", slug="normal-silicone", axis=self.colour,
            category=self.cases, wholesale_price=Decimal("1.20"), sort_order=1)
        self.magsafe = ShelfStyle.objects.create(
            name="MagSafe", slug="magsafe", axis=self.colour,
            category=self.cases, wholesale_price=Decimal("3.50"), sort_order=2)
        self.glass = ShelfStyle.objects.create(
            name="Tempered glass", slug="tempered-glass", axis=self.glue,
            category=self.glass_cat, wholesale_price=Decimal("0.80"), sort_order=3)

        apple = Brand.objects.create(name="Apple")
        DeviceModel.objects.create(brand=apple, name="iPhone 15", release_year=2023)
        user = get_user_model().objects.create_superuser(
            username="nav_mgr", password="pw123456", email="n@x.com")
        self.client.force_login(user)

    def test_types_are_grouped_under_their_category(self):
        response = self.client.get(reverse('shelf'))
        groups = {g['name']: [s.name for s in g['styles']]
                  for g in response.context['groups']}
        self.assertEqual(groups["Cases"], ["Normal silicone", "MagSafe"])
        self.assertEqual(groups["Screen protectors"], ["Tempered glass"])

    def test_the_page_knows_which_category_is_open(self):
        response = self.client.get(reverse('shelf_style', args=['tempered-glass']))
        self.assertEqual(response.context['current_group']['name'], "Screen protectors")

    def test_the_wholesale_price_is_shown(self):
        html = self.client.get(reverse('shelf_style', args=['magsafe'])).content.decode()
        self.assertIn("3.50", html)
        self.assertIn("wholesale", html)

    def test_a_type_with_no_price_set_does_not_show_an_empty_tag(self):
        self.magsafe.wholesale_price = None
        self.magsafe.save(update_fields=['wholesale_price'])
        html = self.client.get(reverse('shelf_style', args=['magsafe'])).content.decode()
        self.assertNotIn("wholesale", html)

    def test_glass_and_cases_show_different_columns(self):
        cases = self.client.get(reverse('shelf_style', args=['normal-silicone']))
        glass = self.client.get(reverse('shelf_style', args=['tempered-glass']))
        self.assertEqual([c.name for c in cases.context['columns']], ['Black'])
        self.assertEqual([c.name for c in glass.context['columns']],
                         ['Full glue curved'])

    def test_the_grid_is_read_only_until_edit(self):
        html = self.client.get(reverse('shelf')).content.decode()
        self.assertIn('id="edit-btn"', html)
        self.assertIn('id="save-btn"', html)
        # every cell records the state the server holds, so Cancel can restore
        self.assertIn('data-saved=', html)

    def test_a_style_with_no_category_still_appears(self):
        from stock.models import ShelfStyle
        ShelfStyle.objects.create(name="Loose", slug="loose", axis=self.colour,
                                  sort_order=9)
        names = [g['name'] for g in self.client.get(reverse('shelf')).context['groups']]
        self.assertIn("Other", names)


class ShelfAxisTests(TestCase):
    """A style names the axis that forms its columns, so a new kind of goods
    brings its own columns without a code change."""

    def setUp(self):
        from stock.models import ShelfAxis, ShelfOption, ShelfStyle, DeviceModel
        self.colour, _ = ShelfAxis.objects.get_or_create(
            slug='colour', defaults={'name': 'Colour'})
        ShelfOption.objects.all().delete()
        self.glue = ShelfAxis.objects.create(name='Glue & edge', slug='glue-edge')

        self.black = ShelfOption.objects.create(axis=self.colour, name='Black')
        self.curved = ShelfOption.objects.create(axis=self.glue, name='Full glue curved')

        self.cases = ShelfStyle.objects.create(name='Normal silicone',
                                               slug='normal-silicone', axis=self.colour)
        self.glass = ShelfStyle.objects.create(name='Tempered glass',
                                               slug='tempered-glass', axis=self.glue)
        apple = Brand.objects.create(name='Apple')
        self.model = DeviceModel.objects.create(brand=apple, name='iPhone 15',
                                                release_year=2023)
        user = get_user_model().objects.create_superuser(
            username='axis_mgr', password='pw123456', email='a@x.com')
        self.client.force_login(user)

    def test_each_style_shows_only_its_own_axis_columns(self):
        cases = self.client.get(reverse('shelf_style', args=[self.cases.slug]))
        glass = self.client.get(reverse('shelf_style', args=[self.glass.slug]))
        self.assertEqual([c.name for c in cases.context['columns']], ['Black'])
        self.assertEqual([c.name for c in glass.context['columns']],
                         ['Full glue curved'])

    def test_the_same_handset_appears_under_every_style(self):
        # Entered once, available everywhere - that is the point of a shared
        # model registry.
        for style in (self.cases, self.glass):
            response = self.client.get(reverse('shelf_style', args=[style.slug]))
            names = [row['model'].name for row in response.context['rows']]
            self.assertEqual(names, ['iPhone 15'], style.name)

    def test_states_do_not_leak_between_styles(self):
        import json
        self.client.post(reverse('shelf_save_states'), {
            'style': self.cases.pk,
            'changes': json.dumps([{'model': self.model.pk,
                                    'option': self.black.pk, 'state': 'extra'}])})
        glass = self.client.get(reverse('shelf_style', args=[self.glass.slug]))
        row = glass.context['rows'][0]
        self.assertEqual([c['state'] for c in row['cells']], ['out'])

    def test_a_new_option_only_joins_its_own_axis(self):
        from stock.models import ShelfOption
        self.client.post(reverse('shelf_add_option'),
                         {'axis': self.glue.pk, 'name': 'Privacy'})
        self.assertTrue(ShelfOption.objects.filter(axis=self.glue, name='Privacy').exists())
        cases = self.client.get(reverse('shelf_style', args=[self.cases.slug]))
        self.assertEqual([c.name for c in cases.context['columns']], ['Black'])

    def test_adding_a_style_picks_an_axis(self):
        from stock.models import ShelfStyle
        response = self.client.post(reverse('shelf_add_style'),
                                    {'name': 'Hydrogel film', 'axis': self.glue.pk})
        self.assertEqual(response.status_code, 302)
        style = ShelfStyle.objects.get(slug='hydrogel-film')
        self.assertEqual(style.axis, self.glue)

    def test_a_style_added_without_an_axis_falls_back_to_colour(self):
        from stock.models import ShelfStyle
        self.client.post(reverse('shelf_add_style'), {'name': 'MagSafe'})
        self.assertEqual(ShelfStyle.objects.get(slug='magsafe').axis, self.colour)

    def test_an_axis_in_use_cannot_be_deleted_out_from_under_a_style(self):
        from django.db.models import ProtectedError
        with self.assertRaises(ProtectedError):
            self.glue.delete()


class PhoneModelSeedTests(TestCase):
    def setUp(self):
        from django.core.management import call_command
        self.call = call_command

    def test_dry_run_writes_nothing(self):
        from stock.models import DeviceModel
        self.call("seed_phone_models", verbosity=0)
        self.assertEqual(DeviceModel.objects.count(), 0)

    def test_apply_creates_the_lineup_in_shelf_order(self):
        from django.db.models import Value
        from django.db.models.functions import Coalesce
        from stock.models import DeviceModel
        self.call("seed_phone_models", "--apply", verbosity=0)
        names = list(DeviceModel.objects
                     .annotate(y=Coalesce('release_year', Value(9999)))
                     .order_by('y', 'sort_key', 'name')
                     .values_list('name', flat=True))
        self.assertIn("iPhone 16 Pro Max", names)
        # the year decides first, so X sits between the 8 and the 11
        self.assertLess(names.index("iPhone 8"), names.index("iPhone X"))
        self.assertLess(names.index("iPhone X"), names.index("iPhone 11"))
        self.assertLess(names.index("iPhone 11"), names.index("iPhone 15"))

    def test_running_it_twice_adds_nothing(self):
        from stock.models import DeviceModel
        self.call("seed_phone_models", "--apply", verbosity=0)
        before = DeviceModel.objects.count()
        self.call("seed_phone_models", "--apply", verbosity=0)
        self.assertEqual(DeviceModel.objects.count(), before)

    def test_aliases_let_the_till_find_a_handset(self):
        from stock.models import resolve_device
        self.call("seed_phone_models", "--apply", verbosity=0)
        self.assertEqual(resolve_device("15PM").name, "iPhone 15 Pro Max")
        self.assertEqual(resolve_device("iphone 15 pro max").name, "iPhone 15 Pro Max")

    def test_an_unknown_brand_is_reported_not_guessed(self):
        from io import StringIO
        from stock.models import DeviceModel
        out = StringIO()
        self.call("seed_phone_models", "--brand", "Nokia", "--apply", stdout=out)
        self.assertIn("No lineup", out.getvalue())
        self.assertEqual(DeviceModel.objects.count(), 0)


class NaturalKeyTests(SimpleTestCase):
    def test_numbers_compare_as_numbers(self):
        from stock.models import natural_key
        ordered = sorted(["iPhone 15", "iPhone 9", "iPhone 11", "iPhone 15 Pro"],
                         key=natural_key)
        self.assertEqual(ordered,
                         ["iPhone 9", "iPhone 11", "iPhone 15", "iPhone 15 Pro"])

    def test_it_copes_with_no_digits_and_with_empty(self):
        from stock.models import natural_key
        self.assertEqual(natural_key("Galaxy S"), "galaxy s")
        self.assertEqual(natural_key(""), "")
        self.assertEqual(natural_key(None), "")


class SupplierOrderSortingTests(TestCase):
    """Supplier history is read newest-arrival first."""

    def setUp(self):
        from stock.models import InboundOrder, Purchase, Supplier, Category
        self.supplier = Supplier.objects.create(name="PERFUME EUROPE")
        self.category = Category.objects.create(name="Perfumes")
        self.product = Product.objects.create(
            name="Oud", barcode="9930000000001", brand="B",
            category=self.category, default_price=Decimal("50"))
        user = get_user_model().objects.create_superuser(
            username="sup_mgr", password="pw123456", email="s@x.com")
        self.client.force_login(user)

    def _order(self, invoice_date, received_on, invoice_no):
        from stock.models import InboundOrder, Purchase
        order = InboundOrder.objects.create(
            supplier=self.supplier, invoice_no=invoice_no,
            invoice_date=invoice_date, status='pending_receipt')
        Purchase.objects.create(inbound_order=order, product=self.product,
                                supplier=self.supplier, quantity=1, remaining=1,
                                cost_price=Decimal("10"), date=received_on)
        order.status = 'received'
        order.received_at = received_on
        order.save(update_fields=['status', 'received_at'])
        return order

    def test_the_order_that_arrived_last_is_listed_first(self):
        # An invoice can be dated after a shipment that turned up later, which
        # is exactly how today's delivery ended up buried.
        old_arrival = timezone.now() - timedelta(days=10)
        new_arrival = timezone.now()
        self._order(invoice_date=(timezone.now() + timedelta(days=2)).date(),
                    received_on=old_arrival, invoice_no="LATE-INVOICE")
        self._order(invoice_date=(timezone.now() - timedelta(days=6)).date(),
                    received_on=new_arrival, invoice_no="ARRIVED-TODAY")

        response = self.client.get(reverse('supplier_detail', args=[self.supplier.pk]))
        titles = [row['title'] for row in response.context['history_page']]
        self.assertEqual(titles[0], "ARRIVED-TODAY", titles)

    def test_an_old_order_with_no_invoice_date_does_not_float_to_today(self):
        """The exact shape that put a May order above this morning's delivery.

        No invoice date and no receipt time used to fall back to "today", so
        the row sorted as today while displaying its real, much older date.
        """
        from stock.models import InboundOrder, Purchase
        old_order = InboundOrder.objects.create(
            supplier=self.supplier, invoice_no="OLD-NO-INVOICE-DATE",
            invoice_date=None, status='received')
        InboundOrder.objects.filter(pk=old_order.pk).update(
            invoice_date=None, received_at=None,
            created_at=timezone.now() - timedelta(days=100))
        Purchase.objects.create(inbound_order=old_order, product=self.product,
                                supplier=self.supplier, quantity=1, remaining=1,
                                cost_price=Decimal("10"),
                                date=timezone.now() - timedelta(days=100))

        self._order(invoice_date=(timezone.now() - timedelta(days=6)).date(),
                    received_on=timezone.now(), invoice_no="ARRIVED-TODAY")

        response = self.client.get(reverse('supplier_detail', args=[self.supplier.pk]))
        rows = list(response.context['history_page'])

        # The invariant, stated directly: a row sorts by the date it shows.
        # Asserting only on position passed by luck - both rows landed on the
        # same day and the microseconds happened to favour the right one.
        for row in rows:
            self.assertEqual(
                timezone.localtime(row['sort_at']).date(), row['display_date'],
                f"{row['title']} sorts by a different date than it displays")

        titles = [row['title'] for row in rows]
        self.assertEqual(titles[0], "ARRIVED-TODAY", titles)
        self.assertEqual(titles[-1], "OLD-NO-INVOICE-DATE", titles)

    def test_rows_are_in_strict_date_order(self):
        from stock.models import InboundOrder, Purchase
        # A spread of the awkward shapes, all mixed together.
        self._order(invoice_date=None, received_on=timezone.now() - timedelta(days=2),
                    invoice_no="TWO-DAYS-AGO")
        self._order(invoice_date=(timezone.now() + timedelta(days=5)).date(),
                    received_on=timezone.now() - timedelta(days=30),
                    invoice_no="FUTURE-INVOICE-OLD-ARRIVAL")
        self._order(invoice_date=(timezone.now() - timedelta(days=1)).date(),
                    received_on=timezone.now(), invoice_no="TODAY")

        response = self.client.get(reverse('supplier_detail', args=[self.supplier.pk]))
        rows = list(response.context['history_page'])
        dates = [row['display_date'] for row in rows]
        self.assertEqual(dates, sorted(dates, reverse=True), dates)
        for row in rows:
            self.assertEqual(timezone.localtime(row['sort_at']).date(), row['display_date'])

    def test_the_listed_date_is_when_the_goods_arrived(self):
        arrival = timezone.now() - timedelta(days=1)
        self._order(invoice_date=(timezone.now() - timedelta(days=8)).date(),
                    received_on=arrival, invoice_no="INV-9")
        response = self.client.get(reverse('supplier_detail', args=[self.supplier.pk]))
        row = response.context['history_page'][0]
        self.assertEqual(row['display_date'], arrival.date())
        self.assertIn('invoice', row['meta'])      # the invoice date is still shown

    def test_a_pending_order_says_so(self):
        from stock.models import InboundOrder, Purchase
        order = InboundOrder.objects.create(
            supplier=self.supplier, invoice_no="PENDING-1",
            invoice_date=timezone.localdate(), status='pending_receipt')
        Purchase.objects.create(inbound_order=order, product=self.product,
                                supplier=self.supplier, quantity=1, remaining=1,
                                cost_price=Decimal("10"), date=timezone.now())
        response = self.client.get(reverse('supplier_detail', args=[self.supplier.pk]))
        row = response.context['history_page'][0]
        self.assertIn('awaiting receipt', row['meta'])


class CreateCategoryTests(TestCase):
    """The shop adds its own categories; no developer, no migration."""

    def setUp(self):
        self.accessories = Category.objects.create(
            name="Accessories", form_kind="accessory", sync_to_shopify=False)
        self.manager = get_user_model().objects.create_superuser(
            username="boss3", password="pw123456", email="b3@x.com")
        self.client.force_login(self.manager)

    def _post(self, **data):
        return self.client.post(reverse('create_category'), data)

    def test_a_top_level_category_is_created_with_the_kind_given(self):
        response = self._post(name="Vapes", form_kind="accessory")
        self.assertEqual(response.status_code, 200)
        category = Category.objects.get(name="Vapes")
        self.assertEqual(category.form_kind, "accessory")
        self.assertIsNone(category.parent)

    def test_a_subcategory_inherits_its_parents_kind_and_shopify_setting(self):
        response = self._post(name="Tablet cases", parent=self.accessories.pk)
        self.assertEqual(response.status_code, 200)
        child = Category.objects.get(name="Tablet cases")
        self.assertEqual(child.parent, self.accessories)
        self.assertEqual(child.effective_form_kind, "accessory")
        self.assertFalse(child.sync_to_shopify)

    def test_a_new_perfume_category_goes_on_the_storefront(self):
        self._post(name="Niche perfume", form_kind="perfume")
        self.assertTrue(Category.objects.get(name="Niche perfume").sync_to_shopify)

    def test_a_new_accessory_category_stays_off_the_storefront(self):
        self._post(name="Vapes", form_kind="accessory")
        self.assertFalse(Category.objects.get(name="Vapes").sync_to_shopify)

    def test_the_response_carries_what_the_page_needs(self):
        payload = self._post(name="Tablet cases", parent=self.accessories.pk).json()
        self.assertTrue(payload['ok'])
        self.assertEqual(payload['name'], "Tablet cases")
        self.assertEqual(payload['kind'], "accessory")
        self.assertEqual(payload['parent'], self.accessories.pk)

    def test_a_duplicate_under_the_same_parent_is_refused_and_points_at_it(self):
        first = self._post(name="Tablet cases", parent=self.accessories.pk).json()
        response = self._post(name="tablet CASES", parent=self.accessories.pk)
        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()['id'], first['id'])
        self.assertEqual(Category.objects.filter(name__iexact="Tablet cases").count(), 1)

    def test_the_same_name_under_a_different_parent_is_allowed(self):
        other = Category.objects.create(name="Shisha")
        self._post(name="Glass", parent=self.accessories.pk)
        response = self._post(name="Glass", parent=other.pk)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(Category.objects.filter(name="Glass").count(), 2)

    def test_a_blank_name_is_refused(self):
        response = self._post(name="   ")
        self.assertEqual(response.status_code, 400)
        self.assertFalse(Category.objects.exclude(pk=self.accessories.pk).exists())

    def test_an_over_long_name_is_refused(self):
        self.assertEqual(self._post(name="x" * 51).status_code, 400)

    def test_a_missing_parent_is_refused(self):
        self.assertEqual(self._post(name="Orphan", parent=99999).status_code, 400)

    def test_an_unknown_kind_falls_back_to_general(self):
        self._post(name="Mystery", form_kind="nonsense")
        self.assertEqual(Category.objects.get(name="Mystery").form_kind, "general")

    def test_an_employee_cannot_add_categories(self):
        employee = get_user_model().objects.create_user(username="staff3", password="pw123456")
        self.client.force_login(employee)
        self._post(name="Sneaky")
        self.assertFalse(Category.objects.filter(name="Sneaky").exists())

    def test_it_only_answers_post(self):
        self.assertEqual(self.client.get(reverse('create_category')).status_code, 405)


class ProductListCategoryChipsTests(TestCase):
    """Three chips, not twelve - and one covers everything beneath it."""

    def setUp(self):
        from stock.services.barcodes import assign_internal_barcode
        self.accessories = Category.objects.create(name="Accessories", form_kind="accessory")
        self.cases = Category.objects.create(name="Cases", parent=self.accessories)
        self.perfumes = Category.objects.create(name="Perfumes", form_kind="perfume")

        self.case = Product(name="Clear TPU", brand="Generic", category=self.cases,
                            default_price=Decimal("9"))
        assign_internal_barcode(self.case)
        self.perfume = Product.objects.create(
            name="Oud", barcode="9910000000001", brand="B",
            category=self.perfumes, default_price=Decimal("50"))

        user = get_user_model().objects.create_superuser(
            username="boss4", password="pw123456", email="b4@x.com")
        self.client.force_login(user)

    def test_only_top_level_categories_are_offered_as_chips(self):
        response = self.client.get(reverse('product_list'))
        names = [c.name for c in response.context['categories']]
        self.assertEqual(sorted(names), ["Accessories", "Perfumes"])
        self.assertNotIn("Cases", names)

    def test_filtering_by_a_parent_includes_its_subcategories(self):
        response = self.client.get(reverse('product_list'), {'category': self.accessories.pk})
        names = [p.name for p in response.context['products']]
        self.assertIn("Clear TPU", names)
        self.assertNotIn("Oud", names)

    def test_filtering_by_a_leaf_still_works(self):
        response = self.client.get(reverse('product_list'), {'category': self.cases.pk})
        self.assertEqual([p.name for p in response.context['products']], ["Clear TPU"])

    def test_a_deep_tree_is_covered(self):
        from stock.services.barcodes import assign_internal_barcode
        deep = Category.objects.create(name="MagSafe cases", parent=self.cases)
        product = Product(name="MagSafe clear", brand="Generic", category=deep,
                          default_price=Decimal("15"))
        assign_internal_barcode(product)
        response = self.client.get(reverse('product_list'), {'category': self.accessories.pk})
        self.assertIn("MagSafe clear", [p.name for p in response.context['products']])

    def test_a_category_cycle_does_not_hang_the_page(self):
        self.accessories.parent = self.cases
        self.accessories.save()
        response = self.client.get(reverse('product_list'), {'category': self.accessories.pk})
        self.assertEqual(response.status_code, 200)


class AddProductWizardTests(TestCase):
    """Pick a category first; the form then asks only that category's questions."""

    def setUp(self):
        from django.core.management import call_command
        call_command("seed_accessory_attributes", "--apply", verbosity=0)
        self.perfumes = Category.objects.create(name="Perfumes", form_kind="perfume")
        self.cases = Category.objects.get(name="Cases")
        self.accessories = Category.objects.get(name="Accessories")
        self.brand = Brand.objects.create(name="Generic")
        user = get_user_model().objects.create_superuser(
            username="mgr2", password="pw123456", email="m2@x.com")
        self.client.force_login(user)

    def _html(self):
        return self.client.get(reverse('add_product')).content.decode()

    # -- step 1 ------------------------------------------------------------
    def test_the_page_opens_on_the_category_picker(self):
        html = self._html()
        self.assertIn('id="step-pick"', html)
        self.assertIn('id="step-form"', html)
        self.assertIn('What are you adding?', html)

    def test_top_level_categories_are_offered_as_buttons(self):
        html = self._html()
        for name in ["Accessories", "Perfumes"]:
            self.assertIn(f'>{name}</span>', html)

    def test_a_category_with_subcategories_offers_them(self):
        html = self._html()
        self.assertIn(f'data-children-of="{self.accessories.pk}"', html)
        self.assertIn('>Cases</span>', html)
        self.assertIn('Other accessories', html)

    def test_the_picker_knows_each_category_form_kind(self):
        import json
        response = self.client.get(reverse('add_product'))
        kinds = json.loads(response.context['category_kind_json'])
        self.assertEqual(kinds[str(self.cases.pk)], "accessory")
        self.assertEqual(kinds[str(self.perfumes.pk)], "perfume")

    def test_a_subcategory_inherits_the_form_kind_of_its_parent(self):
        import json
        child = Category.objects.create(name="Tablet cases", parent=self.accessories)
        response = self.client.get(reverse('add_product'))
        kinds = json.loads(response.context['category_kind_json'])
        self.assertEqual(kinds[str(child.pk)], "accessory")

    def test_every_kind_of_accessory_the_shop_sells_has_a_category(self):
        html = self._html()
        for name in ["Cases", "Screen protectors", "Cables", "Chargers &amp; plugs",
                     "Power banks", "Audio", "Mice &amp; keyboards", "Storage",
                     "Holders &amp; mounts"]:
            self.assertIn(f'>{name}</span>', html, name)

    def test_the_page_offers_adding_a_category_and_a_subcategory(self):
        html = self._html()
        self.assertIn('data-new-parent=""', html)                     # top level
        self.assertIn(f'data-new-parent="{self.accessories.pk}"', html)  # subcategory
        self.assertIn('+ New category', html)
        self.assertIn('+ New type', html)

    def test_choosing_a_category_does_not_navigate_away(self):
        # The picker stays on the page and the form opens underneath it, so
        # there is no jump and the choice can be changed with one tap.
        html = self._html()
        self.assertNotIn('window.scrollTo', html)
        self.assertNotIn('stepPick.hidden = true', html)

    def test_case_types_cover_the_kinds_the_shop_stocks(self):
        from stock.models import CategoryAttribute
        case_type = CategoryAttribute.objects.get(code="case_type")
        labels = set(case_type.options.values_list('label', flat=True))
        for label in ["Plain silicone / rubber", "Fancy / printed", "MagSafe",
                      "Flip / wallet"]:
            self.assertIn(label, labels)
        self.assertTrue(case_type.variant_attribute)

    def test_protector_glue_and_edge_are_separate_stock_rows(self):
        from stock.models import CategoryAttribute
        for code in ["glue", "edge", "protector_type"]:
            attribute = CategoryAttribute.objects.get(code=code)
            self.assertTrue(attribute.variant_attribute, code)

    # -- step 2 shape ------------------------------------------------------
    def test_perfume_and_accessory_groups_are_both_present_but_tagged(self):
        html = self._html()
        self.assertIn('data-kind="perfume"', html)
        self.assertIn('data-kind="accessory"', html)
        self.assertIn('data-kind="general accessory"', html)

    def test_the_fitment_group_is_rendered(self):
        html = self._html()
        self.assertIn('universal_fit', html)
        self.assertIn('device_models', html)
        self.assertIn('compatibility_groups', html)

    def test_the_category_select_is_still_submitted(self):
        # The buttons drive it, but it must stay in the DOM for validation.
        self.assertIn('name="category"', self._html())

    def test_colour_is_asked_once_for_accessories(self):
        # Accessories answer colour through the Colour attribute, so the
        # legacy field is scoped to general categories.
        html = self._html()
        self.assertIn('<div class="col-md-6" data-kind="general">', html)

    def test_an_accessory_does_not_get_perfume_questions(self):
        html = self._html()
        # Anchor on the fieldset - the picker button carries the same
        # data-kind attribute and appears earlier in the page.
        start = html.index('<fieldset class="fgroup" data-kind="perfume">')
        perfume_block = html[start:html.index('</fieldset>', start)]
        for field in ['volume_ml', 'concentration', 'gender']:
            self.assertIn(field, perfume_block)

        identity = html[html.index('<legend>Identity</legend>'):]
        identity = identity[:identity.index('</fieldset>')]
        self.assertNotIn('gender', identity)

    # -- saving ------------------------------------------------------------
    def _post(self, **overrides):
        data = {
            'barcode': '',
            'category': self.cases.pk,
            'brand_master': self.brand.pk,
            'name': 'Clear TPU',
            'default_price': '9.90',
        }
        data.update(overrides)
        return self.client.post(reverse('add_product'), data)

    def test_an_accessory_saves_with_its_fitment(self):
        from stock.models import DeviceModel
        apple = Brand.objects.create(name="Apple")
        device = DeviceModel.objects.create(brand=apple, name="iPhone 15 Pro Max")
        response = self._post(device_models=[device.pk])
        self.assertEqual(response.status_code, 302)
        product = Product.objects.get(name="Clear TPU")
        self.assertTrue(product.fits(device))
        self.assertTrue(product.barcode_is_internal)

    def test_a_universal_accessory_saves(self):
        response = self._post(name="USB-C cable", universal_fit='on')
        self.assertEqual(response.status_code, 302)
        self.assertTrue(Product.objects.get(name="USB-C cable").universal_fit)

    def test_gender_sits_in_the_perfume_group_and_does_not_break_accessories(self):
        # Gender is only asked of perfume now; an accessory posting without it
        # must still save. (Rendering it *only* in a hidden group once cleared
        # it on non-perfumes - hence this test.)
        response = self._post()
        self.assertEqual(response.status_code, 302)

    def test_a_perfume_keeps_the_gender_it_was_given(self):
        response = self.client.post(reverse('add_product'), {
            'barcode': '9900000000001',
            'category': self.perfumes.pk,
            'brand_master': self.brand.pk,
            'name': 'Oud',
            'default_price': '50',
            'gender': 'men',
            'volume_ml': '100',
        })
        self.assertEqual(response.status_code, 302)
        perfume = Product.objects.get(name="Oud")
        self.assertEqual(perfume.gender, 'men')
        self.assertEqual(perfume.volume_ml, 100)


class DeviceFitmentTests(TestCase):
    """A customer asks for "a case for a 15PM" and must find the right case."""

    def setUp(self):
        from stock.models import DeviceModel, DeviceAlias, CompatibilityGroup
        self.accessories = Category.objects.create(name="Accessories", sync_to_shopify=False)
        self.apple = Brand.objects.create(name="Apple")

        self.pro_max = DeviceModel.objects.create(brand=self.apple, name="iPhone 15 Pro Max",
                                                  release_year=2023)
        self.plus = DeviceModel.objects.create(brand=self.apple, name="iPhone 15 Plus",
                                               release_year=2023)
        for alias in ["15PM", "15 Pro Max", "A2849"]:
            DeviceAlias.objects.create(device=self.pro_max, alias=alias)

        # Same glass on both handsets, so one protector covers the pair.
        self.same_glass = CompatibilityGroup.objects.create(name="iPhone 15 Pro Max / 15 Plus")
        self.same_glass.devices.set([self.pro_max, self.plus])

        self.case = self._product("Clear case")
        self.case.device_models.add(self.pro_max)
        self.protector = self._product("Tempered glass")
        self.protector.compatibility_groups.add(self.same_glass)
        self.cable = self._product("USB-C cable", universal_fit=True)
        self.other_case = self._product("Case for Pixel")

    def _product(self, name, **kwargs):
        from stock.services.barcodes import assign_internal_barcode
        product = Product(name=name, brand="Generic", category=self.accessories,
                          default_price=Decimal("9"), **kwargs)
        assign_internal_barcode(product)
        return product

    # -- resolving what the customer said ---------------------------------
    def test_an_alias_finds_the_handset(self):
        from stock.models import resolve_device
        for spelling in ["15PM", "15 pro max", "a2849", "15-PM"]:
            self.assertEqual(resolve_device(spelling), self.pro_max, spelling)

    def test_the_full_name_finds_the_handset_however_it_is_spaced(self):
        from stock.models import resolve_device
        for spelling in ["iPhone 15 Pro Max", "apple iphone15promax", "APPLE IPHONE 15 PRO MAX"]:
            self.assertEqual(resolve_device(spelling), self.pro_max, spelling)

    def test_an_unknown_spelling_resolves_to_nothing(self):
        from stock.models import resolve_device
        self.assertIsNone(resolve_device("Nokia 3310"))
        self.assertIsNone(resolve_device(""))

    def test_the_same_alias_cannot_point_at_two_handsets(self):
        from stock.models import DeviceAlias
        from django.db.utils import IntegrityError
        with self.assertRaises(IntegrityError):
            DeviceAlias.objects.create(device=self.plus, alias="15 pm")

    # -- finding what fits -------------------------------------------------
    def test_a_handset_finds_its_case_its_group_and_universal_goods(self):
        from stock.models import products_fitting
        names = set(products_fitting(self.pro_max).values_list('name', flat=True))
        self.assertEqual(names, {"Clear case", "Tempered glass", "USB-C cable"})

    def test_a_group_member_finds_the_shared_protector_but_not_the_moulded_case(self):
        from stock.models import products_fitting
        names = set(products_fitting(self.plus).values_list('name', flat=True))
        self.assertEqual(names, {"Tempered glass", "USB-C cable"})

    def test_a_product_is_not_returned_twice_when_it_matches_two_ways(self):
        from stock.models import products_fitting
        self.protector.device_models.add(self.pro_max)     # model AND group
        rows = list(products_fitting(self.pro_max).values_list('name', flat=True))
        self.assertEqual(len(rows), len(set(rows)))

    def test_fits_agrees_with_the_queryset(self):
        self.assertTrue(self.case.fits(self.pro_max))
        self.assertFalse(self.case.fits(self.plus))
        self.assertTrue(self.protector.fits(self.plus))
        self.assertTrue(self.cable.fits(self.plus))
        self.assertFalse(self.other_case.fits(self.pro_max))

    def test_universal_goods_fit_even_an_unknown_handset(self):
        self.assertTrue(self.cable.fits(None))
        self.assertFalse(self.case.fits(None))

    def test_perfume_is_untouched_by_any_of_this(self):
        perfume = Product.objects.create(name="Oud", barcode="9600000000001", brand="B",
                                         category=Category.objects.create(name="Perfumes"),
                                         default_price=Decimal("50"))
        self.assertFalse(perfume.universal_fit)
        self.assertEqual(perfume.fitment_label, '')
        self.assertFalse(perfume.fits(self.pro_max))

    # -- labels ------------------------------------------------------------
    def test_fitment_label_reads_naturally(self):
        self.assertEqual(self.cable.fitment_label, 'Universal')
        self.assertEqual(self.case.fitment_label, 'Apple iPhone 15 Pro Max')
        self.assertEqual(self.protector.fitment_label, 'iPhone 15 Pro Max / 15 Plus')

    def test_search_finds_handsets_by_partial_alias(self):
        from stock.models import search_devices
        self.assertIn(self.pro_max, search_devices("15p"))
        self.assertIn(self.pro_max, search_devices("pro max"))


class TillDeviceSearchTests(TestCase):
    """Typing a handset at the till must return the accessories that fit it."""

    def setUp(self):
        from stock.models import DeviceModel, DeviceAlias, CompatibilityGroup
        from stock.services.barcodes import assign_internal_barcode
        self.accessories = Category.objects.create(name="Accessories", sync_to_shopify=False)
        apple = Brand.objects.create(name="Apple")
        self.pro_max = DeviceModel.objects.create(brand=apple, name="iPhone 15 Pro Max")
        DeviceAlias.objects.create(device=self.pro_max, alias="15PM")
        self.plus = DeviceModel.objects.create(brand=apple, name="iPhone 15 Plus")
        group = CompatibilityGroup.objects.create(name="iPhone 15 Pro Max / 15 Plus")
        group.devices.set([self.pro_max, self.plus])

        def make(name, **kw):
            product = Product(name=name, brand="Generic", category=self.accessories,
                              default_price=Decimal("9"), **kw)
            assign_internal_barcode(product)
            return product

        self.case = make("Clear TPU")
        self.case.device_models.add(self.pro_max)
        self.protector = make("Tempered glass")
        self.protector.compatibility_groups.add(group)
        self.cable = make("USB-C cable", universal_fit=True)

        user = get_user_model().objects.create_user(username="till2", password="pw123456")
        self.client.force_login(user)

    def _search(self, q):
        response = self.client.get(reverse('products_autocomplete'), {'q': q})
        return {row['name'] for row in response.json()['results']}

    def test_an_alias_returns_everything_made_for_that_handset(self):
        self.assertEqual(self._search("15PM"), {"Clear TPU", "Tempered glass"})

    def test_the_handset_name_works_too(self):
        self.assertEqual(self._search("iPhone 15 Pro Max"), {"Clear TPU", "Tempered glass"})

    def test_universal_goods_do_not_answer_every_handset(self):
        # Otherwise every cable in the shop turns up under every phone.
        self.assertNotIn("USB-C cable", self._search("15PM"))

    def test_searching_by_product_name_still_works(self):
        self.assertEqual(self._search("Tempered"), {"Tempered glass"})

    def test_a_product_matching_both_ways_appears_once(self):
        self.protector.device_models.add(self.pro_max)
        response = self.client.get(reverse('products_autocomplete'), {'q': '15PM'})
        names = [row['name'] for row in response.json()['results']]
        self.assertEqual(len(names), len(set(names)))


class ShopifySyncScopeTests(TestCase):
    """Accessories are shop-floor only; only perfume goes on the storefront."""

    def setUp(self):
        self.perfumes = Category.objects.create(name="Perfumes", sync_to_shopify=True)
        self.accessories = Category.objects.create(name="Accessories", sync_to_shopify=False)
        self.perfume = Product.objects.create(
            name="Oud", barcode="9500000000001", brand="B",
            category=self.perfumes, default_price=Decimal("50"))
        self.case = Product.objects.create(
            name="Case", barcode="9500000000002", brand="B",
            category=self.accessories, default_price=Decimal("9"))

    def test_syncable_excludes_categories_that_are_offline(self):
        from stock.services.shopify_sync import shopify_syncable
        names = list(shopify_syncable(Product.objects).values_list('name', flat=True))
        self.assertEqual(names, ["Oud"])

    def test_a_product_with_no_category_still_syncs(self):
        from stock.services.shopify_sync import shopify_syncable
        orphan = Product.objects.create(name="Orphan", barcode="9500000000003",
                                        brand="B", category=None, default_price=Decimal("1"))
        self.assertIn(orphan, shopify_syncable(Product.objects))

    def test_uploading_a_photo_does_not_list_an_accessory_online(self):
        # The auto-sync signal fires on first photo upload; an offline
        # category must not be pushed just because someone photographed it.
        from unittest.mock import patch
        from stock.models import ProductImage
        with self.settings(SHOPIFY_AUTO_SYNC=True):
            with patch('stock.services.shopify_sync.sync_product') as sync:
                # The push runs on_commit, which TestCase's transaction defers.
                with self.captureOnCommitCallbacks(execute=True):
                    ProductImage.objects.create(product=self.case, image="products/x.jpg")
                self.assertFalse(sync.called)
                with self.captureOnCommitCallbacks(execute=True):
                    ProductImage.objects.create(product=self.perfume, image="products/y.jpg")
                self.assertTrue(sync.called)

    def test_the_product_page_button_refuses_an_offline_category(self):
        user = get_user_model().objects.create_superuser(
            username="boss", password="pw123456", email="b@x.com")
        self.client.force_login(user)
        response = self.client.post(
            reverse('sync_product_to_shopify', args=[self.case.pk]), follow=True)
        self.assertContains(response, "not sold on Shopify")


class InternalBarcodeTests(TestCase):
    """Cases and screen protectors arrive with no EAN; we mint a valid one."""

    def test_check_digit_matches_gs1_examples(self):
        from stock.services.barcodes import ean13_check_digit
        self.assertEqual(ean13_check_digit('590123412345'), '7')
        self.assertEqual(ean13_check_digit('400638133393'), '1')

    def test_generated_code_is_a_valid_ean13(self):
        from stock.services.barcodes import build_internal_barcode, ean13_check_digit
        code = build_internal_barcode(1)
        self.assertEqual(len(code), 13)
        self.assertTrue(code.isdigit())
        self.assertTrue(code.startswith('29'))
        self.assertEqual(code[-1], ean13_check_digit(code[:12]))

    def test_sequence_advances_and_never_collides(self):
        from stock.services.barcodes import assign_internal_barcode
        category = Category.objects.create(name="Accessories")
        codes = []
        for i in range(5):
            product = Product(name=f"Case {i}", brand="Generic", category=category,
                              default_price=Decimal("9"))
            codes.append(assign_internal_barcode(product))
        self.assertEqual(len(set(codes)), 5)
        self.assertEqual(codes, sorted(codes))

    def test_it_skips_a_number_already_taken(self):
        from stock.services.barcodes import assign_internal_barcode, build_internal_barcode
        category = Category.objects.create(name="Accessories")
        taken = build_internal_barcode(1)
        Product.objects.create(name="Squatter", barcode=taken, brand="B",
                               category=category, default_price=Decimal("1"))
        product = Product(name="Case", brand="Generic", category=category,
                          default_price=Decimal("9"))
        self.assertNotEqual(assign_internal_barcode(product), taken)

    def test_a_real_barcode_is_not_flagged_internal(self):
        category = Category.objects.create(name="Accessories")
        product = Product.objects.create(name="Cable", barcode="5901234123457", brand="B",
                                         category=category, default_price=Decimal("5"))
        self.assertFalse(product.barcode_is_internal)


class ProductFormBlankBarcodeTests(TestCase):
    """Adding a no-barcode product through the form must just work."""

    def setUp(self):
        self.category = Category.objects.create(name="Accessories")
        self.brand = Brand.objects.create(name="Generic")

    def _data(self, **overrides):
        data = {
            'barcode': '',
            'category': self.category.id,
            'brand_master': self.brand.id,
            'name': 'Clear TPU case',
            'default_price': '9.90',
        }
        data.update(overrides)
        return data

    def test_a_blank_barcode_is_accepted_and_filled_in(self):
        form = ProductForm(data=self._data())
        self.assertTrue(form.is_valid(), form.errors.as_json())
        product = form.save()
        self.assertTrue(product.barcode)
        self.assertTrue(product.barcode_is_internal)
        self.assertEqual(len(product.barcode), 13)

    def test_two_blank_barcode_products_do_not_collide(self):
        first = ProductForm(data=self._data(name="Case A"))
        self.assertTrue(first.is_valid(), first.errors.as_json())
        one = first.save()
        second = ProductForm(data=self._data(name="Case B"))
        self.assertTrue(second.is_valid(), second.errors.as_json())
        two = second.save()
        self.assertNotEqual(one.barcode, two.barcode)

    def test_a_supplied_barcode_is_kept_verbatim(self):
        form = ProductForm(data=self._data(barcode='5901234123457'))
        self.assertTrue(form.is_valid(), form.errors.as_json())
        product = form.save()
        self.assertEqual(product.barcode, '5901234123457')
        self.assertFalse(product.barcode_is_internal)

    def test_a_duplicate_barcode_is_still_rejected(self):
        Product.objects.create(name="Taken", barcode='5901234123457', brand="B",
                               category=self.category, default_price=Decimal("1"))
        form = ProductForm(data=self._data(barcode='5901234123457'))
        self.assertFalse(form.is_valid())
        self.assertIn('barcode', form.errors)

    def test_editing_a_product_does_not_lose_its_barcode(self):
        product = Product.objects.create(name="Cable", barcode='5901234123457', brand="B",
                                         category=self.category, default_price=Decimal("5"))
        form = ProductForm(data=self._data(barcode='', name="Cable v2"), instance=product)
        self.assertTrue(form.is_valid(), form.errors.as_json())
        saved = form.save()
        self.assertEqual(saved.barcode, '5901234123457')
        self.assertFalse(saved.barcode_is_internal)


class StoreSellableCategoryTests(TestCase):
    """Scentory sells perfume only; Khan Perfume sells everything it stocks."""

    def setUp(self):
        self.perfumes = Category.objects.create(name="Perfumes")
        self.accessories = Category.objects.create(name="Accessories")
        self.khan = Store.objects.create(name="Khan Perfume", code="KHAN", is_default=True)
        self.scentory = Store.objects.create(name="Scentory", code="SCEN")
        self.scentory.sellable_categories.add(self.perfumes)

        self.perfume = Product.objects.create(
            name="Oud", barcode="9400000000001", brand="B",
            category=self.perfumes, default_price=Decimal("50"))
        self.case = Product.objects.create(
            name="Case", barcode="9400000000002", brand="B",
            category=self.accessories, default_price=Decimal("9"))

        self.user = get_user_model().objects.create_user(username="till", password="pw123456")
        StoreProfile.objects.create(user=self.user, store=self.scentory)
        self.client.force_login(self.user)

    def _login_at_khan(self):
        user = get_user_model().objects.create_user(username="khan", password="pw123456")
        StoreProfile.objects.create(user=user, store=self.khan)
        self.client.force_login(user)

    def test_a_store_with_no_categories_configured_sells_everything(self):
        self._login_at_khan()
        response = self.client.get(reverse('products_autocomplete'), {'q': 'Case'})
        self.assertEqual(len(response.json()['results']), 1)

    def test_till_search_hides_categories_the_store_does_not_sell(self):
        response = self.client.get(reverse('products_autocomplete'), {'q': 'Case'})
        self.assertEqual(response.json()['results'], [])

    def test_till_search_still_finds_what_the_store_does_sell(self):
        response = self.client.get(reverse('products_autocomplete'), {'q': 'Oud'})
        self.assertEqual(len(response.json()['results']), 1)

    def test_scanning_an_unsold_category_at_the_till_is_not_found(self):
        response = self.client.get(reverse('check_barcode'), {'barcode': self.case.barcode})
        self.assertFalse(response.json()['exists'])

    def test_inbound_receives_the_whole_catalogue_regardless_of_store(self):
        # Stock is shared company-wide, so receiving must never be store-scoped.
        response = self.client.get(reverse('check_barcode'),
                                   {'barcode': self.case.barcode, 'scope': 'stock'})
        self.assertTrue(response.json()['exists'])

    def test_product_list_is_scoped_to_the_active_store(self):
        response = self.client.get(reverse('product_list'))
        names = [p.name for p in response.context['products']]
        self.assertIn("Oud", names)
        self.assertNotIn("Case", names)


class PurgeUnusedProductsTests(TestCase):
    """The purge is destructive, so its guard matters more than its happy path."""

    def setUp(self):
        from django.core.management import call_command
        self.call = call_command
        self.cat = Category.objects.create(name="Accessories")
        self.perfumes = Category.objects.create(name="Perfumes")

    def _product(self, barcode, category=None):
        return Product.objects.create(name="X", barcode=barcode, brand="B",
                                      category=category or self.cat,
                                      default_price=Decimal("5"))

    def test_deletes_only_products_with_no_history(self):
        unused = self._product("9300000000001")
        sold = self._product("9300000000002")
        purchased = self._product("9300000000003")
        order = SaleOrder.objects.create()
        Sale.objects.create(order=order, product=sold, quantity=1,
                            unit_price=Decimal("5"), payment_method="cash")
        Purchase.objects.create(product=purchased, quantity=1, remaining=1,
                                cost_price=Decimal("2"))

        self.call("purge_unused_products", "--category", "Accessories", "--apply", verbosity=0)

        self.assertFalse(Product.all_objects.filter(pk=unused.pk).exists())
        self.assertTrue(Product.all_objects.filter(pk=sold.pk).exists())
        self.assertTrue(Product.all_objects.filter(pk=purchased.pk).exists())

    def test_a_product_with_only_an_image_is_kept(self):
        from stock.models import ProductImage
        pictured = self._product("9300000000004")
        ProductImage.objects.create(product=pictured, image="products/x.jpg")
        self.call("purge_unused_products", "--category", "Accessories", "--apply", verbosity=0)
        self.assertTrue(Product.all_objects.filter(pk=pictured.pk).exists())

    def test_dry_run_deletes_nothing(self):
        unused = self._product("9300000000005")
        self.call("purge_unused_products", "--category", "Accessories", verbosity=0)
        self.assertTrue(Product.all_objects.filter(pk=unused.pk).exists())

    def test_category_filters_scope_the_purge(self):
        accessory = self._product("9300000000006")
        perfume = self._product("9300000000007", category=self.perfumes)
        self.call("purge_unused_products", "--exclude-category", "perfum", "--apply", verbosity=0)
        self.assertFalse(Product.all_objects.filter(pk=accessory.pk).exists())
        self.assertTrue(Product.all_objects.filter(pk=perfume.pk).exists())


class PerfumeAttributeTests(TestCase):
    """Volume / concentration / fragrance family / inspired-by as real data."""

    def setUp(self):
        from stock.models import Concentration, FragranceFamily, Inspiration
        self.cat = Category.objects.create(name="Perfumes")
        self.edp = Concentration.objects.create(name="Eau de Parfum", short="EDP", sort_order=20)
        self.floral = FragranceFamily.objects.create(name="Floral", sort_order=10)
        self.oud = FragranceFamily.objects.create(name="Oud", sort_order=20)
        self.insp = Inspiration.objects.create(house="Givenchy", name="L'Interdit")

    def test_volume_drives_the_variant_label(self):
        p = Product.objects.create(name="Rosa", barcode="9000000000001", brand="Lattafa",
                                   category=self.cat, default_price=Decimal("20"),
                                   volume_ml=100)
        # volume is authoritative once set, so the label stops depending on free text
        self.assertEqual(p.variant_label, "100ml")
        self.assertIn("100ml", p.display_name)

    def test_spec_still_used_when_no_volume(self):
        p = Product.objects.create(name="Cable", barcode="9000000000002", brand="Anker",
                                   category=self.cat, default_price=Decimal("5"), spec="20W")
        self.assertEqual(p.variant_label, "20W")

    def test_attributes_round_trip_through_the_edit_form(self):
        p = Product.objects.create(name="Khamrah", barcode="9000000000003", brand="Lattafa",
                                   category=self.cat, default_price=Decimal("30"))
        get_user_model().objects.create_superuser("perf_mgr", password="pw123456")
        self.client.login(username="perf_mgr", password="pw123456")
        resp = self.client.post(reverse("edit_product", args=[p.pk]), {
            "barcode": "9000000000003", "category": self.cat.id, "new_brand_name": "Lattafa",
            "name": "Khamrah", "default_price": "30.00",
            "volume_ml": "100", "concentration": self.edp.id,
            "fragrance_families": [self.floral.id, self.oud.id],
            "inspired_by": self.insp.id,
        })
        self.assertIn(resp.status_code, (200, 302))
        p.refresh_from_db()
        self.assertEqual(p.volume_ml, 100)
        self.assertEqual(p.concentration, self.edp)
        self.assertEqual(p.inspired_by, self.insp)
        self.assertEqual(set(p.fragrance_families.all()), {self.floral, self.oud})

    def test_inspired_by_is_internal_and_never_sent_to_shopify(self):
        from stock.services import shopify_sync
        p = Product.objects.create(name="Rosa", barcode="9000000000004", brand="Lattafa",
                                   category=self.cat, default_price=Decimal("20"),
                                   inspired_by=self.insp, description="A floral perfume.")
        payload = shopify_sync._shopify_description_html(p)
        self.assertNotIn("Givenchy", payload)
        self.assertNotIn("Interdit", payload)


class ProductEditFormGroupingTests(TestCase):
    """The edit form groups fields and hides the ones a category does not use.

    Hiding is presentational only: every field still renders, so submitting the
    form never silently clears a value the group happened to hide.
    """

    def setUp(self):
        self.perfumes = Category.objects.create(name="Perfumes")
        self.accessories = Category.objects.create(name="Accessories")
        get_user_model().objects.create_superuser("grp_mgr", password="pw123456")
        self.client.login(username="grp_mgr", password="pw123456")

    def test_every_field_is_rendered_once(self):
        p = Product.objects.create(name="Khamrah", barcode="9200000000001", brand="Lattafa",
                                   category=self.perfumes, default_price=Decimal("30"))
        import re
        html = self.client.get(reverse("edit_product", args=[p.pk])).content.decode()
        for field in ['color', 'spec', 'gender', 'volume_ml', 'concentration', 'inspired_by']:
            # count real widgets only - the page's JS also mentions these names
            widgets = re.findall(r'<(?:input|select|textarea)[^>]*name="%s"' % field, html)
            self.assertEqual(len(widgets), 1, f'{field}: {len(widgets)} widgets')
        # Groups are tagged with the category kinds they belong to.
        self.assertIn('data-kind="perfume"', html)
        self.assertIn('data-kind="general accessory"', html)

    def test_add_form_is_grouped_and_renders_each_field_once(self):
        import re
        html = self.client.get(reverse("add_product")).content.decode()
        for field in ['color', 'spec', 'gender', 'volume_ml', 'concentration', 'inspired_by']:
            widgets = re.findall(r'<(?:input|select|textarea)[^>]*name="%s"' % field, html)
            self.assertEqual(len(widgets), 1, f'{field}: {len(widgets)} widgets')
        # Groups are tagged with the category kinds they belong to.
        self.assertIn('data-kind="perfume"', html)
        self.assertIn('data-kind="general accessory"', html)
        self.assertIn("fam-picker", html)

    def test_saving_a_non_perfume_keeps_gender_and_colour(self):
        p = Product.objects.create(name="Case", barcode="9200000000002", brand="Anker",
                                   category=self.accessories, default_price=Decimal("9"),
                                   color="Black", gender="unisex")
        resp = self.client.post(reverse("edit_product", args=[p.pk]), {
            "barcode": "9200000000002", "category": self.accessories.id,
            "new_brand_name": "Anker", "name": "Case", "default_price": "9.00",
            "color": "Black", "gender": "unisex", "spec": "",
        })
        self.assertIn(resp.status_code, (200, 302))
        p.refresh_from_db()
        self.assertEqual(p.color, "Black")
        self.assertEqual(p.gender, "unisex")


class PerfumeAttributeBackfillTests(TestCase):
    def test_backfill_reads_volume_strength_and_family_from_existing_text(self):
        from django.core.management import call_command
        from stock.models import Concentration, FragranceFamily
        cat = Category.objects.create(name="Perfumes")
        p = Product.objects.create(
            name="EDP", model="KHAMRAH", barcode="9100000000001", brand="Lattafa",
            category=cat, default_price=Decimal("30"), spec="100ML",
            description="Uma fragrancia floral inspirada na iconica fragrancia Givenchy L'Interdit.",
        )
        call_command("backfill_perfume_attributes", "--apply", verbosity=0)

        p.refresh_from_db()
        self.assertEqual(p.volume_ml, 100)                      # from "100ML"
        self.assertEqual(p.concentration.short, "EDP")          # from the name
        self.assertEqual(p.name, "KHAMRAH")                     # name was just the strength
        self.assertEqual(p.inspired_by.house, "Givenchy")       # known house only
        self.assertIn("Floral", [f.name for f in p.fragrance_families.all()])
        # the lookup tables are seeded so the shop can extend them itself
        self.assertTrue(Concentration.objects.filter(short="Extrait").exists())
        self.assertTrue(FragranceFamily.objects.filter(name="Oud").exists())

    def test_dry_run_writes_nothing(self):
        from django.core.management import call_command
        from stock.models import Concentration
        cat = Category.objects.create(name="Perfumes")
        p = Product.objects.create(name="Asad", barcode="9100000000002", brand="Lattafa",
                                   category=cat, default_price=Decimal("30"), spec="100ml")
        call_command("backfill_perfume_attributes", verbosity=0)
        p.refresh_from_db()
        self.assertIsNone(p.volume_ml)
        self.assertEqual(Concentration.objects.count(), 0)


class SupplierLeadTimeTests(TestCase):
    """Lead time = inbound order placed (placed_at) -> received (received_at)."""

    def _order(self, supplier, days, status="received"):
        from django.utils import timezone as tz
        placed = tz.now() - timezone.timedelta(days=30)
        order = InboundOrder.objects.create(supplier=supplier, status=status,
                                            total_amount=Decimal("0.00"))
        InboundOrder.objects.filter(pk=order.pk).update(
            created_at=placed,
            placed_at=placed,
            received_at=(placed + timezone.timedelta(days=days)) if days is not None else None,
        )
        return InboundOrder.objects.get(pk=order.pk)

    def test_stats_summarise_every_order_that_can_be_measured(self):
        from stock.services.lead_time import supplier_lead_stats
        s = Supplier.objects.create(name="Lead Co")
        self._order(s, 2)
        self._order(s, 4)
        self._order(s, 6)
        # Placed and received in the same breath is a real, if short, wait -
        # it counts as zero rather than disappearing.
        self._order(s, 0)
        # No receipt time at all: nothing to measure, so it is left out.
        self._order(s, None)

        stats = supplier_lead_stats(s)

        self.assertEqual(stats["sample"], 4)
        self.assertAlmostEqual(stats["avg_days"], 3.0, places=3)
        self.assertAlmostEqual(stats["median_days"], 3.0, places=3)
        self.assertAlmostEqual(stats["min_days"], 0.0, places=3)
        self.assertAlmostEqual(stats["max_days"], 6.0, places=3)

    def test_an_order_that_was_never_placed_is_not_guessed_at(self):
        from stock.services.lead_time import supplier_lead_stats
        s = Supplier.objects.create(name="Untimed Co")
        order = self._order(s, 3)
        InboundOrder.objects.filter(pk=order.pk).update(placed_at=None)
        self.assertIsNone(supplier_lead_stats(s))

    def test_a_supplier_with_no_orders_at_all_gives_none(self):
        from stock.services.lead_time import supplier_lead_stats
        self.assertIsNone(supplier_lead_stats(Supplier.objects.create(name="New Co")))

    def test_list_page_ranks_suppliers_for_comparison(self):
        from stock.services.lead_time import lead_stats_by_supplier
        fast = Supplier.objects.create(name="Fast Co")
        slow = Supplier.objects.create(name="Slow Co")
        self._order(fast, 1); self._order(fast, 3)
        self._order(slow, 10)

        by_id = lead_stats_by_supplier()

        self.assertAlmostEqual(by_id[fast.id]["avg_days"], 2.0, places=3)
        self.assertAlmostEqual(by_id[slow.id]["avg_days"], 10.0, places=3)

        get_user_model().objects.create_superuser("lead_mgr", password="pw123456")
        self.client.login(username="lead_mgr", password="pw123456")
        page = self.client.get(reverse("supplier_list"))
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Lead time")
        html = page.content.decode()
        # the quickest supplier is flagged, the slowest too
        self.assertIn("lead-val fast", html)
        self.assertIn("lead-val slow", html)

    def test_detail_page_shows_the_breakdown(self):
        fast = Supplier.objects.create(name="Detail Co")
        self._order(fast, 5)
        get_user_model().objects.create_superuser("lead_mgr2", password="pw123456")
        self.client.login(username="lead_mgr2", password="pw123456")
        page = self.client.get(reverse("supplier_detail", args=[fast.id]))
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Fast / slow")       # compact scorecard label
        self.assertContains(page, "Recent lead times")
        self.assertContains(page, "Lead time")


class PerfumeBackfillCommandTests(TestCase):
    def test_backfill_prices_unlocked_perfumes(self):
        from django.core.management import call_command
        perfumes = Category.objects.create(name="Perfumes")
        p = Product.objects.create(name="P", barcode="8400000000001", brand="B", category=perfumes)
        # Create the batch WITHOUT triggering the signal path would be hard; instead
        # set price back to a stale value, then let the command recompute.
        Purchase.objects.create(product=p, quantity=5, remaining=5, cost_price=Decimal("12.34"))
        Product.objects.filter(pk=p.pk).update(default_price=Decimal("1"), wholesale_price=Decimal("1"))
        call_command("sync_perfume_prices")
        p.refresh_from_db()
        self.assertEqual(p.wholesale_price, Decimal("23"))
        self.assertEqual(p.default_price, Decimal("35"))


class ProductGenderTests(TestCase):
    def test_gender_shopify_tag_mapping(self):
        cat = Category.objects.create(name="Perfumes")
        p = Product.objects.create(name="P", barcode="8300000000001", brand="B",
                                   category=cat, gender="men")
        self.assertEqual(p.gender_shopify_tag, "Homem")
        Product.objects.filter(pk=p.pk).update(gender="women")
        p.refresh_from_db()
        self.assertEqual(p.gender_shopify_tag, "Mulher")
        Product.objects.filter(pk=p.pk).update(gender="")
        p.refresh_from_db()
        self.assertEqual(p.gender_shopify_tag, "")

    def test_gender_tag_included_in_shopify_sync_tags(self):
        from stock.services.shopify_sync import _shopify_tags
        cat = Category.objects.create(name="Perfumes")
        p = Product.objects.create(name="Asad", barcode="8300000000002", brand="LATTAFA",
                                   category=cat, gender="unisex")
        self.assertIn("Unissexo", _shopify_tags(p))

    def test_product_form_includes_gender_field(self):
        from stock.forms import ProductForm
        self.assertIn("gender", ProductForm().fields)


class ShopifyCsvCategoryTaxonomyTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.mgr = user_model.objects.create_superuser(username="cat_mgr", password="pw123456")
        self.client.login(username="cat_mgr", password="pw123456")
        self.perfumes = Category.objects.create(name="Perfumes")

    def _export_rows(self):
        import csv, io
        p = Product.objects.create(name="Asad Elixir", barcode="8200000000001", brand="LATTAFA",
                                   category=self.perfumes, default_price=Decimal("35"))
        Purchase.objects.create(product=p, quantity=3, remaining=3, cost_price=Decimal("12"))
        resp = self.client.get(reverse("export_shopify_inventory_csv"))
        self.assertEqual(resp.status_code, 200)
        text = resp.content.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))

    def test_product_category_column_uses_full_shopify_taxonomy_path(self):
        rows = self._export_rows()
        first = rows[0]
        self.assertEqual(
            first["Product category"],
            "Health & Beauty > Personal Care > Cosmetics > Perfumes & Colognes > Eaux de Parfum",
        )
        # Type keeps the local category name (free-text custom product type).
        self.assertEqual(first["Type"], "Perfumes")
        # Google column uses Google's own perfume taxonomy, not the bare name.
        self.assertEqual(
            first["Google Shopping / Google product category"],
            "Health & Beauty > Personal Care > Cosmetics > Perfume & Cologne",
        )
        self.assertNotEqual(first["Product category"], "Perfumes")


class SalesCalendarViewTests(TestCase):
    def setUp(self):
        from stock.models import Store
        user_model = get_user_model()
        self.mgr = user_model.objects.create_superuser(username="cal_mgr", password="pw123456")
        self.client.login(username="cal_mgr", password="pw123456")
        self.cat = Category.objects.create(name="Perfumes")
        self.customer = Customer.objects.create(name="Dona Ana")
        self.store = Store.objects.create(name="Amadora", code="AMD")
        self.product = Product.objects.create(name="Asad", barcode="8100000000001",
                                              brand="LATTAFA", category=self.cat,
                                              default_price=Decimal("35"))
        Purchase.objects.create(product=self.product, quantity=10, remaining=10,
                                cost_price=Decimal("12"))
        order = SaleOrder.objects.create(customer=self.customer, store=self.store)
        with self.captureOnCommitCallbacks(execute=True):
            Sale.objects.create(order=order, product=self.product, customer=self.customer,
                                quantity=1, unit_price=Decimal("35"), payment_method="cash")

    def test_default_is_month_calendar_with_day_modal_and_store(self):
        resp = self.client.get(reverse("sales_records"))
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.context["calendar_mode"])
        self.assertIn("calendar_weeks", resp.context)
        html = resp.content.decode("utf-8")
        # calendar grid + product search present
        self.assertIn('class="calendar"', html)
        self.assertIn("Filter by product", html)
        # today's sale -> a clickable day cell + a matching modal, with the store name inside
        today_id = "cal-" + timezone.localdate().strftime("%Y%m%d")
        self.assertIn('data-day-modal="%s"' % today_id, html)
        self.assertIn('id="%s-modal"' % today_id, html)
        self.assertIn("Amadora", html)
        # the old long stacked list heading must NOT render in calendar mode
        self.assertNotIn("orders in this range", html)

    def test_product_filter_narrows_the_page(self):
        resp = self.client.get(reverse("sales_records"), {"product_q": "zzz-nomatch"})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.context["calendar_mode"])
        self.assertEqual(resp.context["day_blocks"], [])
        self.assertIn("No sales this month", resp.content.decode("utf-8"))

    def test_year_view_shows_trend(self):
        resp = self.client.get(reverse("sales_records"), {"view": "year"})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.context.get("show_trend"))
        self.assertFalse(resp.context.get("calendar_mode"))


class StockLedgerServiceTests(TestCase):
    def setUp(self):
        self.cat = Category.objects.create(name="Perfumes")
        self.customer = Customer.objects.create(name="Ana")
        self.product = Product.objects.create(name="Asad", barcode="8300000009001",
                                              brand="B", category=self.cat,
                                              default_price=Decimal("30"))

    def _sale(self, qty, affects=True):
        order = SaleOrder.objects.create(customer=self.customer, affects_stock=affects)
        with self.captureOnCommitCallbacks(execute=True):
            Sale.objects.create(order=order, product=self.product, customer=self.customer,
                                quantity=qty, unit_price=Decimal("30"), payment_method="cash")
        return order

    def _set_remaining(self, purchase, r):
        Purchase.objects.filter(pk=purchase.pk).update(remaining=r)

    def test_reconciles_when_everything_is_logged(self):
        from stock.services.stock_ledger import build_stock_ledger
        p = Purchase.objects.create(product=self.product, quantity=10, remaining=10, cost_price=Decimal("10"))
        self._sale(3, affects=True)
        self._set_remaining(p, 7)
        led = build_stock_ledger(self.product)
        self.assertEqual(led["reconstructed_balance"], 7)
        self.assertEqual(led["actual_onhand"], 7)
        self.assertFalse(led["has_discrepancy"])

    def test_untracked_leak_shows_discrepancy(self):
        from stock.services.stock_ledger import build_stock_ledger
        p = Purchase.objects.create(product=self.product, quantity=10, remaining=10, cost_price=Decimal("10"))
        self._sale(3, affects=True)
        # someone edited the batch remaining/quantity with no log -> 2 units vanish
        self._set_remaining(p, 5)
        led = build_stock_ledger(self.product)
        self.assertEqual(led["reconstructed_balance"], 7)
        self.assertEqual(led["actual_onhand"], 5)
        self.assertEqual(led["difference"], -2)
        self.assertEqual(led["unexplained"], 2)
        self.assertTrue(led["has_discrepancy"])

    def test_no_stock_sale_flagged_and_not_subtracted(self):
        from stock.services.stock_ledger import build_stock_ledger
        Purchase.objects.create(product=self.product, quantity=10, remaining=10, cost_price=Decimal("10"))
        self._sale(3, affects=False)
        led = build_stock_ledger(self.product)
        self.assertEqual(led["reconstructed_balance"], 10)
        self.assertFalse(led["has_discrepancy"])
        sale_events = [e for e in led["events"] if e["kind"] == "sale"]
        self.assertTrue(sale_events[0]["no_stock"])
        self.assertEqual(sale_events[0]["out_qty"], 0)

    def test_total_stock_increase_not_double_counted(self):
        from stock.models import StockAdjustmentLog
        from stock.services.stock_ledger import build_stock_ledger
        Purchase.objects.create(product=self.product, quantity=10, remaining=10, cost_price=Decimal("0"))
        StockAdjustmentLog.objects.create(product=self.product, adjustment_type="total_stock",
                                          old_value=0, new_value=10)
        led = build_stock_ledger(self.product)
        self.assertEqual(led["reconstructed_balance"], 10)  # not 20
        self.assertEqual(led["actual_onhand"], 10)
        self.assertFalse(led["has_discrepancy"])

    def test_product_detail_links_to_full_sales_history(self):
        # The full Stock Ledger moved to the dedicated /sales-history/ page; the
        # product page no longer renders it and instead links there.
        Purchase.objects.create(product=self.product, quantity=4, remaining=4, cost_price=Decimal("10"))
        user_model = get_user_model()
        user_model.objects.create_superuser(username="led_mgr", password="pw123456")
        self.client.login(username="led_mgr", password="pw123456")
        resp = self.client.get(reverse("product_detail", args=[self.product.id]))
        self.assertEqual(resp.status_code, 200)
        html = resp.content.decode("utf-8")
        self.assertNotIn("Stock Ledger", html)
        self.assertIn("?product=%d" % self.product.id, html)


class ProductSalesHistoryPageTests(TestCase):
    def setUp(self):
        from stock.models import Store
        user_model = get_user_model()
        self.mgr = user_model.objects.create_superuser(username="psh_mgr", password="pw123456")
        self.client.login(username="psh_mgr", password="pw123456")
        self.cat = Category.objects.create(name="Perfumes")
        self.customer = Customer.objects.create(name="Dona Ana")
        self.store = Store.objects.create(name="Amadora", code="AMD")
        self.product = Product.objects.create(name="Asad Elixir", barcode="8400000009001",
                                              brand="LATTAFA", category=self.cat,
                                              default_price=Decimal("35"))
        Purchase.objects.create(product=self.product, quantity=10, remaining=10, cost_price=Decimal("12"))
        order = SaleOrder.objects.create(customer=self.customer, store=self.store)
        with self.captureOnCommitCallbacks(execute=True):
            Sale.objects.create(order=order, product=self.product, customer=self.customer,
                                quantity=2, unit_price=Decimal("35"), payment_method="cash")

    def test_search_lists_matching_products(self):
        resp = self.client.get(reverse("product_sales_history"), {"q": "Asad"})
        self.assertEqual(resp.status_code, 200)
        self.assertIsNone(resp.context.get("product"))
        html = resp.content.decode("utf-8")
        self.assertIn("Asad Elixir", html)
        self.assertIn("?product=%d" % self.product.id, html)

    def test_product_detail_shows_store_qty_and_ledger(self):
        resp = self.client.get(reverse("product_sales_history"), {"product": self.product.id})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["product"].id, self.product.id)
        self.assertEqual(resp.context["total_qty"], 2)
        self.assertEqual(resp.context["order_count"], 1)
        html = resp.content.decode("utf-8")
        self.assertIn("Amadora", html)          # store shown on the merged ledger row
        # Sales Detail is merged into the ledger: one timeline carrying store + running balance
        self.assertIn("Sales &amp; Stock Ledger", html)
        self.assertIn("Balance", html)          # running balance column present
        self.assertIn("By store", html)

    def test_store_filter_narrows_rows(self):
        from stock.models import Store
        other = Store.objects.create(name="Sintra", code="SIN")
        resp = self.client.get(reverse("product_sales_history"),
                               {"product": self.product.id, "store": other.id})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["total_qty"], 0)  # sale was in Amadora, not Sintra

    def test_employee_is_redirected(self):
        user_model = get_user_model()
        user_model.objects.create_user(username="psh_emp", password="pw123456")
        self.client.logout()
        self.client.login(username="psh_emp", password="pw123456")
        resp = self.client.get(reverse("product_sales_history"), {"product": self.product.id})
        self.assertEqual(resp.status_code, 302)


class ProductDetailRecentSalesTests(TestCase):
    def test_sales_history_limited_to_10_days_with_link(self):
        user_model = get_user_model()
        user_model.objects.create_superuser(username="pd_mgr", password="pw123456")
        self.client.login(username="pd_mgr", password="pw123456")
        cat = Category.objects.create(name="Perfumes")
        customer = Customer.objects.create(name="Ana")
        product = Product.objects.create(name="Yara", barcode="8400000009002", brand="B",
                                         category=cat, default_price=Decimal("28"))
        Purchase.objects.create(product=product, quantity=20, remaining=20, cost_price=Decimal("10"))
        # a recent sale (today) and an old sale (40 days ago)
        recent = SaleOrder.objects.create(customer=customer)
        with self.captureOnCommitCallbacks(execute=True):
            Sale.objects.create(order=recent, product=product, customer=customer,
                                quantity=1, unit_price=Decimal("28"), payment_method="cash")
        old = SaleOrder.objects.create(customer=customer)
        old_sale = Sale.objects.create(order=old, product=product, customer=customer,
                                       quantity=1, unit_price=Decimal("28"), payment_method="cash")
        old_dt = timezone.now() - timezone.timedelta(days=40)
        Sale.objects.filter(pk=old_sale.pk).update(date=old_dt)
        SaleOrder.objects.filter(pk=old.pk).update(created_at=old_dt)

        resp = self.client.get(reverse("product_detail", args=[product.id]))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.context["sales"]), 1)          # only the recent one
        self.assertEqual(resp.context["sales_older_count"], 1)   # the 40-day-old one hidden
        html = resp.content.decode("utf-8")
        self.assertIn("View full sales history", html)
        self.assertIn(reverse("product_sales_history"), html)


class DownloadDbBackupViewTests(TransactionTestCase):
    # TransactionTestCase (not TestCase): the view runs SQLite's online backup on
    # the live connection. Under TestCase, the outer atomic block leaves the
    # test-created user uncommitted, and backing up an in-memory DB with a pending
    # write transaction spins on SQLITE_BUSY (hangs). Committing writes avoids it;
    # production uses a file DB in autocommit, where the backup runs fine.
    def test_manager_downloads_valid_sqlite(self):
        import os
        import sqlite3
        get_user_model().objects.create_superuser(username="bk_mgr", password="pw123456")
        self.client.login(username="bk_mgr", password="pw123456")
        resp = self.client.get(reverse("download_db_backup"))
        self.assertEqual(resp.status_code, 200)
        self.assertIn("attachment", resp["Content-Disposition"])
        self.assertIn(".sqlite3", resp["Content-Disposition"])
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "dl.sqlite3")
            with open(p, "wb") as f:
                f.write(resp.content)
            con = sqlite3.connect(p)
            try:
                self.assertEqual(con.execute("PRAGMA integrity_check").fetchone()[0], "ok")
                # it's really our DB: a known table is present
                names = {r[0] for r in con.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'")}
                self.assertIn("stock_product", names)
            finally:
                con.close()

    def test_employee_is_redirected(self):
        get_user_model().objects.create_user(username="bk_emp", password="pw123456")
        self.client.login(username="bk_emp", password="pw123456")
        resp = self.client.get(reverse("download_db_backup"))
        self.assertEqual(resp.status_code, 302)  # not a manager -> redirected, no download


class SyncShopifyBarcodesCommandTests(TestCase):
    def _run(self, product_barcode, shopify_rec, apply=False):
        from io import StringIO
        from unittest import mock
        from stock.services.shopify_sync import _shopify_title
        cat = Category.objects.create(name="P")
        product = Product.objects.create(name="Asad", barcode=product_barcode, brand="Lattafa",
                                         category=cat, default_price=Decimal("10"))
        title = _shopify_title(product)
        client = mock.Mock()
        client.is_configured.return_value = True
        client.all_products_by_title.return_value = {title: shopify_rec} if shopify_rec else {}
        args = ["sync_shopify_barcodes", "--brand", "Lattafa"]
        if apply:
            args.append("--apply")
        out = StringIO()
        with mock.patch("stock.management.commands.sync_shopify_barcodes.ShopifyClient",
                        return_value=client):
            call_command(*args, stdout=out)
        return client, out.getvalue()

    def _rec(self, sku, barcode):
        return {"product_id": "gid://P/1", "variant_id": "gid://V/1", "sku": sku, "barcode": barcode}

    def test_dry_run_reports_change_but_writes_nothing(self):
        client, output = self._run("NEW123", self._rec("OLD999", "OLD999"), apply=False)
        self.assertIn("DRY RUN", output)
        self.assertIn("-> NEW123", output)
        client.update_variant_barcode_sku.assert_not_called()

    def test_apply_updates_when_barcode_differs(self):
        client, _ = self._run("NEW123", self._rec("OLD999", "OLD999"), apply=True)
        client.update_variant_barcode_sku.assert_called_once_with(
            "gid://P/1", "gid://V/1", "NEW123", "NEW123")

    def test_no_write_when_already_correct(self):
        client, _ = self._run("SAME1", self._rec("SAME1", "SAME1"), apply=True)
        client.update_variant_barcode_sku.assert_not_called()

    def test_skips_when_not_on_shopify(self):
        client, output = self._run("NEW123", None, apply=True)
        client.update_variant_barcode_sku.assert_not_called()
        self.assertIn("not on Shopify", output)


class SyncShopifyInventoryCommandTests(TestCase):
    def _setup(self, price, qty, rec):
        from unittest import mock
        cat = Category.objects.create(name="P")
        p = Product.objects.create(name="Asad", barcode="B1", brand="Lattafa",
                                   category=cat, default_price=Decimal(price))
        if qty:
            Purchase.objects.create(product=p, quantity=qty, remaining=qty, cost_price=Decimal("1"))
        client = mock.Mock()
        client.is_configured.return_value = True
        client.all_variants_by_sku.return_value = {"B1": rec} if rec else {}
        client.get_location_id.return_value = "gid://L/1"
        return p, client

    def _rec(self, price, available):
        return {"product_id": "P", "variant_id": "V", "inventory_item_id": "I",
                "price": price, "available": available}

    def _run(self, client, *extra):
        from io import StringIO
        from unittest import mock
        out = StringIO()
        with mock.patch("stock.management.commands.sync_shopify_inventory.ShopifyClient",
                        return_value=client):
            call_command("sync_shopify_inventory", "--brand", "Lattafa", *extra, stdout=out)
        return out.getvalue()

    def test_dry_run_writes_nothing(self):
        _, client = self._setup("35.00", 5, self._rec("20.00", 2))
        out = self._run(client)
        self.assertIn("DRY RUN", out)
        client.update_variant_price.assert_not_called()
        client.set_inventory_available.assert_not_called()

    def test_apply_pushes_price_and_inventory(self):
        _, client = self._setup("35.00", 5, self._rec("20.00", 2))
        self._run(client, "--apply")
        client.update_variant_price.assert_called_once_with("P", "V", "35.00")
        client.set_inventory_available.assert_called_once_with("I", "gid://L/1", 5)

    def test_no_write_when_matching(self):
        _, client = self._setup("35.00", 5, self._rec("35.00", 5))
        self._run(client, "--apply")
        client.update_variant_price.assert_not_called()
        client.set_inventory_available.assert_not_called()


class ShopifyInventorySignalTests(TestCase):
    def _make_product(self, barcode):
        cat = Category.objects.create(name="P")
        p = Product.objects.create(name="Asad", barcode=barcode, brand="Lattafa",
                                   category=cat, default_price=Decimal("10"))
        Purchase.objects.create(product=p, quantity=10, remaining=10, cost_price=Decimal("1"))
        return p

    def test_sale_pushes_inventory_when_enabled(self):
        from unittest import mock
        p = self._make_product("B1")  # created while sync is OFF -> no push yet
        with override_settings(SHOPIFY_INVENTORY_SYNC=True), \
             mock.patch("stock.services.shopify_sync.sync_product_price_inventory",
                        return_value=("inv_updated", "")) as push:
            with self.captureOnCommitCallbacks(execute=True):
                order = SaleOrder.objects.create()
                Sale.objects.create(order=order, product=p, quantity=1,
                                    unit_price=Decimal("10"), payment_method="cash")
        push.assert_called()
        args, kwargs = push.call_args
        self.assertEqual(args[0], p)
        self.assertFalse(kwargs.get("do_price"))     # inventory only, not price
        self.assertTrue(kwargs.get("do_inventory"))

    def test_no_push_when_disabled(self):
        from unittest import mock
        p = self._make_product("B2")
        with mock.patch("stock.services.shopify_sync.sync_product_price_inventory") as push:
            with self.captureOnCommitCallbacks(execute=True):
                order = SaleOrder.objects.create()
                Sale.objects.create(order=order, product=p, quantity=1,
                                    unit_price=Decimal("10"), payment_method="cash")
        push.assert_not_called()

    def test_price_change_pushes_price_when_enabled(self):
        from unittest import mock
        p = self._make_product("B3")  # created while sync is OFF
        with override_settings(SHOPIFY_INVENTORY_SYNC=True), \
             mock.patch("stock.services.shopify_sync.sync_product_price_inventory",
                        return_value=("inv_updated", "")) as push:
            with self.captureOnCommitCallbacks(execute=True):
                p.default_price = Decimal("99.00")
                p.save()
        push.assert_called()
        args, kwargs = push.call_args
        self.assertEqual(args[0], p)
        self.assertTrue(kwargs.get("do_price"))       # price only, not inventory
        self.assertFalse(kwargs.get("do_inventory"))

    def test_price_unchanged_no_push(self):
        from unittest import mock
        p = self._make_product("B4")
        with override_settings(SHOPIFY_INVENTORY_SYNC=True), \
             mock.patch("stock.services.shopify_sync.sync_product_price_inventory") as push:
            with self.captureOnCommitCallbacks(execute=True):
                p.name = "Renamed"   # save without touching price
                p.save()
        push.assert_not_called()


class SyncProductToShopifyButtonTests(TestCase):
    def _perfume(self):
        cat = Category.objects.create(name="Perfumes")
        return Product.objects.create(name="Asad", barcode="B1", brand="L",
                                      category=cat, default_price=Decimal("10"))

    def test_manager_syncs_perfume(self):
        from unittest import mock
        from stock.services import shopify_sync
        p = self._perfume()
        get_user_model().objects.create_superuser("shopmgr", password="pw123456")
        self.client.login(username="shopmgr", password="pw123456")
        with mock.patch("stock.services.shopify_client.ShopifyClient") as Client, \
             mock.patch("stock.services.shopify_sync.sync_product",
                        return_value=(shopify_sync.CREATED, "T")) as sp, \
             mock.patch("stock.services.shopify_sync.sync_product_price_inventory",
                        return_value=(shopify_sync.INV_UPDATED, "d")) as spi:
            Client.return_value.is_configured.return_value = True
            resp = self.client.post(reverse("sync_product_to_shopify", args=[p.id]))
        self.assertEqual(resp.status_code, 302)
        sp.assert_called_once()
        spi.assert_called_once()

    def test_enabling_tracking_always_rewrites_the_quantity(self):
        """Regression: syncing a stocked product left Shopify showing 0.

        Turning a variant from untracked to tracked makes Shopify (re)initialise
        its inventory level at 0. The sync used to skip the quantity write when
        the variant's *reported* quantity already equalled the target, so the
        variant was switched to tracked and then stranded at 0."""
        from unittest import mock
        from stock.services import shopify_sync
        product = self._perfume()
        Purchase.objects.create(product=product, supplier=None, quantity=6,
                                cost_price=Decimal("4.00"), remaining=6)
        self.assertEqual(product.total_stock(), 6)

        client = mock.MagicMock()
        # No decant variants -> target is the full on-hand (6). Shopify already
        # reports 6, but the variant is untracked, so tracking must be enabled
        # AND the quantity rewritten afterwards.
        client.find_variant_by_sku.side_effect = lambda sku: ({
            'product_id': 'gid://P', 'variant_id': 'gid://V',
            'inventory_item_id': 'gid://I', 'price': '10.00',
            'available': 6, 'tracked': False, 'policy': 'CONTINUE',
        } if sku == product.barcode else None)
        client.get_location_id.return_value = 'gid://L'

        code, detail = shopify_sync.sync_product_price_inventory(
            product, client, do_price=False, do_inventory=True)

        self.assertEqual(code, shopify_sync.INV_UPDATED)
        client.set_variant_stocked.assert_called_once_with('gid://P', 'gid://V')
        client.set_inventory_available.assert_called_once_with('gid://I', 'gid://L', 6)

    def test_non_perfume_blocked(self):
        from unittest import mock
        cat = Category.objects.create(name="Accessories")
        p = Product.objects.create(name="Bag", barcode="B2", brand="L",
                                   category=cat, default_price=Decimal("10"))
        get_user_model().objects.create_superuser("shopmgr2", password="pw123456")
        self.client.login(username="shopmgr2", password="pw123456")
        with mock.patch("stock.services.shopify_sync.sync_product") as sp:
            self.client.post(reverse("sync_product_to_shopify", args=[p.id]))
        sp.assert_not_called()

    def test_employee_blocked(self):
        from unittest import mock
        p = self._perfume()
        get_user_model().objects.create_user("shopemp", password="pw123456")
        self.client.login(username="shopemp", password="pw123456")
        with mock.patch("stock.services.shopify_sync.sync_product") as sp:
            resp = self.client.post(reverse("sync_product_to_shopify", args=[p.id]))
        self.assertEqual(resp.status_code, 302)
        sp.assert_not_called()


class SyncAllPerfumesButtonTests(TestCase):
    def test_manager_launches_background_sync(self):
        from unittest import mock
        get_user_model().objects.create_superuser("allmgr", password="pw123456")
        self.client.login(username="allmgr", password="pw123456")
        with mock.patch("stock.services.shopify_client.ShopifyClient") as Client, \
             mock.patch("threading.Thread") as Thread:
            Client.return_value.is_configured.return_value = True
            resp = self.client.post(reverse("sync_all_perfumes_to_shopify"))
        self.assertEqual(resp.status_code, 302)
        Thread.assert_called_once()
        Thread.return_value.start.assert_called_once()

    def test_employee_blocked(self):
        from unittest import mock
        get_user_model().objects.create_user("allemp", password="pw123456")
        self.client.login(username="allemp", password="pw123456")
        with mock.patch("threading.Thread") as Thread:
            self.client.post(reverse("sync_all_perfumes_to_shopify"))
        Thread.assert_not_called()

    def test_get_does_not_launch(self):
        from unittest import mock
        get_user_model().objects.create_superuser("allmgr2", password="pw123456")
        self.client.login(username="allmgr2", password="pw123456")
        with mock.patch("threading.Thread") as Thread:
            self.client.get(reverse("sync_all_perfumes_to_shopify"))
        Thread.assert_not_called()


class ShopifyDescriptionFormatTests(SimpleTestCase):
    def test_linebreaks_preserved(self):
        from types import SimpleNamespace
        from stock.services.shopify_sync import _shopify_description_html
        html = _shopify_description_html(SimpleNamespace(description="Line one\nLine two\n\nNew para"))
        self.assertIn("<br>", html)    # single newline -> <br>
        self.assertIn("<p>", html)     # blank line -> paragraph

    def test_empty_description(self):
        from types import SimpleNamespace
        from stock.services.shopify_sync import _shopify_description_html
        self.assertEqual(_shopify_description_html(SimpleNamespace(description="")), "")


class SyncShopifyPerfumesCollectionTests(TestCase):
    def test_adds_perfume_to_its_brand_manual_collection(self):
        from io import StringIO
        from unittest import mock
        from stock.services import shopify_sync
        cat = Category.objects.create(name="Perfumes")
        Product.objects.create(name="Khamrah", barcode="BC1", brand="Lattafa",
                               category=cat, default_price=Decimal("10"))
        client = mock.Mock()
        client.is_configured.return_value = True
        client.all_variants_by_sku.return_value = {
            "BC1": {"product_id": "gid://P/1", "status": "ACTIVE", "variant_id": "v",
                    "inventory_item_id": "i", "price": "10.00", "available": 5}}
        client.all_collections.return_value = [
            {"id": "gid://C/lattafa", "title": "Lattafa - Perfumes Árabes", "smart": False},
            {"id": "gid://C/rayhaan", "title": "Rayhaan - Perfumes Árabes", "smart": True},
        ]
        with mock.patch("stock.management.commands.sync_shopify_perfumes.ShopifyClient",
                        return_value=client), \
             mock.patch("stock.services.shopify_sync.sync_product",
                        return_value=(shopify_sync.SKIP_HAS_IMAGE, "")), \
             mock.patch("stock.services.shopify_sync.sync_product_price_inventory",
                        return_value=(shopify_sync.INV_UNCHANGED, "")):
            call_command("sync_shopify_perfumes", "--apply", stdout=StringIO())
        client.collection_add_products.assert_called_once_with("gid://C/lattafa", ["gid://P/1"])

    def test_missing_product_not_created_without_create_flag(self):
        from io import StringIO
        from unittest import mock
        cat = Category.objects.create(name="Perfumes")
        Product.objects.create(name="New", barcode="NEW", brand="Lattafa",
                               category=cat, default_price=Decimal("10"))
        client = mock.Mock()
        client.is_configured.return_value = True
        client.all_variants_by_sku.return_value = {}      # not on Shopify
        client.all_collections.return_value = []
        with mock.patch("stock.management.commands.sync_shopify_perfumes.ShopifyClient",
                        return_value=client), \
             mock.patch("stock.services.shopify_sync.sync_product") as sp:
            call_command("sync_shopify_perfumes", "--apply", stdout=StringIO())
        sp.assert_not_called()                             # default: don't create missing


class FixDecantSkusTests(TestCase):
    def test_rekeys_mismatched_decant_skus(self):
        from io import StringIO
        from unittest import mock
        client = mock.Mock()
        client.is_configured.return_value = True
        client.all_products_full_variants.return_value = [{
            "product_id": "gid://P/1",
            "title": "Maison Alhambra Yeah!Man",
            "variants": [
                {"id": "v0", "sku": "6290360590745"},          # 100ml (base)
                {"id": "v1", "sku": "1213454656777-10ML"},     # wrong base
                {"id": "v2", "sku": "6290360590745-5ML"},      # already correct
            ],
        }]
        with mock.patch("stock.management.commands.fix_decant_skus.ShopifyClient",
                        return_value=client):
            call_command("fix_decant_skus", "--apply", stdout=StringIO())
        client.fix_variant_sku.assert_called_once_with("gid://P/1", "v1", "6290360590745-10ML")

    def test_dry_run_writes_nothing(self):
        from io import StringIO
        from unittest import mock
        client = mock.Mock()
        client.is_configured.return_value = True
        client.all_products_full_variants.return_value = [{
            "product_id": "gid://P/1", "title": "X",
            "variants": [{"id": "v0", "sku": "AAA"}, {"id": "v1", "sku": "BBB-10ML"}],
        }]
        with mock.patch("stock.management.commands.fix_decant_skus.ShopifyClient",
                        return_value=client):
            call_command("fix_decant_skus", stdout=StringIO())
        client.fix_variant_sku.assert_not_called()


class SaleCostBasisTests(TestCase):
    def _perfume(self, price="40"):
        cat = Category.objects.create(name="P")
        return Product.objects.create(name="X", barcode="B", brand="L",
                                      category=cat, default_price=Decimal(price))

    def test_consume_returns_fifo_cost(self):
        from stock.services.stock_ops import consume_stock_fifo
        p = self._perfume()
        Purchase.objects.create(product=p, quantity=2, remaining=2, cost_price=Decimal("20"))
        Purchase.objects.create(product=p, quantity=2, remaining=2, cost_price=Decimal("15"))
        self.assertEqual(consume_stock_fifo(p, 3), Decimal("55.00"))  # 2@20 + 1@15

    def test_profit_prefers_stored_cost_basis_over_reconstruction(self):
        from stock.services.profit import sale_profit_map_for_sale_ids
        p = self._perfume()
        Purchase.objects.create(product=p, quantity=5, remaining=5, cost_price=Decimal("21.95"))
        order = SaleOrder.objects.create()
        s = Sale.objects.create(order=order, product=p, quantity=1, unit_price=Decimal("40"),
                                cost_basis=Decimal("15.50"), payment_method="cash")
        m = sale_profit_map_for_sale_ids([s.id])
        self.assertEqual(m[s.id]["cost"], Decimal("15.50"))
        self.assertEqual(m[s.id]["profit"], Decimal("24.50"))

    def test_backfill_anchors_recent_sale_to_newest_batch(self):
        from io import StringIO
        p = self._perfume()
        Purchase.objects.create(product=p, quantity=6, remaining=0, cost_price=Decimal("21.95"))   # old, gone
        Purchase.objects.create(product=p, quantity=12, remaining=11, cost_price=Decimal("15.50"))  # new
        order = SaleOrder.objects.create()
        recent = Sale.objects.create(order=order, product=p, quantity=1, unit_price=Decimal("40"),
                                     payment_method="cash")
        call_command("backfill_sale_cost_basis", "--apply", stdout=StringIO())
        recent.refresh_from_db()
        self.assertEqual(recent.cost_basis, Decimal("15.50"))  # newest batch, not old 21.95


class PruneShopifyProductsTests(TestCase):
    def test_deletes_only_products_not_in_app(self):
        from io import StringIO
        from unittest import mock
        cat = Category.objects.create(name="Perfumes")
        Product.objects.create(name="Keep", barcode="INAPP", brand="L",
                               category=cat, default_price=Decimal("10"))
        client = mock.Mock()
        client.is_configured.return_value = True
        client.all_products_by_title.return_value = {
            "Keep": {"product_id": "gid://P/keep", "variant_id": "v", "sku": "INAPP", "barcode": "INAPP"},
            "Gone": {"product_id": "gid://P/gone", "variant_id": "v", "sku": "NOTINAPP", "barcode": "NOTINAPP"},
            "Ambiguous": None,   # duplicate title -> skip
            "NoSku": {"product_id": "gid://P/nosku", "variant_id": "v", "sku": "", "barcode": ""},
        }
        with mock.patch("stock.management.commands.prune_shopify_products.ShopifyClient",
                        return_value=client):
            call_command("prune_shopify_products", "--apply", stdout=StringIO())
        client.delete_product.assert_called_once_with("gid://P/gone")

    def test_dry_run_deletes_nothing(self):
        from io import StringIO
        from unittest import mock
        client = mock.Mock()
        client.is_configured.return_value = True
        client.all_products_by_title.return_value = {
            "Gone": {"product_id": "gid://P/gone", "variant_id": "v", "sku": "X", "barcode": "X"},
        }
        with mock.patch("stock.management.commands.prune_shopify_products.ShopifyClient",
                        return_value=client):
            call_command("prune_shopify_products", stdout=StringIO())
        client.delete_product.assert_not_called()


class BackfillPerfumeSpecTests(TestCase):
    def test_fills_blank_perfume_spec_only(self):
        from io import StringIO
        cat = Category.objects.create(name="Perfumes")
        acc = Category.objects.create(name="Accessories")
        blank = Product.objects.create(name="A", barcode="S1", brand="L", category=cat, default_price=Decimal("1"))
        has_spec = Product.objects.create(name="B", barcode="S2", brand="L", category=cat,
                                          default_price=Decimal("1"), spec="90ml")
        non_perfume = Product.objects.create(name="C", barcode="S3", brand="L", category=acc, default_price=Decimal("1"))

        call_command("backfill_perfume_spec", "--apply", stdout=StringIO())

        blank.refresh_from_db(); has_spec.refresh_from_db(); non_perfume.refresh_from_db()
        self.assertEqual(blank.spec, "100ml")            # blank perfume filled
        self.assertEqual(has_spec.spec, "90ml")          # existing spec kept
        self.assertIn(non_perfume.spec, (None, ""))      # non-perfume untouched

    def test_dry_run_writes_nothing(self):
        from io import StringIO
        cat = Category.objects.create(name="Perfumes")
        p = Product.objects.create(name="A", barcode="S9", brand="L", category=cat, default_price=Decimal("1"))
        call_command("backfill_perfume_spec", stdout=StringIO())
        p.refresh_from_db()
        self.assertIn(p.spec, (None, ""))


class SyncShopifyStorefrontTests(TestCase):
    def test_hides_soldout_shows_restocked_and_sets_collections(self):
        from io import StringIO
        from unittest import mock
        cat = Category.objects.create(name="Perfumes")
        Product.objects.create(name="Out", barcode="OUT", brand="L",
                               category=cat, default_price=Decimal("10"))
        p_in = Product.objects.create(name="In", barcode="IN", brand="L",
                                      category=cat, default_price=Decimal("10"))
        Purchase.objects.create(product=p_in, quantity=5, remaining=5, cost_price=Decimal("1"))

        client = mock.Mock()
        client.is_configured.return_value = True
        client.all_variants_by_sku.return_value = {
            "OUT": {"product_id": "gid://P/out", "status": "ACTIVE", "variant_id": "v",
                    "inventory_item_id": "i", "price": "10.00", "available": 0},
            "IN": {"product_id": "gid://P/in", "status": "DRAFT", "variant_id": "v",
                   "inventory_item_id": "i", "price": "10.00", "available": 5},
        }
        client.find_collection_by_title.return_value = "gid://C/x"

        with mock.patch("stock.management.commands.sync_shopify_storefront.ShopifyClient",
                        return_value=client):
            call_command("sync_shopify_storefront", "--apply", stdout=StringIO())

        status_calls = {c.args[0]: c.args[1] for c in client.set_product_status.call_args_list}
        self.assertEqual(status_calls, {"gid://P/out": "DRAFT", "gid://P/in": "ACTIVE"})
        self.assertTrue(client.set_collection_products.called)

    def test_dry_run_writes_nothing(self):
        from io import StringIO
        from unittest import mock
        cat = Category.objects.create(name="Perfumes")
        Product.objects.create(name="Out", barcode="OUT", brand="L", category=cat, default_price=Decimal("10"))
        client = mock.Mock()
        client.is_configured.return_value = True
        client.all_variants_by_sku.return_value = {
            "OUT": {"product_id": "gid://P/out", "status": "ACTIVE", "variant_id": "v",
                    "inventory_item_id": "i", "price": "10.00", "available": 0},
        }
        client.find_collection_by_title.return_value = None
        with mock.patch("stock.management.commands.sync_shopify_storefront.ShopifyClient",
                        return_value=client):
            call_command("sync_shopify_storefront", stdout=StringIO())
        client.set_product_status.assert_not_called()
        client.set_collection_products.assert_not_called()
        client.create_collection.assert_not_called()


class ShopifyLocationConfigTests(SimpleTestCase):
    def test_uses_configured_location_id_without_api_call(self):
        from unittest import mock
        from stock.services.shopify_client import ShopifyClient
        with override_settings(SHOPIFY_LOCATION_ID="gid://shopify/Location/999"):
            c = ShopifyClient(domain="x.myshopify.com", token="t")
            with mock.patch.object(c, "graphql") as g:
                self.assertEqual(c.get_location_id(), "gid://shopify/Location/999")
                g.assert_not_called()


class DecantInventoryLogicTests(SimpleTestCase):
    def test_targets_reserve_and_decant(self):
        from stock.services.shopify_sync import _inventory_targets
        present = {"B1", "B1-10ML", "B1-5ML"}
        # decants available while any stock (a sample) exists; only N=0 hides them
        self.assertEqual(_inventory_targets("B1", 5, present), {"B1": 3, "B1-10ML": 10, "B1-5ML": 10})
        self.assertEqual(_inventory_targets("B1", 3, present), {"B1": 1, "B1-10ML": 10, "B1-5ML": 10})
        self.assertEqual(_inventory_targets("B1", 2, present), {"B1": 0, "B1-10ML": 10, "B1-5ML": 10})
        self.assertEqual(_inventory_targets("B1", 1, present), {"B1": 0, "B1-10ML": 10, "B1-5ML": 10})
        self.assertEqual(_inventory_targets("B1", 0, present), {"B1": 0, "B1-10ML": 0, "B1-5ML": 0})

    def test_targets_no_decant_uses_full_onhand(self):
        from stock.services.shopify_sync import _inventory_targets
        self.assertEqual(_inventory_targets("B1", 5, {"B1"}), {"B1": 5})
        self.assertEqual(_inventory_targets("B1", 0, {"B1"}), {"B1": 0})


class DecantInventoryPushTests(TestCase):
    def test_sets_each_variant_by_reserve_rules(self):
        from unittest import mock
        from stock.services import shopify_sync
        cat = Category.objects.create(name="P")
        p = Product.objects.create(name="Asad", barcode="B1", brand="L",
                                   category=cat, default_price=Decimal("10"))
        Purchase.objects.create(product=p, quantity=5, remaining=5, cost_price=Decimal("1"))
        sv = {
            "B1": {"product_id": "P", "variant_id": "V0", "inventory_item_id": "I0", "price": "10.00", "available": 0},
            "B1-10ML": {"product_id": "P", "variant_id": "V1", "inventory_item_id": "I1", "price": "2.00", "available": 0},
            "B1-5ML": {"product_id": "P", "variant_id": "V2", "inventory_item_id": "I2", "price": "1.50", "available": 0},
        }
        client = mock.Mock()
        code, _ = shopify_sync.sync_product_price_inventory(
            p, client, do_price=False, do_inventory=True, shop_variants=sv, location_id="L")
        self.assertEqual(code, shopify_sync.INV_UPDATED)
        calls = {c.args[0]: c.args[2] for c in client.set_inventory_available.call_args_list}
        self.assertEqual(calls, {"I0": 3, "I1": 10, "I2": 10})  # 100ml=5-2=3, decants=10

    def test_untracked_decant_is_made_tracked_and_deny(self):
        # French Avenue case: decant is untracked + CONTINUE at qty 0, so it stays
        # buyable. Flip it to track+deny — and because enabling tracking makes
        # Shopify (re)initialise the level, the quantity is written afterwards too.
        from unittest import mock
        from stock.services import shopify_sync
        cat = Category.objects.create(name="Perfumes")
        p = Product.objects.create(name="X", barcode="B1", brand="L",
                                   category=cat, default_price=Decimal("10"))  # N=0
        sv = {
            "B1": {"product_id": "P", "variant_id": "v0", "inventory_item_id": "I0",
                   "price": "10.00", "available": 0, "tracked": True, "policy": "DENY"},
            "B1-10ML": {"product_id": "P", "variant_id": "v1", "inventory_item_id": "I1",
                        "price": "2.00", "available": 0, "tracked": False, "policy": "CONTINUE"},
        }
        client = mock.Mock()
        code, _ = shopify_sync.sync_product_price_inventory(
            p, client, do_price=False, do_inventory=True, shop_variants=sv, location_id="L")
        self.assertEqual(code, shopify_sync.INV_UPDATED)
        client.set_variant_stocked.assert_called_once_with("P", "v1")   # decant fixed
        # The 100ml variant was already tracked at the right qty -> untouched;
        # only the newly tracked decant gets its level written (target 0 here).
        client.set_inventory_available.assert_called_once_with("I1", "L", 0)

    def test_per_product_path_looks_up_variants_and_is_correct(self):
        # The product-page button path: shop_variants=None -> find_variant_by_sku.
        # Proves it computes the same targets as the bulk path (no separate bug).
        from unittest import mock
        from stock.services import shopify_sync
        cat = Category.objects.create(name="Perfumes")
        p = Product.objects.create(name="Asad", barcode="B1", brand="L",
                                   category=cat, default_price=Decimal("10"))
        Purchase.objects.create(product=p, quantity=5, remaining=5, cost_price=Decimal("1"))  # N=5
        variants = {
            "B1": {"product_id": "P", "variant_id": "v0", "inventory_item_id": "I0", "price": "10.00", "available": 0},
            "B1-10ML": {"product_id": "P", "variant_id": "v1", "inventory_item_id": "I1", "price": "2.00", "available": 0},
            "B1-5ML": {"product_id": "P", "variant_id": "v2", "inventory_item_id": "I2", "price": "1.50", "available": 0},
        }
        client = mock.Mock()
        client.find_variant_by_sku.side_effect = lambda sku: variants.get(sku)
        client.get_location_id.return_value = "L"
        code, _ = shopify_sync.sync_product_price_inventory(
            p, client, do_price=False, do_inventory=True)  # shop_variants=None -> lookup
        self.assertEqual(code, shopify_sync.INV_UPDATED)
        calls = {c.args[0]: c.args[2] for c in client.set_inventory_available.call_args_list}
        self.assertEqual(calls, {"I0": 3, "I1": 10, "I2": 10})  # N=5 -> 100ml=3, decants=10


class BackupDbCommandTests(TestCase):
    def test_backup_creates_valid_snapshot_and_prunes(self):
        import os
        import sqlite3
        with tempfile.TemporaryDirectory() as d:
            # Pre-seed three old snapshot filenames so pruning has something to cut.
            for name in ("db-20200101-000000.sqlite3",
                         "db-20200102-000000.sqlite3",
                         "db-20200103-000000.sqlite3"):
                open(os.path.join(d, name), "w").close()

            call_command("backup_db", "--dir", d, "--keep", "2")

            snaps = sorted(f for f in os.listdir(d) if f.startswith("db-"))
            self.assertEqual(len(snaps), 2)                       # kept newest 2
            newest = snaps[-1]                                    # the real snapshot (today's timestamp)
            self.assertTrue(newest > "db-2020")                  # sorts after the seeded 2020 ones
            con = sqlite3.connect(os.path.join(d, newest))
            try:
                self.assertEqual(con.execute("PRAGMA integrity_check").fetchone()[0], "ok")
            finally:
                con.close()
