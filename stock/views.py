# stock/views.py
import csv
from decimal import Decimal,ROUND_HALF_UP
import hashlib
import json
from io import BytesIO
from collections import defaultdict
from datetime import timedelta, datetime, date
import qrcode
import base64
from calendar import monthrange, Calendar
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q, Sum, Count, Max
from django.db.models.functions import Coalesce
from django.http import (
    HttpResponse, JsonResponse, FileResponse
)
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.text import slugify
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST
from django.urls import reverse
from urllib.parse import urlencode

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,KeepTogether,HRFlowable

from django.db.models import Sum, OuterRef, Subquery, IntegerField, F, DecimalField, DateTimeField, ExpressionWrapper, Q,Value
from io import BytesIO as _BytesIO

from .forms import (
    ProductForm, CustomerForm, ARInvoiceForm, ARPaymentForm,
    SaleOrderCorrectionForm, SupplierForm, EmployeeAccountForm,
    PrintProfileForm, InboundOrderEditForm, InboundPurchaseFormSet, DirectPurchaseEditForm,
    InboundReceiveForm, InboundPendingFormSet, StoreForm,
)
from .models import (
    Product, Purchase, Sale, Supplier, Customer, ProductImage,
    Category, SaleOrder, InboundOrder, InboundPendingItem, ARInvoice, ARItem,
    Brand, SaleOrderChangeLog, AttendanceRecord, PrintProfile,
    StockAdjustmentLog, SaleOrderPayment, Store, StoreProfile,
)
from .permissions import (
    has_admin_access,
    has_manager_access,
    has_order_reconciliation_access,
    has_sales_sensitive_access,
    admin_required,
    manager_required,
)
from .stores import (
    ACTIVE_STORE_SESSION_KEY,
    ALL_STORES,
    available_stores,
    can_switch_store,
    resolve_active_store,
    scope_sales_by_store,
    store_for_new_sale,
)
from .services.cloudinary_urls import product_image_cdn_url
from .services.dashboard import (
    build_monthly_dashboard_snapshot,
    build_period_comparison,
    build_target_progress,
    build_yearly_sales_overview,
    order_tender_amounts,
    resolve_dashboard_month,
    resolve_year,
)
from .services.inventory import build_inventory_snapshot
from .services.order_corrections import (
    delete_sale_order_correction,
    save_sale_order_correction,
)
from .services.profit import sale_profit_map_for_sale_ids
from .services.stock_ops import consume_stock_fifo
from .services.stock_ledger import build_stock_ledger
from .services import rebuild_all_daily_summaries

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


def build_product_label(product):
    if not product:
        return 'Unknown product'
    return product.display_name


def get_product_image_url(product):
    if not product:
        return ''

    images_manager = getattr(product, 'images', None)
    if images_manager is None:
        return ''

    for image in images_manager.all():
        image_file = getattr(image, 'image', None)
        if not image_file:
            continue
        try:
            return image_file.url
        except ValueError:
            continue
    return ''


def build_brand_series_map():
    mapping = {}
    for brand in Brand.objects.order_by('name').prefetch_related('series'):
        mapping[str(brand.id)] = [
            {'id': series.id, 'name': series.name}
            for series in brand.series.all().order_by('name')
        ]
    return mapping


def build_brand_catalog():
    catalog = []
    for brand in Brand.objects.order_by('name').prefetch_related('categories'):
        catalog.append({
            'id': brand.id,
            'name': brand.name,
            'category_ids': [str(category.id) for category in brand.categories.all().order_by('name')],
        })
    return catalog


def annotate_catalog_metrics(queryset):
    stock_subquery = (
        Purchase.objects
        .filter(product=OuterRef('pk'))
        .values('product')
        .annotate(total=Sum('remaining'))
        .values('total')[:1]
    )
    sold_subquery = (
        Sale.objects
        .filter(product=OuterRef('pk'))
        .values('product')
        .annotate(total=Sum('quantity'))
        .values('total')[:1]
    )

    return queryset.annotate(
        image_count=Count('images', distinct=True),
        total_stock=Coalesce(Subquery(stock_subquery, output_field=IntegerField()), Value(0)),
        total_sold=Coalesce(Subquery(sold_subquery, output_field=IntegerField()), Value(0)),
    )


def customer_catalog_case(value):
    value = (value or '').strip()
    if not value:
        return ''
    return value.title()


def build_customer_product_title(product):
    model_part = (getattr(product, 'model', '') or '').strip()
    name_part = (getattr(product, 'name', '') or '').strip()
    title_parts = []
    if model_part:
        title_parts.append(model_part)
    if name_part and name_part.lower() != model_part.lower():
        title_parts.append(name_part)
    return customer_catalog_case(' - '.join(title_parts) or name_part or build_product_label(product))


def get_catalog_availability_parts(stock_val):
    stock_val = int(stock_val or 0)
    if stock_val >= 4:
        return 'Available now', 'in-stock'
    if stock_val > 0:
        return 'Low stock', 'low-stock'
    return 'Currently unavailable', 'out-stock'


SHOPIFY_PRODUCT_CSV_HEADERS = [
    'Title',
    'URL handle',
    'Description',
    'Vendor',
    'Product category',
    'Type',
    'Tags',
    'Published on online store',
    'Status',
    'SKU',
    'Barcode',
    'Option1 name',
    'Option1 value',
    'Option1 Linked To',
    'Option2 name',
    'Option2 value',
    'Option2 Linked To',
    'Option3 name',
    'Option3 value',
    'Option3 Linked To',
    'Price',
    'Compare-at price',
    'Cost per item',
    'Charge tax',
    'Tax code',
    'Unit price total measure',
    'Unit price total measure unit',
    'Unit price base measure',
    'Unit price base measure unit',
    'Inventory tracker',
    'Inventory quantity',
    'Continue selling when out of stock',
    'Weight value (grams)',
    'Weight unit for display',
    'Requires shipping',
    'Fulfillment service',
    'Product image URL',
    'Image position',
    'Image alt text',
    'Variant image URL',
    'Gift card',
    'SEO title',
    'SEO description',
    'Color (product.metafields.shopify.color-pattern)',
    'Google Shopping / Google product category',
    'Google Shopping / Gender',
    'Google Shopping / Age group',
    'Google Shopping / Manufacturer part number (MPN)',
    'Google Shopping / Ad group name',
    'Google Shopping / Ads labels',
    'Google Shopping / Condition',
    'Google Shopping / Custom product',
    'Google Shopping / Custom label 0',
    'Google Shopping / Custom label 1',
    'Google Shopping / Custom label 2',
    'Google Shopping / Custom label 3',
    'Google Shopping / Custom label 4',
]

# The CSV "Product category" column feeds Shopify's *Standard Product Taxonomy*, which
# expects a valid taxonomy path — NOT our local category name. A bare name like
# "Perfumes" is unmatched, so Shopify silently drops it and ML-auto-categorizes the
# product (which mis-filed everything under a Pet node). Map our category → the full
# Shopify taxonomy path so imports land correctly. Unknown categories stay blank.
SHOPIFY_TAXONOMY_BY_CATEGORY = {
    'Perfumes': 'Health & Beauty > Personal Care > Cosmetics > Perfumes & Colognes > Eaux de Parfum',
}
# The Google Shopping column uses Google's *separate* product taxonomy, also unmatched
# by a bare "Perfumes"; map it to Google's perfume node.
GOOGLE_CATEGORY_BY_CATEGORY = {
    'Perfumes': 'Health & Beauty > Personal Care > Cosmetics > Perfume & Cologne',
}


def _clean_csv_text(value):
    return (str(value).strip() if value is not None else '')


def _build_shopify_base_title(product):
    seen = set()
    parts = []
    for value in [product.brand, product.model, product.name]:
        cleaned = _clean_csv_text(value)
        key = cleaned.lower()
        if cleaned and key not in seen:
            parts.append(cleaned)
            seen.add(key)
    return ' - '.join(parts) or product.display_name or product.barcode


def _build_unique_shopify_handle(raw_title, fallback, used_handles):
    base = slugify(raw_title, allow_unicode=False).strip('-')
    if not base:
        base = slugify(fallback, allow_unicode=False).strip('-') or f'product-{fallback}'
    base = (base[:220].strip('-') or f'product-{fallback}')
    handle = base
    suffix = 2
    while handle in used_handles:
        suffix_text = f'-{suffix}'
        handle = f'{base[:255 - len(suffix_text)]}{suffix_text}'
        suffix += 1
    used_handles.add(handle)
    return handle


def _format_shopify_money(value):
    if value is None:
        return ''
    return str(Decimal(value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))


def _first_clean_product_value(group_products, attr):
    for product in group_products:
        cleaned = _clean_csv_text(getattr(product, attr, ''))
        if cleaned:
            return cleaned
    return ''


def _build_shopify_tags(group_products):
    seen = set()
    tags = []
    for product in group_products:
        values = [
            getattr(product.category, 'name', ''),
            product.brand,
            product.model,
            product.name,
            product.spec,
            product.color,
            product.gender_shopify_tag,
        ]
        for value in values:
            cleaned = _clean_csv_text(value)
            key = cleaned.lower()
            if cleaned and key not in seen:
                tags.append(cleaned)
                seen.add(key)
    return ', '.join(tags)


def _build_shopify_color_metafield(group_products):
    colors = []
    seen = set()
    for product in group_products:
        color = _clean_csv_text(product.color)
        key = color.lower()
        if color and key not in seen:
            colors.append(color)
            seen.add(key)
    return '; '.join(colors)


def _build_shopify_image_url(request, product):
    image_url = get_product_image_url(product)
    if not image_url:
        return ''
    # Shopify fetches this URL from its own servers, so the local absolute URL
    # (a LAN address) is only a fallback for when Cloudinary is not configured.
    return product_image_cdn_url(product) or request.build_absolute_uri(image_url)


def _first_shopify_image_url(request, group_products):
    for product in group_products:
        image_url = _build_shopify_image_url(request, product)
        if image_url:
            return image_url
    return ''


def _apply_shopify_option_columns(row, product, group_products):
    has_spec = any(_clean_csv_text(item.spec) for item in group_products)
    has_color = any(_clean_csv_text(item.color) for item in group_products)
    spec_value = _clean_csv_text(product.spec) or 'Default'
    color_value = _clean_csv_text(product.color) or 'Default'

    if has_spec:
        row['Option1 name'] = 'Spec'
        row['Option1 value'] = spec_value
        if has_color:
            row['Option2 name'] = 'Color'
            row['Option2 value'] = color_value
            row['Option2 Linked To'] = 'product.metafields.shopify.color-pattern'
        return

    if has_color:
        row['Option1 name'] = 'Color'
        row['Option1 value'] = color_value
        row['Option1 Linked To'] = 'product.metafields.shopify.color-pattern'
        return

    if len(group_products) > 1:
        row['Option1 name'] = 'SKU'
        row['Option1 value'] = product.barcode
        return

    row['Option1 name'] = 'Title'
    row['Option1 value'] = 'Default Title'


def get_month_bounds(month_value=None):
    today = timezone.localdate()
    if month_value:
        try:
            year_str, month_str = month_value.split('-', 1)
            year = int(year_str)
            month = int(month_str)
            month_start = today.replace(year=year, month=month, day=1)
        except (TypeError, ValueError):
            month_start = today.replace(day=1)
    else:
        month_start = today.replace(day=1)

    last_day = monthrange(month_start.year, month_start.month)[1]
    month_end = month_start.replace(day=last_day)
    return month_start, month_end


def format_duration_hours(duration):
    if not duration:
        return "0.0h"
    total_seconds = max(duration.total_seconds(), 0)
    hours = total_seconds / 3600
    return f"{hours:.1f}h"


def append_attendance_note(existing_note, new_note, prefix):
    new_note = (new_note or '').strip()
    if not new_note:
        return existing_note or ''
    if existing_note:
        return f"{existing_note}\n{prefix}: {new_note}"
    return f"{prefix}: {new_note}"


def normalize_recorded_at(recorded_at):
    if timezone.is_naive(recorded_at):
        return timezone.make_aware(recorded_at, timezone.get_current_timezone())
    return timezone.localtime(recorded_at)


def comparable_recorded_at(recorded_at):
    if recorded_at is None:
        return None
    return normalize_recorded_at(recorded_at)


def group_purchases_by_recorded_second(purchases):
    grouped = defaultdict(list)
    for purchase in purchases:
        group_key = comparable_recorded_at(purchase.date).replace(microsecond=0)
        grouped[group_key].append(purchase)
    return grouped


def annotate_inbound_formset_runtime(formset):
    for line_form in formset.forms:
        purchase = getattr(line_form, 'instance', None)
        if not purchase or not getattr(purchase, 'pk', None):
            continue
        purchase.sold_units = max((purchase.quantity or 0) - (purchase.remaining or 0), 0)
        purchase.line_total = (purchase.cost_price or Decimal('0.00')) * (purchase.quantity or 0)


# -----------------------------
# 入库
# -----------------------------
def _pending_inbound_orders():
    return (
        InboundOrder.objects
        .filter(status='pending_receipt')
        .select_related('supplier')
        .prefetch_related('pending_items__product__images')
        .order_by('-created_at')
    )


def _build_pending_reviews():
    """For each pending order build the editable review form + formset rendered
    inside a per-order modal on the inbound page."""
    reviews = []
    for order in _pending_inbound_orders():
        formset = InboundPendingFormSet(instance=order, prefix='lines')
        lines = []
        for line_form in formset.forms:
            item = line_form.instance
            item.image_url = get_product_image_url(item.product) if item.pk else ''
            item.display_name = build_product_label(item.product) if item.pk else ''
            lines.append({'form': line_form, 'item': item})
        reviews.append({
            'order': order,
            'form': InboundReceiveForm(instance=order),
            'formset': formset,
            'lines': lines,
        })
    return reviews


@login_required
@manager_required
def inbound_view(request):

    today = timezone.localdate()
    context = {
        'today': today,
        'pending_reviews': _build_pending_reviews(),
    }

    if request.method != 'POST':
        return render(request, 'stock/inbound.html', context)

    # 前端传过来的字段（多行明细 + 可选发票信息）
    items_json  = (request.POST.get('items_json') or '[]').strip()
    supplier_id = request.POST.get('supplier') or None
    invoice_no  = (request.POST.get('invoice_no') or '').strip()
    invoice_date = (request.POST.get('invoice_date') or '').strip()

    # 解析购物车
    try:
        items = json.loads(items_json)
        assert isinstance(items, list) and items, 'Cart is empty.'
    except Exception:
        context['error'] = 'Cart is empty or invalid.'
        return render(request, 'stock/inbound.html', context)

    supplier = Supplier.objects.filter(id=supplier_id).first() if supplier_id else None

    # 简易幂等保护：相同负载 8 秒内仅处理一次
    idem_raw = '|'.join([
        request.path,
        items_json,
        str(supplier_id or ''),
        invoice_no,
        invoice_date,
    ])
    idem_key = 'idem:' + hashlib.sha256(idem_raw.encode('utf-8', 'ignore')).hexdigest()
    if not cache.add(idem_key, 1, timeout=8):
        # 重复提交，直接返回提示，避免重复入库
        try:
            items_tmp = json.loads(items_json)
            total_tmp = sum(Decimal(str(r.get('cost_price', 0))) * int(r.get('qty', 0)) for r in items_tmp)
        except Exception:
            total_tmp = Decimal('0.00')
        context['success'] = f"Duplicate ignored. Total €{total_tmp:.2f}."
        return render(request, 'stock/inbound.html', context)

    try:
        with transaction.atomic():
            order_total = Decimal('0.00')

            if supplier:
                # 有供应商：创建“暂定（pending_receipt）”入库单，先不入库、不产生库存，
                # 行项目存为 InboundPendingItem，待确认收货后再转为 Purchase。
                order = InboundOrder.objects.create(
                    supplier=supplier,
                    invoice_no=invoice_no or None,
                    invoice_date=invoice_date or timezone.localdate(),
                    status='pending_receipt',
                    total_amount=Decimal('0.00'),
                )

                for row in items:
                    barcode = str(row.get('barcode', '')).strip()
                    try:
                        qty = int(row.get('qty', 0))
                        cost_price = Decimal(str(row.get('cost_price', 0)))
                    except Exception:
                        raise ValueError('Quantity/cost must be numbers.')

                    if not barcode or qty < 1 or cost_price <= 0:
                        raise ValueError('Invalid item line.')

                    product = Product.objects.get(barcode=barcode)

                    InboundPendingItem.objects.create(
                        inbound_order=order,
                        product=product,
                        quantity=qty,
                        cost_price=cost_price,
                    )
                    order_total += cost_price * qty

                # 回写整单总额
                order.total_amount = order_total
                order.save(update_fields=['total_amount'])

                context['success'] = (
                    f'Pending order #{order.id} created for {supplier.name}, €{order_total:.2f}. '
                    f'Confirm receipt below when the goods arrive.'
                )
            else:
                # 没有 supplier：普通入库，不创建 InboundOrder
                for row in items:
                    barcode = str(row.get('barcode', '')).strip()
                    try:
                        qty = int(row.get('qty', 0))
                        cost_price = Decimal(str(row.get('cost_price', 0)))
                    except Exception:
                        raise ValueError('Quantity/cost must be numbers.')

                    if not barcode or qty < 1 or cost_price <= 0:
                        raise ValueError('Invalid item line.')

                    product = Product.objects.get(barcode=barcode)

                    Purchase.objects.create(
                        inbound_order=None,     # 关键：不归入入库单
                        product=product,
                        supplier=None,          # 普通入库不记录供应商
                        quantity=qty,
                        remaining=qty,
                        cost_price=cost_price,
                        date=timezone.now(),
                    )
                    order_total += cost_price * qty

                context['success'] = (
                    f'Stock added (no supplier). Total €{order_total:.2f}.'
                )

    except Product.DoesNotExist:
        context['error'] = 'One product not found by barcode.'
    except ValueError as e:
        context['error'] = str(e)
    except Exception as e:
        context['error'] = 'Unexpected error: ' + str(e)

    # 刷新待收货列表（可能刚新增了一张暂定单）
    context['pending_reviews'] = _build_pending_reviews()
    return render(request, 'stock/inbound.html', context)


@login_required
def suppliers_autocomplete(request):
    q = (request.GET.get('q') or '').strip()
    qs = Supplier.objects.all()
    if q:
        qs = qs.filter(name__icontains=q)
    results = [{'id': s.id, 'name': s.name} for s in qs.order_by('name')[:10]]
    return JsonResponse({'results': results})


@login_required
@manager_required
def inbound_receive_view(request, order_id):
    """复核 / 编辑 / 确认收货 / 取消 一张暂定（pending_receipt）入库单。

    确认收货时，把 InboundPendingItem 行转换为 Purchase（remaining=quantity，
    date=now），并将订单状态置为 received。
    """
    order = get_object_or_404(
        InboundOrder.objects
        .filter(status='pending_receipt')
        .select_related('supplier')
        .prefetch_related('pending_items__product__images'),
        id=order_id,
    )

    if request.method == 'POST':
        action = (request.POST.get('action') or 'save').strip()

        if action == 'cancel':
            deleted_id = order.id
            order.delete()
            messages.success(request, f'Pending order #{deleted_id} cancelled.')
            return redirect('inbound')

        form = InboundReceiveForm(request.POST, instance=order)
        formset = InboundPendingFormSet(request.POST, instance=order, prefix='lines')

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                order = form.save()
                formset.save()

                remaining_items = list(order.pending_items.select_related('product').all())
                if not remaining_items:
                    deleted_id = order.id
                    order.delete()
                    messages.success(request, f'Pending order #{deleted_id} cancelled (no products left).')
                    return redirect('inbound')

                order.total_amount = sum(
                    ((it.cost_price or Decimal('0.00')) * it.quantity for it in remaining_items),
                    Decimal('0.00'),
                )
                order.save(update_fields=['total_amount'])

                if action == 'receive':
                    now = timezone.now()
                    for it in remaining_items:
                        Purchase.objects.create(
                            inbound_order=order,
                            product=it.product,
                            supplier=order.supplier,
                            quantity=it.quantity,
                            remaining=it.quantity,
                            cost_price=it.cost_price,
                            date=now,
                        )
                    order.status = 'received'
                    order.received_at = now
                    order.save(update_fields=['status', 'received_at'])
                    order.pending_items.all().delete()
                    messages.success(
                        request,
                        f'Order #{order.id} received. {len(remaining_items)} product(s) added to stock.'
                    )
                    return redirect('inbound')

            messages.success(request, f'Pending order #{order.id} saved.')
            return redirect('inbound_receive', order_id=order.id)

        messages.error(request, 'Please fix the highlighted fields.')

    # 复核/确认在 Inbound 页面的弹窗中完成；GET 或校验失败都回到 Inbound 页。
    return redirect('inbound')


# -----------------------------
# 出库（一次订单多个产品 + 每行可选支付方式）
# 前端通过 items_json 提交：[{barcode, qty, price, payment}, ...]
# 可选 customer_id
# -----------------------------
@login_required
def outbound_view(request):
    # A sale must land on one concrete store. Managers viewing the "All stores"
    # aggregate have no unambiguous target, so block selling until they pick a
    # store (employees are always locked to their home store, never blocked).
    _store, store_is_all = resolve_active_store(request)
    if request.method != 'POST':
        return render(request, 'stock/outbound.html', {
            'outbound_blocked': store_is_all,
            'active_store_name': _store.name if _store else '',
        })

    if store_is_all:
        return render(request, 'stock/outbound.html', {
            'outbound_blocked': True,
            'error': 'Select a specific store before selling — you are viewing "All stores". '
                     'Switch to the store you are selling from, then try again.',
        })

    items_json = request.POST.get('items_json', '[]')
    payments_json = request.POST.get('payments_json', '[]')
    customer_id = request.POST.get('customer_id') or None
    customer = Customer.objects.filter(id=customer_id).first() if customer_id else None

    # 简易幂等保护：相同负载 8 秒内仅处理一次
    idem_raw = '|'.join([
        request.path,
        items_json,
        payments_json,
        str(customer_id or ''),
    ])
    idem_key = 'idem:' + hashlib.sha256(idem_raw.encode('utf-8', 'ignore')).hexdigest()
    if not cache.add(idem_key, 1, timeout=8):
        try:
            items_tmp = json.loads(items_json)
            total_tmp = sum(Decimal(str(r.get('price', 0))) * int(r.get('qty', 0)) for r in items_tmp)
            qty_tmp = sum(int(r.get('qty', 0)) for r in items_tmp)
        except Exception:
            total_tmp = Decimal('0.00'); qty_tmp = 0
        return render(request, 'stock/outbound.html', {
            'success': f'Duplicate ignored. {qty_tmp} item(s), €{total_tmp:.2f}.'
        })

    try:
        items = json.loads(items_json)
        assert isinstance(items, list) and len(items) > 0
    except Exception:
        return render(request, 'stock/outbound.html', {'error': 'Cart is empty or invalid.'})

    valid_methods = {'cash', 'card', 'mbway'}
    method_priority = {'cash': 0, 'card': 1, 'mbway': 2}

    # 先解析行项目并计算订单总额（price 已是折后单价）
    parsed_items = []
    order_revenue = Decimal('0.00')
    order_qty = 0
    line_payment_totals = defaultdict(Decimal)  # 旧格式（按行支付）回退用
    try:
        for row in items:
            barcode = str(row.get('barcode', '')).strip()
            try:
                qty = int(row.get('qty', 0))
                price = Decimal(str(row.get('price', 0)))
            except Exception:
                raise ValueError('Quantity/price must be numbers.')
            if not barcode or qty < 1 or price <= 0:
                raise ValueError('Invalid product line.')
            line_payment = str(row.get('payment', '')).strip()  # 旧格式可选
            parsed_items.append({'barcode': barcode, 'qty': qty, 'price': price, 'payment': line_payment})
            line_total = price * qty
            order_revenue += line_total
            order_qty += qty
            if line_payment in valid_methods:
                line_payment_totals[line_payment] += line_total
    except ValueError as e:
        return render(request, 'stock/outbound.html', {'error': str(e)})

    # 解析订单级支付拆分（新格式）；缺省时回退到按行支付聚合（旧格式）
    try:
        raw_payments = json.loads(payments_json) if payments_json else []
    except Exception:
        raw_payments = []

    payment_totals = defaultdict(Decimal)
    if isinstance(raw_payments, list) and raw_payments:
        for entry in raw_payments:
            method = str(entry.get('method', '')).strip()
            try:
                amount = Decimal(str(entry.get('amount', 0)))
            except Exception:
                return render(request, 'stock/outbound.html', {'error': 'Payment amount must be a number.'})
            if method not in valid_methods or amount <= 0:
                continue
            payment_totals[method] += amount
    elif line_payment_totals:
        payment_totals = line_payment_totals  # 旧格式回退

    if not payment_totals:
        return render(request, 'stock/outbound.html', {'error': 'Please choose at least one payment method.'})

    paid_total = sum(payment_totals.values(), Decimal('0.00'))
    if abs(paid_total - order_revenue) > Decimal('0.01'):
        return render(request, 'stock/outbound.html', {
            'error': f'Payments (€{paid_total:.2f}) must add up to the total (€{order_revenue:.2f}).'
        })

    # 主支付方式：金额最大者（并列时按 cash→card→mbway）
    primary_method = min(
        payment_totals.items(),
        key=lambda kv: (-kv[1], method_priority.get(kv[0], 9)),
    )[0]

    active_store = store_for_new_sale(request)
    try:
        with transaction.atomic():
            order = SaleOrder.objects.create(customer=customer, store=active_store)

            for row in parsed_items:
                try:
                    product = Product.objects.get(barcode=row['barcode'])
                except Product.DoesNotExist:
                    raise ValueError(f"Product not found by barcode: {row['barcode']}")

                # FIFO 扣减（原子条件更新；select_for_update 在 SQLite 上是 no-op，见 stock_ops）
                # 返回本次消耗的真实成本，作为这笔销售的成本基准存下来。
                line_cost = consume_stock_fifo(product, row['qty'])

                # 行支付方式：旧格式用行内值，新格式（订单级拆分）用主支付方式
                line_method = row['payment'] if row['payment'] in valid_methods else primary_method
                Sale.objects.create(
                    order=order,
                    product=product,
                    customer=customer,
                    store=active_store,
                    quantity=row['qty'],
                    unit_price=row['price'],
                    cost_basis=line_cost,
                    payment_method=line_method,
                    date=timezone.now(),
                )

            # 订单级支付拆分（权威记录）
            for method, amount in payment_totals.items():
                SaleOrderPayment.objects.create(order=order, method=method, amount=amount)

    except ValidationError as e:
        return render(request, 'stock/outbound.html', {'error': '; '.join(e.messages)})
    except ValueError as e:
        return render(request, 'stock/outbound.html', {'error': str(e)})
    except Exception as e:
        return render(request, 'stock/outbound.html', {'error': 'Unexpected error: ' + str(e)})

    pay_summary = ', '.join(
        f'{label} €{payment_totals[code]:.2f}'
        for code, label in (('cash', 'Cash'), ('card', 'Card'), ('mbway', 'MBWay'))
        if payment_totals.get(code)
    )
    return render(request, 'stock/outbound.html', {
        'success': f'Sale completed. {order_qty} item(s), €{order_revenue:.2f} ({pay_summary}).'
    })


# -----------------------------
# 产品列表
# -----------------------------
@login_required
def product_list_view(request):
    show_sales_sensitive = has_sales_sensitive_access(request.user)
    query = request.GET.get('q', '').strip()
    selected_category = request.GET.get('category')
    selected_brand = request.GET.get('brand')
    stock_status = (request.GET.get('stock_status') or '').strip()
    sort_by = (request.GET.get('sort') or 'latest').strip()
    state = get_filtered_product_list_state(
        show_sales_sensitive=show_sales_sensitive,
        query=query,
        selected_category=selected_category,
        selected_brand=selected_brand,
        stock_status=stock_status,
        sort_by=sort_by,
    )

    products_qs = state['products_qs']
    categories = state['categories']
    brand_options = state['brand_options']
    selected_category_int = state['selected_category']
    selected_brand_id = state['selected_brand']
    sort_by = state['sort_by']
    active_filter_labels = state['active_filter_labels']

    total_types = products_qs.count()
    in_stock_types = products_qs.filter(total_stock__gt=0).count()

    paginator = Paginator(products_qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    for p in page_obj:
        p.display_price = p.default_price
        p.primary_image_url = get_product_image_url(p)

    total_stock = sum(p.total_stock for p in page_obj)
    query_params = request.GET.copy()
    query_params.pop('page', None)
    context = {
        'products': page_obj,
        'page_obj': page_obj,
        'query': query,
        'categories': categories,
        'brands': brand_options,
        'selected_category': selected_category_int,
        'selected_brand': selected_brand_id,
        'stock_status': stock_status,
        'sort_by': sort_by,
        'total_types': total_types,
        'in_stock_types': in_stock_types,
        'total_stock': total_stock,
        'filter_querystring': query_params.urlencode(),
        'active_filter_labels': active_filter_labels,
        'can_manage': has_manager_access(request.user),
        'show_sales_sensitive': show_sales_sensitive,
        'export_defaults': {
            'price_mode': 'retail',
            'only_in_stock': stock_status == 'in_stock',
            'include_images': True,
        },
        'shopify_export_defaults': {
            'only_in_stock': stock_status == 'in_stock',
        },
    }
    return render(request, 'stock/product_list.html', context)


def get_filtered_product_list_state(*, show_sales_sensitive, query, selected_category, selected_brand, stock_status, sort_by):
    recent_sales_start = timezone.now() - timedelta(days=30)
    stock_subquery = (
        Purchase.objects
        .filter(product=OuterRef('pk'))
        .values('product')
        .annotate(total=Sum('remaining'))
        .values('total')[:1]
    )
    total_sales_subquery = (
        Sale.objects
        .filter(product=OuterRef('pk'))
        .values('product')
        .annotate(total=Sum('quantity'))
        .values('total')[:1]
    )
    recent_sales_subquery = (
        Sale.objects
        .filter(product=OuterRef('pk'), date__gte=recent_sales_start)
        .values('product')
        .annotate(total=Sum('quantity'))
        .values('total')[:1]
    )
    last_sold_subquery = (
        Sale.objects
        .filter(product=OuterRef('pk'))
        .order_by('-date')
        .values('date')[:1]
    )

    products_qs = Product.objects.annotate(
        total_stock=Coalesce(Subquery(stock_subquery, output_field=IntegerField()), Value(0), output_field=IntegerField()),
    )
    if show_sales_sensitive:
        products_qs = products_qs.annotate(
            total_sold_qty=Coalesce(Subquery(total_sales_subquery, output_field=IntegerField()), Value(0), output_field=IntegerField()),
            recent_sold_qty=Coalesce(
                Subquery(recent_sales_subquery, output_field=IntegerField()),
                Value(0),
                output_field=IntegerField(),
            ),
            last_sold_at=Subquery(last_sold_subquery, output_field=DateTimeField()),
        )
    else:
        products_qs = products_qs.annotate(
            total_sold_qty=Value(0, output_field=IntegerField()),
            recent_sold_qty=Value(0, output_field=IntegerField()),
            last_sold_at=Value(None, output_field=DateTimeField()),
        )
    products_qs = (
        products_qs
        .select_related('category', 'brand_master', 'series_master')
        .prefetch_related('images')
    )

    if query:
        products_qs = products_qs.filter(
            Q(name__icontains=query) |
            Q(barcode__icontains=query) |
            Q(brand__icontains=query) |
            Q(model__icontains=query) |
            Q(spec__icontains=query) |
            Q(color__icontains=query)
        )

    if selected_category:
        products_qs = products_qs.filter(category_id=selected_category)

    brand_options = (
        Brand.objects
        .filter(products__isnull=False)
        .order_by('name')
        .distinct()
    )
    if selected_category:
        brand_options = brand_options.filter(products__category_id=selected_category).distinct()

    selected_brand_id = None
    selected_brand_obj = None
    if selected_brand and str(selected_brand).isdigit():
        selected_brand_id = int(selected_brand)
        selected_brand_obj = Brand.objects.filter(id=selected_brand_id).first()
        if selected_brand_obj:
            products_qs = products_qs.filter(
                Q(brand_master_id=selected_brand_id) |
                Q(brand__iexact=selected_brand_obj.name)
            )

    if stock_status == 'in_stock':
        products_qs = products_qs.filter(total_stock__gt=0)
    elif stock_status == 'low_stock':
        products_qs = products_qs.filter(total_stock__gt=0, total_stock__lte=5)
    elif stock_status == 'out_of_stock':
        products_qs = products_qs.filter(total_stock__lte=0)

    allowed_sort_options = {'latest', 'name_asc', 'stock_desc', 'price_desc', 'price_asc'}
    if show_sales_sensitive:
        allowed_sort_options.update({'sales_desc', 'recent_sales_desc'})
    if sort_by not in allowed_sort_options:
        sort_by = 'latest'

    sort_options = {
        'latest': ['-id'],
        'name_asc': ['brand', 'model', 'name', 'id'],
        'stock_desc': ['-total_stock', 'brand', 'model', 'name'],
        'sales_desc': ['-total_sold_qty', '-recent_sold_qty', 'brand', 'model', 'name'],
        'recent_sales_desc': ['-recent_sold_qty', '-total_sold_qty', 'brand', 'model', 'name'],
        'price_desc': ['-default_price', 'brand', 'model', 'name'],
        'price_asc': ['default_price', 'brand', 'model', 'name'],
    }
    products_qs = products_qs.order_by(*sort_options.get(sort_by, sort_options['latest']))

    categories = Category.objects.all()

    active_filter_labels = []
    if query:
        active_filter_labels.append({'label': f'Search: {query}'})
    if selected_category and str(selected_category).isdigit():
        category_obj = categories.filter(id=int(selected_category)).first()
        if category_obj:
            active_filter_labels.append({'label': f'Category: {category_obj.name}'})
    if selected_brand_obj:
        active_filter_labels.append({'label': f'Brand: {selected_brand_obj.name}'})
    if stock_status:
        stock_label_map = {
            'in_stock': 'In stock',
            'low_stock': 'Low stock',
            'out_of_stock': 'Out of stock',
        }
        active_filter_labels.append({'label': f'Stock: {stock_label_map.get(stock_status, stock_status)}'})

    sort_label_map = {
        'latest': 'Latest added',
        'name_asc': 'Name A-Z',
        'stock_desc': 'Stock high to low',
        'sales_desc': 'Best selling',
        'recent_sales_desc': 'Best selling 30 days',
        'price_desc': 'Retail price high to low',
        'price_asc': 'Retail price low to high',
    }
    active_filter_labels.append({'label': f'Sort: {sort_label_map.get(sort_by, "Latest added")}'})

    return {
        'products_qs': products_qs,
        'categories': categories,
        'brand_options': brand_options,
        'selected_category': int(selected_category) if selected_category and str(selected_category).isdigit() else None,
        'selected_brand': selected_brand_id,
        'selected_brand_obj': selected_brand_obj,
        'stock_status': stock_status,
        'sort_by': sort_by,
        'active_filter_labels': active_filter_labels,
    }


@login_required
def export_product_list_excel(request):
    show_sales_sensitive = has_sales_sensitive_access(request.user)
    query = request.GET.get('q', '').strip()
    selected_category = request.GET.get('category')
    selected_brand = request.GET.get('brand')
    stock_status = (request.GET.get('stock_status') or '').strip()
    sort_by = (request.GET.get('sort') or 'latest').strip()
    price_mode = (request.GET.get('price_mode') or 'retail').strip()
    only_in_stock = (request.GET.get('only_in_stock') or '').strip().lower() in {'1', 'true', 'on', 'yes'}
    include_images = (request.GET.get('include_images') or '').strip().lower() in {'1', 'true', 'on', 'yes'}

    if price_mode not in {'retail', 'wholesale', 'both'}:
        price_mode = 'retail'

    state = get_filtered_product_list_state(
        show_sales_sensitive=show_sales_sensitive,
        query=query,
        selected_category=selected_category,
        selected_brand=selected_brand,
        stock_status=stock_status,
        sort_by=sort_by,
    )
    products_qs = state['products_qs']
    if only_in_stock:
        products_qs = products_qs.filter(total_stock__gt=0)

    products = list(products_qs)
    for product in products:
        product.export_title = build_customer_product_title(product)
        # Append the Specification (volume, e.g. "100ml") after the product name.
        spec = (product.spec or '').strip()
        if spec:
            product.export_title = f'{product.export_title} {spec}'
        product.export_brand = customer_catalog_case((product.brand or '').strip()) or 'No Brand'
        product.export_model = customer_catalog_case((product.model or '').strip()) or 'Other Selections'
        product.export_category_name = customer_catalog_case(getattr(product.category, 'name', ''))
        product.export_availability, _ = get_catalog_availability_parts(product.total_stock)

    base_params = request.GET.copy()
    for key in ['price_mode', 'only_in_stock', 'include_images']:
        base_params.pop(key, None)
    fallback_url = reverse('product_list')
    if base_params:
        fallback_url = f"{fallback_url}?{base_params.urlencode()}"

    if not products:
        messages.warning(request, 'No products matched the current filters and export options.')
        return redirect(fallback_url)

    brand_groups = {}
    for product in products:
        brand_groups.setdefault(product.export_brand, []).append(product)

    wb = Workbook()
    title_fill = PatternFill('solid', fgColor='FFF6EFE4')
    header_fill = PatternFill('solid', fgColor='FFE8EEF6')
    section_fill = PatternFill('solid', fgColor='FFF7F4EE')
    thin = Side(border_style='thin', color='FFD7DFE8')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    wrap = Alignment(wrap_text=True, vertical='top')
    center = Alignment(horizontal='center', vertical='center')
    sheet_names = set()

    def unique_sheet_title(raw_title):
        cleaned = (raw_title or 'Products').strip()
        for char in ['\\', '/', '*', '[', ']', ':', '?']:
            cleaned = cleaned.replace(char, ' ')
        cleaned = cleaned[:31] or 'Products'
        candidate = cleaned
        suffix = 2
        while candidate in sheet_names:
            suffix_text = f" {suffix}"
            candidate = f"{cleaned[:31-len(suffix_text)]}{suffix_text}"
            suffix += 1
        sheet_names.add(candidate)
        return candidate

    filter_summary_bits = [item['label'] for item in state['active_filter_labels']]
    if only_in_stock:
        filter_summary_bits.append('Export: only in-stock products')
    filter_summary_bits.append(
        {
            'retail': 'Price: retail only',
            'wholesale': 'Price: wholesale only',
            'both': 'Price: retail + wholesale',
        }[price_mode]
    )
    filter_summary_bits.append('Images: yes' if include_images else 'Images: no')
    filter_summary = ' | '.join(filter_summary_bits)

    first_sheet = True
    for brand_name, brand_products in brand_groups.items():
        if first_sheet:
            ws = wb.active
            first_sheet = False
        else:
            ws = wb.create_sheet()
        ws.title = unique_sheet_title(brand_name)

        columns = []
        if include_images:
            columns.append(('Image', 16))
        columns.extend([
            ('Product', 34),
            ('EAN', 16),
            ('Category', 18),
        ])
        if price_mode in {'retail', 'both'}:
            columns.append(('Retail Price', 15))
        if price_mode in {'wholesale', 'both'}:
            columns.append(('Wholesale Price', 17))
        columns.append(('Availability', 20))

        for idx, (_, width) in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        max_col = len(columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.cell(row=1, column=1, value=f'{brand_name} | Customer Product List')
        ws.cell(row=1, column=1).fill = title_fill
        ws.cell(row=1, column=1).font = Font(bold=True, size=15)
        ws.cell(row=1, column=1).alignment = Alignment(vertical='center')

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        ws.cell(row=2, column=1, value=filter_summary)
        ws.cell(row=2, column=1).alignment = wrap
        ws.cell(row=2, column=1).font = Font(color='FF5B6675')

        current_row = 4
        model_groups = {}
        for product in brand_products:
            model_groups.setdefault(product.export_model, []).append(product)

        for model_name, model_products in model_groups.items():
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_col)
            ws.cell(row=current_row, column=1, value=model_name)
            ws.cell(row=current_row, column=1).fill = section_fill
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
            current_row += 1

            for col_idx, (header, _) in enumerate(columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = center
            current_row += 1

            for product in model_products:
                row_idx = current_row
                col_idx = 1

                if include_images:
                    ws.cell(row=row_idx, column=col_idx).border = border
                    ws.cell(row=row_idx, column=col_idx).alignment = center
                    try:
                        first_image = product.images.all()[0]
                        pil = PILImage.open(first_image.image.path)
                        pil.thumbnail((120, 120))
                        tmp = _BytesIO()
                        pil.save(tmp, format='PNG')
                        tmp.seek(0)
                        xlimg = XLImage(tmp)
                        ws.add_image(xlimg, f'{get_column_letter(col_idx)}{row_idx}')
                        ws.row_dimensions[row_idx].height = 88
                    except Exception:
                        pass
                    col_idx += 1

                ws.cell(row=row_idx, column=col_idx, value=product.export_title).alignment = wrap
                ws.cell(row=row_idx, column=col_idx).border = border
                col_idx += 1

                # EAN (barcode) — stored as text so long codes keep every digit.
                ean_cell = ws.cell(row=row_idx, column=col_idx, value=(product.barcode or '-'))
                ean_cell.border = border
                ean_cell.alignment = center
                ean_cell.number_format = '@'
                col_idx += 1

                ws.cell(row=row_idx, column=col_idx, value=product.export_category_name or '-').alignment = wrap
                ws.cell(row=row_idx, column=col_idx).border = border
                col_idx += 1

                if price_mode in {'retail', 'both'}:
                    retail_cell = ws.cell(row=row_idx, column=col_idx, value=float(product.default_price) if product.default_price is not None else None)
                    retail_cell.number_format = '"EUR" #,##0.00'
                    retail_cell.border = border
                    col_idx += 1

                if price_mode in {'wholesale', 'both'}:
                    wholesale_cell = ws.cell(row=row_idx, column=col_idx, value=float(product.wholesale_price) if product.wholesale_price is not None else None)
                    wholesale_cell.number_format = '"EUR" #,##0.00'
                    wholesale_cell.border = border
                    col_idx += 1

                availability_cell = ws.cell(row=row_idx, column=col_idx, value=product.export_availability)
                availability_cell.border = border
                availability_cell.alignment = center
                current_row += 1

            current_row += 1

        ws.freeze_panes = 'A5'

    export_buffer = _BytesIO()
    wb.save(export_buffer)
    export_buffer.seek(0)
    ts = timezone.localtime().strftime('%Y%m%d_%H%M')
    filename = f'Customer_Product_List_{ts}.xlsx'
    response = FileResponse(export_buffer, as_attachment=True, filename=filename)
    response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@login_required
@manager_required
def export_shopify_inventory_csv(request):
    show_sales_sensitive = has_sales_sensitive_access(request.user)
    query = request.GET.get('q', '').strip()
    selected_category = request.GET.get('category')
    selected_brand = request.GET.get('brand')
    stock_status = (request.GET.get('stock_status') or '').strip()
    sort_by = (request.GET.get('sort') or 'latest').strip()
    only_in_stock = (request.GET.get('only_in_stock') or '').strip().lower() in {'1', 'true', 'on', 'yes'}

    state = get_filtered_product_list_state(
        show_sales_sensitive=show_sales_sensitive,
        query=query,
        selected_category=selected_category,
        selected_brand=selected_brand,
        stock_status=stock_status,
        sort_by=sort_by,
    )
    products_qs = state['products_qs']
    if only_in_stock:
        products_qs = products_qs.filter(total_stock__gt=0)

    products = list(products_qs)
    base_params = request.GET.copy()
    for key in ['only_in_stock']:
        base_params.pop(key, None)
    fallback_url = reverse('product_list')
    if base_params:
        fallback_url = f"{fallback_url}?{base_params.urlencode()}"

    if not products:
        messages.warning(request, 'No products matched the current filters and Shopify export options.')
        return redirect(fallback_url)

    product_groups = {}
    for product in products:
        key = (
            _clean_csv_text(product.brand).lower(),
            _clean_csv_text(product.model).lower(),
            _clean_csv_text(product.name).lower(),
            product.category_id,
        )
        product_groups.setdefault(key, []).append(product)

    ts = timezone.localtime().strftime('%Y%m%d_%H%M')
    filename = f'Shopify_Product_Inventory_{ts}.csv'
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')

    writer = csv.DictWriter(response, fieldnames=SHOPIFY_PRODUCT_CSV_HEADERS, lineterminator='\n')
    writer.writeheader()

    used_handles = set()
    for group_products in product_groups.values():
        first_product = group_products[0]
        base_title = _build_shopify_base_title(first_product)
        handle = _build_unique_shopify_handle(base_title, first_product.barcode or first_product.id, used_handles)
        product_description = _first_clean_product_value(group_products, 'description')
        product_category = _clean_csv_text(getattr(first_product.category, 'name', ''))
        shopify_taxonomy = SHOPIFY_TAXONOMY_BY_CATEGORY.get(product_category, '')
        google_category = GOOGLE_CATEGORY_BY_CATEGORY.get(product_category, '')
        product_image_url = _first_shopify_image_url(request, group_products)
        color_metafield = _build_shopify_color_metafield(group_products)
        tags = _build_shopify_tags(group_products)

        for index, product in enumerate(group_products):
            current_stock = int(getattr(product, 'total_stock', 0) or 0)
            variant_image_url = _build_shopify_image_url(request, product)
            row = {header: '' for header in SHOPIFY_PRODUCT_CSV_HEADERS}

            row.update({
                'URL handle': handle,
                'SKU': product.barcode,
                'Barcode': product.barcode,
                'Price': _format_shopify_money(product.default_price),
                'Cost per item': _format_shopify_money(product.current_fifo_cost_price()),
                'Charge tax': 'TRUE',
                'Inventory tracker': 'shopify',
                'Inventory quantity': str(current_stock),
                'Continue selling when out of stock': 'DENY',
                'Weight unit for display': 'g',
                'Requires shipping': 'TRUE',
                'Fulfillment service': 'manual',
                'Variant image URL': variant_image_url,
            })

            if index == 0:
                row.update({
                    'Title': base_title,
                    'Description': product_description,
                    'Vendor': _clean_csv_text(first_product.brand),
                    'Product category': shopify_taxonomy,
                    'Type': product_category,
                    'Tags': tags,
                    'Published on online store': 'TRUE',
                    'Status': 'Active',
                    'Product image URL': product_image_url,
                    'Image position': '1' if product_image_url else '',
                    'Image alt text': base_title if product_image_url else '',
                    'Gift card': 'FALSE',
                    'SEO title': base_title,
                    'SEO description': product_description,
                    'Color (product.metafields.shopify.color-pattern)': color_metafield,
                    'Google Shopping / Google product category': google_category,
                    'Google Shopping / Condition': 'New',
                    'Google Shopping / Custom product': 'FALSE',
                })

            _apply_shopify_option_columns(row, product, group_products)

            writer.writerow(row)

    return response


# -----------------------------
# 产品增/改/详情
# -----------------------------
@login_required
def add_product_view(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, can_edit_prices=has_manager_access(request.user))
        if form.is_valid():
            product = form.save()
            for img in request.FILES.getlist('images'):
                ProductImage.objects.create(product=product, image=img)
            return redirect('product_detail', pk=product.pk)
    else:
        form = ProductForm(can_edit_prices=has_manager_access(request.user))
    return render(request, 'stock/add_product.html', {
        'form': form,
        'brand_series_map_json': json.dumps(build_brand_series_map()),
        'brand_catalog_json': json.dumps(build_brand_catalog()),
    })


@login_required
def product_detail_view(request, pk):
    product = get_object_or_404(
        Product.objects.select_related('category', 'brand_master', 'series_master').prefetch_related('images'),
        pk=pk
    )

    purchases = product.purchase_set.select_related('supplier').order_by('-date')

    # Sales History on this page shows only the last 10 days; the rest lives on the
    # dedicated product sales-history page (linked via full_sales_url).
    sales_base = (
        product.sale_set
        .annotate(business_at=Coalesce('order__created_at', 'date'))
        .select_related('customer', 'order', 'order__store', 'store', 'order__customer')
        .order_by('-business_at', '-id')
    )
    sales_total_count = sales_base.count()
    recent_cutoff = timezone.localdate() - timedelta(days=10)
    sales = list(sales_base.filter(business_at__date__gte=recent_cutoff))
    sales_older_count = sales_total_count - len(sales)
    full_sales_url = f"{reverse('product_sales_history')}?product={product.id}"

    total_stock = sum(p.remaining for p in purchases if p.remaining > 0)
    fifo_purchase = product.purchase_set.filter(remaining__gt=0).order_by('date').first()
    fifo_price = fifo_purchase.cost_price if fifo_purchase else None

    # Per-supplier cost comparison/trend (purchases is ordered -date, so first seen = latest).
    supplier_cost_map = {}
    for p in purchases:
        entry = supplier_cost_map.setdefault(p.supplier_id, {
            'supplier_id': p.supplier_id,
            'name': p.supplier.name if p.supplier else 'No supplier',
            'batches': 0, 'units': 0, 'costs': [],
            'last_cost': None, 'last_date': None,
        })
        entry['batches'] += 1
        entry['units'] += p.quantity
        if p.cost_price is not None:
            entry['costs'].append(p.cost_price)
        if entry['last_date'] is None:
            entry['last_cost'] = p.cost_price
            entry['last_date'] = p.date

    supplier_costs = []
    for entry in supplier_cost_map.values():
        costs = entry.pop('costs')
        entry['min_cost'] = min(costs) if costs else None
        entry['max_cost'] = max(costs) if costs else None
        entry['avg_cost'] = (sum(costs) / len(costs)) if costs else None
        supplier_costs.append(entry)
    supplier_costs.sort(key=lambda e: (e['last_cost'] is None, e['last_cost'] or Decimal('0.00')))

    priced = [e for e in supplier_costs if e['last_cost'] is not None]
    cheapest_supplier_id = min(priced, key=lambda e: e['last_cost'])['supplier_id'] if len(priced) > 1 else None

    is_manager = has_manager_access(request.user)
    # The full reconstructed Stock Ledger lives on the dedicated /sales-history/
    # page (merged with sales detail); the product page just links there.
    is_perfume = bool(product.category and 'perfum' in (product.category.name or '').lower())

    return render(request, 'stock/product_detail.html', {
        'product': product,
        'purchases': purchases,
        'sales': sales,
        'sales_older_count': sales_older_count,
        'full_sales_url': full_sales_url,
        'total_stock': total_stock,
        'fifo_price': fifo_price,
        'supplier_costs': supplier_costs,
        'cheapest_supplier_id': cheapest_supplier_id,
        'show_sensitive': is_manager,
        'is_perfume': is_perfume,
        'show_sales_sensitive': has_sales_sensitive_access(request.user),
    })


@login_required
def product_sales_history_view(request):
    """Dedicated, deep sales page for one product: search by name, then show the
    complete sales detail (store / date-time / order / customer / qty / price /
    payment / profit*) plus the full stock ledger and per-store / per-month rollups."""
    if not has_manager_access(request.user):
        return redirect('sales_records')

    q = request.GET.get('q', '').strip()
    product_id = request.GET.get('product', '').strip()
    start_str = request.GET.get('start_date', '').strip()
    end_str = request.GET.get('end_date', '').strip()
    store_id = request.GET.get('store', '').strip()
    show_profit = bool(getattr(request.user, 'is_superuser', False))
    payment_labels = dict(Sale.PAYMENT_METHOD_CHOICES)

    context = {
        'q': q,
        'show_profit': show_profit,
        'stores': Store.objects.order_by('name'),
        'selected_store_id': store_id,
        'start_date': start_str,
        'end_date': end_str,
    }

    product = None
    if product_id.isdigit():
        product = (
            Product.objects.filter(pk=int(product_id))
            .select_related('category', 'brand_master').prefetch_related('images').first()
        )

    if not product:
        matches = []
        if q:
            matches = list(
                Product.objects.filter(
                    Q(name__icontains=q) | Q(barcode__icontains=q)
                    | Q(model__icontains=q) | Q(brand__icontains=q)
                ).annotate(sold=Coalesce(Sum('sale__quantity'), 0)).order_by('brand', 'name')[:50]
            )
        context['matches'] = matches
        return render(request, 'stock/product_sales_history.html', context)

    context['product'] = product
    ledger = build_stock_ledger(product)
    context['stock_ledger'] = ledger
    context['ledger_detailed'] = True  # merged Sales Detail + Stock Ledger view
    context['on_hand'] = sum(p.remaining for p in product.purchase_set.all())

    # Profit for the whole ledger (every sale, unfiltered) computed once and reused
    # both for the merged ledger rows and the filtered KPI totals below.
    ledger_sale_ids = [e['id'] for e in ledger['events'] if e['kind'] == 'sale']
    profit_map = sale_profit_map_for_sale_ids(ledger_sale_ids) if show_profit else {}
    if show_profit:
        for e in ledger['events']:
            if e['kind'] == 'sale':
                e['profit'] = profit_map.get(e['id'], {}).get('profit')

    sales_qs = (
        Sale.objects.filter(product=product)
        .annotate(business_at=Coalesce('order__created_at', 'date'))
        .select_related('order', 'order__store', 'order__customer', 'customer', 'store')
        .order_by('-business_at', '-id')
    )

    def _parse(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except Exception:
            return None

    sd, ed = _parse(start_str), _parse(end_str)
    if sd:
        sales_qs = sales_qs.filter(business_at__date__gte=sd)
    if ed:
        sales_qs = sales_qs.filter(business_at__date__lte=ed)
    if store_id.isdigit():
        sid = int(store_id)
        sales_qs = sales_qs.filter(Q(order__store_id=sid) | Q(order__isnull=True, store_id=sid))

    sales = list(sales_qs)

    rows, store_agg, month_agg, order_ids = [], {}, {}, set()
    total_qty, total_rev, total_profit = 0, Decimal('0.00'), Decimal('0.00')
    for s in sales:
        order = s.order
        line_total = (s.unit_price or Decimal('0.00')) * s.quantity
        store = order.store if (order and order.store_id) else s.store
        store_name = store.name if store else '—'
        cust = order.customer if (order and order.customer_id) else s.customer
        at = timezone.localtime(s.business_at) if s.business_at else None
        profit = profit_map.get(s.id, {}).get('profit', Decimal('0.00')) if show_profit else Decimal('0.00')
        rows.append({
            'at': at,
            'store': store_name,
            'order_id': order.id if order else None,
            'detail_url': reverse('sale_order_detail', args=[order.id]) if order else None,
            'customer': cust.name if cust else 'Walk-in / No customer',
            'customer_url': reverse('customer_detail', args=[cust.id]) if cust else None,
            'qty': s.quantity,
            'unit_price': s.unit_price,
            'line_total': line_total,
            'payment': payment_labels.get(s.payment_method, (s.payment_method or 'Other').title()),
            'profit': profit,
            'affects_stock': order.affects_stock if order else True,
        })
        total_qty += s.quantity
        total_rev += line_total
        total_profit += profit
        order_ids.add(order.id if order else f'legacy-{s.id}')
        st = store_agg.setdefault(store_name, {'qty': 0, 'rev': Decimal('0.00')})
        st['qty'] += s.quantity
        st['rev'] += line_total
        if at:
            mm = month_agg.setdefault(at.strftime('%Y-%m'), {'qty': 0, 'rev': Decimal('0.00')})
            mm['qty'] += s.quantity
            mm['rev'] += line_total

    context.update({
        'rows': rows,
        'total_qty': total_qty,
        'total_rev': total_rev,
        'total_profit': total_profit,
        'order_count': len(order_ids),
        'avg_unit': (total_rev / total_qty) if total_qty else Decimal('0.00'),
        'store_breakdown': sorted(
            ({'store': k, **v} for k, v in store_agg.items()), key=lambda r: -r['rev']),
        'month_breakdown': sorted(
            ({'month': k, **v} for k, v in month_agg.items()), key=lambda r: r['month'], reverse=True),
        'first_sale': rows[-1]['at'] if rows else None,
        'last_sale': rows[0]['at'] if rows else None,
    })
    return render(request, 'stock/product_sales_history.html', context)


@login_required
def download_db_backup(request):
    """Manager-only one-click download of a consistent SQLite snapshot.

    Lets the owner grab a full backup on demand (e.g. straight onto a USB by
    pointing the browser's download folder at it). Uses SQLite's online backup
    API against the live connection, so it's consistent even mid-write. The whole
    database is a single file, so this file *is* the backup — restoring is just
    putting it back and reloading."""
    if not has_manager_access(request.user):
        return redirect('dashboard')
    from django.db import connection
    if connection.vendor != 'sqlite':
        return HttpResponse('Backup is only supported on SQLite.', status=400)

    import sqlite3
    import tempfile
    import os
    fd, tmp_path = tempfile.mkstemp(suffix='.sqlite3')
    os.close(fd)
    try:
        connection.ensure_connection()
        target = sqlite3.connect(tmp_path)
        try:
            connection.connection.backup(target)
        finally:
            target.close()
        with open(tmp_path, 'rb') as fh:
            data = fh.read()
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    stamp = timezone.localtime().strftime('%Y%m%d-%H%M%S')
    resp = HttpResponse(data, content_type='application/octet-stream')
    resp['Content-Disposition'] = f'attachment; filename="scentory-db-{stamp}.sqlite3"'
    return resp


@login_required
def sync_all_perfumes_to_shopify(request):
    """Manager-only product-list button: sync every perfume to Shopify. The job
    (200+ products) runs in a detached background process so the request returns
    immediately; progress goes to logs/shopify_perfumes_sync.log."""
    if request.method != 'POST' or not has_manager_access(request.user):
        return redirect('product_list')

    from .services.shopify_client import ShopifyClient
    if not ShopifyClient().is_configured():
        messages.error(request, 'Shopify is not configured (SHOPIFY_ADMIN_TOKEN missing).')
        return redirect('product_list')

    import os
    import threading
    from django.conf import settings
    from django.core.management import call_command
    log_dir = os.path.join(settings.BASE_DIR, 'logs')
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, 'shopify_perfumes_sync.log')

    def _run():
        # Run the command in-process in a background thread (avoids the uwsgi
        # 'which python' problem of spawning a subprocess). Output -> log file.
        try:
            # buffering=1 -> line-buffered, so progress is visible live in the log
            # and survives if the worker is recycled mid-run.
            with open(log_path, 'a', encoding='utf-8', buffering=1) as f:
                f.write(f'\n=== sync started {timezone.now():%Y-%m-%d %H:%M} ===\n')
                call_command('sync_shopify_perfumes', apply=True, stdout=f, stderr=f)
        except Exception as exc:  # noqa: BLE001 — record any crash to the log
            try:
                with open(log_path, 'a', encoding='utf-8') as f:
                    f.write(f'\nSync crashed: {exc}\n')
            except OSError:
                pass

    threading.Thread(target=_run, daemon=True).start()
    messages.success(request, 'Syncing all perfumes to Shopify in the background — '
                              'this takes a few minutes. Progress: logs/shopify_perfumes_sync.log')
    return redirect('product_list')


@login_required
def sync_product_to_shopify(request, pk):
    """Manager-only button on the product page (perfumes only): push this product
    to Shopify — create it if missing, then set its price + decant-aware inventory.
    Handy for listing a new perfume, or a one-off manual re-sync."""
    product = get_object_or_404(Product, pk=pk)
    if request.method != 'POST' or not has_manager_access(request.user):
        return redirect('product_detail', pk=pk)
    if not (product.category and 'perfum' in (product.category.name or '').lower()):
        messages.error(request, 'Shopify sync is for perfumes only.')
        return redirect('product_detail', pk=pk)

    from .services.shopify_client import ShopifyClient
    from .services import shopify_sync
    client = ShopifyClient()
    if not client.is_configured():
        messages.error(request, 'Shopify is not configured (SHOPIFY_ADMIN_TOKEN missing).')
        return redirect('product_detail', pk=pk)

    try:
        code, detail = shopify_sync.sync_product(product, client, create_missing=True, status='ACTIVE')
        if code == shopify_sync.SKIP_NO_BARCODE:
            messages.error(request, 'Product has no barcode — cannot match it on Shopify.')
            return redirect('product_detail', pk=pk)
        if code == shopify_sync.ERROR:
            messages.error(request, f'Shopify: {detail}')
            return redirect('product_detail', pk=pk)
        if code == shopify_sync.CREATED:
            messages.success(request, f'Created on Shopify: {detail}.')

        inv_code, inv_detail = shopify_sync.sync_product_price_inventory(
            product, client, do_price=True, do_inventory=True)
    except Exception as exc:  # keep the button from 500-ing on a Shopify hiccup
        messages.error(request, f'Shopify sync failed: {exc}')
        return redirect('product_detail', pk=pk)

    if inv_code == shopify_sync.INV_UPDATED:
        messages.success(request, f'Synced to Shopify — {inv_detail}.')
    elif inv_code == shopify_sync.INV_UNCHANGED:
        messages.info(request, 'Already up to date on Shopify.')
    elif inv_code == shopify_sync.INV_NOT_IN_SHOPIFY:
        messages.warning(request, 'Not found on Shopify to set price/inventory.')
    elif inv_code == shopify_sync.INV_ERROR:
        messages.error(request, f'Inventory sync: {inv_detail}')
    return redirect('product_detail', pk=pk)


@login_required
def edit_product_view(request, pk):
    product = get_object_or_404(
        Product.objects.select_related('category', 'brand_master', 'series_master'),
        pk=pk
    )

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product,
                           can_edit_prices=has_manager_access(request.user))
        if form.is_valid():
            form.save()

            # ✅ 同时兼容 images / images[] 两种 name
            new_files = []
            new_files.extend(request.FILES.getlist('images'))
            new_files.extend(request.FILES.getlist('images[]'))

            saved_count = 0
            for f in new_files:
                if not f or not getattr(f, 'size', 0):
                    continue
                ProductImage.objects.create(product=product, image=f)
                saved_count += 1

            if saved_count:
                messages.success(request, f"Product saved. {saved_count} image(s) uploaded.")
            else:
                # 如果用户确实选择了文件，但后端没接到，多半是前端 name/表单问题
                if request.POST.get('had_image_selected') == '1':
                    messages.warning(
                        request,
                        "Product saved, but no images were received. Please re-select the files and try again."
                    )
                else:
                    messages.success(request, "Product saved.")

            # ✅ 保留当前 URL 上的查询串（包含 page、q、category 等）
            qs = request.META.get('QUERY_STRING', '')
            detail_url = reverse('product_detail', args=[pk])
            if qs:
                detail_url = f"{detail_url}?{qs}"
            return redirect(detail_url)
        else:
            messages.error(request, "Please fix the highlighted fields.")
    else:
        form = ProductForm(instance=product, can_edit_prices=has_manager_access(request.user))

    images = product.images.all()
    purchases = product.purchase_set.all().order_by('-date')
    purchase_count = product.purchase_set.count()
    sale_count = product.sale_set.count()
    can_delete_product = purchase_count == 0 and sale_count == 0
    can_force_delete_product = request.user.is_superuser
    # total available stock (sum of remaining from all batches)
    total_stock = purchases.aggregate(s=Coalesce(Sum('remaining'), Value(0), output_field=IntegerField()))['s'] or 0

    return render(request, 'stock/edit_product.html', {
        'form': form,
        'product': product,
        'images': images,
        'purchases': purchases,
        'total_stock': total_stock,
        'purchase_count': purchase_count,
        'sale_count': sale_count,
        'can_delete_product': can_delete_product,
        'can_force_delete_product': can_force_delete_product,
        'brand_series_map_json': json.dumps(build_brand_series_map()),
        'brand_catalog_json': json.dumps(build_brand_catalog()),
    })


@require_POST
@login_required
@manager_required
def delete_product_view(request, pk):
    product = get_object_or_404(Product, pk=pk)
    purchase_count = product.purchase_set.count()
    sale_count = product.sale_set.count()
    query_string = request.POST.get('return_querystring', '').strip()
    force_delete = request.POST.get('force_delete') == '1'

    if (purchase_count or sale_count) and not (force_delete and request.user.is_superuser):
        blockers = []
        if purchase_count:
            blockers.append(f'{purchase_count} purchase record{"s" if purchase_count != 1 else ""}')
        if sale_count:
            blockers.append(f'{sale_count} sales record{"s" if sale_count != 1 else ""}')
        messages.error(
            request,
            f'Cannot delete "{product.display_name}" because it still has '
            + ' and '.join(blockers)
            + '.'
        )
        edit_url = reverse('edit_product', args=[product.pk])
        if query_string:
            edit_url = f'{edit_url}?{query_string}'
        return redirect(edit_url)

    product_name = product.display_name
    sale_dates = list(product.sale_set.dates('date', 'day'))
    affected_order_ids = list(
        product.sale_set.exclude(order_id__isnull=True).values_list('order_id', flat=True).distinct()
    )
    affected_inbound_ids = list(
        product.purchase_set.exclude(inbound_order_id__isnull=True).values_list('inbound_order_id', flat=True).distinct()
    )
    image_count = product.images.count()

    with transaction.atomic():
        for image in list(product.images.all()):
            if image.image:
                image.image.delete(save=False)

        product.delete()

        empty_order_count = 0
        if affected_order_ids:
            empty_orders = list(
                SaleOrder.objects.filter(id__in=affected_order_ids)
                .annotate(item_count=Count('items'))
                .filter(item_count=0)
            )
            empty_order_count = len(empty_orders)
            if empty_orders:
                SaleOrder.objects.filter(id__in=[order.id for order in empty_orders]).delete()

        empty_inbound_count = 0
        if affected_inbound_ids:
            inbound_orders = list(InboundOrder.objects.filter(id__in=affected_inbound_ids).prefetch_related('items'))
            empty_inbound_ids = []
            for inbound_order in inbound_orders:
                items = list(inbound_order.items.all())
                if not items:
                    empty_inbound_ids.append(inbound_order.id)
                    continue
                inbound_order.total_amount = sum(
                    (Decimal(item.quantity) * item.cost_price for item in items),
                    Decimal('0.00'),
                )
                inbound_order.save(update_fields=['total_amount'])
            if empty_inbound_ids:
                empty_inbound_count = len(empty_inbound_ids)
                InboundOrder.objects.filter(id__in=empty_inbound_ids).delete()

    if force_delete and request.user.is_superuser and (purchase_count or sale_count):
        messages.success(
            request,
            f'Product "{product_name}" force-deleted with {purchase_count} purchase record'
            f'{"s" if purchase_count != 1 else ""}, {sale_count} sales record'
            f'{"s" if sale_count != 1 else ""}, and {image_count} image'
            f'{"s" if image_count != 1 else ""} removed.'
            + (
                f' Cleaned up {empty_order_count} empty order{"s" if empty_order_count != 1 else ""}'
                f' and {empty_inbound_count} empty inbound order{"s" if empty_inbound_count != 1 else ""}.'
                if empty_order_count or empty_inbound_count else ''
            )
        )
    else:
        messages.success(request, f'Product "{product_name}" deleted.')

    product_list_url = reverse('product_list')
    if query_string:
        product_list_url = f'{product_list_url}?{query_string}'
    return redirect(product_list_url)


# -----------------------------
# 前端查条码（用于入/出库页）
# 这里返回库存、最近进价等；前端想隐藏展示也没问题
# -----------------------------
@login_required
def dashboard_view(request):
    isManger = has_manager_access(request.user)
    show_sales_sensitive = has_sales_sensitive_access(request.user)
    show_order_financials = has_order_reconciliation_access(request.user)
    show_profit = bool(request.user.is_superuser)

    active_store, store_is_all = resolve_active_store(request)
    store_scope = None if store_is_all else active_store
    store_cache_key = store_scope.id if store_scope else 'all'

    site_url = request.build_absolute_uri('/')

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=6,
        border=2,
    )
    qr.add_data(site_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

    today = timezone.localdate()
    selected_cat_ids = request.GET.getlist('cat')
    selected_cat_ids_int = [int(x) for x in selected_cat_ids if str(x).isdigit()]

    all_categories = Category.objects.order_by('name')
    perfume_category = (
        all_categories
        .filter(Q(name__iexact='Perfumes') | Q(name__icontains='perfume'))
        .first()
    )
    dashboard_defaulted_to_perfume = False
    if not selected_cat_ids_int and perfume_category:
        selected_cat_ids_int = [perfume_category.id]
        dashboard_defaulted_to_perfume = True
    categories = all_categories
    selected_cats = list(categories.filter(id__in=selected_cat_ids_int))

    month_context = resolve_dashboard_month(request.GET.get('month', '').strip(), today)

    # Short-TTL cache for the monthly snapshot (the heavy part: FIFO profit replay,
    # inventory/AR scans). Today's live order list below is computed fresh, so the
    # ~60s staleness only affects the monthly overview, not today's operations.
    cache_scope = ','.join(str(cid) for cid in sorted(selected_cat_ids_int))
    snapshot_cache_key = (
        f"dash:snap:{month_context['month_value']}:{cache_scope}:"
        f"{int(show_profit)}:{month_context['period_end'].isoformat()}:store{store_cache_key}"
    )
    monthly_overview = cache.get(snapshot_cache_key)
    if monthly_overview is None:
        monthly_overview = build_monthly_dashboard_snapshot(
            month_start=month_context['month_start'],
            period_end=month_context['period_end'],
            selected_category_ids=selected_cat_ids_int,
            show_profit=show_profit,
            store=store_scope,
        )
        cache.set(snapshot_cache_key, monthly_overview, 300)

    current_headline = {
        'sales_amount': monthly_overview['month_sales_amount'],
        'profit': monthly_overview['month_sales_profit'],
        'order_count': monthly_overview['month_order_count'],
        'avg_ticket': monthly_overview['month_avg_ticket'],
    }
    comparison_cache_key = f"dash:cmp:{month_context['month_value']}:{cache_scope}:{int(show_profit)}:{month_context['period_end'].isoformat()}:store{store_cache_key}"
    mom_comparison = cache.get(comparison_cache_key)
    if mom_comparison is None:
        mom_comparison = build_period_comparison(
            month_context, current_headline,
            selected_category_ids=selected_cat_ids_int,
            show_profit=show_profit,
            store=store_scope,
        )
        cache.set(comparison_cache_key, mom_comparison, 300)

    target_progress = build_target_progress(
        month_context, monthly_overview['month_sales_amount'],
        selected_category_ids=selected_cat_ids_int,
    )

    def build_dashboard_url(month_value=None, category_ids=None):
        params = []
        if month_value:
            params.append(('month', month_value))
        for category_id in category_ids or []:
            params.append(('cat', str(category_id)))
        query = urlencode(params, doseq=True)
        base_url = reverse('dashboard')
        return f'{base_url}?{query}' if query else base_url

    for point in monthly_overview['chart_data']:
        point['detail_url'] = (
            f"{reverse('sales_records')}?"
            f"{urlencode({'start_date': point['date'].isoformat(), 'end_date': point['date'].isoformat()})}"
        )

    selected_cat_chips = []
    for category in selected_cats:
        remaining_ids = [cat_id for cat_id in selected_cat_ids_int if cat_id != category.id]
        selected_cat_chips.append({
            'name': category.name,
            'remove_url': build_dashboard_url(month_context['month_value'], remaining_ids),
        })

    payment_labels = dict(Sale.PAYMENT_METHOD_CHOICES)

    # Employees have no access to the manager-only /today page, so give them
    # today's sales here. Managers keep the clean dashboard + /today, so this is
    # only computed for employees (no profit replay — they never see profit).
    sale_orders_today = []
    today_sales_qty = 0
    total_sales_today = Decimal('0.00')
    today_payment_breakdown = []
    if not isManger:
        orders_qs = (
            SaleOrder.objects
            .filter(created_at__date=today)
            .select_related('customer')
            .prefetch_related('items__product', 'items__product__images', 'payments')
            .order_by('-created_at')
            .distinct()
        )
        orders_qs = scope_sales_by_store(orders_qs, store_scope, store_is_all)
        if selected_cat_ids_int:
            orders_qs = orders_qs.filter(items__product__category_id__in=selected_cat_ids_int).distinct()

        today_payment_totals = defaultdict(Decimal)
        today_payment_qty = defaultdict(int)
        for order in orders_qs:
            view_total_qty = 0
            view_total_amount = Decimal('0.00')
            items_today = []
            for item in order.items.all():
                if selected_cat_ids_int and item.product.category_id not in selected_cat_ids_int:
                    continue
                item.line_total = (item.unit_price or Decimal('0.00')) * item.quantity
                item.payment_label = payment_labels.get(item.payment_method, (item.payment_method or 'Other').title())
                item.image_url = get_product_image_url(item.product)
                items_today.append(item)
                view_total_qty += item.quantity
                view_total_amount += item.line_total
                today_sales_qty += item.quantity
                total_sales_today += item.line_total
                today_payment_qty[item.payment_method] += item.quantity
            if items_today:
                # Order-level tender (split payments), scaled to the filtered subtotal.
                for method, amount in order_tender_amounts(order, view_total_amount, items_today).items():
                    today_payment_totals[method] += amount
                order.view_total_qty = view_total_qty
                order.view_total_amount = view_total_amount
                order.view_items_today = items_today
                sale_orders_today.append(order)

        for code, amount in sorted(today_payment_totals.items(), key=lambda kv: (-kv[1], kv[0] or '')):
            today_payment_breakdown.append({
                'code': code,
                'label': payment_labels.get(code, (code or 'Other').title()),
                'amount': amount,
                'qty': today_payment_qty.get(code, 0),
                'share_pct': (amount / total_sales_today * Decimal('100')) if total_sales_today else Decimal('0.00'),
            })

    # Low-stock section — cached (per store + category); the sold subquery over all
    # products is the main per-load cost, so recompute at most every few minutes.
    THUMB = 'c_fill,w_96,h_96,q_auto,f_auto'
    low_cache_key = f"dash:low:{cache_scope}:store{store_cache_key}"
    low_data = cache.get(low_cache_key)
    if low_data is None:
        stock_subq = (
            Purchase.objects.filter(product=OuterRef('pk')).values('product')
            .annotate(s=Sum('remaining')).values('s')[:1]
        )
        sold_subq = (
            Sale.objects.filter(product=OuterRef('pk')).values('product')
            .annotate(s=Sum('quantity')).values('s')[:1]
        )
        prod_base = Product.objects
        if selected_cat_ids_int:
            prod_base = prod_base.filter(category_id__in=selected_cat_ids_int)
        low_qs = (
            prod_base
            .annotate(stock=Subquery(stock_subq, output_field=IntegerField()),
                      sold=Subquery(sold_subq, output_field=IntegerField()))
            .annotate(stock=Coalesce(F('stock'), 0), sold=Coalesce(F('sold'), 0))
            .filter(stock__lt=5)
            .only('id', 'brand', 'model', 'name', 'category')
            .prefetch_related('images')
        )
        low_by_brand = defaultdict(list)
        for product in low_qs:
            stock = int(product.stock or 0)
            low_by_brand[(product.brand or 'Unknown brand')].append({
                'id': product.id,
                'display_name': product.display_name,
                'brand': product.brand or 'Unknown brand',
                'stock': stock,
                'sold': int(product.sold or 0),
                # small CDN thumbnail (edge-cached, resized) with local fallback
                'image_url': product_image_cdn_url(product, THUMB) or get_product_image_url(product),
            })
        for _brand, _products in low_by_brand.items():
            _products.sort(key=lambda it: (-it['sold'], it['stock'], it['display_name'].lower()))
        blocks = [
            {'brand': brand, 'products': products, 'count': len(products),
             'out_count': sum(1 for it in products if it['stock'] == 0)}
            for brand, products in low_by_brand.items()
        ]
        blocks.sort(key=lambda it: (-it['count'], it['brand'].lower()))
        all_low = [it for b in blocks for it in b['products']]
        focus = sorted(all_low, key=lambda it: (
            0 if it['stock'] == 0 else 1, it['stock'], -it['sold'], it['display_name'].lower()))[:8]
        low_data = {
            'blocks': blocks,
            'count': len(all_low),
            'out_of_stock_count': sum(1 for it in all_low if it['stock'] == 0),
            'low_stock_1_2_count': sum(1 for it in all_low if 1 <= it['stock'] <= 2),
            'low_stock_3_4_count': sum(1 for it in all_low if 3 <= it['stock'] <= 4),
            'focus': focus,
        }
        cache.set(low_cache_key, low_data, 300)

    low_brand_blocks = low_data['blocks']
    low_stock_count = low_data['count']
    out_of_stock_count = low_data['out_of_stock_count']
    low_stock_1_2_count = low_data['low_stock_1_2_count']
    low_stock_3_4_count = low_data['low_stock_3_4_count']
    low_stock_focus_products = low_data['focus']
    low_stock_primary_blocks = low_brand_blocks[:6]
    low_stock_extra_blocks = low_brand_blocks[6:]
    low_stock_extra_brand_count = len(low_stock_extra_blocks)
    low_stock_extra_product_count = sum(block['count'] for block in low_stock_extra_blocks)

    prod_for_totals = Product.objects.all()
    if selected_cat_ids_int:
        prod_for_totals = prod_for_totals.filter(category_id__in=selected_cat_ids_int)

    total_products = prod_for_totals.count()
    total_stock = monthly_overview['current_stock_units']


    return render(request, 'stock/dashboard.html', {
        'categories': categories,
        'selected_cat_ids': [str(i) for i in selected_cat_ids_int],
        'selected_cats': selected_cats,
        'selected_cat_chips': selected_cat_chips,
        'selected_month': month_context['month_value'],
        'month_label': month_context['month_label'],
        'month_scope_label': month_context['scope_label'],
        'period_start': month_context['month_start'],
        'period_end': month_context['period_end'],
        'prev_month_url': build_dashboard_url(month_context['prev_month_value'], selected_cat_ids_int),
        'current_month_url': build_dashboard_url(today.strftime('%Y-%m'), selected_cat_ids_int),
        'next_month_url': build_dashboard_url(month_context['next_month_value'], selected_cat_ids_int) if month_context['next_month_value'] else '',
        'sales_records_month_url': (
            f"{reverse('sales_records')}?"
            f"{urlencode({'start_date': month_context['month_start'].isoformat(), 'end_date': month_context['period_end'].isoformat()})}"
        ),
        'today': today,
        'today_sales_qty': today_sales_qty,
        'today_order_count': len(sale_orders_today),
        'total_sales_today': total_sales_today,
        'today_payment_breakdown': today_payment_breakdown,
        'sale_orders_today': sale_orders_today,
        'low_brand_blocks': low_brand_blocks,
        'low_stock_primary_blocks': low_stock_primary_blocks,
        'low_stock_extra_blocks': low_stock_extra_blocks,
        'low_stock_extra_brand_count': low_stock_extra_brand_count,
        'low_stock_extra_product_count': low_stock_extra_product_count,
        'low_stock_focus_products': low_stock_focus_products,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'low_stock_1_2_count': low_stock_1_2_count,
        'low_stock_3_4_count': low_stock_3_4_count,
        'total_products': total_products,
        'total_stock': total_stock,
        'site_url': site_url,
        'qr_code_base64': qr_base64,
        'isManger': isManger,
        'dashboard_defaulted_to_perfume': dashboard_defaulted_to_perfume,
        'dashboard_default_category_name': perfume_category.name if perfume_category else '',
        'employee_daily_only': not isManger,
        'show_purchase_metrics': isManger,
        'show_sales_sensitive': show_sales_sensitive,
        'show_order_financials': show_order_financials,
        'show_profit': show_profit,
        'ar_scope_note': bool(selected_cat_ids_int),
        'mom_comparison': mom_comparison,
        'target_progress': target_progress,
        'store_is_all': store_is_all,
        **monthly_overview,
    })


@login_required
def yearly_sales_view(request):
    """Sales Trend was merged into the Sales records page (its no-range state is
    the yearly overview). Keep this route as a redirect for old links."""
    params = []
    year_value = (request.GET.get('year') or '').strip()
    if year_value:
        params.append(('year', year_value))
    for category_id in request.GET.getlist('cat'):
        params.append(('cat', str(category_id)))
    query = urlencode(params, doseq=True)
    base_url = reverse('sales_records')
    return redirect(f'{base_url}?{query}' if query else base_url)


@login_required
def check_barcode(request):
    barcode = request.GET.get('barcode', '').strip()
    try:
        product = Product.objects.select_related('brand_master', 'series_master').prefetch_related('images').get(barcode=barcode)
        # 当前库存（可用不用于表单，仅供前端显示时参考；这里不必返回也可）
        stock = product.purchase_set.filter(remaining__gt=0).aggregate(s=Sum('remaining'))['s'] or 0
        # 最近一次进价
        last_purchase = product.purchase_set.order_by('-date').first()
        last_cost = last_purchase.cost_price if last_purchase else None
        image_url = ''
        for image in product.images.all():
            try:
                image_url = image.image.url
            except ValueError:
                image_url = ''
            break
        display_name = product.display_name

        return JsonResponse({
            'exists': True,
            'name': display_name,
            'display_name': display_name,
            'brand': product.brand,
            'model': product.model or '',
            'product_name': product.name,
            'spec': product.spec or '',
            'color': product.color or '',
            'stock': int(stock),
            'last_cost': float(last_cost) if last_cost is not None else None,
            'retail_price': float(product.default_price) if product.default_price is not None else None,
            'wholesale_price': float(product.wholesale_price) if product.wholesale_price is not None else None,
            'image_url': image_url,
        })
    except Product.DoesNotExist:
        return JsonResponse({'exists': False})



# -----------------------------
# 销售汇总列表/详情
# -----------------------------

@login_required
def sale_order_detail_view(request, order_id):
    order_qs = SaleOrder.objects.select_related('customer')
    if not has_manager_access(request.user):
        active_store, store_is_all = resolve_active_store(request)
        order_qs = scope_sales_by_store(order_qs, active_store, store_is_all)
    order = get_object_or_404(order_qs, id=order_id)
    items = list(
        order.items
        .select_related('product', 'product__brand_master', 'product__series_master')
        .order_by('id')
    )

    show_sensitive = has_manager_access(request.user)
    show_sales_sensitive = has_sales_sensitive_access(request.user)
    show_order_financials = has_order_reconciliation_access(request.user)
    can_print_receipt = request.user.is_authenticated
    print_mode = request.GET.get('print') == '1'
    print_layout = 'pos' if request.GET.get('layout') == 'pos' else 'a4'
    show_receipt_amounts = show_order_financials

    total_qty = sum(i.quantity for i in items)
    total_amount = sum((i.unit_price * i.quantity for i in items), Decimal('0.00'))

    from collections import defaultdict
    pay_break = defaultdict(Decimal)
    for i in items:
        i.display_name = build_product_label(i.product)
        i.line_total = (i.unit_price or Decimal('0.00')) * i.quantity
        i.image_url = get_product_image_url(i.product)
        pay_break[i.payment_method] += i.unit_price * i.quantity

    shop_profile = PrintProfile.get_for_store(order.store)
    back_url = request.GET.get('next', '').strip() or reverse('sale_order_detail', args=[order.id])

    context = {
        'order': order,
        'items': items,
        'total_qty': total_qty,
        'total_amount': total_amount,
        'pay_break': dict(pay_break),
        'can_admin_correct': has_admin_access(request.user),
        'can_print_receipt': can_print_receipt,
        'print_mode': print_mode,
        'print_layout': print_layout,
        'show_receipt_amounts': show_receipt_amounts,
        'show_sensitive': show_sensitive,
        'show_sales_sensitive': show_sales_sensitive,
        'show_order_financials': show_order_financials,
        'print_a4_url': f"{reverse('sale_order_detail', args=[order.id])}?print=1&layout=a4",
        'print_pos_url': f"{reverse('sale_order_detail', args=[order.id])}?print=1&layout=pos",
        'edit_print_profile_url': f"{reverse('print_profile_edit')}?{urlencode({'next': back_url})}",
        'shop': {
            'name': shop_profile.name,
            'nif': shop_profile.nif,
            'phone': shop_profile.phone,
            'address': shop_profile.address,
            'email': shop_profile.email,
            'footer_note': shop_profile.footer_note,
        },
    }
    return render(request, 'stock/sale_order_detail.html', context)


@login_required
@admin_required
def print_profile_edit_view(request):
    active_store, store_is_all = resolve_active_store(request)
    target_store = active_store or Store.get_default()
    profile = PrintProfile.get_for_store(target_store)
    next_url = request.GET.get('next', '').strip() or request.POST.get('next', '').strip() or reverse('dashboard')

    if request.method == 'POST':
        form = PrintProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, f'Print header updated for {target_store.name if target_store else "the shop"}.')
            return redirect(next_url)
        messages.error(request, 'Please fix the highlighted fields.')
    else:
        form = PrintProfileForm(instance=profile)

    return render(request, 'stock/print_profile_form.html', {
        'form': form,
        'next_url': next_url,
        'target_store': target_store,
        'store_is_all': store_is_all,
    })


def _correction_cart_item(product, *, qty, price, payment, is_split=False, payment_split=None):
    return {
        'product_id': product.id,
        'barcode': product.barcode,
        'title': build_product_label(product),
        'image_url': get_product_image_url(product),
        'retail': float(product.default_price) if product.default_price is not None else None,
        'wholesale': float(product.wholesale_price) if product.wholesale_price is not None else None,
        'qty': qty,
        'price': price,
        'payment': payment,
        'isSplit': is_split,
        'paymentSplit': payment_split,
    }


def _build_correction_cart(order):
    """Rich cart payload (with product info) preloaded into the JS cart on edit.

    The authoritative payment split lives in the order-level ``SaleOrderPayment``
    rows (each ``Sale`` line only stores its *primary* method). Reconstruct a
    valid per-line split by greedily allocating the order-level payment pool
    across the lines, so a split order reloads — and re-saves — without
    collapsing back to a single method.
    """
    if not order:
        return []

    valid_methods = ('cash', 'card', 'mbway')
    lines = [
        item for item in order.items.select_related('product').order_by('id')
        if item.product
    ]

    pool = defaultdict(Decimal)
    for payment in order.payments.all():
        if payment.method in valid_methods and payment.amount:
            pool[payment.method] += payment.amount
    if not pool:
        # No order-level tender recorded: fall back to each line's own method.
        for item in lines:
            method = item.payment_method if item.payment_method in valid_methods else 'cash'
            pool[method] += (item.unit_price or Decimal('0')) * item.quantity

    cart = []
    for item in lines:
        remaining = (item.unit_price or Decimal('0')) * item.quantity
        line_map = {}
        for method in valid_methods:
            if remaining <= 0:
                break
            available = pool.get(method, Decimal('0'))
            if available <= 0:
                continue
            take = min(available, remaining)  # each method is visited once per line
            line_map[method] = take
            pool[method] = available - take
            remaining -= take
        if remaining > Decimal('0.01'):
            # Pool underflow (data drift): cover the rest with the line's method.
            fallback = item.payment_method if item.payment_method in valid_methods else 'cash'
            line_map[fallback] = line_map.get(fallback, Decimal('0')) + remaining

        methods_used = [m for m in valid_methods if line_map.get(m, Decimal('0')) > 0]
        is_split = len(methods_used) > 1
        payment_split = {m: float(round(line_map[m], 2)) for m in methods_used} if is_split else None
        if methods_used:
            primary = max(methods_used, key=lambda m: line_map[m])
        else:
            primary = item.payment_method if item.payment_method in valid_methods else 'cash'

        cart.append(_correction_cart_item(
            item.product, qty=item.quantity, price=float(item.unit_price),
            payment=primary, is_split=is_split, payment_split=payment_split,
        ))
    return cart


def _enrich_correction_items(raw_items):
    """Rebuild the rich cart from submitted minimal items (used on validation error)."""
    ids = []
    for row in raw_items:
        try:
            ids.append(int(row.get('product_id')))
        except (TypeError, ValueError):
            continue
    products = {p.id: p for p in Product.objects.filter(id__in=ids)}
    cart = []
    for row in raw_items:
        try:
            product = products[int(row.get('product_id'))]
            qty = int(row.get('qty', 0))
            price = float(row.get('price', 0))
        except (TypeError, ValueError, KeyError):
            continue
        cart.append(_correction_cart_item(product, qty=qty, price=price, payment=str(row.get('payment', '')).strip()))
    return cart


def _parse_correction_cart(request):
    """Parse + validate items_json/payments_json from the correction POST.

    Returns ``(line_items, payment_totals, order_revenue)`` mirroring the POS
    outbound contract. Raises ``ValueError`` with a user-facing message.
    """
    valid_methods = {'cash', 'card', 'mbway'}
    method_priority = {'cash': 0, 'card': 1, 'mbway': 2}

    try:
        items = json.loads(request.POST.get('items_json') or '[]')
    except (ValueError, TypeError):
        raise ValueError('Could not read the order lines.')
    if not isinstance(items, list) or not items:
        raise ValueError('Add at least one product line before saving.')

    parsed = []
    order_revenue = Decimal('0.00')
    line_payment_totals = defaultdict(Decimal)
    for row in items:
        try:
            product_id = int(row.get('product_id'))
            qty = int(row.get('qty', 0))
            price = Decimal(str(row.get('price', 0)))
        except (TypeError, ValueError, ArithmeticError):
            raise ValueError('Quantity and price must be numbers.')
        if qty < 1 or price <= 0:
            raise ValueError('Each line needs a product, quantity of at least 1, and a price above 0.')
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            raise ValueError('A selected product no longer exists.')
        line_payment = str(row.get('payment', '')).strip()
        line_total = price * qty
        order_revenue += line_total
        if line_payment in valid_methods:
            line_payment_totals[line_payment] += line_total
        parsed.append({'product': product, 'quantity': qty, 'unit_price': price, 'payment': line_payment})

    try:
        raw_payments = json.loads(request.POST.get('payments_json') or '[]')
    except (ValueError, TypeError):
        raw_payments = []

    payment_totals = defaultdict(Decimal)
    if isinstance(raw_payments, list) and raw_payments:
        for entry in raw_payments:
            method = str(entry.get('method', '')).strip()
            try:
                amount = Decimal(str(entry.get('amount', 0)))
            except (TypeError, ValueError, ArithmeticError):
                raise ValueError('Payment amount must be a number.')
            if method in valid_methods and amount > 0:
                payment_totals[method] += amount
    elif line_payment_totals:
        payment_totals = line_payment_totals

    if not payment_totals:
        raise ValueError('Choose a payment method for the order.')

    paid_total = sum(payment_totals.values(), Decimal('0.00'))
    if abs(paid_total - order_revenue) > Decimal('0.01'):
        raise ValueError(f'Payments (EUR {paid_total:.2f}) must add up to the order total (EUR {order_revenue:.2f}).')

    primary_method = min(payment_totals.items(), key=lambda kv: (-kv[1], method_priority.get(kv[0], 9)))[0]
    line_items = [
        {
            'product': row['product'],
            'quantity': row['quantity'],
            'unit_price': row['unit_price'],
            'payment_method': row['payment'] if row['payment'] in valid_methods else primary_method,
        }
        for row in parsed
    ]
    return line_items, dict(payment_totals), order_revenue


@login_required
@admin_required
def sale_order_correction_center_view(request):
    query = request.GET.get('q', '').strip()
    start_date = parse_date((request.GET.get('start_date') or '').strip())
    end_date = parse_date((request.GET.get('end_date') or '').strip())

    active_store, store_is_all = resolve_active_store(request)
    orders_qs = scope_sales_by_store(
        SaleOrder.objects
        .select_related('customer', 'store')
        .prefetch_related('items')
        .order_by('-created_at', '-id'),
        active_store, store_is_all,
    )

    if query:
        filters = (
            Q(customer__name__icontains=query) |
            Q(customer__nif__icontains=query) |
            Q(note__icontains=query)
        )
        if query.isdigit():
            filters |= Q(id=int(query))
        orders_qs = orders_qs.filter(filters)

    if start_date:
        orders_qs = orders_qs.filter(created_at__date__gte=start_date)
    if end_date:
        orders_qs = orders_qs.filter(created_at__date__lte=end_date)

    paginator = Paginator(orders_qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Group the current page's orders by day (newest first) so the list reads as
    # dated sections rather than one flat run; each day carries a small subtotal.
    order_date_groups = []
    current_group = None
    for order in page_obj.object_list:
        items = list(order.items.all())
        order.view_total_qty = sum(item.quantity for item in items)
        order.view_total_amount = sum((item.quantity * item.unit_price for item in items), Decimal('0.00'))
        order.view_store = order.store.name if order.store_id else '—'

        order_day = timezone.localtime(order.created_at).date()
        if current_group is None or current_group['date'] != order_day:
            current_group = {'date': order_day, 'orders': [], 'count': 0, 'amount': Decimal('0.00')}
            order_date_groups.append(current_group)
        current_group['orders'].append(order)
        current_group['count'] += 1
        current_group['amount'] += order.view_total_amount

    recent_logs = (
        SaleOrderChangeLog.objects
        .select_related('changed_by', 'order')
        .order_by('-created_at', '-id')[:12]
    )

    return render(request, 'stock/sale_order_correction_center.html', {
        'page_obj': page_obj,
        'order_date_groups': order_date_groups,
        'store_is_all': store_is_all,
        'recent_logs': recent_logs,
        'query': query,
        'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', ''),
    })


def _sale_order_correction_view(request, order=None):
    order = order or None
    initial_cart = _build_correction_cart(order)

    if request.method == 'POST':
        form = SaleOrderCorrectionForm(request.POST, instance=order)

        if order and 'delete_order' in request.POST:
            reason = (request.POST.get('reason') or '').strip()
            if not reason:
                form.add_error('reason', 'Reason is required before deleting an order.')
            if not form.errors:
                try:
                    delete_sale_order_correction(
                        order=order,
                        changed_by=request.user,
                        reason=reason,
                    )
                except ValidationError as exc:
                    form.add_error(None, '; '.join(exc.messages))
                else:
                    messages.success(request, f'Order #{order.id} deleted and stock restored.')
                    return redirect('sale_order_correction_center')
        else:
            # Preserve the submitted cart so a validation error does not wipe the work.
            try:
                submitted_raw = json.loads(request.POST.get('items_json') or '[]')
                if isinstance(submitted_raw, list) and submitted_raw:
                    initial_cart = _enrich_correction_items(submitted_raw)
            except (ValueError, TypeError):
                pass

            if form.is_valid():
                try:
                    line_items, payment_totals, _ = _parse_correction_cart(request)
                    _active_stores = available_stores()
                    _raw_store = request.POST.get('store')
                    selected_store = next(
                        (s for s in _active_stores if str(s.id) == str(_raw_store)), None
                    )
                    if selected_store is None:
                        selected_store = (order.store if (order and order.store_id) else None) or store_for_new_sale(request)
                    affects_stock = ('affects_stock' in request.POST) if order is None else None
                    saved_order = save_sale_order_correction(
                        order=order,
                        customer=form.cleaned_data.get('customer'),
                        note=form.cleaned_data.get('note'),
                        order_datetime=form.cleaned_data['order_datetime'],
                        line_items=line_items,
                        payment_totals=payment_totals,
                        changed_by=request.user,
                        reason=form.cleaned_data['reason'],
                        store=selected_store,
                        affects_stock=affects_stock,
                    )
                except ValueError as exc:
                    messages.error(request, str(exc))
                except ValidationError as exc:
                    form.add_error(None, '; '.join(exc.messages))
                else:
                    action_label = 'created' if order is None else 'updated'
                    messages.success(request, f'Order #{saved_order.id} {action_label}.')
                    return redirect('sale_order_correction_edit', order_id=saved_order.id)
            else:
                messages.error(request, 'Please fix the highlighted fields.')
    else:
        form = SaleOrderCorrectionForm(instance=order)

    change_logs = []
    if order and order.pk:
        change_logs = order.change_logs.select_related('changed_by').order_by('-created_at', '-id')[:12]

    stores_for_template = available_stores()
    _raw_selected = request.POST.get('store') if request.method == 'POST' else None
    selected_store_id = None
    if _raw_selected:
        _match = next((s for s in stores_for_template if str(s.id) == str(_raw_selected)), None)
        selected_store_id = _match.id if _match else None
    if selected_store_id is None:
        _default_store = (order.store if (order and order.store_id) else None) or store_for_new_sale(request)
        selected_store_id = _default_store.id if _default_store else None

    affects_stock_checked = ('affects_stock' in request.POST) if request.method == 'POST' else True

    return render(request, 'stock/sale_order_correction_form.html', {
        'form': form,
        'order': order,
        'is_create': order is None,
        'change_logs': change_logs,
        'initial_cart': initial_cart,
        'store_options': stores_for_template,
        'selected_store_id': selected_store_id,
        'affects_stock_checked': affects_stock_checked,
    })


@login_required
@admin_required
def sale_order_correction_create_view(request):
    return _sale_order_correction_view(request, order=None)


@login_required
@admin_required
def sale_order_correction_edit_view(request, order_id):
    order = get_object_or_404(
        SaleOrder.objects.select_related('customer').prefetch_related('items__product'),
        id=order_id,
    )
    return _sale_order_correction_view(request, order=order)


@login_required
@manager_required
def supplier_list_view(request):
    query = (request.GET.get('q') or '').strip()
    country_filter = (request.GET.get('country') or '').strip()
    suppliers_qs = (
        Supplier.objects
        .prefetch_related('product_types')
        .annotate(
            purchase_count=Count('purchase', distinct=True),
            inbound_count=Count('inboundorder', distinct=True),
            last_purchase_at=Max('purchase__date'),
        )
        .order_by('country', 'name')
    )

    if query:
        suppliers_qs = suppliers_qs.filter(
            Q(name__icontains=query) |
            Q(phone__icontains=query) |
            Q(country__icontains=query) |
            Q(address__icontains=query) |
            Q(product_types__name__icontains=query)
        ).distinct()
    if country_filter == '__none__':
        suppliers_qs = suppliers_qs.filter(Q(country__isnull=True) | Q(country=''))
    elif country_filter:
        suppliers_qs = suppliers_qs.filter(country=country_filter)

    suppliers = list(suppliers_qs)
    supplier_count = len(suppliers)
    linked_suppliers = sum(1 for supplier in suppliers if supplier.purchase_count or supplier.inbound_count)

    # 按国家分组（未填国家排在最后）
    groups_map = {}
    for supplier in suppliers:
        groups_map.setdefault((supplier.country or '').strip(), []).append(supplier)
    supplier_groups = [
        {'country': key or 'No country', 'suppliers': groups_map[key]}
        for key in sorted(groups_map, key=lambda k: (k == '', k.lower()))
    ]

    country_options = list(
        Supplier.objects.exclude(country__isnull=True).exclude(country='')
        .values_list('country', flat=True).distinct().order_by('country')
    )

    return render(request, 'stock/supplier_list.html', {
        'supplier_groups': supplier_groups,
        'query': query,
        'country_filter': country_filter,
        'country_options': country_options,
        'supplier_count': supplier_count,
        'linked_suppliers': linked_suppliers,
        'can_manage': has_manager_access(request.user),
    })


@login_required
@manager_required
def supplier_detail_view(request, supplier_id):
    supplier = get_object_or_404(Supplier.objects.prefetch_related('product_types'), id=supplier_id)
    query = (request.GET.get('q') or '').strip()
    start_date = parse_date((request.GET.get('start_date') or '').strip())
    end_date = parse_date((request.GET.get('end_date') or '').strip())
    next_url = request.get_full_path()

    inbound_orders_qs = (
        InboundOrder.objects
        .filter(Q(supplier=supplier) | Q(items__supplier=supplier))
        .select_related('supplier')
        .prefetch_related('items__product', 'items__product__images', 'items__supplier')
        .order_by('-invoice_date', '-created_at', '-id')
        .distinct()
    )
    direct_purchases_qs = (
        Purchase.objects
        .filter(supplier=supplier, inbound_order__isnull=True)
        .select_related('product')
        .prefetch_related('product__images')
        .order_by('-date', '-id')
    )

    if query:
        inbound_orders_qs = inbound_orders_qs.filter(
            Q(invoice_no__icontains=query) |
            Q(note__icontains=query) |
            Q(items__product__barcode__icontains=query) |
            Q(items__product__name__icontains=query) |
            Q(items__product__brand__icontains=query) |
            Q(items__product__model__icontains=query)
        ).distinct()
        direct_purchases_qs = direct_purchases_qs.filter(
            Q(product__barcode__icontains=query) |
            Q(product__name__icontains=query) |
            Q(product__brand__icontains=query) |
            Q(product__model__icontains=query)
        )

    if start_date:
        inbound_orders_qs = inbound_orders_qs.filter(created_at__date__gte=start_date)
        direct_purchases_qs = direct_purchases_qs.filter(date__date__gte=start_date)
    if end_date:
        inbound_orders_qs = inbound_orders_qs.filter(created_at__date__lte=end_date)
        direct_purchases_qs = direct_purchases_qs.filter(date__date__lte=end_date)

    history_rows = []
    history_total_amount = Decimal('0.00')
    history_total_qty = 0
    product_ids = set()
    last_purchase_at = None

    for order in inbound_orders_qs:
        if order.supplier_id == supplier.id:
            items = list(order.items.all())
        else:
            items = [item for item in order.items.all() if item.supplier_id == supplier.id]
        if not items:
            continue

        row_amount = Decimal('0.00')
        row_qty = 0
        for item in items:
            item.line_total = (item.cost_price or Decimal('0.00')) * item.quantity
            item.image_url = get_product_image_url(item.product)
            item.display_name = build_product_label(item.product)
            row_amount += item.line_total
            row_qty += item.quantity
            product_ids.add(item.product_id)

        sort_at = order.invoice_date or timezone.localdate()
        created_anchor = comparable_recorded_at(order.created_at or timezone.now())
        sort_dt = comparable_recorded_at(
            created_anchor.replace(year=sort_at.year, month=sort_at.month, day=sort_at.day)
        )
        history_rows.append({
            'kind': 'Inbound order',
            'title': order.invoice_no or f'Inbound Order #{order.id}',
            'sort_at': sort_dt,
            'display_date': order.invoice_date or order.created_at.date(),
            'meta': f'{len(items)} items',
            'amount': row_amount,
            'qty': row_qty,
            'items': items,
            'note': order.note or '',
            'edit_url': f"{reverse('inbound_order_edit', args=[order.id])}?{urlencode({'next': next_url})}",
        })
        history_total_amount += row_amount
        history_total_qty += row_qty
        if not last_purchase_at or sort_dt > comparable_recorded_at(last_purchase_at):
            last_purchase_at = sort_dt

    direct_groups = group_purchases_by_recorded_second(direct_purchases_qs)

    for group_key, items in direct_groups.items():
        group_key = comparable_recorded_at(group_key)
        row_amount = Decimal('0.00')
        row_qty = 0
        for item in items:
            item.line_total = (item.cost_price or Decimal('0.00')) * item.quantity
            item.image_url = get_product_image_url(item.product)
            item.display_name = build_product_label(item.product)
            item.edit_url = f"{reverse('direct_purchase_edit', args=[item.id])}?{urlencode({'next': next_url})}"
            row_amount += item.line_total
            row_qty += item.quantity
            product_ids.add(item.product_id)

        history_rows.append({
            'kind': 'Direct purchases',
            'title': group_key.strftime('%Y-%m-%d %H:%M:%S'),
            'sort_at': group_key,
            'display_date': group_key.date(),
            'meta': f'{len(items)} direct purchases',
            'amount': row_amount,
            'qty': row_qty,
            'items': items,
            'note': '',
        })
        history_total_amount += row_amount
        history_total_qty += row_qty
        if not last_purchase_at or group_key > comparable_recorded_at(last_purchase_at):
            last_purchase_at = group_key

    history_rows.sort(
        key=lambda row: (comparable_recorded_at(row['sort_at']), row['title']),
        reverse=True,
    )


    # ===== Scorecard (all-time, independent of the history filters above) =====
    spend_expr = ExpressionWrapper(
        F('cost_price') * F('quantity'),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )
    today_local = timezone.localdate()
    all_purchases = Purchase.objects.filter(supplier=supplier)
    agg = all_purchases.aggregate(
        spend=Sum(spend_expr),
        units=Sum('quantity'),
        skus=Count('product', distinct=True),
        orders=Count('inbound_order', distinct=True),
        last=Max('date'),
    )
    lifetime_spend = agg['spend'] or Decimal('0.00')
    year_spend = (all_purchases.filter(date__year=today_local.year)
                  .aggregate(s=Sum(spend_expr))['s'] or Decimal('0.00'))
    month_spend = (all_purchases.filter(date__year=today_local.year, date__month=today_local.month)
                   .aggregate(s=Sum(spend_expr))['s'] or Decimal('0.00'))
    direct_batches = all_purchases.filter(inbound_order__isnull=True).count()
    order_events = (agg['orders'] or 0) + direct_batches
    avg_order_value = (lifetime_spend / order_events) if order_events else Decimal('0.00')

    first_purchase = all_purchases.order_by('date').values_list('date', flat=True).first()
    orders_per_month = None
    if first_purchase and agg['last'] and order_events:
        months_active = max(1, (agg['last'].year - first_purchase.year) * 12
                            + (agg['last'].month - first_purchase.month) + 1)
        orders_per_month = order_events / months_active

    top_products = []
    for row in (all_purchases
                .values('product_id', 'product__brand', 'product__model', 'product__name', 'product__barcode')
                .annotate(spend=Sum(spend_expr), qty=Sum('quantity'))
                .order_by('-spend')[:5]):
        label = ' - '.join(p for p in [row['product__brand'], row['product__model'], row['product__name']] if p) or row['product__barcode']
        top_products.append({
            'product_id': row['product_id'],
            'label': label,
            'spend': row['spend'] or Decimal('0.00'),
            'qty': row['qty'] or 0,
        })

    # Lead time: order placed (created_at) → received (received_at), over received orders.
    lead_days = []
    for created, received in (InboundOrder.objects
                              .filter(supplier=supplier, status='received', received_at__isnull=False)
                              .values_list('created_at', 'received_at')):
        if created and received and received >= created:
            lead_days.append((received - created).total_seconds() / 86400.0)
    avg_lead_days = (sum(lead_days) / len(lead_days)) if lead_days else None

    scorecard = {
        'lifetime_spend': lifetime_spend,
        'year_spend': year_spend,
        'month_spend': month_spend,
        'units': agg['units'] or 0,
        'skus': agg['skus'] or 0,
        'order_events': order_events,
        'inbound_orders': agg['orders'] or 0,
        'direct_batches': direct_batches,
        'avg_order_value': avg_order_value,
        'orders_per_month': orders_per_month,
        'last_purchase': agg['last'],
        'avg_lead_days': avg_lead_days,
        'lead_sample': len(lead_days),
        'top_products': top_products,
    }

    history_page = Paginator(history_rows, 20).get_page(request.GET.get('page'))

    return render(request, 'stock/supplier_detail.html', {
        'supplier': supplier,
        'query': query,
        'history_page': history_page,
        'history_count': len(history_rows),
        'total_amount': history_total_amount,
        'total_qty': history_total_qty,
        'product_count': len(product_ids),
        'last_purchase_at': last_purchase_at,
        'scorecard': scorecard,
        'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', ''),
    })


@login_required
@manager_required
def inbound_order_edit_view(request, order_id):
    order = get_object_or_404(
        InboundOrder.objects.select_related('supplier').prefetch_related('items__product', 'items__product__images'),
        id=order_id,
    )
    next_url = request.GET.get('next', '').strip() or request.POST.get('next', '').strip() or reverse('sales_records')

    if request.method == 'POST':
        form = InboundOrderEditForm(request.POST, instance=order)
        formset = InboundPurchaseFormSet(request.POST, instance=order, prefix='lines')

        if form.is_valid() and formset.is_valid():
            recorded_at = normalize_recorded_at(form.cleaned_data['recorded_at'])
            with transaction.atomic():
                order = form.save(commit=False)
                order.created_at = recorded_at
                order.save()

                for line_form in formset.forms:
                    cleaned = getattr(line_form, 'cleaned_data', None)
                    if not cleaned:
                        continue

                    purchase = line_form.instance
                    sold_units = max((purchase.quantity or 0) - (purchase.remaining or 0), 0)

                    if cleaned.get('DELETE'):
                        purchase.delete()
                        continue

                    new_quantity = cleaned['quantity']
                    purchase.quantity = new_quantity
                    purchase.remaining = new_quantity - sold_units
                    purchase.cost_price = cleaned['cost_price']
                    purchase.supplier = order.supplier
                    purchase.date = recorded_at
                    purchase.inbound_order = order
                    purchase.save()

                remaining_items = list(order.items.all())
                if not remaining_items:
                    deleted_order_id = order.id
                    order.delete()
                    transaction.on_commit(rebuild_all_daily_summaries)
                    messages.success(request, f'Order #{deleted_order_id} deleted (no products left).')
                    return redirect(next_url)

                order.total_amount = sum(
                    ((item.cost_price or Decimal('0.00')) * item.quantity for item in remaining_items),
                    Decimal('0.00'),
                )
                order.save(update_fields=['total_amount', 'supplier', 'invoice_no', 'invoice_date', 'note', 'created_at'])
                transaction.on_commit(rebuild_all_daily_summaries)

            messages.success(request, f'Order #{order.id} updated.')
            return redirect(next_url)

        messages.error(request, 'Please fix the highlighted fields.')
    else:
        form = InboundOrderEditForm(instance=order)
        formset = InboundPurchaseFormSet(instance=order, prefix='lines')

    annotate_inbound_formset_runtime(formset)
    line_summary = []
    for line_form in formset.forms:
        purchase = line_form.instance
        purchase.image_url = get_product_image_url(purchase.product)
        purchase.display_name = build_product_label(purchase.product)
        line_summary.append(purchase)

    return render(request, 'stock/inbound_order_edit.html', {
        'order': order,
        'form': form,
        'formset': formset,
        'next_url': next_url,
        'line_count': len(line_summary),
    })


@login_required
@manager_required
def direct_purchase_edit_view(request, purchase_id):
    purchase = get_object_or_404(
        Purchase.objects.select_related('product', 'supplier', 'inbound_order').prefetch_related('product__images'),
        id=purchase_id,
    )
    if purchase.inbound_order_id:
        next_url = request.GET.get('next', '').strip() or reverse('sales_records')
        return redirect(f"{reverse('inbound_order_edit', args=[purchase.inbound_order_id])}?{urlencode({'next': next_url})}")

    next_url = request.GET.get('next', '').strip() or request.POST.get('next', '').strip() or reverse('sales_records')
    sold_units = max((purchase.quantity or 0) - (purchase.remaining or 0), 0)

    if request.method == 'POST':
        if 'delete_purchase' in request.POST:
            if sold_units > 0:
                messages.error(request, "This purchase already has sold units and can't be deleted.")
            else:
                purchase_label = build_product_label(purchase.product)
                with transaction.atomic():
                    purchase.delete()
                    transaction.on_commit(rebuild_all_daily_summaries)
                messages.success(request, f'Purchase "{purchase_label}" deleted.')
                return redirect(next_url)

        form = DirectPurchaseEditForm(request.POST, instance=purchase)
        if form.is_valid():
            recorded_at = normalize_recorded_at(form.cleaned_data['recorded_at'])
            with transaction.atomic():
                purchase = form.save(commit=False)
                purchase.quantity = form.cleaned_data['quantity']
                purchase.remaining = purchase.quantity - sold_units
                purchase.date = recorded_at
                purchase.inbound_order = None
                purchase.save()
                transaction.on_commit(rebuild_all_daily_summaries)
            messages.success(request, f'Purchase "{build_product_label(purchase.product)}" updated.')
            return redirect(next_url)
        messages.error(request, 'Please fix the highlighted fields.')
    else:
        form = DirectPurchaseEditForm(instance=purchase)

    purchase.sold_units = sold_units
    purchase.image_url = get_product_image_url(purchase.product)
    purchase.display_name = build_product_label(purchase.product)

    return render(request, 'stock/direct_purchase_edit.html', {
        'purchase': purchase,
        'form': form,
        'next_url': next_url,
    })


@login_required
@manager_required
def supplier_create_view(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save()
            messages.success(request, f'Supplier "{supplier.name}" created.')
            return redirect('supplier_list')
    else:
        form = SupplierForm()

    return render(request, 'stock/supplier_form.html', {
        'form': form,
        'supplier': None,
        'is_create': True,
    })


@login_required
@manager_required
def supplier_edit_view(request, supplier_id):
    supplier = get_object_or_404(Supplier.objects.prefetch_related('product_types'), id=supplier_id)

    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            supplier = form.save()
            messages.success(request, f'Supplier "{supplier.name}" updated.')
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)

    return render(request, 'stock/supplier_form.html', {
        'form': form,
        'supplier': supplier,
        'is_create': False,
    })


@login_required
@manager_required
@require_POST
def supplier_delete_view(request, supplier_id):
    supplier = get_object_or_404(Supplier, id=supplier_id)
    supplier_name = supplier.name
    supplier.delete()
    messages.success(request, f'Supplier "{supplier_name}" deleted. Past records were kept.')
    return redirect('supplier_list')


@login_required
@admin_required
def employee_list_view(request):
    query = (request.GET.get('q') or '').strip()
    UserModel = get_user_model()
    employees_qs = (
        UserModel.objects
        .filter(is_superuser=False)
        .annotate(last_clock_in_at=Max('attendance_records__clock_in_at'))
        .order_by('-is_active', '-is_staff', 'username')
    )
    if query:
        employees_qs = employees_qs.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )

    # Scope to the active store's staff (All-stores shows everyone).
    active_store, store_is_all = resolve_active_store(request)
    if not store_is_all and active_store:
        employees_qs = employees_qs.filter(store_profile__store=active_store)

    employees = list(employees_qs.select_related('store_profile__store'))
    open_shifts = {
        record.user_id: record
        for record in AttendanceRecord.objects
        .filter(user__is_superuser=False, clock_out_at__isnull=True)
        .select_related('user')
        .order_by('-clock_in_at')
    }

    for employee in employees:
        employee.role_label = 'Manager' if employee.is_staff else 'Employee'
        employee.open_shift = open_shifts.get(employee.id)

    return render(request, 'stock/employee_list.html', {
        'employees': employees,
        'query': query,
        'active_count': sum(1 for employee in employees if employee.is_active),
        'manager_count': sum(1 for employee in employees if employee.is_staff),
        'checked_in_count': len(open_shifts),
    })


@login_required
@admin_required
def employee_create_view(request):
    if request.method == 'POST':
        form = EmployeeAccountForm(request.POST)
        if form.is_valid():
            user = form.save()
            active_store, store_is_all = resolve_active_store(request)
            home_store = active_store if (not store_is_all and active_store) else Store.get_default()
            StoreProfile.objects.update_or_create(user=user, defaults={'store': home_store})
            messages.success(request, f'Employee account "{user.username}" created.')
            return redirect('employee_list')
    else:
        form = EmployeeAccountForm(initial={'is_active': True, 'role': 'employee'})

    return render(request, 'stock/employee_form.html', {
        'form': form,
        'employee': None,
        'is_create': True,
    })


@login_required
@admin_required
def employee_edit_view(request, user_id):
    UserModel = get_user_model()
    employee = get_object_or_404(UserModel, id=user_id, is_superuser=False)

    if request.method == 'POST':
        form = EmployeeAccountForm(request.POST, instance=employee)
        if form.is_valid():
            employee = form.save()
            messages.success(request, f'Employee account "{employee.username}" updated.')
            return redirect('employee_list')
    else:
        form = EmployeeAccountForm(instance=employee)

    return render(request, 'stock/employee_form.html', {
        'form': form,
        'employee': employee,
        'is_create': False,
    })


@login_required
@admin_required
@require_POST
def employee_toggle_active_view(request, user_id):
    UserModel = get_user_model()
    employee = get_object_or_404(UserModel, id=user_id, is_superuser=False)
    employee.is_active = not employee.is_active
    employee.save(update_fields=['is_active'])
    state_label = 'activated' if employee.is_active else 'paused'
    messages.success(request, f'Employee account "{employee.username}" {state_label}.')
    return redirect('employee_list')


@login_required
@admin_required
@require_POST
def employee_delete_view(request, user_id):
    UserModel = get_user_model()
    employee = get_object_or_404(UserModel, id=user_id, is_superuser=False)
    username = employee.username
    employee.delete()
    messages.success(request, f'Employee account "{username}" deleted.')
    return redirect('employee_list')


@login_required
@manager_required
def attendance_view(request):
    today = timezone.localdate()
    open_shift = (
        AttendanceRecord.objects
        .filter(user=request.user, clock_out_at__isnull=True)
        .order_by('-clock_in_at', '-id')
        .first()
    )

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        note = (request.POST.get('note') or '').strip()

        if action == 'check_in':
            if open_shift:
                messages.warning(request, 'You already have an open shift. Please check out first.')
            else:
                AttendanceRecord.objects.create(
                    user=request.user,
                    clock_in_at=timezone.now(),
                    note=append_attendance_note('', note, 'Clock-in note'),
                )
                messages.success(request, 'Check-in recorded.')
        elif action == 'check_out':
            if not open_shift:
                messages.warning(request, 'No open shift found for checkout.')
            else:
                open_shift.clock_out_at = timezone.now()
                open_shift.note = append_attendance_note(open_shift.note, note, 'Clock-out note')
                open_shift.save(update_fields=['clock_out_at', 'note'])
                messages.success(request, 'Check-out recorded.')
        return redirect('attendance')

    month_value = (request.GET.get('month') or '').strip()
    team_query = (request.GET.get('team_q') or '').strip()
    month_start, month_end = get_month_bounds(month_value)
    selected_month = month_start.strftime('%Y-%m')

    my_records = list(
        AttendanceRecord.objects
        .filter(user=request.user, clock_in_at__date__gte=month_start, clock_in_at__date__lte=month_end)
        .order_by('-clock_in_at', '-id')
    )
    my_total_duration = timedelta()
    for record in my_records:
        record.view_duration = record.worked_duration
        record.view_duration_label = format_duration_hours(record.view_duration)
        my_total_duration += record.view_duration

    context = {
        'open_shift': open_shift,
        'today': today,
        'selected_month': selected_month,
        'month_start': month_start,
        'month_end': month_end,
        'my_records': my_records,
        'my_shift_count': len(my_records),
        'my_total_hours_label': format_duration_hours(my_total_duration),
        'show_team_section': has_manager_access(request.user),
    }

    if has_manager_access(request.user):
        team_records_qs = (
            AttendanceRecord.objects
            .select_related('user')
            .filter(user__is_superuser=False, clock_in_at__date__gte=month_start, clock_in_at__date__lte=month_end)
            .order_by('-clock_in_at', '-id')
        )
        active_store, store_is_all = resolve_active_store(request)
        if not store_is_all and active_store is not None:
            team_records_qs = team_records_qs.filter(user__store_profile__store=active_store)
        if team_query:
            team_records_qs = team_records_qs.filter(
                Q(user__username__icontains=team_query) |
                Q(user__first_name__icontains=team_query) |
                Q(user__last_name__icontains=team_query)
            )

        team_records = list(team_records_qs[:120])
        team_summary = {}
        for record in team_records:
            record.view_duration = record.worked_duration
            record.view_duration_label = format_duration_hours(record.view_duration)
            summary = team_summary.setdefault(record.user_id, {
                'user': record.user,
                'shift_count': 0,
                'total_duration': timedelta(),
                'open_shift': False,
                'last_clock_in_at': record.clock_in_at,
            })
            summary['shift_count'] += 1
            summary['total_duration'] += record.view_duration
            if record.is_open:
                summary['open_shift'] = True
            if record.clock_in_at > summary['last_clock_in_at']:
                summary['last_clock_in_at'] = record.clock_in_at

        team_members = []
        for summary in team_summary.values():
            summary['total_hours_label'] = format_duration_hours(summary['total_duration'])
            team_members.append(summary)
        team_members.sort(key=lambda item: (not item['open_shift'], item['user'].username.lower()))

        context.update({
            'team_query': team_query,
            'team_records': team_records,
            'team_members': team_members,
            'team_open_shift_count': sum(1 for item in team_members if item['open_shift']),
        })

    return render(request, 'stock/attendance.html', context)



# -----------------------------
# 导出（任意时间范围的销售 + 采购）
# -----------------------------
@login_required
@manager_required
def export_sales_purchases_pdf(request):
    start_str = request.GET.get('start_date', '').strip()
    end_str   = request.GET.get('end_date', '').strip()
    start_date = end_date = None
    try:
        if start_str: start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
    except: start_date = None
    try:
        if end_str: end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    except: end_date = None

    if not start_date and not end_date:
        buf = BytesIO(); buf.write(b'No date selected'); buf.seek(0)
        return FileResponse(buf, as_attachment=True, filename='sales_report.pdf')

    if start_date and not end_date: end_date = start_date
    if end_date and not start_date: start_date = end_date
    if start_date > end_date: start_date, end_date = end_date, start_date

    # ====== 查询 ======
    sales_qs = (
        Sale.objects
        .filter(date__date__gte=start_date, date__date__lte=end_date)
        .select_related('order', 'product', 'customer')
        .order_by('date', 'order_id', 'id')
    )

    # ====== 分组（日期 → 订单）======
    day_map = defaultdict(lambda: {
        'orders': defaultdict(lambda: {
            'order': None,
            'items': [],
            'total_qty': 0,
            'total_amount': Decimal('0.00'),
            'pay_break': defaultdict(Decimal),
        }),
        'totals': {'qty': 0, 'amount': Decimal('0.00')},
        'pay_break': defaultdict(Decimal),
        'order_count': 0,
    })

    for s in sales_qs:
        d = s.date.date()
        o = s.order
        oid = o.id if o else 0
        pack = day_map[d]['orders'][oid]
        if pack['order'] is None:
            if o:
                o.cust_name = o.customer.name if o.customer else '-'
                o.created_hm = o.created_at.strftime('%H:%M') if o.created_at else s.date.strftime('%H:%M')
            pack['order'] = o
            day_map[d]['order_count'] += 1

        pack['items'].append(s)
        pack['total_qty'] += s.quantity
        sub = (s.unit_price or Decimal('0')) * s.quantity
        pack['total_amount'] += sub
        pack['pay_break'][(s.payment_method or '').lower()] += sub

        day_map[d]['totals']['qty']    += s.quantity
        day_map[d]['totals']['amount'] += sub
        day_map[d]['pay_break'][(s.payment_method or '').lower()] += sub

    # ====== 样式 ======
    styles  = getSampleStyleSheet()
    title   = ParagraphStyle('title', parent=styles['Title'], spaceAfter=4)
    h2      = ParagraphStyle('h2', parent=styles['Heading2'], spaceBefore=6, spaceAfter=6)
    h3      = ParagraphStyle('h3', parent=styles['Heading3'], spaceBefore=6, spaceAfter=6)
    p       = ParagraphStyle('p',  parent=styles['Normal'], fontSize=9, leading=11)
    p_bold  = ParagraphStyle('p_bold', parent=p, fontName='Helvetica-Bold')
    p_small = ParagraphStyle('p_small', parent=styles['Normal'], fontSize=8, leading=9.6)

    # Summary 专用大字样式（非表格）
    big_num = ParagraphStyle('big_num', parent=styles['Normal'], fontName='Helvetica-Bold',
                             fontSize=18, leading=22, textColor=colors.HexColor('#0f172a'))
    big_lab = ParagraphStyle('big_lab', parent=styles['Normal'], fontName='Helvetica',
                             fontSize=10, leading=12, textColor=colors.HexColor('#64748b'))
    line_val = ParagraphStyle('line_val', parent=styles['Normal'], fontName='Helvetica-Bold',
                              fontSize=12, leading=15, textColor=colors.HexColor('#0f172a'))

    def money(x: Decimal) -> str:
        x = (x or Decimal('0')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        return f"{x:.2f}"

    # ====== 文档与页眉页脚 ======
    buf = BytesIO()
    PAGE = landscape(A4)
    left, right, top, bottom = 22, 22, 50, 30
    doc = SimpleDocTemplate(
        buf, pagesize=PAGE, leftMargin=left, rightMargin=right,
        topMargin=top, bottomMargin=bottom, title='Sales Report'
    )
    usable_w = PAGE[0] - left - right

    THEAD_BG = colors.HexColor('#e2e8f0')
    THEAD_TX = colors.HexColor('#0f172a')
    GRID     = colors.HexColor('#cbd5e1')
    ALT1     = colors.white
    ALT2     = colors.HexColor('#f8fafc')

    date_filter = f"{start_date} → {end_date}" if start_date != end_date else f"{start_date}"
    gen_time = datetime.now().strftime('%Y-%m-%d %H:%M')

    def header_footer(c, d):
        c.saveState()
        c.setFont("Helvetica-Bold", 9)
        c.setFillColorRGB(0.06, 0.09, 0.16)
        c.drawString(d.leftMargin, d.height + d.topMargin - 18, "Sales Report")
        c.setFont("Helvetica", 8)
        c.drawRightString(d.pagesize[0]-d.rightMargin, d.height + d.topMargin - 18,
                          f"Range: {date_filter}   |   Generated: {gen_time}")
        c.setFont("Helvetica", 8)
        c.setFillColorRGB(0.34, 0.41, 0.5)
        c.drawRightString(d.pagesize[0]-d.rightMargin, d.bottomMargin - 12, f"Page {d.page}")
        c.restoreState()

    # ====== 组件 ======
    def order_items_table(items):
        colw = [
            0.56*usable_w,  # Product
            0.08*usable_w,  # Qty
            0.12*usable_w,  # Unit
            0.14*usable_w,  # Subtotal
            0.10*usable_w,  # Payment
        ]
        rows = [[
            Paragraph("Product", p_bold),
            Paragraph("Qty",  p_bold),
            Paragraph("Unit (€)", p_bold),
            Paragraph("Subtotal (€)", p_bold),
            Paragraph("Payment", p_bold),
        ]]
        for s in items:
            prod = s.product
            brand = getattr(prod, 'brand', '') or ''
            model = getattr(prod, 'model', '') or ''
            name  = getattr(prod, 'name', '') or ''
            line  = " - ".join([x for x in [brand, model, name] if x]) or '—'
            rows.append([
                Paragraph(build_product_label(prod) or line, p_small),
                Paragraph(str(s.quantity), p_small),
                Paragraph(money(s.unit_price or Decimal('0')), p_small),
                Paragraph(money((s.unit_price or Decimal('0')) * s.quantity), p_small),
                Paragraph((s.payment_method or '').capitalize() or '—', p_small),
            ])

        t = Table(rows, colWidths=colw, repeatRows=1, hAlign='LEFT')
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), THEAD_BG),
            ('TEXTCOLOR', (0,0), (-1,0), THEAD_TX),
            ('LINEABOVE', (0,0), (-1,0), 0.6, GRID),
            ('LINEBELOW', (0,0), (-1,0), 0.6, GRID),
            ('GRID', (0,0), (-1,-1), 0.25, GRID),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('LEADING', (0,1), (-1,-1), 10),
            ('ALIGN', (1,1), (3,-1), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [ALT1, ALT2]),
            ('LEFTPADDING', (0,0), (-1,-1), 5),
            ('RIGHTPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        return t

    def day_summary_box(day_dict):
        orders = day_dict['order_count']
        qty    = day_dict['totals']['qty']
        amt    = day_dict['totals']['amount']
        avg    = (amt/qty) if qty else Decimal('0')

        pay = day_dict['pay_break']
        cash   = pay.get('cash',   Decimal('0'))
        card   = pay.get('card',   Decimal('0'))
        mbway  = pay.get('mbway',  Decimal('0'))
        others = sum(v for k, v in pay.items() if k not in {'cash', 'card', 'mbway'})

        # 统一 8 列，整页宽度分配；数值单元用更大字号
        colw = [0.10*usable_w, 0.15*usable_w, 0.10*usable_w, 0.15*usable_w,
                0.14*usable_w, 0.10*usable_w, 0.16*usable_w, 0.10*usable_w]

        data = [
            # ── 第1行：核心指标（标签/值交替）
            [Paragraph("Total (€)", p_bold),  Paragraph(f"€ {money(amt)}", ParagraphStyle('big', parent=p_bold, fontSize=12, leading=14)),
            Paragraph("Orders", p_bold),     Paragraph(str(orders),       ParagraphStyle('big', parent=p_bold, fontSize=12, leading=14)),
            Paragraph("Items",  p_bold),     Paragraph(str(qty),          ParagraphStyle('big', parent=p_bold, fontSize=12, leading=14)),
            Paragraph("Avg Ticket (€)", p_bold), Paragraph(f"€ {money(avg)}", ParagraphStyle('big', parent=p_bold, fontSize=12, leading=14))],
            # ── 第2行：支付方式拆分
            [Paragraph("CASH (€)", p_small),  Paragraph(money(cash),  p_small),
            Paragraph("Card (€)", p_small),  Paragraph(money(card),  p_small),
            Paragraph("MBWay (€)", p_small), Paragraph(money(mbway), p_small),
            Paragraph("Other (€)", p_small), Paragraph(money(others), p_small)],
        ]

        t = Table(data, colWidths=colw, hAlign='LEFT')
        t.setStyle(TableStyle([
            # 表头行（第1行）背景更亮、边框更粗，突出 Total 区
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
            ('LINEABOVE',  (0,0), (-1,0), 0.8, GRID),
            ('LINEBELOW',  (0,0), (-1,0), 0.8, GRID),
            # 整表边框与网格
            ('BOX',        (0,0), (-1,-1), 0.6, GRID),
            ('INNERGRID',  (0,0), (-1,-1), 0.25, GRID),
            # 标签单元（偶数列 0,2,4,6）底色微弱区分
            ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#eef2ff')),
            ('BACKGROUND', (2,0), (2,-1), colors.HexColor('#eef2ff')),
            ('BACKGROUND', (4,0), (4,-1), colors.HexColor('#eef2ff')),
            ('BACKGROUND', (6,0), (6,-1), colors.HexColor('#eef2ff')),
            # 值单元（第1行）字号更大、颜色更深
            ('FONTSIZE',   (1,0), (1,0), 12),
            ('FONTSIZE',   (3,0), (3,0), 12),
            ('FONTSIZE',   (5,0), (5,0), 12),
            ('FONTSIZE',   (7,0), (7,0), 12),
            ('TEXTCOLOR',  (1,0), (7,0), colors.HexColor('#0f172a')),
            ('FONTNAME',   (1,0), (7,0), 'Helvetica-Bold'),
            # 内边距
            ('LEFTPADDING',  (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING',   (0,0), (-1,-1), 4),
            ('BOTTOMPADDING',(0,0), (-1,-1), 4),
            # 对齐：金额/数量值右对齐更易比较
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
            ('ALIGN', (3,0), (3,-1), 'RIGHT'),
            ('ALIGN', (5,0), (5,-1), 'RIGHT'),
            ('ALIGN', (7,0), (7,-1), 'RIGHT'),
            # 垂直对齐
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))

        return [t]

    def order_header_row(o, row):
        head = Table([[
            Paragraph(f"Order # {o.id if o else '-'}", p_bold),
            Paragraph(f"Time: {getattr(o, 'created_hm', '--:--')}", p),
            Paragraph(f"Customer: {getattr(o, 'cust_name', '-')}", p),
            Paragraph(f"Total: € {money(row['total_amount'])}", p_bold),
            Paragraph(" | ".join(f"{k.capitalize()} € {money(v)}" for k,v in row['pay_break'].items()) or "-", p),
        ]], colWidths=[0.17*usable_w, 0.13*usable_w, 0.34*usable_w, 0.16*usable_w, 0.20*usable_w])
        head.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#edf2f7')),
            ('BOX', (0,0), (-1,-1), 0.5, GRID),
            ('INNERGRID', (0,0), (-1,-1), 0.25, GRID),
            ('FONTNAME', (0,0), (0,0), 'Helvetica-Bold'),
            ('FONTNAME', (3,0), (3,0), 'Helvetica-Bold'),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        return head

    # ====== 生成内容 ======
    elements = [
        Paragraph("📋 Sales Report", title),
        Paragraph(f"🗓 Date Range: {date_filter}", h3),
        Spacer(1, 8),
    ]

    grand_orders = 0
    grand_qty    = 0
    grand_amt    = Decimal('0')
    pay_all      = defaultdict(Decimal)

    first_page = True
    for d in sorted(day_map.keys(), reverse=True):
        if not first_page:
            elements.append(PageBreak())
        first_page = False

        # 防止页眉重叠
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(f"📅 {d}", h2))

        day = day_map[d]
        grand_orders += day['order_count']
        grand_qty    += day['totals']['qty']
        grand_amt    += day['totals']['amount']
        for k,v in day['pay_break'].items():
            pay_all[k] += v

        # 日期摘要
        elements += day_summary_box(day)
        elements.append(Spacer(1, 6))

        # 订单（id 倒序）
        for oid, row in sorted(day['orders'].items(), key=lambda kv: kv[0], reverse=True):
            o = row['order']
            head = order_header_row(o, row)
            items_tbl = order_items_table(row['items'])
            elements.append(KeepTogether([head, Spacer(1, 3), items_tbl, Spacer(1, 8)]))

    # ====== 最后一页：总体 Summary（非表格，突出显示） ======
    elements.append(PageBreak())
    elements.append(Paragraph("Overall Summary", h2))
    elements.append(HRFlowable(width="100%", thickness=0.6, color=colors.HexColor('#cbd5e1')))
    elements.append(Spacer(1, 8))

    avg_ticket = (grand_amt / grand_qty) if grand_qty else Decimal('0')
    cash  = pay_all.get('cash',  Decimal('0'))
    card  = pay_all.get('card',  Decimal('0'))
    mbway = pay_all.get('mbway', Decimal('0'))
    others = [(k, v) for k, v in pay_all.items() if k not in {'cash','card','mbway'}]

    # 关键指标（大号）
    elements += [
        Paragraph("Total Amount (€)", big_lab),
        Paragraph(f"€ {money(grand_amt)}", big_num),
        Spacer(1, 6),
        Paragraph("Orders", big_lab),
        Paragraph(f"{grand_orders}", big_num),
        Spacer(1, 6),
        Paragraph("Items", big_lab),
        Paragraph(f"{grand_qty}", big_num),
        Spacer(1, 10),
        Paragraph("Average Ticket (€)", big_lab),
        Paragraph(f"€ {money(avg_ticket)}", line_val),
        Spacer(1, 14),
        HRFlowable(width="100%", thickness=0.6, color=colors.HexColor('#e2e8f0')),
        Spacer(1, 10),
        Paragraph("Payments Breakdown", big_lab),
        Paragraph(f"CASH: € {money(cash)}", line_val),
        Paragraph(f"Card: € {money(card)}", line_val),
        Paragraph(f"MBWay: € {money(mbway)}", line_val),
    ]
    if others:
        other_line = " · ".join(f"{k.capitalize()}: € {money(v)}" for k,v in sorted(others))
        elements.append(Paragraph(other_line, line_val))

    # 导出
    doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)
    buf.seek(0)
    fname = f"Sales_{start_date}_to_{end_date}.pdf" if start_date != end_date else f"Sales_{start_date}.pdf"
    return FileResponse(buf, as_attachment=True, filename=fname)


@login_required
def set_active_store(request):
    """Managers/admins switch the active store (or 'All stores'). Employees are
    locked to their home store and this is a no-op for them."""
    if request.method == 'POST' and can_switch_store(request.user):
        value = (request.POST.get('store') or '').strip()
        if value == ALL_STORES:
            request.session[ACTIVE_STORE_SESSION_KEY] = ALL_STORES
        elif value.isdigit() and Store.objects.filter(id=int(value), is_active=True).exists():
            request.session[ACTIVE_STORE_SESSION_KEY] = int(value)
    return redirect(request.META.get('HTTP_REFERER') or 'dashboard')


@login_required
@admin_required
def store_list_view(request):
    stores = (
        Store.objects
        .annotate(
            staff_count=Count('staff', distinct=True),
            order_count=Count('sale_orders', distinct=True),
        )
        .order_by('-is_default', 'name')
    )
    return render(request, 'stock/store_list.html', {
        'stores': stores,
        'store_count': stores.count(),
    })


def _store_form_view(request, store):
    if request.method == 'POST':
        form = StoreForm(request.POST, instance=store)
        if form.is_valid():
            saved = form.save()
            # Keep exactly one default store at all times.
            if saved.is_default:
                Store.objects.exclude(pk=saved.pk).update(is_default=False)
            if not Store.objects.filter(is_default=True).exists():
                Store.objects.filter(pk=saved.pk).update(is_default=True)
            messages.success(request, f'Store "{saved.name}" saved.')
            return redirect('store_list')
    else:
        form = StoreForm(instance=store)

    return render(request, 'stock/store_form.html', {
        'form': form,
        'store': store,
        'is_create': store is None,
    })


@login_required
@admin_required
def store_create_view(request):
    return _store_form_view(request, None)


@login_required
@admin_required
def store_edit_view(request, store_id):
    store = get_object_or_404(Store, id=store_id)
    return _store_form_view(request, store)


@login_required
@admin_required
@require_POST
def store_delete_view(request, store_id):
    store = get_object_or_404(Store, id=store_id)
    if store.is_default:
        messages.error(request, 'The default store cannot be deleted. Set another store as default first.')
    elif store.sale_orders.exists() or store.sales.exists() or store.staff.exists() or store.ar_invoices.exists():
        messages.error(request, 'This store has sales, invoices or staff. Deactivate it instead of deleting.')
    else:
        name = store.name
        store.delete()
        messages.success(request, f'Store "{name}" deleted.')
    return redirect('store_list')


@login_required
@manager_required
def daily_summary_view(request):
    """A focused, store-scoped end-of-day summary: today's orders, KPIs, payment
    mix, top products and stock received. Defaults to today; ?date= views a day."""
    show_profit = bool(getattr(request.user, 'is_superuser', False))
    show_order_financials = has_order_reconciliation_access(request.user)
    active_store, store_is_all = resolve_active_store(request)
    payment_labels = dict(Sale.PAYMENT_METHOD_CHOICES)

    today = timezone.localdate()
    day = parse_date((request.GET.get('date') or '').strip()) or today
    if day > today:
        day = today

    orders_qs = scope_sales_by_store(
        SaleOrder.objects
        .filter(created_at__date=day)
        .select_related('customer', 'store')
        .prefetch_related('items__product', 'items__product__images', 'payments'),
        active_store, store_is_all,
    ).order_by('-created_at', '-id')  # most recent sale first

    profit_map = {}
    if show_profit:
        sale_ids_qs = scope_sales_by_store(
            Sale.objects.filter(order__created_at__date=day), active_store, store_is_all,
        )
        profit_map = sale_profit_map_for_sale_ids(sale_ids_qs.values_list('id', flat=True))

    total_amount = Decimal('0.00')
    total_qty = 0
    total_profit = Decimal('0.00')
    pay_totals = defaultdict(Decimal)
    product_stats = {}
    orders = []

    for order in orders_qs:
        items = list(order.items.all())
        order_total = Decimal('0.00')
        order_qty = 0
        order_profit = Decimal('0.00')

        for item in items:
            item.display_name = build_product_label(item.product)
            item.image_url = get_product_image_url(item.product)
            item.line_total = (item.unit_price or Decimal('0.00')) * item.quantity
            item.payment_label = payment_labels.get(item.payment_method, (item.payment_method or 'Other').title())
            item.line_profit = profit_map.get(item.id, {}).get('profit', Decimal('0.00')) if show_profit else Decimal('0.00')
            order_total += item.line_total
            order_qty += item.quantity
            order_profit += item.line_profit

            stats = product_stats.setdefault(item.product_id, {
                'product': item.product, 'name': item.display_name,
                'qty': 0, 'revenue': Decimal('0.00'),
            })
            stats['qty'] += item.quantity
            stats['revenue'] += item.line_total

        # Payment mix from the authoritative order-level tender (SaleOrderPayment)
        # so split payments (card + cash) are shown, not just the primary method.
        order_pay = order_tender_amounts(order, order_total, items)

        order.summary_total = order_total
        order.summary_qty = order_qty
        order.summary_profit = order_profit
        order.summary_customer = order.customer.name if order.customer_id else 'Walk-in'
        order.summary_store = order.store.name if order.store_id else '—'
        order.summary_time = timezone.localtime(order.created_at).strftime('%H:%M')
        order.summary_pay = [
            {'label': payment_labels.get(m, (m or 'Other').title()), 'amount': a}
            for m, a in sorted(order_pay.items(), key=lambda kv: (-kv[1], kv[0] or '')) if m
        ]
        order.summary_items = items
        order.detail_url = reverse('sale_order_detail', args=[order.id])
        orders.append(order)

        total_amount += order_total
        total_qty += order_qty
        total_profit += order_profit
        for method, amount in order_pay.items():
            if method:
                pay_totals[method] += amount

    order_count = len(orders)
    avg_order = (total_amount / order_count) if order_count else Decimal('0.00')

    pay_total_sum = sum(pay_totals.values())
    payment_rows = [
        {
            'label': payment_labels.get(method, (method or 'Other').title()),
            'amount': amount,
            'pct': round(float(amount) / float(pay_total_sum) * 100, 1) if pay_total_sum else 0,
        }
        for method, amount in sorted(pay_totals.items(), key=lambda kv: (-kv[1], kv[0] or ''))
    ]
    payment_chart = [{'label': row['label'], 'amount': float(row['amount'])} for row in payment_rows]

    stats_list = sorted(product_stats.values(), key=lambda s: (s['revenue'], s['qty']), reverse=True)
    max_revenue = max((s['revenue'] for s in stats_list), default=Decimal('0.00'))
    top_products = [
        {
            'product': s['product'], 'name': s['name'], 'qty': s['qty'], 'revenue': s['revenue'],
            'bar': round(float(s['revenue']) / float(max_revenue) * 100, 1) if max_revenue else 0,
        }
        for s in stats_list[:8]
    ]

    # Stock received that day (inventory is shared across stores).
    received_units = 0
    received_cost = Decimal('0.00')
    received_orders = 0
    for inbound in InboundOrder.objects.filter(created_at__date=day).prefetch_related('items'):
        received_orders += 1
        for item in inbound.items.all():
            received_units += item.quantity
            received_cost += (item.cost_price or Decimal('0.00')) * item.quantity
    for purchase in Purchase.objects.filter(inbound_order__isnull=True, date__date=day):
        received_units += purchase.quantity
        received_cost += (purchase.cost_price or Decimal('0.00')) * purchase.quantity

    next_day = day + timedelta(days=1)
    return render(request, 'stock/daily_summary.html', {
        'day': day,
        'is_today': day == today,
        'prev_day_url': f"{reverse('daily_summary')}?date={(day - timedelta(days=1)).isoformat()}",
        'next_day_url': f"{reverse('daily_summary')}?date={next_day.isoformat()}" if next_day <= today else '',
        'today_url': reverse('daily_summary'),
        'full_records_url': f"{reverse('sales_records')}?start_date={day.isoformat()}&end_date={day.isoformat()}",
        'orders': orders,
        'order_count': order_count,
        'total_amount': total_amount,
        'total_qty': total_qty,
        'total_profit': total_profit,
        'avg_order': avg_order,
        'payment_rows': payment_rows,
        'payment_chart': payment_chart,
        'top_products': top_products,
        'received_orders': received_orders,
        'received_units': received_units,
        'received_cost': received_cost,
        'show_profit': show_profit,
        'show_order_financials': show_order_financials,
        'store_is_all': store_is_all,
    })


def _employee_sales_day_view(request):
    """Employee Sales: a single day's orders, amounts only, no charts."""
    day = timezone.localdate()
    day_str = (request.GET.get('date') or '').strip()
    if day_str:
        try:
            day = datetime.strptime(day_str, '%Y-%m-%d').date()
        except ValueError:
            day = timezone.localdate()

    active_store, store_is_all = resolve_active_store(request)

    order_str = (request.GET.get('order') or '').strip()
    if order_str:
        oq = SaleOrder.objects.filter(id=order_str) if order_str.isdigit() else SaleOrder.objects.none()
        found = scope_sales_by_store(oq, active_store, store_is_all).first()
        if found:
            return redirect('sale_order_detail', order_id=found.id)
        messages.warning(request, f'Order #{order_str} not found.')

    orders_qs = (
        SaleOrder.objects
        .filter(created_at__date=day)
        .select_related('customer')
        .prefetch_related('items', 'payments')
        .order_by('-created_at', '-id')
    )
    # Reuse the shared store scoping (unfiltered when store is None / "all stores").
    orders_qs = scope_sales_by_store(orders_qs, active_store, store_is_all)

    payment_labels = {'cash': 'Cash', 'card': 'Card', 'mbway': 'MBWay'}
    orders = []
    day_total = Decimal('0.00')
    for order in orders_qs:
        items = list(order.items.all())
        total = sum((i.quantity * (i.unit_price or Decimal('0.00')) for i in items), Decimal('0.00'))
        qty = sum(i.quantity for i in items)
        methods = [payment_labels.get(p.method, (p.method or '').title()) for p in order.payments.all()]
        orders.append({
            'order_id': order.id,
            'created_hhmm': timezone.localtime(order.created_at).strftime('%H:%M'),
            'customer_name': order.customer.name if order.customer_id else 'Walk-in / No customer',
            'item_count': qty,
            'payment_label': ', '.join(dict.fromkeys(methods)) or '-',
            'total_amount': total,
            'items': [
                {
                    'name': build_product_label(i.product),
                    'qty': i.quantity,
                    'unit_price': i.unit_price or Decimal('0.00'),
                    'line_total': (i.unit_price or Decimal('0.00')) * i.quantity,
                }
                for i in items
            ],
        })
        day_total += total

    return render(request, 'stock/sales_records_employee.html', {
        'day': day,
        'date_value': day.strftime('%Y-%m-%d'),
        'orders': orders,
        'order_count': len(orders),
        'day_total': day_total,
    })


@login_required
def record_view(request):
    if not has_manager_access(request.user):
        return _employee_sales_day_view(request)
    start_str = request.GET.get('start_date', '').strip()
    end_str = request.GET.get('end_date', '').strip()
    view_mode = request.GET.get('view', '').strip()
    month_str = request.GET.get('month', '').strip()
    product_q = request.GET.get('product_q', '').strip()
    show_sensitive = has_manager_access(request.user)
    show_sales_sensitive = has_sales_sensitive_access(request.user)
    show_order_financials = has_order_reconciliation_access(request.user)
    show_profit = bool(getattr(request.user, 'is_superuser', False))

    empty_context = {
        'start_date': '',
        'end_date': '',
        'date_filter': None,
        'day_blocks': [],
        'purchase_day_blocks': [],
        'total_sales_amount': Decimal('0.00'),
        'total_sales_qty': 0,
        'total_sales_orders': 0,
        'avg_order_amount': Decimal('0.00'),
        'payment_totals': {},
        'total_sales_profit': Decimal('0.00'),
        'total_purchase_amount': Decimal('0.00'),
        'total_purchase_qty': 0,
        'total_purchase_orders': 0,
        'show_purchases': False,
        'show_sales_sensitive': show_sales_sensitive,
        'show_order_financials': show_order_financials,
        'show_profit': show_profit,
    }

    start_date = end_date = None
    if start_str:
        try:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        except Exception:
            start_date = None
    if end_str:
        try:
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
        except Exception:
            end_date = None

    if view_mode == 'year':
        # Yearly trend overview (drill into a month calendar for detail).
        view_today = timezone.localdate()
        year = resolve_year(request.GET.get('year'), view_today)
        selected_cat_ids_int = [int(x) for x in request.GET.getlist('cat') if str(x).isdigit()]
        active_store, store_is_all = resolve_active_store(request)
        overview = build_yearly_sales_overview(
            year, view_today,
            selected_category_ids=selected_cat_ids_int,
            show_profit=show_profit,
            store=None if store_is_all else active_store,
        )

        base_url = reverse('sales_records')

        def _records_url(extra):
            params = list(extra.items()) + [('cat', str(cid)) for cid in selected_cat_ids_int]
            return f"{base_url}?{urlencode(params)}"

        year_chart = []
        for row in overview['monthly_rows']:
            month = row['month']
            # Drill into the month's calendar view (not a flat range list).
            drill_url = _records_url({'month': f'{year:04d}-{month:02d}'})
            row['detail_url'] = drill_url
            year_chart.append({'label': row['label'], 'amount': float(row['amount']), 'url': drill_url})

        year_nav = {
            'prev': _records_url({'view': 'year', 'year': overview['prev_year']}),
            'next': _records_url({'view': 'year', 'year': overview['next_year']}) if overview['next_year'] else '',
            'current': _records_url({'view': 'year', 'year': view_today.year}),
        }

        return render(request, 'stock/sales_records.html', {
            **empty_context,
            'show_trend': True,
            'year_nav': year_nav,
            'year_chart': year_chart,
            **overview,
        })

    calendar_mode = False
    cal_year = cal_month = None
    if not start_date and not end_date:
        # Default landing: this month's sales calendar (drill a day open in a modal).
        calendar_mode = True
        today_local = timezone.localdate()
        cal_year, cal_month = today_local.year, today_local.month
        if month_str:
            try:
                yy, mm = month_str.split('-')[:2]
                cal_year, cal_month = int(yy), int(mm)
                date(cal_year, cal_month, 1)
            except Exception:
                cal_year, cal_month = today_local.year, today_local.month
        start_date = date(cal_year, cal_month, 1)
        end_date = date(cal_year, cal_month, monthrange(cal_year, cal_month)[1])

    if start_date and not end_date:
        end_date = start_date
    if end_date and not start_date:
        start_date = end_date
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    payment_labels = dict(Sale.PAYMENT_METHOD_CHOICES)
    next_url = request.get_full_path()

    def build_product_label(product):
        return globals()['build_product_label'](product)

    def get_product_image_url(product):
        if not product:
            return ''
        for image in product.images.all():
            image_file = getattr(image, 'image', None)
            if not image_file:
                continue
            try:
                return image_file.url
            except ValueError:
                continue
        return ''

    def annotate_purchase_items(items):
        annotated_items = list(items)
        for item in annotated_items:
            item.display_name = build_product_label(item.product)
            item.line_total = (item.cost_price or Decimal('0.00')) * item.quantity
            item.image_url = get_product_image_url(item.product)
        return annotated_items

    def build_payment_breakdown(source_breakdown):
        return [
            {
                'code': method,
                'label': payment_labels.get(method, (method or 'Other').title()),
                'amount': amount,
            }
            for method, amount in sorted(source_breakdown.items(), key=lambda item: (-item[1], item[0] or ''))
            if method
        ]

    sales_qs = (
        Sale.objects
        .annotate(business_at=Coalesce('order__created_at', 'date'))
        .filter(
            Q(order__created_at__date__gte=start_date, order__created_at__date__lte=end_date) |
            Q(order__isnull=True, date__date__gte=start_date, date__date__lte=end_date)
        )
        .select_related('order__customer', 'order__store', 'product', 'customer')
        .prefetch_related('product__images')
        .order_by('-business_at', '-order_id', '-id')
    )
    if product_q:
        sales_qs = sales_qs.filter(
            Q(product__name__icontains=product_q)
            | Q(product__barcode__icontains=product_q)
            | Q(product__model__icontains=product_q)
            | Q(product__brand__icontains=product_q)
        )
    active_store, store_is_all = resolve_active_store(request)
    sales_qs = scope_sales_by_store(sales_qs, active_store, store_is_all)
    sale_profit_map = sale_profit_map_for_sale_ids(sales_qs.values_list('id', flat=True)) if show_profit else {}

    day_totals_map = defaultdict(lambda: {
        'totals': {'qty': 0, 'amount': Decimal('0.00'), 'profit': Decimal('0.00')},
        'pay_break': defaultdict(Decimal),
    })
    grouped_orders = defaultdict(dict)
    product_stats = {}

    for sale in sales_qs:
        business_at = sale.business_at or sale.date
        local_business_at = timezone.localtime(business_at)
        day = local_business_at.date()
        order = sale.order
        order_key = order.id if order else f'legacy-{sale.id}'
        row = grouped_orders[day].get(order_key)

        if row is None:
            customer_obj = order.customer if order and order.customer_id else sale.customer
            created_at = timezone.localtime(order.created_at) if order and order.created_at else local_business_at
            row = {
                'order': order,
                'order_id': order.id if order else None,
                'detail_url': reverse('sale_order_detail', args=[order.id]) if order else None,
                'customer_name': customer_obj.name if customer_obj else 'Walk-in / No customer',
                'customer_detail_url': reverse('customer_detail', args=[customer_obj.id]) if customer_obj else None,
                'store_name': order.store.name if order and order.store_id else '',
                'created_at': created_at,
                'created_hhmm': created_at.strftime('%H:%M'),
                'items_today': [],
                'total_qty': 0,
                'total_amount': Decimal('0.00'),
                'total_profit': Decimal('0.00'),
                'pay_break': defaultdict(Decimal),
            }
            grouped_orders[day][order_key] = row

        sale.display_name = build_product_label(sale.product)
        sale.line_total = (sale.unit_price or Decimal('0.00')) * sale.quantity
        sale.payment_label = payment_labels.get(sale.payment_method, (sale.payment_method or 'Other').title())
        sale.image_url = get_product_image_url(sale.product)
        sale.line_cost = sale_profit_map.get(sale.id, {}).get('cost', Decimal('0.00')) if show_profit else Decimal('0.00')
        sale.line_profit = sale_profit_map.get(sale.id, {}).get('profit', Decimal('0.00')) if show_profit else Decimal('0.00')

        row['items_today'].append(sale)
        row['total_qty'] += sale.quantity
        row['total_amount'] += sale.line_total
        row['total_profit'] += sale.line_profit
        row['pay_break'][sale.payment_method] += sale.line_total

        day_totals_map[day]['totals']['qty'] += sale.quantity
        day_totals_map[day]['totals']['amount'] += sale.line_total
        day_totals_map[day]['totals']['profit'] += sale.line_profit
        day_totals_map[day]['pay_break'][sale.payment_method] += sale.line_total

        stats = product_stats.get(sale.product_id)
        if stats is None:
            stats = product_stats[sale.product_id] = {
                'product': sale.product,
                'name': sale.display_name,
                'qty': 0,
                'revenue': Decimal('0.00'),
                'profit': Decimal('0.00'),
            }
        stats['qty'] += sale.quantity
        stats['revenue'] += sale.line_total
        stats['profit'] += sale.line_profit

    day_blocks = []
    for day in sorted(grouped_orders.keys(), reverse=True):
        orders = []
        for row in sorted(
            grouped_orders[day].values(),
            key=lambda item: (item['created_at'], item['order_id'] or 0),
            reverse=True
        ):
            row['items_today'].sort(
                key=lambda item: (
                    item.display_name.lower(),
                    item.payment_method or '',
                    item.id,
                )
            )
            row['pay_break_display'] = build_payment_breakdown(row['pay_break'])

            row['detail_id'] = f"sales-{day.strftime('%Y%m%d')}-{row['order_id'] or row['items_today'][0].id}"
            orders.append(row)

        totals = day_totals_map[day]['totals']
        order_count = len(orders)
        day_blocks.append({
            'date': day,
            'cal_id': f"cal-{day.strftime('%Y%m%d')}",
            'orders': orders,
            'totals': totals,
            'order_count': order_count,
            'avg_order_amount': (totals['amount'] / order_count) if order_count else Decimal('0.00'),
            'pay_break': dict(day_totals_map[day]['pay_break']),
            'pay_break_display': build_payment_breakdown(day_totals_map[day]['pay_break']),
        })

    total_sales_amount = sum(block['totals']['amount'] for block in day_blocks) if day_blocks else Decimal('0.00')
    total_sales_qty = sum(block['totals']['qty'] for block in day_blocks) if day_blocks else 0
    total_sales_orders = sum(block['order_count'] for block in day_blocks) if day_blocks else 0
    total_sales_profit = sum(block['totals']['profit'] for block in day_blocks) if day_blocks else Decimal('0.00')
    avg_order_amount = (total_sales_amount / total_sales_orders) if total_sales_orders else Decimal('0.00')
    date_filter = f'{start_date} to {end_date}' if start_date != end_date else f'{start_date}'

    payment_totals = defaultdict(Decimal)
    for block in day_blocks:
        for method, amount in block['pay_break'].items():
            if method:
                payment_totals[method.lower()] += amount

    inbound_orders = (
        InboundOrder.objects
        .filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
        .select_related('supplier')
        .prefetch_related('items__product', 'items__product__images', 'items__supplier')
        .order_by('-created_at')
    )

    purchase_map = defaultdict(lambda: {
        'purchases': [],
        'totals': {'qty': 0, 'amount': Decimal('0.00')},
    })

    for order in inbound_orders:
        day = order.created_at.date()
        items = annotate_purchase_items(order.items.all())

        total_qty = sum(item.quantity for item in items)
        total_cost = sum((item.line_total for item in items), Decimal('0.00'))
        effective_supplier = order.supplier or next((item.supplier for item in items if item.supplier_id), None)

        purchase_map[day]['purchases'].append({
            'purchase': order,
            'items': items,
            'supplier_name': effective_supplier.name if effective_supplier else 'No supplier linked',
            'has_supplier': bool(effective_supplier),
            'identity_label': 'Supplier' if effective_supplier else 'Supplier Missing',
            'header_badge': f'Inbound #{order.id}',
            'created_at': order.created_at,
            'total_qty': total_qty,
            'total_cost': total_cost,
            'detail_id': f"purchase-{day.strftime('%Y%m%d')}-{order.id}",
            'edit_url': f"{reverse('inbound_order_edit', args=[order.id])}?{urlencode({'next': next_url})}",
        })

        purchase_map[day]['totals']['qty'] += total_qty
        purchase_map[day]['totals']['amount'] += total_cost

    standalone_purchases = (
        Purchase.objects
        .filter(inbound_order__isnull=True, date__date__gte=start_date, date__date__lte=end_date)
        .select_related('product', 'supplier')
        .prefetch_related('product__images')
        .order_by('-date', '-id')
    )

    standalone_by_day = defaultdict(list)
    for purchase in standalone_purchases:
        day = purchase.date.date()
        standalone_by_day[day].append(purchase)

    for day, purchase_items in standalone_by_day.items():
        items = annotate_purchase_items(purchase_items)
        for item in items:
            item.edit_url = f"{reverse('direct_purchase_edit', args=[item.id])}?{urlencode({'next': next_url})}"
        total_qty = sum(item.quantity for item in items)
        total_cost = sum((item.line_total for item in items), Decimal('0.00'))
        supplier_names = sorted({item.supplier.name for item in items if item.supplier_id})
        created_at = max(item.date for item in items)

        if len(supplier_names) == 1:
            supplier_name = supplier_names[0]
            identity_label = 'Supplier'
            has_supplier = True
        elif len(supplier_names) > 1:
            supplier_name = 'Mixed suppliers'
            identity_label = 'Purchases'
            has_supplier = True
        else:
            supplier_name = 'No supplier linked'
            identity_label = 'Purchases'
            has_supplier = False

        purchase_map[day]['purchases'].append({
            'purchase': None,
            'items': items,
            'supplier_name': supplier_name,
            'has_supplier': has_supplier,
            'identity_label': identity_label,
            'header_badge': 'Purchases',
            'created_at': created_at,
            'total_qty': total_qty,
            'total_cost': total_cost,
            'detail_id': f"purchase-standalone-{day.strftime('%Y%m%d')}",
        })

        purchase_map[day]['totals']['qty'] += total_qty
        purchase_map[day]['totals']['amount'] += total_cost

    purchase_day_blocks = []
    for day in sorted(purchase_map.keys(), reverse=True):
        rows = sorted(
            purchase_map[day]['purchases'],
            key=lambda item: item['created_at'],
            reverse=True,
        )
        purchase_day_blocks.append({
            'date': day,
            'purchases': rows,
            'totals': purchase_map[day]['totals'],
            'purchase_count': len(rows),
        })

    total_purchase_amount = sum(block['totals']['amount'] for block in purchase_day_blocks) if purchase_day_blocks else Decimal('0.00')
    total_purchase_qty = sum(block['totals']['qty'] for block in purchase_day_blocks) if purchase_day_blocks else 0
    total_purchase_orders = sum(block['purchase_count'] for block in purchase_day_blocks) if purchase_day_blocks else 0

    # Daily trend (ascending) merging sales money-in, purchases money-out, and profit.
    sales_by_day = {block['date']: block['totals']['amount'] for block in day_blocks}
    profit_by_day = {block['date']: block['totals']['profit'] for block in day_blocks}
    purchases_by_day = {block['date']: block['totals']['amount'] for block in purchase_day_blocks}
    trend_data = [
        {
            'label': day.strftime('%m-%d'),
            'sales': float(sales_by_day.get(day, 0) or 0),
            'purchases': float(purchases_by_day.get(day, 0) or 0),
            'profit': float(profit_by_day.get(day, 0) or 0),
        }
        for day in sorted(set(sales_by_day) | set(purchases_by_day))
    ]

    # Payment-method split for the donut (floats; gated like the amounts).
    payment_chart = [
        {'label': payment_labels.get(method, (method or 'Other').title()), 'amount': float(amount)}
        for method, amount in sorted(payment_totals.items(), key=lambda kv: (-kv[1], kv[0] or ''))
        if method
    ]

    # Best-selling products across the whole range.
    stats_list = sorted(product_stats.values(), key=lambda s: (s['revenue'], s['qty']), reverse=True)
    max_revenue = max((s['revenue'] for s in stats_list), default=Decimal('0.00'))
    top_products = [
        {
            'product': s['product'],
            'name': s['name'],
            'qty': s['qty'],
            'revenue': s['revenue'],
            'profit': s['profit'],
            'bar': round(float(s['revenue']) / float(max_revenue) * 100, 1) if max_revenue else 0,
        }
        for s in stats_list[:8]
    ]

    # Month-calendar presentation (default landing): a grid of day cells with per-day
    # order count + amount; each active day opens a modal (rendered from day_blocks).
    calendar_context = {'calendar_mode': calendar_mode}
    if calendar_mode:
        by_date = {block['date']: block for block in day_blocks}
        today_local = timezone.localdate()
        weeks = []
        for week in Calendar(firstweekday=0).monthdatescalendar(cal_year, cal_month):
            cells = []
            for cell_date in week:
                block = by_date.get(cell_date)
                cells.append({
                    'day': cell_date.day,
                    'in_month': cell_date.month == cal_month,
                    'is_today': cell_date == today_local,
                    'has_sales': bool(block),
                    'order_count': block['order_count'] if block else 0,
                    'qty': block['totals']['qty'] if block else 0,
                    'amount': block['totals']['amount'] if block else Decimal('0.00'),
                    'cal_id': f"cal-{cell_date.strftime('%Y%m%d')}",
                })
            weeks.append(cells)

        prev_month = (start_date - timedelta(days=1)).replace(day=1)
        next_month_first = (end_date + timedelta(days=1))

        def _cal_url(**extra):
            params = {}
            if product_q:
                params['product_q'] = product_q
            params.update(extra)
            return f"{reverse('sales_records')}?{urlencode(params)}" if params else reverse('sales_records')

        calendar_context.update({
            'calendar_weeks': weeks,
            'calendar_month_label': start_date.strftime('%B %Y'),
            'calendar_weekdays': ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
            'prev_month_url': _cal_url(month=prev_month.strftime('%Y-%m')),
            'next_month_url': _cal_url(month=next_month_first.strftime('%Y-%m')),
            'this_month_url': _cal_url(),
            'clear_product_url': f"{reverse('sales_records')}?{urlencode({'month': start_date.strftime('%Y-%m')})}",
            'year_view_url': f"{reverse('sales_records')}?view=year",
            'product_q': product_q,
            'calendar_month_value': start_date.strftime('%Y-%m'),
        })

    return render(request, 'stock/sales_records.html', {
        **calendar_context,
        'trend_data': trend_data,
        'payment_chart': payment_chart,
        'top_products': top_products,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'date_filter': date_filter,
        'day_blocks': day_blocks,
        'purchase_day_blocks': purchase_day_blocks,
        'total_sales_amount': total_sales_amount,
        'total_sales_qty': total_sales_qty,
        'total_sales_orders': total_sales_orders,
        'avg_order_amount': avg_order_amount,
        'payment_totals': dict(payment_totals),
        'total_sales_profit': total_sales_profit,
        'total_purchase_amount': total_purchase_amount,
        'total_purchase_qty': total_purchase_qty,
        'total_purchase_orders': total_purchase_orders,
        'show_purchases': show_sensitive,
        'show_sales_sensitive': show_sales_sensitive,
        'show_order_financials': show_order_financials,
        'show_profit': show_profit,
    })


# -----------------------------
# 其他：图片删除、客户相关、自动完成、目录页
# -----------------------------
@require_POST
@login_required
@manager_required
def delete_product_image(request, image_id):
    image = get_object_or_404(ProductImage, id=image_id)
    product_id = image.product.id
    image.delete()

    # 若是 AJAX 请求，返回 JSON；否则跳转回编辑页
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    return redirect('edit_product', pk=product_id)


@login_required
def check_customer(request):
    query = request.GET.get('q', '').strip()
    if not query:
        return JsonResponse({'exists': False})

    customer = Customer.objects.filter(
        Q(nif__icontains=query) | Q(name__icontains=query)
    ).first()

    if customer:
        return JsonResponse({
            'exists': True,
            'id': customer.id,
            'name': customer.name,
            'phone': customer.phone,
            'email': customer.email,
        })

    return JsonResponse({'exists': False})


@login_required
def customers_autocomplete(request):
    query = request.GET.get('q', '').strip()
    results = []

    if query:
        customers = Customer.objects.filter(
            Q(name__icontains=query) | Q(nif__icontains=query)
        ).order_by('name')[:10]

        results = [{'id': c.id, 'name': c.name, 'nif': c.nif} for c in customers]

    return JsonResponse({'results': results})


@require_POST
@login_required
def add_customer(request):
    nif = (request.POST.get('nif') or '').strip()
    name = (request.POST.get('name') or '').strip()
    phone = (request.POST.get('phone') or '').strip() or None
    email = (request.POST.get('email') or '').strip() or None
    notes = (request.POST.get('notes') or '').strip() or None

    # 基本必填校验
    if not nif or not name:
        return JsonResponse({'success': False, 'error': 'NIF and name are required.'})

    # NIF 必须为 9 位数字
    if not nif.isdigit() or len(nif) != 9:
        return JsonResponse({'success': False, 'error': 'NIF must be exactly 9 digits.'})

    # 检查重复
    if Customer.objects.filter(nif=nif).exists():
        return JsonResponse({'success': False, 'error': 'Customer with this NIF already exists.'})

    # 创建客户
    customer = Customer.objects.create(
        nif=nif, name=name, phone=phone, email=email, notes=notes
    )

    return JsonResponse({
        'success': True,
        'id': customer.id,
        'name': customer.name,
        'phone': customer.phone,
        'email': customer.email
    })

def _employee_customer_search_view(request):
    """Employee Customers: search + add only; no full list, no history."""
    query = (request.GET.get('q') or '').strip()
    customers = []
    if query:
        customers = list(
            Customer.objects.filter(
                Q(name__icontains=query)
                | Q(phone__icontains=query)
                | Q(email__icontains=query)
                | Q(nif__icontains=query)
            ).order_by('name')[:50]
        )
    return render(request, 'stock/customer_search_employee.html', {
        'query': query,
        'customers': customers,
    })


@login_required
def customer_search_view(request):
    if not has_manager_access(request.user):
        return _employee_customer_search_view(request)
    query = request.GET.get('q', '').strip()
    show_sensitive = has_manager_access(request.user)
    show_sales_sensitive = has_sales_sensitive_access(request.user)
    money_field = DecimalField(max_digits=12, decimal_places=2)
    recent_cutoff = timezone.now() - timedelta(days=60)

    active_store, store_is_all = resolve_active_store(request)
    store_q = Q() if (store_is_all or active_store is None) else Q(store=active_store)

    order_count_subquery = (
        SaleOrder.objects
        .filter(store_q, customer=OuterRef('pk'))
        .values('customer')
        .annotate(total=Count('id'))
        .values('total')[:1]
    )
    spent_subquery = (
        Sale.objects
        .filter(store_q, order__customer=OuterRef('pk'))
        .values('order__customer')
        .annotate(
            total=Sum(
                ExpressionWrapper(
                    F('unit_price') * F('quantity'),
                    output_field=money_field,
                )
            )
        )
        .values('total')[:1]
    )
    balance_subquery = (
        ARInvoice.objects
        .filter(store_q, customer=OuterRef('pk'))
        .values('customer')
        .annotate(
            total=Sum(
                ExpressionWrapper(
                    F('total_amount') - F('amount_paid'),
                    output_field=money_field,
                )
            )
        )
        .values('total')[:1]
    )
    last_order_subquery = (
        SaleOrder.objects
        .filter(store_q, customer=OuterRef('pk'))
        .order_by('-created_at')
        .values('created_at')[:1]
    )

    customers_qs = (
        Customer.objects
        .annotate(
            total_orders=Coalesce(Subquery(order_count_subquery, output_field=IntegerField()), Value(0)),
            total_spent=Coalesce(Subquery(spent_subquery, output_field=money_field), Value(Decimal('0.00'))),
            balance_due=Coalesce(Subquery(balance_subquery, output_field=money_field), Value(Decimal('0.00'))),
            last_order_at=Subquery(last_order_subquery, output_field=DateTimeField()),
        )
    )

    if query:
        customers_qs = customers_qs.filter(
            Q(nif__icontains=query)
            | Q(name__icontains=query)
            | Q(phone__icontains=query)
            | Q(email__icontains=query)
        )

    # Under a specific store, show only customers who have bought there.
    if not store_is_all and active_store is not None:
        customers_qs = customers_qs.filter(total_orders__gt=0)

    customers_qs = customers_qs.order_by('-last_order_at', '-total_orders', 'name')

    customer_summary = customers_qs.aggregate(
        customer_count=Count('id'),
        active_count=Count('id', filter=Q(last_order_at__gte=recent_cutoff)),
        with_orders_count=Count('id', filter=Q(total_orders__gt=0)),
        open_balance_count=Count('id', filter=Q(balance_due__gt=0)),
        open_balance_total=Coalesce(Sum('balance_due'), Value(Decimal('0.00'))),
    )
    customer_summary['new_count'] = max(
        0,
        (customer_summary['customer_count'] or 0) - (customer_summary['with_orders_count'] or 0),
    )
    _active = customer_summary['active_count'] or 0
    _returning = customer_summary['with_orders_count'] or 0
    activity_breakdown = [
        {'label': 'Active (60d)', 'value': _active},
        {'label': 'Quiet', 'value': max(0, _returning - _active)},
        {'label': 'No orders', 'value': customer_summary['new_count']},
    ]

    paginator = Paginator(customers_qs, 18)
    page_obj = paginator.get_page(request.GET.get('page'))
    customers = []
    for customer in page_obj.object_list:
        balance_due = customer.balance_due or Decimal('0.00')
        total_orders = customer.total_orders or 0
        last_order_at = customer.last_order_at
        if balance_due > 0 and show_sensitive:
            customer.activity_label = 'Open balance'
            customer.activity_class = 'warning'
        elif total_orders == 0:
            customer.activity_label = 'New'
            customer.activity_class = 'neutral'
        elif last_order_at and last_order_at >= recent_cutoff:
            customer.activity_label = 'Active'
            customer.activity_class = 'success'
        else:
            customer.activity_label = 'Quiet'
            customer.activity_class = 'neutral'
        customers.append(customer)

    return render(request, 'stock/customer_search.html', {
        'query': query,
        'customers': customers,
        'page_obj': page_obj,
        'customer_summary': customer_summary,
        'activity_breakdown': activity_breakdown,
        'show_sensitive': show_sensitive,
        'show_sales_sensitive': show_sales_sensitive,
    })


def _employee_customer_orders_view(request, customer_id):
    """Employee: a customer's orders only (for reconciliation) - no analytics."""
    customer = get_object_or_404(Customer, id=customer_id)
    active_store, store_is_all = resolve_active_store(request)
    orders_qs = (
        SaleOrder.objects.filter(customer=customer)
        .prefetch_related('items', 'payments')
        .order_by('-created_at', '-id')
    )
    orders_qs = scope_sales_by_store(orders_qs, active_store, store_is_all)
    payment_labels = {'cash': 'Cash', 'card': 'Card', 'mbway': 'MBWay'}
    orders = []
    for order in orders_qs:
        items = list(order.items.all())
        total = sum((i.quantity * (i.unit_price or Decimal('0.00')) for i in items), Decimal('0.00'))
        methods = [payment_labels.get(p.method, (p.method or '').title()) for p in order.payments.all()]
        orders.append({
            'order_id': order.id,
            'created_at': timezone.localtime(order.created_at).strftime('%Y-%m-%d %H:%M'),
            'item_count': sum(i.quantity for i in items),
            'payment_label': ', '.join(dict.fromkeys(methods)) or '-',
            'total_amount': total,
            'items': [
                {
                    'name': build_product_label(i.product),
                    'qty': i.quantity,
                    'unit_price': i.unit_price or Decimal('0.00'),
                    'line_total': (i.unit_price or Decimal('0.00')) * i.quantity,
                }
                for i in items
            ],
        })
    return render(request, 'stock/customer_orders_employee.html', {
        'customer': customer,
        'orders': orders,
    })


@login_required
def customer_detail_view(request, customer_id):
    if not has_manager_access(request.user):
        return _employee_customer_orders_view(request, customer_id)
    customer = get_object_or_404(Customer, id=customer_id)
    show_sensitive = has_manager_access(request.user)
    show_sales_sensitive = has_sales_sensitive_access(request.user)
    show_profit = bool(getattr(request.user, 'is_superuser', False))

    payment_labels = dict(Sale.PAYMENT_METHOD_CHOICES)

    def build_product_label(product):
        return globals()['build_product_label'](product)

    def build_payment_breakdown(source_breakdown):
        return [
            {
                'code': method,
                'label': payment_labels.get(method, (method or 'Other').title()),
                'amount': amount,
            }
            for method, amount in sorted(source_breakdown.items(), key=lambda item: (-item[1], item[0] or ''))
            if method
        ]

    # Date-range scope (presets + custom). Default = all time.
    today = timezone.now().date()
    preset = (request.GET.get('preset') or '').strip()
    if preset == 'month':
        range_key, range_start, range_end = 'month', today.replace(day=1), today
    elif preset == 'year':
        range_key, range_start, range_end = 'year', today.replace(month=1, day=1), today
    elif preset == 'all':
        range_key, range_start, range_end = 'all', None, None
    else:
        range_start = parse_date((request.GET.get('start_date') or '').strip())
        range_end = parse_date((request.GET.get('end_date') or '').strip())
        range_key = 'custom' if (range_start or range_end) else 'all'

    if range_key == 'all':
        range_label = 'All time'
    elif range_key == 'month':
        range_label = today.strftime('%B %Y')
    elif range_key == 'year':
        range_label = str(today.year)
    else:
        range_label = f"{range_start.isoformat() if range_start else '…'} → {range_end.isoformat() if range_end else '…'}"

    active_store, store_is_all = resolve_active_store(request)
    orders_qs = scope_sales_by_store(
        SaleOrder.objects
        .filter(customer=customer)
        .prefetch_related('items__product')
        .order_by('-created_at'),
        active_store, store_is_all,
    )
    if range_start:
        orders_qs = orders_qs.filter(created_at__date__gte=range_start)
    if range_end:
        orders_qs = orders_qs.filter(created_at__date__lte=range_end)

    customer_sale_ids = list(
        Sale.objects.filter(order__in=orders_qs).values_list('id', flat=True)
    )
    sale_profit_map = sale_profit_map_for_sale_ids(customer_sale_ids) if show_profit else {}

    total_spent = Decimal('0.00')
    total_items = 0
    total_orders = 0
    total_profit = Decimal('0.00')
    largest_order_total = Decimal('0.00')
    month_map = defaultdict(lambda: {
        'days': defaultdict(lambda: {
            'orders': [],
            'totals': {'amount': Decimal('0.00'), 'qty': 0, 'profit': Decimal('0.00')},
        }),
        'totals': {'amount': Decimal('0.00'), 'qty': 0, 'profit': Decimal('0.00'), 'orders': 0},
        'payments': defaultdict(Decimal),
    })
    payment_totals = defaultdict(Decimal)
    product_stats = {}
    order_dts = []

    for order in orders_qs:
        items = list(order.items.all())
        pay_break = defaultdict(Decimal)
        order_total = Decimal('0.00')
        order_qty = 0
        order_profit = Decimal('0.00')

        for item in items:
            item.display_name = build_product_label(item.product)
            item.image_url = get_product_image_url(item.product)
            item.total_price = (item.unit_price or Decimal('0.00')) * (item.quantity or 0)
            item.payment_label = payment_labels.get(item.payment_method, (item.payment_method or 'Other').title())
            item.line_profit = sale_profit_map.get(item.id, {}).get('profit', Decimal('0.00')) if show_profit else Decimal('0.00')
            order_total += item.total_price
            order_qty += item.quantity or 0
            order_profit += item.line_profit
            pay_break[item.payment_method] += item.total_price

            stats = product_stats.get(item.product_id)
            if stats is None:
                stats = product_stats[item.product_id] = {
                    'product': item.product,
                    'name': item.display_name,
                    'image_url': item.image_url,
                    'qty': 0,
                    'spend': Decimal('0.00'),
                }
            stats['qty'] += item.quantity or 0
            stats['spend'] += item.total_price

        order.display_total_amount = order_total
        order.display_total_qty = order_qty
        order.display_total_profit = order_profit
        order.display_pay_break = build_payment_breakdown(pay_break)
        order.items_list = items
        order.detail_url = reverse('sale_order_detail', args=[order.id])
        order.detail_id = f'customer-order-{order.id}'

        total_spent += order_total
        total_items += order_qty
        total_orders += 1
        total_profit += order_profit
        order_dts.append(order.created_at)
        if order_total > largest_order_total:
            largest_order_total = order_total

        day = order.created_at.date()
        month_key = day.replace(day=1)
        month_entry = month_map[month_key]
        day_entry = month_entry['days'][day]
        day_entry['orders'].append(order)
        day_entry['totals']['amount'] += order_total
        day_entry['totals']['qty'] += order_qty
        day_entry['totals']['profit'] += order_profit
        month_entry['totals']['amount'] += order_total
        month_entry['totals']['qty'] += order_qty
        month_entry['totals']['profit'] += order_profit
        month_entry['totals']['orders'] += 1
        for method, amount in pay_break.items():
            month_entry['payments'][method] += amount
            payment_totals[method] += amount

    month_blocks = []
    for month_key in sorted(month_map.keys(), reverse=True):
        month_entry = month_map[month_key]
        day_blocks = []
        for day in sorted(month_entry['days'].keys(), reverse=True):
            day_entry = month_entry['days'][day]
            day_entry['orders'].sort(key=lambda item: item.created_at, reverse=True)
            day_blocks.append({
                'date': day,
                'orders': day_entry['orders'],
                'totals': day_entry['totals'],
                'order_count': len(day_entry['orders']),
            })

        month_blocks.append({
            'month': month_key,
            'month_label': month_key.strftime('%B %Y'),
            'day_blocks': day_blocks,
            'totals': month_entry['totals'],
            'payment_breakdown': build_payment_breakdown(month_entry['payments']),
        })

    # Monthly spend trend (oldest -> newest) for the chart.
    spend_trend = [
        {
            'label': month_key.strftime('%b %Y'),
            'amount': float(month_map[month_key]['totals']['amount']),
        }
        for month_key in sorted(month_map.keys())
    ]

    # Customer-level payment-method mix across all orders.
    pay_total_sum = sum(payment_totals.values())
    payment_mix = [
        {
            'label': payment_labels.get(method, (method or 'Other').title()),
            'amount': float(amount),
            'pct': round(float(amount) / float(pay_total_sum) * 100, 1) if pay_total_sum else 0,
        }
        for method, amount in sorted(payment_totals.items(), key=lambda kv: (-kv[1], kv[0] or ''))
        if method
    ]

    # Top products bought by this customer (by spend when allowed, else units).
    stats_list = list(product_stats.values())
    if show_sales_sensitive:
        stats_list.sort(key=lambda s: (s['spend'], s['qty']), reverse=True)
        max_metric = max((s['spend'] for s in stats_list), default=Decimal('0.00'))
    else:
        stats_list.sort(key=lambda s: (s['qty'], s['spend']), reverse=True)
        max_metric = max((s['qty'] for s in stats_list), default=0)
    top_products = []
    for stats in stats_list[:8]:
        metric = stats['spend'] if show_sales_sensitive else stats['qty']
        top_products.append({
            'product': stats['product'],
            'name': stats['name'],
            'image_url': stats['image_url'],
            'qty': stats['qty'],
            'spend': stats['spend'],
            'bar': round(float(metric) / float(max_metric) * 100, 1) if max_metric else 0,
        })

    # Purchase cadence KPIs.
    cadence = {'since': None, 'avg_gap_days': None, 'orders_per_month': None}
    if order_dts:
        first_dt = min(order_dts)
        last_dt = max(order_dts)
        cadence['since'] = first_dt
        if total_orders > 1:
            span_days = (last_dt - first_dt).days
            cadence['avg_gap_days'] = round(span_days / (total_orders - 1), 1)
            span_months = (last_dt.year - first_dt.year) * 12 + (last_dt.month - first_dt.month) + 1
            cadence['orders_per_month'] = round(total_orders / span_months, 1) if span_months else None

    today = timezone.now().date()
    ar_qs = scope_sales_by_store(
        ARInvoice.objects.filter(customer=customer).order_by('-created_at'),
        active_store, store_is_all,
    )

    if ar_qs.exists():
        agg = ar_qs.aggregate(
            total_due=Sum('total_amount'),
            total_paid=Sum('amount_paid'),
        )
        total_due = agg['total_due'] or Decimal('0.00')
        total_paid = agg['total_paid'] or Decimal('0.00')
    else:
        total_due = total_paid = Decimal('0.00')

    ar_totals = {
        'total_due': total_due,
        'total_paid': total_paid,
        'balance': total_due - total_paid,
    }

    ar_list = []
    for inv in ar_qs:
        inv.balance_val = inv.balance
        inv.is_overdue = bool(inv.due_date and inv.balance_val > 0 and inv.due_date < today)
        ar_list.append(inv)

    ar_open = [inv for inv in ar_list if inv.balance_val > 0]
    show_ar = show_sensitive and (ar_totals['balance'] > 0 and len(ar_open) > 0)

    ar_counts = {
        'unpaid': sum(1 for inv in ar_open if inv.status == 'unpaid'),
        'partial': sum(1 for inv in ar_open if inv.status == 'partial'),
        'paid': sum(1 for inv in ar_list if inv.status == 'paid'),
    }
    ar_overdue = {'count': sum(1 for inv in ar_open if inv.is_overdue)}
    ar_next_due = scope_sales_by_store(
        ARInvoice.objects
        .filter(customer=customer, status__in=['unpaid', 'partial'], due_date__isnull=False)
        .order_by('due_date'),
        active_store, store_is_all,
    ).first()
    last_order_at = orders_qs.first().created_at if total_orders else None
    average_order_value = (
        (total_spent / total_orders).quantize(Decimal('0.01'))
        if total_orders else Decimal('0.00')
    )

    return render(request, 'stock/customer_detail.html', {
        'customer': customer,
        'month_blocks': month_blocks,
        'spend_trend': spend_trend,
        'payment_mix': payment_mix,
        'top_products': top_products,
        'cadence': cadence,
        'range_key': range_key,
        'range_label': range_label,
        'range_active': bool(range_start or range_end),
        'start_date': range_start.isoformat() if range_start else '',
        'end_date': range_end.isoformat() if range_end else '',
        'total_spent': total_spent,
        'total_items': total_items,
        'total_orders': total_orders,
        'total_profit': total_profit,
        'largest_order_total': largest_order_total,
        'average_order_value': average_order_value,
        'last_order_at': last_order_at,
        'show_sensitive': show_sensitive,
        'show_sales_sensitive': show_sales_sensitive,
        'show_profit': show_profit,
        'show_ar': show_ar,
        'ar_totals': ar_totals,
        'ar_counts': ar_counts,
        'ar_overdue': ar_overdue,
        'ar_next_due': ar_next_due,
        'ar_open': ar_open,
    })


@login_required
@manager_required
def edit_customer_view(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)

    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            return redirect('customer_detail', customer_id=customer.id)
    else:
        form = CustomerForm(instance=customer)

    return render(request, 'stock/edit_customer.html', {
        'form': form,
        'customer': customer
    })

@login_required
@manager_required
def delete_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    has_sales = Sale.objects.filter(customer=customer).exists()

    if has_sales:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Cannot delete: this customer has sales records.'})
        messages.error(request, 'Cannot delete: this customer has sales records.')
        return redirect('customer_detail', customer_id=pk)

    if request.method == 'POST':
        name = customer.name
        customer.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'Customer "{name}" deleted.'})
        messages.success(request, f'Customer "{name}" deleted.')
        return redirect('customer_search')

@login_required
def products_autocomplete(request):
    q = request.GET.get('q', '').strip()
    product_id = request.GET.get('product_id', '').strip()
    results = []
    matches = Product.objects.none()
    base_qs = (
        Product.objects
        .select_related('brand_master', 'series_master')
        .prefetch_related('images')
    )

    if product_id.isdigit():
        matches = base_qs.filter(id=int(product_id))[:1]
    elif q:
        matches = (
            base_qs
            .filter(
                Q(name__icontains=q) |
                Q(brand__icontains=q) |
                Q(model__icontains=q) |
                Q(spec__icontains=q) |
                Q(color__icontains=q) |
                Q(barcode__icontains=q)
            )
            .order_by('brand', 'model', 'name', 'spec', 'color')[:12]
        )

    for p in matches:
        results.append({
            'id': p.id,
            'name': p.name,
            'display_name': p.display_name,
            'brand': p.brand or '',
            'model': p.model or '',
            'spec': p.spec or '',
            'color': p.color or '',
            'barcode': p.barcode,
            'retail_price': float(p.default_price) if p.default_price is not None else None,
            'wholesale_price': float(p.wholesale_price) if p.wholesale_price is not None else None,
            'image_url': get_product_image_url(p),
        })
    return JsonResponse({'results': results})



# -----------------------------
# 目录页（只展示有图片的产品；品牌/型号分组 + 库存&销量排序）
# -----------------------------

@login_required
@manager_required
def catalog_view(request):
    query = request.GET.get('q', '').strip()
    category_id = request.GET.get('category', '').strip()
    selected_category_name = ''

    qs = (
        annotate_catalog_metrics(
            Product.objects
            .filter(images__isnull=False)
            .distinct()
        )
        .select_related('category', 'brand_master', 'series_master')
        .prefetch_related('images')
    )

    if query:
        qs = qs.filter(
            Q(name__icontains=query) |
            Q(brand__icontains=query) |
            Q(model__icontains=query) |
            Q(spec__icontains=query) |
            Q(color__icontains=query) |
            Q(barcode__icontains=query)
        )

    if category_id:
        qs = qs.filter(category_id=category_id)
        selected_category_name = (
            Category.objects.filter(id=category_id).values_list('name', flat=True).first() or ''
        )
        selected_category_name = customer_catalog_case(selected_category_name)

    hot_ids = set(qs.order_by('-total_sold').values_list('id', flat=True)[:5])

    from collections import defaultdict as _dd
    brand_bucket = _dd(lambda: {
        'count': 0,
        'models': _dd(lambda: {'products': [], 'model_sold': 0}),
        'outro': {'products': [], 'model_sold': 0},
    })

    all_products = []
    total_stock_units = 0
    total_sales_units = 0

    for p in qs:
        stock_val = int(p.total_stock or 0)
        p.stock_val = stock_val
        p.stock_state = 2 if stock_val >= 4 else (1 if stock_val > 0 else 0)
        p.sales_val = int(p.total_sold or 0)
        p.primary_image = p.images.first()
        title_parts = []
        model_part = (p.model or '').strip()
        name_part = (p.name or '').strip()
        if model_part:
            title_parts.append(model_part)
        if name_part and name_part.lower() != model_part.lower():
            title_parts.append(name_part)
        p.catalog_title = customer_catalog_case(' - '.join(title_parts) or name_part or p.display_name)
        p.catalog_spec = customer_catalog_case(p.spec)
        p.catalog_color = customer_catalog_case(p.color)
        p.catalog_category_name = customer_catalog_case(getattr(p.category, 'name', ''))
        if stock_val >= 4:
            p.availability_label = 'Available now'
            p.availability_class = 'in-stock'
        elif stock_val > 0:
            p.availability_label = 'Low stock'
            p.availability_class = 'low-stock'
        else:
            p.availability_label = 'Currently unavailable'
            p.availability_class = 'out-stock'
        all_products.append(p)
        total_stock_units += p.stock_val
        total_sales_units += p.sales_val

        brand = (p.brand or '').strip() or '—'
        model_key = (p.model or '').strip()

        if model_key:
            brand_bucket[brand]['models'][model_key]['products'].append(p)
            brand_bucket[brand]['models'][model_key]['model_sold'] += p.sales_val
        else:
            brand_bucket[brand]['outro']['products'].append(p)
            brand_bucket[brand]['outro']['model_sold'] += p.sales_val

        brand_bucket[brand]['count'] += 1

    def product_sort_key(prod):
        return (-prod.stock_state, -prod.stock_val, -prod.sales_val, (prod.display_name or '').lower())

    brands = []
    for brand_name, data in brand_bucket.items():
        model_blocks = []

        for model_name, mdata in sorted(
            data['models'].items(),
            key=lambda kv: (-kv[1]['model_sold'], kv[0].lower())
        ):
            prods = sorted(mdata['products'], key=product_sort_key)
            model_blocks.append({
                'model': model_name,
                'display_model': customer_catalog_case(model_name),
                'products': prods,
                'has_stock': any(prod.stock_val > 0 for prod in prods),
            })

        if data['outro']['products']:
            outro_sorted = sorted(
                data['outro']['products'],
                key=lambda p: (-p.sales_val, -p.stock_state, -p.stock_val, (p.display_name or '').lower())
            )
            model_blocks.append({
                'model': '',
                'display_model': '',
                'products': outro_sorted,
                'has_stock': any(prod.stock_val > 0 for prod in outro_sorted),
            })

        brands.append({
            'brand': brand_name,
            'display_brand': customer_catalog_case(brand_name),
            'count': data['count'],
            'models': model_blocks,
            'has_stock': any(block['has_stock'] for block in model_blocks),
        })

    brands.sort(key=lambda b: (-b['count'], b['brand'].lower()))
    categories = Category.objects.all()
    hot_products = sorted(
        [product for product in all_products if product.id in hot_ids],
        key=lambda product: (-(1 if product.stock_val > 0 else 0), -product.sales_val, -product.stock_val, (product.display_name or '').lower()),
    )

    return render(request, 'stock/catalog.html', {
        'brands': brands,
        'categories': categories,
        'query': query,
        'selected_category': category_id,
        'selected_category_name': selected_category_name,
        'hot_ids': list(hot_ids),
        'hot_products': hot_products,
        'total_catalog_products': len(all_products),
        'total_catalog_brands': len(brands),
        'total_catalog_stock': total_stock_units,
        'total_catalog_sales': total_sales_units,
        'can_manage': has_manager_access(request.user),
    })

# -----------------------------
# 导出：目录（按品牌分表）到 Excel
# -----------------------------
@login_required
@manager_required
def export_catalog_excel(request):
    q = request.GET.get('q', '').strip()
    category_id = (request.GET.get('category') or '').strip()

    qs = (
        annotate_catalog_metrics(
            Product.objects
            .filter(images__isnull=False)
            .distinct()
        )
        .select_related('category', 'brand_master', 'series_master')
        .prefetch_related('images')
    )

    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(brand__icontains=q) |
            Q(model__icontains=q) |
            Q(spec__icontains=q) |
            Q(color__icontains=q) |
            Q(barcode__icontains=q)
        )
    if category_id:
        qs = qs.filter(category_id=category_id)

    from collections import defaultdict as _dd
    brand_bucket = _dd(lambda: {
        'count': 0,
        'models': _dd(lambda: {'products': [], 'model_sold': 0}),
        'outro': {'products': [], 'model_sold': 0},
    })

    def _stock_state(v: int) -> int:
        return 2 if v > 5 else (1 if v > 0 else 0)

    for p in qs:
        p.stock_val = int(p.total_stock or 0)
        p.stock_state = _stock_state(p.stock_val)
        p.sales_val = int(getattr(p, 'total_sold', 0) or 0)

        brand = (p.brand or '').strip() or '—'
        model_key = (p.model or '').strip()
        if model_key:
            brand_bucket[brand]['models'][model_key]['products'].append(p)
            brand_bucket[brand]['models'][model_key]['model_sold'] += p.sales_val
        else:
            brand_bucket[brand]['outro']['products'].append(p)
            brand_bucket[brand]['outro']['model_sold'] += p.sales_val
        brand_bucket[brand]['count'] += 1

    if not brand_bucket:
        return HttpResponse('No products to export for current filters.', content_type='text/plain')

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_fill = PatternFill('solid', fgColor='FFE2E8F0')
    thin = Side(border_style='thin', color='FFCBD5E1')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    wrap = Alignment(wrap_text=True, vertical='top')

    brands_sorted = sorted(
        [{'brand': b, 'count': data['count']} for b, data in brand_bucket.items()],
        key=lambda x: (-x['count'], x['brand'].lower())
    )

    def _product_sort_key(prod):
        return (-prod.stock_state, -prod.stock_val, -prod.sales_val, (prod.display_name or '').lower())

    for b in brands_sorted:
        brand_name = b['brand']
        data = brand_bucket[brand_name]

        ws = wb.create_sheet(title=(brand_name[:30].replace('/', '-') or 'Sem Marca'))
        ws.freeze_panes = 'A2'
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 60

        headers = ['Foto', 'Nome', 'Preço (€)', 'Stock', 'Descrição']
        ws.append(headers)
        for col in ['A','B','C','D','E']:
            cell = ws[f"{col}1"]
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment(vertical='center')

        model_blocks = []
        for model_name, mdata in data['models'].items():
            prods_sorted = sorted(mdata['products'], key=_product_sort_key)
            model_blocks.append({'model': model_name, 'products': prods_sorted, 'model_sold': mdata['model_sold']})
        model_blocks.sort(key=lambda kv: (-kv['model_sold'], kv['model'].lower()))

        outro_sorted = []
        if data['outro']['products']:
            outro_sorted = sorted(
                data['outro']['products'],
                key=lambda p: (-p.sales_val, -p.stock_state, -p.stock_val, (p.display_name or '').lower())
            )

        row = 2
        def _append(prod):
            nonlocal row
            # Nome
            parts = []
            if (prod.model or '').strip():
                parts.append(prod.model.strip())
            parts.append((prod.name or '').strip())
            disp = ' · '.join([x for x in parts if x]) or (prod.name or '')
            disp = prod.display_name
            ws.cell(row=row, column=2, value=disp).alignment = wrap

            # Preço
            price_val = getattr(prod, 'default_price', None)
            c_price = ws.cell(row=row, column=3, value=float(price_val) if price_val is not None else None)
            if price_val is not None:
                c_price.number_format = '€#,##0.00'

            # Stock 状态
            s_val = int(getattr(prod, 'total_stock', 0) or 0)
            if s_val > 5:
                s_text, s_color, s_font = 'Em stock', 'FFBBF7D0', Font(color='FF065F46', bold=True)
            elif s_val > 0:
                s_text, s_color, s_font = 'Baixo', 'FFFED7AA', Font(color='FF7C2D12', bold=True)
            else:
                s_text, s_color, s_font = 'Esgotado', 'FFFECACA', Font(color='FF7F1D1D', bold=True)
            c_stock = ws.cell(row=row, column=4, value=s_text)
            c_stock.fill = PatternFill('solid', fgColor=s_color)
            c_stock.font = s_font
            c_stock.border = border
            c_stock.alignment = Alignment(horizontal='center', vertical='center')

            # 描述
            desc = (prod.description or '').strip()
            if len(desc) > 240:
                desc = desc[:240] + '…'
            ws.cell(row=row, column=5, value=desc).alignment = wrap

            # 图片
            try:
                first_img = prod.images.all()[0]
                img_path = first_img.image.path
                pil = PILImage.open(img_path)
                pil.thumbnail((120, 120))
                tmp = _BytesIO(); pil.save(tmp, format='PNG'); tmp.seek(0)
                xlimg = XLImage(tmp)
                ws.add_image(xlimg, f'A{row}')
                ws.row_dimensions[row].height = 95
            except Exception:
                pass

            for c in (2,3,4,5):
                ws.cell(row=row, column=c).border = border
            row += 1

        for blk in model_blocks:
            for prod in blk['products']:
                _append(prod)
        for prod in outro_sorted:
            _append(prod)

    buf = _BytesIO(); wb.save(buf); buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    fname = f"Catalog_{ts}.xlsx"
    resp = FileResponse(buf, as_attachment=True, filename=fname)
    resp["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return resp

# -----------------------------
# 欠账
# -----------------------------
@login_required
@manager_required
def ar_new_view(request):
    # 读取 GET 里的预选客户
    pre_id = (request.GET.get('customer') or '').strip()
    pre_customer = None
    if pre_id.isdigit():
        pre_customer = Customer.objects.filter(id=int(pre_id)).only('id', 'name', 'nif').first()

    if request.method == 'POST':
        form = ARInvoiceForm(request.POST)
        if form.is_valid():
            invoice = form.save(commit=False)

            #  兜底：若表单未包含 customer 字段，则用隐藏域 customer 赋值
            if not getattr(invoice, 'customer_id', None):
                post_cust_id = (request.POST.get('customer') or '').strip()
                if post_cust_id.isdigit():
                    invoice.customer_id = int(post_cust_id)

            invoice.store = store_for_new_sale(request)
            invoice.total_amount = Decimal('0.00')
            invoice.save()

            total = Decimal('0.00')
            product_names = request.POST.getlist('product_name')
            quantities = request.POST.getlist('quantity')
            unit_prices = request.POST.getlist('unit_price')

            for name, qty, price in zip(product_names, quantities, unit_prices):
                try:
                    qty = int(qty)
                    price = Decimal(price)
                    ARItem.objects.create(
                        invoice=invoice,
                        product_name=name,
                        quantity=qty,
                        unit_price=price
                    )
                    total += qty * price
                except Exception:
                    continue

            invoice.total_amount = total
            paid = invoice.amount_paid or Decimal('0')
            invoice.status = 'paid' if paid >= total else ('partial' if paid > 0 else 'unpaid')
            invoice.save()

            messages.success(request, f"IOU #{invoice.id} created, total €{total:.2f}.")
            return redirect('ar_detail', invoice_id=invoice.id)
        else:
            messages.error(request, "Please fix the highlighted fields.")
    else:
        # GET 时把预选客户灌入初始值（如果表单包含 customer 字段会生效）
        init = {}
        if pre_customer:
            init['customer'] = pre_customer.id
        form = ARInvoiceForm(initial=init)

    return render(request, 'stock/ar_new.html', {
        'form': form,
        'pre_customer': pre_customer,  # 模板里用来预填隐藏域与搜索框
    })

@login_required
@manager_required
def ar_list_view(request):
    q = request.GET.get('q', '')
    status = request.GET.get('status', '')
    sort = request.GET.get('sort', '-id')
    show_sensitive = has_manager_access(request.user)

    active_store, store_is_all = resolve_active_store(request)
    invoices = scope_sales_by_store(ARInvoice.objects.all(), active_store, store_is_all)
    if q:
        invoices = invoices.filter(customer__name__icontains=q)
    if status:
        invoices = invoices.filter(status=status)

    invoices = invoices.order_by(sort)

    totals = invoices.aggregate(
        total_amount=Sum('total_amount') or 0,
        total_paid=Sum('amount_paid') or 0,
    )
    totals['balance'] = (totals['total_amount'] or 0) - (totals['total_paid'] or 0)

    context = {
        'q': q,
        'status': status,
        'sort': sort,
        'invoices': invoices,
        'totals': totals,
        'show_sensitive': show_sensitive,
    }
    return render(request, 'stock/ar_list.html', context)

@login_required
@manager_required
def ar_detail_view(request, invoice_id):
    invoice = get_object_or_404(ARInvoice.objects.select_related('customer'), id=invoice_id)
    items = invoice.items.all().order_by('id')
    pay_form = ARPaymentForm()
    return render(request, 'stock/ar_detail.html', {
        'invoice': invoice,
        'items': items,
        'pay_form': pay_form,
        'show_sensitive': has_manager_access(request.user),
    })


@require_POST
@login_required
@manager_required
def ar_add_payment_view(request, invoice_id):
    invoice = get_object_or_404(ARInvoice, id=invoice_id)
    form = ARPaymentForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Invalid payment.")
        return redirect('ar_detail', invoice_id=invoice.id)

    payment = form.save(commit=False)
    payment.invoice = invoice
    payment.save()

    # 汇总已付并更新状态
    total_paid = invoice.payments.aggregate(s=Sum('amount'))['s'] or Decimal('0.00')
    invoice.amount_paid = total_paid
    if total_paid >= invoice.total_amount:
        invoice.status = 'paid'
    elif total_paid > 0:
        invoice.status = 'partial'
    else:
        invoice.status = 'unpaid'
    invoice.save()

    messages.success(request, f"Payment €{payment.amount:.2f} recorded.")
    return redirect('ar_detail', invoice_id=invoice.id)

@require_POST
@login_required
@manager_required
def ar_add_items_view(request, invoice_id):
    invoice = get_object_or_404(ARInvoice, id=invoice_id)

    names = request.POST.getlist('product_name')
    quantities = request.POST.getlist('quantity')
    prices = request.POST.getlist('unit_price')

    total_added = Decimal('0.00')

    for name, qty, price in zip(names, quantities, prices):
        if not name.strip():
            continue
        try:
            qty = int(qty)
            price = Decimal(price)
        except:
            continue
        ARItem.objects.create(
            invoice=invoice,
            product_name=name.strip(),
            quantity=qty,
            unit_price=price
        )
        total_added += qty * price

    # 更新总金额
    invoice.total_amount += total_added
    paid = invoice.amount_paid or Decimal('0')
    invoice.status = 'paid' if paid >= invoice.total_amount else ('partial' if paid > 0 else 'unpaid')
    invoice.save()

    messages.success(request, f"{len(names)} item(s) added, +€{total_added:.2f}.")
    return redirect('ar_detail', invoice_id=invoice.id)

# API: 调整单个购买批次的 remaining
@login_required
@require_POST
def api_adjust_purchase_stock(request):
    """
    POST JSON:
    {
      "purchase_id": 456,
      "new_remaining": 50
    }
    """
    if not has_manager_access(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8') if isinstance(request.body, bytes) else request.body)
        purchase_id = int(data.get('purchase_id', 0))
        new_remaining = int(data.get('new_remaining', 0))

        if new_remaining < 0:
            return JsonResponse({'success': False, 'error': 'Stock cannot be negative'})

        purchase = get_object_or_404(Purchase.objects.select_related('product'), pk=purchase_id)
        if new_remaining > purchase.quantity:
            return JsonResponse({'success': False, 'error': 'Remaining stock cannot exceed original quantity'})

        with transaction.atomic():
            old_remaining = purchase.remaining
            # 条件更新：仅当 remaining 未被并发修改时才写入（SQLite 上 select_for_update 是 no-op）
            updated = Purchase.objects.filter(pk=purchase.pk, remaining=old_remaining).update(
                remaining=new_remaining
            )
            if not updated:
                return JsonResponse({'success': False, 'error': 'Stock changed concurrently, please retry.'})

            StockAdjustmentLog.objects.create(
                user=request.user,
                product=purchase.product,
                purchase=purchase,
                adjustment_type='purchase_remaining',
                old_value=old_remaining,
                new_value=new_remaining,
            )
            inventory_snapshot = build_inventory_snapshot(purchase.product)

            from .services.pricing import sync_perfume_price
            sync_perfume_price(purchase.product)

        return JsonResponse({
            'success': True,
            'message': f'Updated remaining from {old_remaining} to {new_remaining}',
            'old_remaining': old_remaining,
            'new_remaining': new_remaining,
            'inventory_snapshot': inventory_snapshot,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# API: 调整产品总库存（按 FIFO 创建/移除批次）
@login_required
@require_POST
def api_adjust_total_stock(request):
    """
    POST JSON:
    {
      "product_id": 271,
      "new_total_stock": 100
    }
    """
    if not has_manager_access(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8') if isinstance(request.body, bytes) else request.body)
        product_id = int(data.get('product_id', 0))
        new_total = int(data.get('new_total_stock', 0))

        if new_total < 0:
            return JsonResponse({'success': False, 'error': 'Stock cannot be negative'})

        product = get_object_or_404(Product, pk=product_id)

        with transaction.atomic():
            current_total = product.total_stock()
            difference = new_total - current_total

            if difference == 0:
                return JsonResponse({
                    'success': True,
                    'message': 'No change needed',
                    'inventory_snapshot': build_inventory_snapshot(product),
                })

            if difference > 0:
                Purchase.objects.create(
                    product=product,
                    supplier=None,
                    quantity=difference,
                    cost_price=Decimal('0.00'),
                    remaining=difference,
                    inbound_order=None
                )
                message = f'Increased stock by {difference}'
            else:
                # 复用 FIFO 原子条件扣减：SQLite 上 select_for_update 是 no-op，
                # consume_stock_fifo 用条件 UPDATE 保证并发安全
                consume_stock_fifo(product, abs(difference))
                message = f'Decreased stock by {abs(difference)}'

            from .services.pricing import sync_perfume_price
            sync_perfume_price(product)

            StockAdjustmentLog.objects.create(
                user=request.user,
                product=product,
                purchase=None,
                adjustment_type='total_stock',
                old_value=current_total,
                new_value=new_total,
            )
            inventory_snapshot = build_inventory_snapshot(product)

        return JsonResponse({
            'success': True,
            'message': message,
            'inventory_snapshot': inventory_snapshot,
        })
    except ValidationError as e:
        return JsonResponse({'success': False, 'error': '; '.join(e.messages)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
